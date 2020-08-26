#  encoding: utf-8

import tushare as ts
import os
import pandas as pd  # https://www.jianshu.com/p/5c0aa1fa19af
import openpyxl 
from openpyxl import load_workbook,Workbook
import time,datetime  # https://www.jb51.net/article/66019.htm
import re
import os
import tempfile

############################## 全局数据初始化块 数据初始化函数  End ##############################

##############  properities 函数 Begin ##############

class Properties:

    def __init__(self, file_name):
        self.file_name = file_name
        self.properties = {}
        try:
            fopen = open(self.file_name, 'r')
            for line in fopen:
                line = line.strip()
                if line.find('=') > 0 and not line.startswith('#'):
                    strs = line.split('=')
                    self.properties[strs[0].strip()] = strs[1].strip()
        except Exception :
            raise
        else:
            fopen.close()

    def has_key(self, key):
        return key in self.properties

    def get(self, key, default_value=''):
        if key in self.properties:
            return self.properties[key]
        return default_value

    def put(self, key, value):
        self.properties[key] = value
        replace_property(self.file_name, key + '=.*', key + '=' + value, True)


def replace_property(file_name, from_regex, to_str, append_on_not_exists=True):
    tmpfile = tempfile.TemporaryFile()

    if os.path.exists(file_name):
        r_open = open(file_name, 'r')
        pattern = re.compile(r'' + from_regex)
        found = None
        for line in r_open:
            if pattern.search(line) and not line.strip().startswith('#'):
                found = True
                line = re.sub(from_regex, to_str, line)
            tmpfile.write(line.encode())
        if not found and append_on_not_exists:
            tmpfile.write(('\n' + to_str).encode())
        r_open.close()
        tmpfile.seek(0)

        content = tmpfile.read()

        if os.path.exists(file_name):
            os.remove(file_name)

        w_open = open(file_name, 'wb')
        w_open.write(content)
        w_open.close()

        tmpfile.close()
    else:
        print ("file %s not found" % file_name)

##############  properities 函数 End ##############


##############  XLSX excel 函数 Begin ##############

def createexcel(filename):    ### 创建本地文件名称为 filename的文件
    wb = Workbook()
    curPath = os.path.abspath('.')
    cur_desktop_path=os.path.join(os.path.expanduser("~"), 'Desktop')
    curDir = cur_desktop_path+os.sep+"zbin"+os.sep+"J0"+os.sep  ## 创建~/Desktop/zbin/J0 这个 这个工作目录
    #curDir = curPath+os.sep
    if not os.path.exists(curDir):
        os.makedirs(curDir)
    #curPath = os.path.dirname(os.path.abspath('.'))
    #t = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
    #suffix = '.xlsx' # 文件类型
    #newfile = t + suffix
    xlsxPath = curDir + filename
    print("curPath curDir=" + curPath)
    print("createexcel zbin_JO_Dir=" + curDir)
    print("createexcel path=" + xlsxPath)
    if not os.path.exists(xlsxPath):
        wb.save(xlsxPath)
        print(" wb.save(xlsxPath)  xlsxPath =" + xlsxPath)
        time.sleep(1)
    return xlsxPath

def getColumnIndex(table, columnName):   ## 返回 table 中 名称为  columnName 的 那列 的索引
    columnIndex = None
    for i in range(table.ncols):
        if(table.cell_value(0, i) == columnName):
            columnIndex = i
            break
    return columnIndex
#封装函数    https://blog.csdn.net/weixin_41267342/article/details/86634007

##############  XLSX excel 函数 End ##############



############## tscode_股票列表的初始化  Begin ##############
#封装函数    https://blog.csdn.net/weixin_41267342/article/details/86634007
def init_tscode_data(book_name, sheet_name,ts_code_set,tscode_name_dict):
    # 读取excel
    wb = openpyxl.load_workbook(book_name)
    # 读取sheet表
    ws = wb[sheet_name]
    for i in range(ws.max_row):
         # 获取下拉框中的所有选择值
         if (i == 0 or i == 1):
             continue
         #print("i="+str(i)+" 总的列数:" + str(ws.max_row)+"  value:"+str(ws.cell(i,2).value))
         tscode_item = str(ws.cell(i,2).value)
         tscode_name_item = str(ws.cell(i,4).value)
         #print("index="+str(i)+" tscode_item = "+ str(tscode_item) + "   tscode_name_item="+str(tscode_name_item))
         ts_code_set.add(str(ws.cell(i,2).value))
         tscode_name_dict[tscode_item]=tscode_name_item



tscode_set = set()    #### 股票代码的集合   000001.SZ   .... 999999.SH
tscode_name_dict = dict()  #### code-name 的 map的集合
init_tscode_data("股票列表.xlsx","股票列表",tscode_set,tscode_name_dict)
############## tscode_股票列表的初始化  End  ##############


############################## 运行属性 Begin ##############################
pd.set_option('display.max_rows', None)   ##  解决纵向出现...
#pd.set_option('display.width', 1000) 
pd.set_option('expand_frame_repr', False)  ##  解决横向出现...
Cur_Abs_Path=os.path.abspath('.')   # 表示当前所处的文件夹的绝对路径
print("当前绝对路径:"+Cur_Abs_Path)
Cur_Ref_Path=os.path.abspath('..')  # 表示当前所处的文件夹上一级文件夹的绝对路径
print("当前父目录绝对路径:"+Cur_Ref_Path)
############################## 运行属性 End ##############################

############################## 时间Date Begin ##############################

now_yyyymmdd=str(time.strftime('%Y%m%d', time.localtime(time.time())))
print("now_yyyymmdd = "+ str(now_yyyymmdd))

############################## 时间Date End ##############################


############################## Prop初始化 Begin ##############################

#
#Thu Aug 13 22:33:45 CST 2020
#rixianhangqing-time_record_date=20200707
#rixianhangqing-time_start_date=20010101

desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')
zbin_path = str(desktop_path)+os.sep+"zbin"
j0_properties_path= str(zbin_path)+os.sep+"J0.properties"
print("desktop_path = "+ str(desktop_path))
print("zbin_path = "+ str(zbin_path))
print("j0_properties_path = "+ str(j0_properties_path))
J0_PROPS =  Properties(j0_properties_path)


############################## Prop初始化 End ##############################


pro = ts.pro_api('43acb9a5ddc2cf73c6c4ea54796748f965457ed57daaa736bb778ea2')
# print(J0_PROPS.get(tree_node_name+"record_date"))           #根据key读取value
# J0_PROPS.put(tree_node_name+"record_date", now_yyyymmdd)       ###  覆盖原有的 key 和 value
#  zukgit
# guanlicengxinchouhechigu_zukgit_website  =   https://tushare.pro/document/2?doc_id=194
createexcel('管理层薪酬和持股.xlsx')
stk_rewards_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0\\管理层薪酬和持股.xlsx')
stk_rewards_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0\\管理层薪酬和持股.xlsx', engine='openpyxl')
stk_rewards_excel_writer.book = stk_rewards_book
stk_rewards_excel_writer.sheets = dict((ws.title, ws) for ws in stk_rewards_book.worksheets)
stk_rewards_data_List = pd.DataFrame()
stk_rewards_item_0_tscode_list = list() 
stk_rewards_item_0 = pro.stk_rewards(ts_code='000001.SZ,000002.SZ,000003.SZ,000004.SZ,000005.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_0 返回数据 row 行数 = "+str(stk_rewards_item_0.shape[0]))
for ts_code_sh in stk_rewards_item_0['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_0_tscode_list.append(stock_name)
stk_rewards_item_0_addname_dataframe=pd.DataFrame()
stk_rewards_item_0_addname_dataframe['cname'] = stk_rewards_item_0_tscode_list
for table_name in stk_rewards_item_0.columns.values.tolist():
    stk_rewards_item_0_addname_dataframe[table_name] = stk_rewards_item_0[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_0_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_1_tscode_list = list() 
stk_rewards_item_1 = pro.stk_rewards(ts_code='000006.SZ,000007.SZ,000008.SZ,000009.SZ,000010.SZ,000011.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_1 返回数据 row 行数 = "+str(stk_rewards_item_1.shape[0]))
for ts_code_sh in stk_rewards_item_1['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_1_tscode_list.append(stock_name)
stk_rewards_item_1_addname_dataframe=pd.DataFrame()
stk_rewards_item_1_addname_dataframe['cname'] = stk_rewards_item_1_tscode_list
for table_name in stk_rewards_item_1.columns.values.tolist():
    stk_rewards_item_1_addname_dataframe[table_name] = stk_rewards_item_1[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_1_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_2_tscode_list = list() 
stk_rewards_item_2 = pro.stk_rewards(ts_code='000012.SZ,000013.SZ,000014.SZ,000015.SZ,000016.SZ,000017.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_2 返回数据 row 行数 = "+str(stk_rewards_item_2.shape[0]))
for ts_code_sh in stk_rewards_item_2['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_2_tscode_list.append(stock_name)
stk_rewards_item_2_addname_dataframe=pd.DataFrame()
stk_rewards_item_2_addname_dataframe['cname'] = stk_rewards_item_2_tscode_list
for table_name in stk_rewards_item_2.columns.values.tolist():
    stk_rewards_item_2_addname_dataframe[table_name] = stk_rewards_item_2[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_2_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_3_tscode_list = list() 
stk_rewards_item_3 = pro.stk_rewards(ts_code='000018.SZ,000019.SZ,000020.SZ,000021.SZ,000023.SZ,000024.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_3 返回数据 row 行数 = "+str(stk_rewards_item_3.shape[0]))
for ts_code_sh in stk_rewards_item_3['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_3_tscode_list.append(stock_name)
stk_rewards_item_3_addname_dataframe=pd.DataFrame()
stk_rewards_item_3_addname_dataframe['cname'] = stk_rewards_item_3_tscode_list
for table_name in stk_rewards_item_3.columns.values.tolist():
    stk_rewards_item_3_addname_dataframe[table_name] = stk_rewards_item_3[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_3_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_4_tscode_list = list() 
stk_rewards_item_4 = pro.stk_rewards(ts_code='000025.SZ,000026.SZ,000027.SZ,000028.SZ,000029.SZ,000030.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_4 返回数据 row 行数 = "+str(stk_rewards_item_4.shape[0]))
for ts_code_sh in stk_rewards_item_4['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_4_tscode_list.append(stock_name)
stk_rewards_item_4_addname_dataframe=pd.DataFrame()
stk_rewards_item_4_addname_dataframe['cname'] = stk_rewards_item_4_tscode_list
for table_name in stk_rewards_item_4.columns.values.tolist():
    stk_rewards_item_4_addname_dataframe[table_name] = stk_rewards_item_4[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_4_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_5_tscode_list = list() 
stk_rewards_item_5 = pro.stk_rewards(ts_code='000031.SZ,000032.SZ,000033.SZ,000034.SZ,000035.SZ,000036.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_5 返回数据 row 行数 = "+str(stk_rewards_item_5.shape[0]))
for ts_code_sh in stk_rewards_item_5['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_5_tscode_list.append(stock_name)
stk_rewards_item_5_addname_dataframe=pd.DataFrame()
stk_rewards_item_5_addname_dataframe['cname'] = stk_rewards_item_5_tscode_list
for table_name in stk_rewards_item_5.columns.values.tolist():
    stk_rewards_item_5_addname_dataframe[table_name] = stk_rewards_item_5[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_5_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_6_tscode_list = list() 
stk_rewards_item_6 = pro.stk_rewards(ts_code='000037.SZ,000038.SZ,000039.SZ,000040.SZ,000042.SZ,000045.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_6 返回数据 row 行数 = "+str(stk_rewards_item_6.shape[0]))
for ts_code_sh in stk_rewards_item_6['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_6_tscode_list.append(stock_name)
stk_rewards_item_6_addname_dataframe=pd.DataFrame()
stk_rewards_item_6_addname_dataframe['cname'] = stk_rewards_item_6_tscode_list
for table_name in stk_rewards_item_6.columns.values.tolist():
    stk_rewards_item_6_addname_dataframe[table_name] = stk_rewards_item_6[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_6_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_7_tscode_list = list() 
stk_rewards_item_7 = pro.stk_rewards(ts_code='000046.SZ,000047.SZ,000048.SZ,000049.SZ,000050.SZ,000055.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_7 返回数据 row 行数 = "+str(stk_rewards_item_7.shape[0]))
for ts_code_sh in stk_rewards_item_7['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_7_tscode_list.append(stock_name)
stk_rewards_item_7_addname_dataframe=pd.DataFrame()
stk_rewards_item_7_addname_dataframe['cname'] = stk_rewards_item_7_tscode_list
for table_name in stk_rewards_item_7.columns.values.tolist():
    stk_rewards_item_7_addname_dataframe[table_name] = stk_rewards_item_7[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_7_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_8_tscode_list = list() 
stk_rewards_item_8 = pro.stk_rewards(ts_code='000056.SZ,000058.SZ,000059.SZ,000060.SZ,000061.SZ,000062.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_8 返回数据 row 行数 = "+str(stk_rewards_item_8.shape[0]))
for ts_code_sh in stk_rewards_item_8['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_8_tscode_list.append(stock_name)
stk_rewards_item_8_addname_dataframe=pd.DataFrame()
stk_rewards_item_8_addname_dataframe['cname'] = stk_rewards_item_8_tscode_list
for table_name in stk_rewards_item_8.columns.values.tolist():
    stk_rewards_item_8_addname_dataframe[table_name] = stk_rewards_item_8[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_8_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_9_tscode_list = list() 
stk_rewards_item_9 = pro.stk_rewards(ts_code='000063.SZ,000065.SZ,000066.SZ,000068.SZ,000069.SZ,000070.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_9 返回数据 row 行数 = "+str(stk_rewards_item_9.shape[0]))
for ts_code_sh in stk_rewards_item_9['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_9_tscode_list.append(stock_name)
stk_rewards_item_9_addname_dataframe=pd.DataFrame()
stk_rewards_item_9_addname_dataframe['cname'] = stk_rewards_item_9_tscode_list
for table_name in stk_rewards_item_9.columns.values.tolist():
    stk_rewards_item_9_addname_dataframe[table_name] = stk_rewards_item_9[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_9_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_10_tscode_list = list() 
stk_rewards_item_10 = pro.stk_rewards(ts_code='000078.SZ,000088.SZ,000089.SZ,000090.SZ,000096.SZ,000099.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_10 返回数据 row 行数 = "+str(stk_rewards_item_10.shape[0]))
for ts_code_sh in stk_rewards_item_10['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_10_tscode_list.append(stock_name)
stk_rewards_item_10_addname_dataframe=pd.DataFrame()
stk_rewards_item_10_addname_dataframe['cname'] = stk_rewards_item_10_tscode_list
for table_name in stk_rewards_item_10.columns.values.tolist():
    stk_rewards_item_10_addname_dataframe[table_name] = stk_rewards_item_10[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_10_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_11_tscode_list = list() 
stk_rewards_item_11 = pro.stk_rewards(ts_code='000100.SZ,000150.SZ,000151.SZ,000153.SZ,000155.SZ,000156.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_11 返回数据 row 行数 = "+str(stk_rewards_item_11.shape[0]))
for ts_code_sh in stk_rewards_item_11['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_11_tscode_list.append(stock_name)
stk_rewards_item_11_addname_dataframe=pd.DataFrame()
stk_rewards_item_11_addname_dataframe['cname'] = stk_rewards_item_11_tscode_list
for table_name in stk_rewards_item_11.columns.values.tolist():
    stk_rewards_item_11_addname_dataframe[table_name] = stk_rewards_item_11[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_11_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_12_tscode_list = list() 
stk_rewards_item_12 = pro.stk_rewards(ts_code='000157.SZ,000158.SZ,000159.SZ,000166.SZ,000301.SZ,000333.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_12 返回数据 row 行数 = "+str(stk_rewards_item_12.shape[0]))
for ts_code_sh in stk_rewards_item_12['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_12_tscode_list.append(stock_name)
stk_rewards_item_12_addname_dataframe=pd.DataFrame()
stk_rewards_item_12_addname_dataframe['cname'] = stk_rewards_item_12_tscode_list
for table_name in stk_rewards_item_12.columns.values.tolist():
    stk_rewards_item_12_addname_dataframe[table_name] = stk_rewards_item_12[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_12_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_13_tscode_list = list() 
stk_rewards_item_13 = pro.stk_rewards(ts_code='000338.SZ,000400.SZ,000401.SZ,000402.SZ,000403.SZ,000404.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_13 返回数据 row 行数 = "+str(stk_rewards_item_13.shape[0]))
for ts_code_sh in stk_rewards_item_13['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_13_tscode_list.append(stock_name)
stk_rewards_item_13_addname_dataframe=pd.DataFrame()
stk_rewards_item_13_addname_dataframe['cname'] = stk_rewards_item_13_tscode_list
for table_name in stk_rewards_item_13.columns.values.tolist():
    stk_rewards_item_13_addname_dataframe[table_name] = stk_rewards_item_13[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_13_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_14_tscode_list = list() 
stk_rewards_item_14 = pro.stk_rewards(ts_code='000405.SZ,000406.SZ,000407.SZ,000408.SZ,000409.SZ,000410.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_14 返回数据 row 行数 = "+str(stk_rewards_item_14.shape[0]))
for ts_code_sh in stk_rewards_item_14['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_14_tscode_list.append(stock_name)
stk_rewards_item_14_addname_dataframe=pd.DataFrame()
stk_rewards_item_14_addname_dataframe['cname'] = stk_rewards_item_14_tscode_list
for table_name in stk_rewards_item_14.columns.values.tolist():
    stk_rewards_item_14_addname_dataframe[table_name] = stk_rewards_item_14[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_14_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_15_tscode_list = list() 
stk_rewards_item_15 = pro.stk_rewards(ts_code='000411.SZ,000412.SZ,000413.SZ,000415.SZ,000416.SZ,000417.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_15 返回数据 row 行数 = "+str(stk_rewards_item_15.shape[0]))
for ts_code_sh in stk_rewards_item_15['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_15_tscode_list.append(stock_name)
stk_rewards_item_15_addname_dataframe=pd.DataFrame()
stk_rewards_item_15_addname_dataframe['cname'] = stk_rewards_item_15_tscode_list
for table_name in stk_rewards_item_15.columns.values.tolist():
    stk_rewards_item_15_addname_dataframe[table_name] = stk_rewards_item_15[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_15_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_16_tscode_list = list() 
stk_rewards_item_16 = pro.stk_rewards(ts_code='000418.SZ,000419.SZ,000420.SZ,000421.SZ,000422.SZ,000423.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_16 返回数据 row 行数 = "+str(stk_rewards_item_16.shape[0]))
for ts_code_sh in stk_rewards_item_16['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_16_tscode_list.append(stock_name)
stk_rewards_item_16_addname_dataframe=pd.DataFrame()
stk_rewards_item_16_addname_dataframe['cname'] = stk_rewards_item_16_tscode_list
for table_name in stk_rewards_item_16.columns.values.tolist():
    stk_rewards_item_16_addname_dataframe[table_name] = stk_rewards_item_16[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_16_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_17_tscode_list = list() 
stk_rewards_item_17 = pro.stk_rewards(ts_code='000425.SZ,000426.SZ,000428.SZ,000429.SZ,000430.SZ,000488.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_17 返回数据 row 行数 = "+str(stk_rewards_item_17.shape[0]))
for ts_code_sh in stk_rewards_item_17['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_17_tscode_list.append(stock_name)
stk_rewards_item_17_addname_dataframe=pd.DataFrame()
stk_rewards_item_17_addname_dataframe['cname'] = stk_rewards_item_17_tscode_list
for table_name in stk_rewards_item_17.columns.values.tolist():
    stk_rewards_item_17_addname_dataframe[table_name] = stk_rewards_item_17[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_17_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_18_tscode_list = list() 
stk_rewards_item_18 = pro.stk_rewards(ts_code='000498.SZ,000501.SZ,000502.SZ,000503.SZ,000504.SZ,000505.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_18 返回数据 row 行数 = "+str(stk_rewards_item_18.shape[0]))
for ts_code_sh in stk_rewards_item_18['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_18_tscode_list.append(stock_name)
stk_rewards_item_18_addname_dataframe=pd.DataFrame()
stk_rewards_item_18_addname_dataframe['cname'] = stk_rewards_item_18_tscode_list
for table_name in stk_rewards_item_18.columns.values.tolist():
    stk_rewards_item_18_addname_dataframe[table_name] = stk_rewards_item_18[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_18_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_19_tscode_list = list() 
stk_rewards_item_19 = pro.stk_rewards(ts_code='000506.SZ,000507.SZ,000508.SZ,000509.SZ,000510.SZ,000511.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_19 返回数据 row 行数 = "+str(stk_rewards_item_19.shape[0]))
for ts_code_sh in stk_rewards_item_19['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_19_tscode_list.append(stock_name)
stk_rewards_item_19_addname_dataframe=pd.DataFrame()
stk_rewards_item_19_addname_dataframe['cname'] = stk_rewards_item_19_tscode_list
for table_name in stk_rewards_item_19.columns.values.tolist():
    stk_rewards_item_19_addname_dataframe[table_name] = stk_rewards_item_19[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_19_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_20_tscode_list = list() 
stk_rewards_item_20 = pro.stk_rewards(ts_code='000513.SZ,000514.SZ,000515.SZ,000516.SZ,000517.SZ,000518.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_20 返回数据 row 行数 = "+str(stk_rewards_item_20.shape[0]))
for ts_code_sh in stk_rewards_item_20['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_20_tscode_list.append(stock_name)
stk_rewards_item_20_addname_dataframe=pd.DataFrame()
stk_rewards_item_20_addname_dataframe['cname'] = stk_rewards_item_20_tscode_list
for table_name in stk_rewards_item_20.columns.values.tolist():
    stk_rewards_item_20_addname_dataframe[table_name] = stk_rewards_item_20[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_20_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_21_tscode_list = list() 
stk_rewards_item_21 = pro.stk_rewards(ts_code='000519.SZ,000520.SZ,000521.SZ,000522.SZ,000523.SZ,000524.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_21 返回数据 row 行数 = "+str(stk_rewards_item_21.shape[0]))
for ts_code_sh in stk_rewards_item_21['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_21_tscode_list.append(stock_name)
stk_rewards_item_21_addname_dataframe=pd.DataFrame()
stk_rewards_item_21_addname_dataframe['cname'] = stk_rewards_item_21_tscode_list
for table_name in stk_rewards_item_21.columns.values.tolist():
    stk_rewards_item_21_addname_dataframe[table_name] = stk_rewards_item_21[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_21_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_22_tscode_list = list() 
stk_rewards_item_22 = pro.stk_rewards(ts_code='000525.SZ,000526.SZ,000527.SZ,000528.SZ,000529.SZ,000530.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_22 返回数据 row 行数 = "+str(stk_rewards_item_22.shape[0]))
for ts_code_sh in stk_rewards_item_22['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_22_tscode_list.append(stock_name)
stk_rewards_item_22_addname_dataframe=pd.DataFrame()
stk_rewards_item_22_addname_dataframe['cname'] = stk_rewards_item_22_tscode_list
for table_name in stk_rewards_item_22.columns.values.tolist():
    stk_rewards_item_22_addname_dataframe[table_name] = stk_rewards_item_22[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_22_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_23_tscode_list = list() 
stk_rewards_item_23 = pro.stk_rewards(ts_code='000531.SZ,000532.SZ,000533.SZ,000534.SZ,000535.SZ,000536.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_23 返回数据 row 行数 = "+str(stk_rewards_item_23.shape[0]))
for ts_code_sh in stk_rewards_item_23['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_23_tscode_list.append(stock_name)
stk_rewards_item_23_addname_dataframe=pd.DataFrame()
stk_rewards_item_23_addname_dataframe['cname'] = stk_rewards_item_23_tscode_list
for table_name in stk_rewards_item_23.columns.values.tolist():
    stk_rewards_item_23_addname_dataframe[table_name] = stk_rewards_item_23[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_23_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_24_tscode_list = list() 
stk_rewards_item_24 = pro.stk_rewards(ts_code='000537.SZ,000538.SZ,000539.SZ,000540.SZ,000541.SZ,000542.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_24 返回数据 row 行数 = "+str(stk_rewards_item_24.shape[0]))
for ts_code_sh in stk_rewards_item_24['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_24_tscode_list.append(stock_name)
stk_rewards_item_24_addname_dataframe=pd.DataFrame()
stk_rewards_item_24_addname_dataframe['cname'] = stk_rewards_item_24_tscode_list
for table_name in stk_rewards_item_24.columns.values.tolist():
    stk_rewards_item_24_addname_dataframe[table_name] = stk_rewards_item_24[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_24_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_25_tscode_list = list() 
stk_rewards_item_25 = pro.stk_rewards(ts_code='000543.SZ,000544.SZ,000545.SZ,000546.SZ,000547.SZ,000548.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_25 返回数据 row 行数 = "+str(stk_rewards_item_25.shape[0]))
for ts_code_sh in stk_rewards_item_25['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_25_tscode_list.append(stock_name)
stk_rewards_item_25_addname_dataframe=pd.DataFrame()
stk_rewards_item_25_addname_dataframe['cname'] = stk_rewards_item_25_tscode_list
for table_name in stk_rewards_item_25.columns.values.tolist():
    stk_rewards_item_25_addname_dataframe[table_name] = stk_rewards_item_25[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_25_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_26_tscode_list = list() 
stk_rewards_item_26 = pro.stk_rewards(ts_code='000549.SZ,000550.SZ,000551.SZ,000552.SZ,000553.SZ,000554.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_26 返回数据 row 行数 = "+str(stk_rewards_item_26.shape[0]))
for ts_code_sh in stk_rewards_item_26['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_26_tscode_list.append(stock_name)
stk_rewards_item_26_addname_dataframe=pd.DataFrame()
stk_rewards_item_26_addname_dataframe['cname'] = stk_rewards_item_26_tscode_list
for table_name in stk_rewards_item_26.columns.values.tolist():
    stk_rewards_item_26_addname_dataframe[table_name] = stk_rewards_item_26[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_26_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_27_tscode_list = list() 
stk_rewards_item_27 = pro.stk_rewards(ts_code='000555.SZ,000556.SZ,000557.SZ,000558.SZ,000559.SZ,000560.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_27 返回数据 row 行数 = "+str(stk_rewards_item_27.shape[0]))
for ts_code_sh in stk_rewards_item_27['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_27_tscode_list.append(stock_name)
stk_rewards_item_27_addname_dataframe=pd.DataFrame()
stk_rewards_item_27_addname_dataframe['cname'] = stk_rewards_item_27_tscode_list
for table_name in stk_rewards_item_27.columns.values.tolist():
    stk_rewards_item_27_addname_dataframe[table_name] = stk_rewards_item_27[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_27_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_28_tscode_list = list() 
stk_rewards_item_28 = pro.stk_rewards(ts_code='000561.SZ,000562.SZ,000563.SZ,000564.SZ,000565.SZ,000566.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_28 返回数据 row 行数 = "+str(stk_rewards_item_28.shape[0]))
for ts_code_sh in stk_rewards_item_28['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_28_tscode_list.append(stock_name)
stk_rewards_item_28_addname_dataframe=pd.DataFrame()
stk_rewards_item_28_addname_dataframe['cname'] = stk_rewards_item_28_tscode_list
for table_name in stk_rewards_item_28.columns.values.tolist():
    stk_rewards_item_28_addname_dataframe[table_name] = stk_rewards_item_28[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_28_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_29_tscode_list = list() 
stk_rewards_item_29 = pro.stk_rewards(ts_code='000567.SZ,000568.SZ,000569.SZ,000570.SZ,000571.SZ,000572.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_29 返回数据 row 行数 = "+str(stk_rewards_item_29.shape[0]))
for ts_code_sh in stk_rewards_item_29['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_29_tscode_list.append(stock_name)
stk_rewards_item_29_addname_dataframe=pd.DataFrame()
stk_rewards_item_29_addname_dataframe['cname'] = stk_rewards_item_29_tscode_list
for table_name in stk_rewards_item_29.columns.values.tolist():
    stk_rewards_item_29_addname_dataframe[table_name] = stk_rewards_item_29[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_29_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_30_tscode_list = list() 
stk_rewards_item_30 = pro.stk_rewards(ts_code='000573.SZ,000576.SZ,000578.SZ,000581.SZ,000582.SZ,000583.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_30 返回数据 row 行数 = "+str(stk_rewards_item_30.shape[0]))
for ts_code_sh in stk_rewards_item_30['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_30_tscode_list.append(stock_name)
stk_rewards_item_30_addname_dataframe=pd.DataFrame()
stk_rewards_item_30_addname_dataframe['cname'] = stk_rewards_item_30_tscode_list
for table_name in stk_rewards_item_30.columns.values.tolist():
    stk_rewards_item_30_addname_dataframe[table_name] = stk_rewards_item_30[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_30_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_31_tscode_list = list() 
stk_rewards_item_31 = pro.stk_rewards(ts_code='000584.SZ,000585.SZ,000586.SZ,000587.SZ,000588.SZ,000589.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_31 返回数据 row 行数 = "+str(stk_rewards_item_31.shape[0]))
for ts_code_sh in stk_rewards_item_31['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_31_tscode_list.append(stock_name)
stk_rewards_item_31_addname_dataframe=pd.DataFrame()
stk_rewards_item_31_addname_dataframe['cname'] = stk_rewards_item_31_tscode_list
for table_name in stk_rewards_item_31.columns.values.tolist():
    stk_rewards_item_31_addname_dataframe[table_name] = stk_rewards_item_31[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_31_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_32_tscode_list = list() 
stk_rewards_item_32 = pro.stk_rewards(ts_code='000590.SZ,000591.SZ,000592.SZ,000593.SZ,000594.SZ,000595.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_32 返回数据 row 行数 = "+str(stk_rewards_item_32.shape[0]))
for ts_code_sh in stk_rewards_item_32['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_32_tscode_list.append(stock_name)
stk_rewards_item_32_addname_dataframe=pd.DataFrame()
stk_rewards_item_32_addname_dataframe['cname'] = stk_rewards_item_32_tscode_list
for table_name in stk_rewards_item_32.columns.values.tolist():
    stk_rewards_item_32_addname_dataframe[table_name] = stk_rewards_item_32[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_32_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_33_tscode_list = list() 
stk_rewards_item_33 = pro.stk_rewards(ts_code='000596.SZ,000597.SZ,000598.SZ,000599.SZ,000600.SZ,000601.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_33 返回数据 row 行数 = "+str(stk_rewards_item_33.shape[0]))
for ts_code_sh in stk_rewards_item_33['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_33_tscode_list.append(stock_name)
stk_rewards_item_33_addname_dataframe=pd.DataFrame()
stk_rewards_item_33_addname_dataframe['cname'] = stk_rewards_item_33_tscode_list
for table_name in stk_rewards_item_33.columns.values.tolist():
    stk_rewards_item_33_addname_dataframe[table_name] = stk_rewards_item_33[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_33_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_34_tscode_list = list() 
stk_rewards_item_34 = pro.stk_rewards(ts_code='000602.SZ,000603.SZ,000605.SZ,000606.SZ,000607.SZ,000608.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_34 返回数据 row 行数 = "+str(stk_rewards_item_34.shape[0]))
for ts_code_sh in stk_rewards_item_34['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_34_tscode_list.append(stock_name)
stk_rewards_item_34_addname_dataframe=pd.DataFrame()
stk_rewards_item_34_addname_dataframe['cname'] = stk_rewards_item_34_tscode_list
for table_name in stk_rewards_item_34.columns.values.tolist():
    stk_rewards_item_34_addname_dataframe[table_name] = stk_rewards_item_34[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_34_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_35_tscode_list = list() 
stk_rewards_item_35 = pro.stk_rewards(ts_code='000609.SZ,000610.SZ,000611.SZ,000612.SZ,000613.SZ,000615.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_35 返回数据 row 行数 = "+str(stk_rewards_item_35.shape[0]))
for ts_code_sh in stk_rewards_item_35['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_35_tscode_list.append(stock_name)
stk_rewards_item_35_addname_dataframe=pd.DataFrame()
stk_rewards_item_35_addname_dataframe['cname'] = stk_rewards_item_35_tscode_list
for table_name in stk_rewards_item_35.columns.values.tolist():
    stk_rewards_item_35_addname_dataframe[table_name] = stk_rewards_item_35[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_35_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_36_tscode_list = list() 
stk_rewards_item_36 = pro.stk_rewards(ts_code='000616.SZ,000617.SZ,000618.SZ,000619.SZ,000620.SZ,000621.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_36 返回数据 row 行数 = "+str(stk_rewards_item_36.shape[0]))
for ts_code_sh in stk_rewards_item_36['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_36_tscode_list.append(stock_name)
stk_rewards_item_36_addname_dataframe=pd.DataFrame()
stk_rewards_item_36_addname_dataframe['cname'] = stk_rewards_item_36_tscode_list
for table_name in stk_rewards_item_36.columns.values.tolist():
    stk_rewards_item_36_addname_dataframe[table_name] = stk_rewards_item_36[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_36_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_37_tscode_list = list() 
stk_rewards_item_37 = pro.stk_rewards(ts_code='000622.SZ,000623.SZ,000625.SZ,000626.SZ,000627.SZ,000628.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_37 返回数据 row 行数 = "+str(stk_rewards_item_37.shape[0]))
for ts_code_sh in stk_rewards_item_37['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_37_tscode_list.append(stock_name)
stk_rewards_item_37_addname_dataframe=pd.DataFrame()
stk_rewards_item_37_addname_dataframe['cname'] = stk_rewards_item_37_tscode_list
for table_name in stk_rewards_item_37.columns.values.tolist():
    stk_rewards_item_37_addname_dataframe[table_name] = stk_rewards_item_37[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_37_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_38_tscode_list = list() 
stk_rewards_item_38 = pro.stk_rewards(ts_code='000629.SZ,000630.SZ,000631.SZ,000632.SZ,000633.SZ,000635.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_38 返回数据 row 行数 = "+str(stk_rewards_item_38.shape[0]))
for ts_code_sh in stk_rewards_item_38['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_38_tscode_list.append(stock_name)
stk_rewards_item_38_addname_dataframe=pd.DataFrame()
stk_rewards_item_38_addname_dataframe['cname'] = stk_rewards_item_38_tscode_list
for table_name in stk_rewards_item_38.columns.values.tolist():
    stk_rewards_item_38_addname_dataframe[table_name] = stk_rewards_item_38[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_38_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_39_tscode_list = list() 
stk_rewards_item_39 = pro.stk_rewards(ts_code='000636.SZ,000637.SZ,000638.SZ,000639.SZ,000650.SZ,000651.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_39 返回数据 row 行数 = "+str(stk_rewards_item_39.shape[0]))
for ts_code_sh in stk_rewards_item_39['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_39_tscode_list.append(stock_name)
stk_rewards_item_39_addname_dataframe=pd.DataFrame()
stk_rewards_item_39_addname_dataframe['cname'] = stk_rewards_item_39_tscode_list
for table_name in stk_rewards_item_39.columns.values.tolist():
    stk_rewards_item_39_addname_dataframe[table_name] = stk_rewards_item_39[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_39_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_40_tscode_list = list() 
stk_rewards_item_40 = pro.stk_rewards(ts_code='000652.SZ,000653.SZ,000655.SZ,000656.SZ,000657.SZ,000658.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_40 返回数据 row 行数 = "+str(stk_rewards_item_40.shape[0]))
for ts_code_sh in stk_rewards_item_40['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_40_tscode_list.append(stock_name)
stk_rewards_item_40_addname_dataframe=pd.DataFrame()
stk_rewards_item_40_addname_dataframe['cname'] = stk_rewards_item_40_tscode_list
for table_name in stk_rewards_item_40.columns.values.tolist():
    stk_rewards_item_40_addname_dataframe[table_name] = stk_rewards_item_40[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_40_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_41_tscode_list = list() 
stk_rewards_item_41 = pro.stk_rewards(ts_code='000659.SZ,000660.SZ,000661.SZ,000662.SZ,000663.SZ,000665.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_41 返回数据 row 行数 = "+str(stk_rewards_item_41.shape[0]))
for ts_code_sh in stk_rewards_item_41['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_41_tscode_list.append(stock_name)
stk_rewards_item_41_addname_dataframe=pd.DataFrame()
stk_rewards_item_41_addname_dataframe['cname'] = stk_rewards_item_41_tscode_list
for table_name in stk_rewards_item_41.columns.values.tolist():
    stk_rewards_item_41_addname_dataframe[table_name] = stk_rewards_item_41[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_41_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_42_tscode_list = list() 
stk_rewards_item_42 = pro.stk_rewards(ts_code='000666.SZ,000667.SZ,000668.SZ,000669.SZ,000670.SZ,000671.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_42 返回数据 row 行数 = "+str(stk_rewards_item_42.shape[0]))
for ts_code_sh in stk_rewards_item_42['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_42_tscode_list.append(stock_name)
stk_rewards_item_42_addname_dataframe=pd.DataFrame()
stk_rewards_item_42_addname_dataframe['cname'] = stk_rewards_item_42_tscode_list
for table_name in stk_rewards_item_42.columns.values.tolist():
    stk_rewards_item_42_addname_dataframe[table_name] = stk_rewards_item_42[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_42_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_43_tscode_list = list() 
stk_rewards_item_43 = pro.stk_rewards(ts_code='000672.SZ,000673.SZ,000675.SZ,000676.SZ,000677.SZ,000678.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_43 返回数据 row 行数 = "+str(stk_rewards_item_43.shape[0]))
for ts_code_sh in stk_rewards_item_43['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_43_tscode_list.append(stock_name)
stk_rewards_item_43_addname_dataframe=pd.DataFrame()
stk_rewards_item_43_addname_dataframe['cname'] = stk_rewards_item_43_tscode_list
for table_name in stk_rewards_item_43.columns.values.tolist():
    stk_rewards_item_43_addname_dataframe[table_name] = stk_rewards_item_43[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_43_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_44_tscode_list = list() 
stk_rewards_item_44 = pro.stk_rewards(ts_code='000679.SZ,000680.SZ,000681.SZ,000682.SZ,000683.SZ,000685.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_44 返回数据 row 行数 = "+str(stk_rewards_item_44.shape[0]))
for ts_code_sh in stk_rewards_item_44['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_44_tscode_list.append(stock_name)
stk_rewards_item_44_addname_dataframe=pd.DataFrame()
stk_rewards_item_44_addname_dataframe['cname'] = stk_rewards_item_44_tscode_list
for table_name in stk_rewards_item_44.columns.values.tolist():
    stk_rewards_item_44_addname_dataframe[table_name] = stk_rewards_item_44[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_44_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_45_tscode_list = list() 
stk_rewards_item_45 = pro.stk_rewards(ts_code='000686.SZ,000687.SZ,000688.SZ,000689.SZ,000690.SZ,000691.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_45 返回数据 row 行数 = "+str(stk_rewards_item_45.shape[0]))
for ts_code_sh in stk_rewards_item_45['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_45_tscode_list.append(stock_name)
stk_rewards_item_45_addname_dataframe=pd.DataFrame()
stk_rewards_item_45_addname_dataframe['cname'] = stk_rewards_item_45_tscode_list
for table_name in stk_rewards_item_45.columns.values.tolist():
    stk_rewards_item_45_addname_dataframe[table_name] = stk_rewards_item_45[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_45_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_46_tscode_list = list() 
stk_rewards_item_46 = pro.stk_rewards(ts_code='000692.SZ,000693.SZ,000695.SZ,000697.SZ,000698.SZ,000699.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_46 返回数据 row 行数 = "+str(stk_rewards_item_46.shape[0]))
for ts_code_sh in stk_rewards_item_46['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_46_tscode_list.append(stock_name)
stk_rewards_item_46_addname_dataframe=pd.DataFrame()
stk_rewards_item_46_addname_dataframe['cname'] = stk_rewards_item_46_tscode_list
for table_name in stk_rewards_item_46.columns.values.tolist():
    stk_rewards_item_46_addname_dataframe[table_name] = stk_rewards_item_46[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_46_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_47_tscode_list = list() 
stk_rewards_item_47 = pro.stk_rewards(ts_code='000700.SZ,000701.SZ,000702.SZ,000703.SZ,000705.SZ,000707.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_47 返回数据 row 行数 = "+str(stk_rewards_item_47.shape[0]))
for ts_code_sh in stk_rewards_item_47['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_47_tscode_list.append(stock_name)
stk_rewards_item_47_addname_dataframe=pd.DataFrame()
stk_rewards_item_47_addname_dataframe['cname'] = stk_rewards_item_47_tscode_list
for table_name in stk_rewards_item_47.columns.values.tolist():
    stk_rewards_item_47_addname_dataframe[table_name] = stk_rewards_item_47[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_47_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_48_tscode_list = list() 
stk_rewards_item_48 = pro.stk_rewards(ts_code='000708.SZ,000709.SZ,000710.SZ,000711.SZ,000712.SZ,000713.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_48 返回数据 row 行数 = "+str(stk_rewards_item_48.shape[0]))
for ts_code_sh in stk_rewards_item_48['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_48_tscode_list.append(stock_name)
stk_rewards_item_48_addname_dataframe=pd.DataFrame()
stk_rewards_item_48_addname_dataframe['cname'] = stk_rewards_item_48_tscode_list
for table_name in stk_rewards_item_48.columns.values.tolist():
    stk_rewards_item_48_addname_dataframe[table_name] = stk_rewards_item_48[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_48_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_49_tscode_list = list() 
stk_rewards_item_49 = pro.stk_rewards(ts_code='000715.SZ,000716.SZ,000717.SZ,000718.SZ,000719.SZ,000720.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_49 返回数据 row 行数 = "+str(stk_rewards_item_49.shape[0]))
for ts_code_sh in stk_rewards_item_49['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_49_tscode_list.append(stock_name)
stk_rewards_item_49_addname_dataframe=pd.DataFrame()
stk_rewards_item_49_addname_dataframe['cname'] = stk_rewards_item_49_tscode_list
for table_name in stk_rewards_item_49.columns.values.tolist():
    stk_rewards_item_49_addname_dataframe[table_name] = stk_rewards_item_49[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_49_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_50_tscode_list = list() 
stk_rewards_item_50 = pro.stk_rewards(ts_code='000721.SZ,000722.SZ,000723.SZ,000725.SZ,000726.SZ,000727.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_50 返回数据 row 行数 = "+str(stk_rewards_item_50.shape[0]))
for ts_code_sh in stk_rewards_item_50['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_50_tscode_list.append(stock_name)
stk_rewards_item_50_addname_dataframe=pd.DataFrame()
stk_rewards_item_50_addname_dataframe['cname'] = stk_rewards_item_50_tscode_list
for table_name in stk_rewards_item_50.columns.values.tolist():
    stk_rewards_item_50_addname_dataframe[table_name] = stk_rewards_item_50[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_50_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_51_tscode_list = list() 
stk_rewards_item_51 = pro.stk_rewards(ts_code='000728.SZ,000729.SZ,000730.SZ,000731.SZ,000732.SZ,000733.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_51 返回数据 row 行数 = "+str(stk_rewards_item_51.shape[0]))
for ts_code_sh in stk_rewards_item_51['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_51_tscode_list.append(stock_name)
stk_rewards_item_51_addname_dataframe=pd.DataFrame()
stk_rewards_item_51_addname_dataframe['cname'] = stk_rewards_item_51_tscode_list
for table_name in stk_rewards_item_51.columns.values.tolist():
    stk_rewards_item_51_addname_dataframe[table_name] = stk_rewards_item_51[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_51_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_52_tscode_list = list() 
stk_rewards_item_52 = pro.stk_rewards(ts_code='000735.SZ,000736.SZ,000737.SZ,000738.SZ,000739.SZ,000748.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_52 返回数据 row 行数 = "+str(stk_rewards_item_52.shape[0]))
for ts_code_sh in stk_rewards_item_52['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_52_tscode_list.append(stock_name)
stk_rewards_item_52_addname_dataframe=pd.DataFrame()
stk_rewards_item_52_addname_dataframe['cname'] = stk_rewards_item_52_tscode_list
for table_name in stk_rewards_item_52.columns.values.tolist():
    stk_rewards_item_52_addname_dataframe[table_name] = stk_rewards_item_52[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_52_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_53_tscode_list = list() 
stk_rewards_item_53 = pro.stk_rewards(ts_code='000750.SZ,000751.SZ,000752.SZ,000753.SZ,000755.SZ,000756.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_53 返回数据 row 行数 = "+str(stk_rewards_item_53.shape[0]))
for ts_code_sh in stk_rewards_item_53['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_53_tscode_list.append(stock_name)
stk_rewards_item_53_addname_dataframe=pd.DataFrame()
stk_rewards_item_53_addname_dataframe['cname'] = stk_rewards_item_53_tscode_list
for table_name in stk_rewards_item_53.columns.values.tolist():
    stk_rewards_item_53_addname_dataframe[table_name] = stk_rewards_item_53[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_53_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_54_tscode_list = list() 
stk_rewards_item_54 = pro.stk_rewards(ts_code='000757.SZ,000758.SZ,000759.SZ,000760.SZ,000761.SZ,000762.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_54 返回数据 row 行数 = "+str(stk_rewards_item_54.shape[0]))
for ts_code_sh in stk_rewards_item_54['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_54_tscode_list.append(stock_name)
stk_rewards_item_54_addname_dataframe=pd.DataFrame()
stk_rewards_item_54_addname_dataframe['cname'] = stk_rewards_item_54_tscode_list
for table_name in stk_rewards_item_54.columns.values.tolist():
    stk_rewards_item_54_addname_dataframe[table_name] = stk_rewards_item_54[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_54_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_55_tscode_list = list() 
stk_rewards_item_55 = pro.stk_rewards(ts_code='000763.SZ,000765.SZ,000766.SZ,000767.SZ,000768.SZ,000769.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_55 返回数据 row 行数 = "+str(stk_rewards_item_55.shape[0]))
for ts_code_sh in stk_rewards_item_55['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_55_tscode_list.append(stock_name)
stk_rewards_item_55_addname_dataframe=pd.DataFrame()
stk_rewards_item_55_addname_dataframe['cname'] = stk_rewards_item_55_tscode_list
for table_name in stk_rewards_item_55.columns.values.tolist():
    stk_rewards_item_55_addname_dataframe[table_name] = stk_rewards_item_55[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_55_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_56_tscode_list = list() 
stk_rewards_item_56 = pro.stk_rewards(ts_code='000776.SZ,000777.SZ,000778.SZ,000779.SZ,000780.SZ,000782.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_56 返回数据 row 行数 = "+str(stk_rewards_item_56.shape[0]))
for ts_code_sh in stk_rewards_item_56['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_56_tscode_list.append(stock_name)
stk_rewards_item_56_addname_dataframe=pd.DataFrame()
stk_rewards_item_56_addname_dataframe['cname'] = stk_rewards_item_56_tscode_list
for table_name in stk_rewards_item_56.columns.values.tolist():
    stk_rewards_item_56_addname_dataframe[table_name] = stk_rewards_item_56[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_56_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_57_tscode_list = list() 
stk_rewards_item_57 = pro.stk_rewards(ts_code='000783.SZ,000785.SZ,000786.SZ,000787.SZ,000788.SZ,000789.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_57 返回数据 row 行数 = "+str(stk_rewards_item_57.shape[0]))
for ts_code_sh in stk_rewards_item_57['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_57_tscode_list.append(stock_name)
stk_rewards_item_57_addname_dataframe=pd.DataFrame()
stk_rewards_item_57_addname_dataframe['cname'] = stk_rewards_item_57_tscode_list
for table_name in stk_rewards_item_57.columns.values.tolist():
    stk_rewards_item_57_addname_dataframe[table_name] = stk_rewards_item_57[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_57_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_58_tscode_list = list() 
stk_rewards_item_58 = pro.stk_rewards(ts_code='000790.SZ,000791.SZ,000792.SZ,000793.SZ,000795.SZ,000796.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_58 返回数据 row 行数 = "+str(stk_rewards_item_58.shape[0]))
for ts_code_sh in stk_rewards_item_58['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_58_tscode_list.append(stock_name)
stk_rewards_item_58_addname_dataframe=pd.DataFrame()
stk_rewards_item_58_addname_dataframe['cname'] = stk_rewards_item_58_tscode_list
for table_name in stk_rewards_item_58.columns.values.tolist():
    stk_rewards_item_58_addname_dataframe[table_name] = stk_rewards_item_58[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_58_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_59_tscode_list = list() 
stk_rewards_item_59 = pro.stk_rewards(ts_code='000797.SZ,000798.SZ,000799.SZ,000800.SZ,000801.SZ,000802.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_59 返回数据 row 行数 = "+str(stk_rewards_item_59.shape[0]))
for ts_code_sh in stk_rewards_item_59['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_59_tscode_list.append(stock_name)
stk_rewards_item_59_addname_dataframe=pd.DataFrame()
stk_rewards_item_59_addname_dataframe['cname'] = stk_rewards_item_59_tscode_list
for table_name in stk_rewards_item_59.columns.values.tolist():
    stk_rewards_item_59_addname_dataframe[table_name] = stk_rewards_item_59[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_59_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_60_tscode_list = list() 
stk_rewards_item_60 = pro.stk_rewards(ts_code='000803.SZ,000805.SZ,000806.SZ,000807.SZ,000809.SZ,000810.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_60 返回数据 row 行数 = "+str(stk_rewards_item_60.shape[0]))
for ts_code_sh in stk_rewards_item_60['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_60_tscode_list.append(stock_name)
stk_rewards_item_60_addname_dataframe=pd.DataFrame()
stk_rewards_item_60_addname_dataframe['cname'] = stk_rewards_item_60_tscode_list
for table_name in stk_rewards_item_60.columns.values.tolist():
    stk_rewards_item_60_addname_dataframe[table_name] = stk_rewards_item_60[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_60_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_61_tscode_list = list() 
stk_rewards_item_61 = pro.stk_rewards(ts_code='000811.SZ,000812.SZ,000813.SZ,000815.SZ,000816.SZ,000817.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_61 返回数据 row 行数 = "+str(stk_rewards_item_61.shape[0]))
for ts_code_sh in stk_rewards_item_61['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_61_tscode_list.append(stock_name)
stk_rewards_item_61_addname_dataframe=pd.DataFrame()
stk_rewards_item_61_addname_dataframe['cname'] = stk_rewards_item_61_tscode_list
for table_name in stk_rewards_item_61.columns.values.tolist():
    stk_rewards_item_61_addname_dataframe[table_name] = stk_rewards_item_61[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_61_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_62_tscode_list = list() 
stk_rewards_item_62 = pro.stk_rewards(ts_code='000818.SZ,000819.SZ,000820.SZ,000821.SZ,000822.SZ,000823.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_62 返回数据 row 行数 = "+str(stk_rewards_item_62.shape[0]))
for ts_code_sh in stk_rewards_item_62['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_62_tscode_list.append(stock_name)
stk_rewards_item_62_addname_dataframe=pd.DataFrame()
stk_rewards_item_62_addname_dataframe['cname'] = stk_rewards_item_62_tscode_list
for table_name in stk_rewards_item_62.columns.values.tolist():
    stk_rewards_item_62_addname_dataframe[table_name] = stk_rewards_item_62[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_62_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_63_tscode_list = list() 
stk_rewards_item_63 = pro.stk_rewards(ts_code='000825.SZ,000826.SZ,000827.SZ,000828.SZ,000829.SZ,000830.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_63 返回数据 row 行数 = "+str(stk_rewards_item_63.shape[0]))
for ts_code_sh in stk_rewards_item_63['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_63_tscode_list.append(stock_name)
stk_rewards_item_63_addname_dataframe=pd.DataFrame()
stk_rewards_item_63_addname_dataframe['cname'] = stk_rewards_item_63_tscode_list
for table_name in stk_rewards_item_63.columns.values.tolist():
    stk_rewards_item_63_addname_dataframe[table_name] = stk_rewards_item_63[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_63_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_64_tscode_list = list() 
stk_rewards_item_64 = pro.stk_rewards(ts_code='000831.SZ,000832.SZ,000833.SZ,000835.SZ,000836.SZ,000837.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_64 返回数据 row 行数 = "+str(stk_rewards_item_64.shape[0]))
for ts_code_sh in stk_rewards_item_64['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_64_tscode_list.append(stock_name)
stk_rewards_item_64_addname_dataframe=pd.DataFrame()
stk_rewards_item_64_addname_dataframe['cname'] = stk_rewards_item_64_tscode_list
for table_name in stk_rewards_item_64.columns.values.tolist():
    stk_rewards_item_64_addname_dataframe[table_name] = stk_rewards_item_64[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_64_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_65_tscode_list = list() 
stk_rewards_item_65 = pro.stk_rewards(ts_code='000838.SZ,000839.SZ,000848.SZ,000850.SZ,000851.SZ,000852.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_65 返回数据 row 行数 = "+str(stk_rewards_item_65.shape[0]))
for ts_code_sh in stk_rewards_item_65['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_65_tscode_list.append(stock_name)
stk_rewards_item_65_addname_dataframe=pd.DataFrame()
stk_rewards_item_65_addname_dataframe['cname'] = stk_rewards_item_65_tscode_list
for table_name in stk_rewards_item_65.columns.values.tolist():
    stk_rewards_item_65_addname_dataframe[table_name] = stk_rewards_item_65[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_65_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_66_tscode_list = list() 
stk_rewards_item_66 = pro.stk_rewards(ts_code='000856.SZ,000858.SZ,000859.SZ,000860.SZ,000861.SZ,000862.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_66 返回数据 row 行数 = "+str(stk_rewards_item_66.shape[0]))
for ts_code_sh in stk_rewards_item_66['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_66_tscode_list.append(stock_name)
stk_rewards_item_66_addname_dataframe=pd.DataFrame()
stk_rewards_item_66_addname_dataframe['cname'] = stk_rewards_item_66_tscode_list
for table_name in stk_rewards_item_66.columns.values.tolist():
    stk_rewards_item_66_addname_dataframe[table_name] = stk_rewards_item_66[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_66_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_67_tscode_list = list() 
stk_rewards_item_67 = pro.stk_rewards(ts_code='000863.SZ,000866.SZ,000868.SZ,000869.SZ,000875.SZ,000876.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_67 返回数据 row 行数 = "+str(stk_rewards_item_67.shape[0]))
for ts_code_sh in stk_rewards_item_67['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_67_tscode_list.append(stock_name)
stk_rewards_item_67_addname_dataframe=pd.DataFrame()
stk_rewards_item_67_addname_dataframe['cname'] = stk_rewards_item_67_tscode_list
for table_name in stk_rewards_item_67.columns.values.tolist():
    stk_rewards_item_67_addname_dataframe[table_name] = stk_rewards_item_67[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_67_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_68_tscode_list = list() 
stk_rewards_item_68 = pro.stk_rewards(ts_code='000877.SZ,000878.SZ,000880.SZ,000881.SZ,000882.SZ,000883.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_68 返回数据 row 行数 = "+str(stk_rewards_item_68.shape[0]))
for ts_code_sh in stk_rewards_item_68['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_68_tscode_list.append(stock_name)
stk_rewards_item_68_addname_dataframe=pd.DataFrame()
stk_rewards_item_68_addname_dataframe['cname'] = stk_rewards_item_68_tscode_list
for table_name in stk_rewards_item_68.columns.values.tolist():
    stk_rewards_item_68_addname_dataframe[table_name] = stk_rewards_item_68[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_68_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_69_tscode_list = list() 
stk_rewards_item_69 = pro.stk_rewards(ts_code='000885.SZ,000886.SZ,000887.SZ,000888.SZ,000889.SZ,000890.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_69 返回数据 row 行数 = "+str(stk_rewards_item_69.shape[0]))
for ts_code_sh in stk_rewards_item_69['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_69_tscode_list.append(stock_name)
stk_rewards_item_69_addname_dataframe=pd.DataFrame()
stk_rewards_item_69_addname_dataframe['cname'] = stk_rewards_item_69_tscode_list
for table_name in stk_rewards_item_69.columns.values.tolist():
    stk_rewards_item_69_addname_dataframe[table_name] = stk_rewards_item_69[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_69_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_70_tscode_list = list() 
stk_rewards_item_70 = pro.stk_rewards(ts_code='000892.SZ,000893.SZ,000895.SZ,000897.SZ,000898.SZ,000899.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_70 返回数据 row 行数 = "+str(stk_rewards_item_70.shape[0]))
for ts_code_sh in stk_rewards_item_70['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_70_tscode_list.append(stock_name)
stk_rewards_item_70_addname_dataframe=pd.DataFrame()
stk_rewards_item_70_addname_dataframe['cname'] = stk_rewards_item_70_tscode_list
for table_name in stk_rewards_item_70.columns.values.tolist():
    stk_rewards_item_70_addname_dataframe[table_name] = stk_rewards_item_70[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_70_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_71_tscode_list = list() 
stk_rewards_item_71 = pro.stk_rewards(ts_code='000900.SZ,000901.SZ,000902.SZ,000903.SZ,000905.SZ,000906.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_71 返回数据 row 行数 = "+str(stk_rewards_item_71.shape[0]))
for ts_code_sh in stk_rewards_item_71['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_71_tscode_list.append(stock_name)
stk_rewards_item_71_addname_dataframe=pd.DataFrame()
stk_rewards_item_71_addname_dataframe['cname'] = stk_rewards_item_71_tscode_list
for table_name in stk_rewards_item_71.columns.values.tolist():
    stk_rewards_item_71_addname_dataframe[table_name] = stk_rewards_item_71[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_71_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_72_tscode_list = list() 
stk_rewards_item_72 = pro.stk_rewards(ts_code='000908.SZ,000909.SZ,000910.SZ,000911.SZ,000912.SZ,000913.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_72 返回数据 row 行数 = "+str(stk_rewards_item_72.shape[0]))
for ts_code_sh in stk_rewards_item_72['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_72_tscode_list.append(stock_name)
stk_rewards_item_72_addname_dataframe=pd.DataFrame()
stk_rewards_item_72_addname_dataframe['cname'] = stk_rewards_item_72_tscode_list
for table_name in stk_rewards_item_72.columns.values.tolist():
    stk_rewards_item_72_addname_dataframe[table_name] = stk_rewards_item_72[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_72_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_73_tscode_list = list() 
stk_rewards_item_73 = pro.stk_rewards(ts_code='000915.SZ,000916.SZ,000917.SZ,000918.SZ,000919.SZ,000920.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_73 返回数据 row 行数 = "+str(stk_rewards_item_73.shape[0]))
for ts_code_sh in stk_rewards_item_73['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_73_tscode_list.append(stock_name)
stk_rewards_item_73_addname_dataframe=pd.DataFrame()
stk_rewards_item_73_addname_dataframe['cname'] = stk_rewards_item_73_tscode_list
for table_name in stk_rewards_item_73.columns.values.tolist():
    stk_rewards_item_73_addname_dataframe[table_name] = stk_rewards_item_73[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_73_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_74_tscode_list = list() 
stk_rewards_item_74 = pro.stk_rewards(ts_code='000921.SZ,000922.SZ,000923.SZ,000925.SZ,000926.SZ,000927.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_74 返回数据 row 行数 = "+str(stk_rewards_item_74.shape[0]))
for ts_code_sh in stk_rewards_item_74['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_74_tscode_list.append(stock_name)
stk_rewards_item_74_addname_dataframe=pd.DataFrame()
stk_rewards_item_74_addname_dataframe['cname'] = stk_rewards_item_74_tscode_list
for table_name in stk_rewards_item_74.columns.values.tolist():
    stk_rewards_item_74_addname_dataframe[table_name] = stk_rewards_item_74[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_74_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_75_tscode_list = list() 
stk_rewards_item_75 = pro.stk_rewards(ts_code='000928.SZ,000929.SZ,000930.SZ,000931.SZ,000932.SZ,000933.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_75 返回数据 row 行数 = "+str(stk_rewards_item_75.shape[0]))
for ts_code_sh in stk_rewards_item_75['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_75_tscode_list.append(stock_name)
stk_rewards_item_75_addname_dataframe=pd.DataFrame()
stk_rewards_item_75_addname_dataframe['cname'] = stk_rewards_item_75_tscode_list
for table_name in stk_rewards_item_75.columns.values.tolist():
    stk_rewards_item_75_addname_dataframe[table_name] = stk_rewards_item_75[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_75_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_76_tscode_list = list() 
stk_rewards_item_76 = pro.stk_rewards(ts_code='000935.SZ,000936.SZ,000937.SZ,000938.SZ,000939.SZ,000948.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_76 返回数据 row 行数 = "+str(stk_rewards_item_76.shape[0]))
for ts_code_sh in stk_rewards_item_76['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_76_tscode_list.append(stock_name)
stk_rewards_item_76_addname_dataframe=pd.DataFrame()
stk_rewards_item_76_addname_dataframe['cname'] = stk_rewards_item_76_tscode_list
for table_name in stk_rewards_item_76.columns.values.tolist():
    stk_rewards_item_76_addname_dataframe[table_name] = stk_rewards_item_76[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_76_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_77_tscode_list = list() 
stk_rewards_item_77 = pro.stk_rewards(ts_code='000949.SZ,000950.SZ,000951.SZ,000952.SZ,000953.SZ,000955.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_77 返回数据 row 行数 = "+str(stk_rewards_item_77.shape[0]))
for ts_code_sh in stk_rewards_item_77['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_77_tscode_list.append(stock_name)
stk_rewards_item_77_addname_dataframe=pd.DataFrame()
stk_rewards_item_77_addname_dataframe['cname'] = stk_rewards_item_77_tscode_list
for table_name in stk_rewards_item_77.columns.values.tolist():
    stk_rewards_item_77_addname_dataframe[table_name] = stk_rewards_item_77[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_77_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_78_tscode_list = list() 
stk_rewards_item_78 = pro.stk_rewards(ts_code='000956.SZ,000957.SZ,000958.SZ,000959.SZ,000960.SZ,000961.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_78 返回数据 row 行数 = "+str(stk_rewards_item_78.shape[0]))
for ts_code_sh in stk_rewards_item_78['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_78_tscode_list.append(stock_name)
stk_rewards_item_78_addname_dataframe=pd.DataFrame()
stk_rewards_item_78_addname_dataframe['cname'] = stk_rewards_item_78_tscode_list
for table_name in stk_rewards_item_78.columns.values.tolist():
    stk_rewards_item_78_addname_dataframe[table_name] = stk_rewards_item_78[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_78_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_79_tscode_list = list() 
stk_rewards_item_79 = pro.stk_rewards(ts_code='000962.SZ,000963.SZ,000965.SZ,000966.SZ,000967.SZ,000968.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_79 返回数据 row 行数 = "+str(stk_rewards_item_79.shape[0]))
for ts_code_sh in stk_rewards_item_79['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_79_tscode_list.append(stock_name)
stk_rewards_item_79_addname_dataframe=pd.DataFrame()
stk_rewards_item_79_addname_dataframe['cname'] = stk_rewards_item_79_tscode_list
for table_name in stk_rewards_item_79.columns.values.tolist():
    stk_rewards_item_79_addname_dataframe[table_name] = stk_rewards_item_79[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_79_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_80_tscode_list = list() 
stk_rewards_item_80 = pro.stk_rewards(ts_code='000969.SZ,000970.SZ,000971.SZ,000972.SZ,000973.SZ,000975.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_80 返回数据 row 行数 = "+str(stk_rewards_item_80.shape[0]))
for ts_code_sh in stk_rewards_item_80['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_80_tscode_list.append(stock_name)
stk_rewards_item_80_addname_dataframe=pd.DataFrame()
stk_rewards_item_80_addname_dataframe['cname'] = stk_rewards_item_80_tscode_list
for table_name in stk_rewards_item_80.columns.values.tolist():
    stk_rewards_item_80_addname_dataframe[table_name] = stk_rewards_item_80[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_80_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_81_tscode_list = list() 
stk_rewards_item_81 = pro.stk_rewards(ts_code='000976.SZ,000977.SZ,000978.SZ,000979.SZ,000980.SZ,000981.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_81 返回数据 row 行数 = "+str(stk_rewards_item_81.shape[0]))
for ts_code_sh in stk_rewards_item_81['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_81_tscode_list.append(stock_name)
stk_rewards_item_81_addname_dataframe=pd.DataFrame()
stk_rewards_item_81_addname_dataframe['cname'] = stk_rewards_item_81_tscode_list
for table_name in stk_rewards_item_81.columns.values.tolist():
    stk_rewards_item_81_addname_dataframe[table_name] = stk_rewards_item_81[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_81_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_82_tscode_list = list() 
stk_rewards_item_82 = pro.stk_rewards(ts_code='000982.SZ,000983.SZ,000985.SZ,000987.SZ,000988.SZ,000989.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_82 返回数据 row 行数 = "+str(stk_rewards_item_82.shape[0]))
for ts_code_sh in stk_rewards_item_82['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_82_tscode_list.append(stock_name)
stk_rewards_item_82_addname_dataframe=pd.DataFrame()
stk_rewards_item_82_addname_dataframe['cname'] = stk_rewards_item_82_tscode_list
for table_name in stk_rewards_item_82.columns.values.tolist():
    stk_rewards_item_82_addname_dataframe[table_name] = stk_rewards_item_82[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_82_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_83_tscode_list = list() 
stk_rewards_item_83 = pro.stk_rewards(ts_code='000990.SZ,000993.SZ,000995.SZ,000996.SZ,000997.SZ,000998.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_83 返回数据 row 行数 = "+str(stk_rewards_item_83.shape[0]))
for ts_code_sh in stk_rewards_item_83['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_83_tscode_list.append(stock_name)
stk_rewards_item_83_addname_dataframe=pd.DataFrame()
stk_rewards_item_83_addname_dataframe['cname'] = stk_rewards_item_83_tscode_list
for table_name in stk_rewards_item_83.columns.values.tolist():
    stk_rewards_item_83_addname_dataframe[table_name] = stk_rewards_item_83[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_83_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_84_tscode_list = list() 
stk_rewards_item_84 = pro.stk_rewards(ts_code='000999.SZ,001696.SZ,001872.SZ,001896.SZ,001914.SZ,001965.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_84 返回数据 row 行数 = "+str(stk_rewards_item_84.shape[0]))
for ts_code_sh in stk_rewards_item_84['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_84_tscode_list.append(stock_name)
stk_rewards_item_84_addname_dataframe=pd.DataFrame()
stk_rewards_item_84_addname_dataframe['cname'] = stk_rewards_item_84_tscode_list
for table_name in stk_rewards_item_84.columns.values.tolist():
    stk_rewards_item_84_addname_dataframe[table_name] = stk_rewards_item_84[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_84_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_85_tscode_list = list() 
stk_rewards_item_85 = pro.stk_rewards(ts_code='001979.SZ,002001.SZ,002002.SZ,002003.SZ,002004.SZ,002005.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_85 返回数据 row 行数 = "+str(stk_rewards_item_85.shape[0]))
for ts_code_sh in stk_rewards_item_85['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_85_tscode_list.append(stock_name)
stk_rewards_item_85_addname_dataframe=pd.DataFrame()
stk_rewards_item_85_addname_dataframe['cname'] = stk_rewards_item_85_tscode_list
for table_name in stk_rewards_item_85.columns.values.tolist():
    stk_rewards_item_85_addname_dataframe[table_name] = stk_rewards_item_85[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_85_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_86_tscode_list = list() 
stk_rewards_item_86 = pro.stk_rewards(ts_code='002006.SZ,002007.SZ,002008.SZ,002009.SZ,002010.SZ,002011.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_86 返回数据 row 行数 = "+str(stk_rewards_item_86.shape[0]))
for ts_code_sh in stk_rewards_item_86['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_86_tscode_list.append(stock_name)
stk_rewards_item_86_addname_dataframe=pd.DataFrame()
stk_rewards_item_86_addname_dataframe['cname'] = stk_rewards_item_86_tscode_list
for table_name in stk_rewards_item_86.columns.values.tolist():
    stk_rewards_item_86_addname_dataframe[table_name] = stk_rewards_item_86[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_86_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_87_tscode_list = list() 
stk_rewards_item_87 = pro.stk_rewards(ts_code='002012.SZ,002013.SZ,002014.SZ,002015.SZ,002016.SZ,002017.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_87 返回数据 row 行数 = "+str(stk_rewards_item_87.shape[0]))
for ts_code_sh in stk_rewards_item_87['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_87_tscode_list.append(stock_name)
stk_rewards_item_87_addname_dataframe=pd.DataFrame()
stk_rewards_item_87_addname_dataframe['cname'] = stk_rewards_item_87_tscode_list
for table_name in stk_rewards_item_87.columns.values.tolist():
    stk_rewards_item_87_addname_dataframe[table_name] = stk_rewards_item_87[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_87_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_88_tscode_list = list() 
stk_rewards_item_88 = pro.stk_rewards(ts_code='002018.SZ,002019.SZ,002020.SZ,002021.SZ,002022.SZ,002023.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_88 返回数据 row 行数 = "+str(stk_rewards_item_88.shape[0]))
for ts_code_sh in stk_rewards_item_88['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_88_tscode_list.append(stock_name)
stk_rewards_item_88_addname_dataframe=pd.DataFrame()
stk_rewards_item_88_addname_dataframe['cname'] = stk_rewards_item_88_tscode_list
for table_name in stk_rewards_item_88.columns.values.tolist():
    stk_rewards_item_88_addname_dataframe[table_name] = stk_rewards_item_88[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_88_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_89_tscode_list = list() 
stk_rewards_item_89 = pro.stk_rewards(ts_code='002024.SZ,002025.SZ,002026.SZ,002027.SZ,002028.SZ,002029.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_89 返回数据 row 行数 = "+str(stk_rewards_item_89.shape[0]))
for ts_code_sh in stk_rewards_item_89['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_89_tscode_list.append(stock_name)
stk_rewards_item_89_addname_dataframe=pd.DataFrame()
stk_rewards_item_89_addname_dataframe['cname'] = stk_rewards_item_89_tscode_list
for table_name in stk_rewards_item_89.columns.values.tolist():
    stk_rewards_item_89_addname_dataframe[table_name] = stk_rewards_item_89[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_89_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_90_tscode_list = list() 
stk_rewards_item_90 = pro.stk_rewards(ts_code='002030.SZ,002031.SZ,002032.SZ,002033.SZ,002034.SZ,002035.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_90 返回数据 row 行数 = "+str(stk_rewards_item_90.shape[0]))
for ts_code_sh in stk_rewards_item_90['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_90_tscode_list.append(stock_name)
stk_rewards_item_90_addname_dataframe=pd.DataFrame()
stk_rewards_item_90_addname_dataframe['cname'] = stk_rewards_item_90_tscode_list
for table_name in stk_rewards_item_90.columns.values.tolist():
    stk_rewards_item_90_addname_dataframe[table_name] = stk_rewards_item_90[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_90_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_91_tscode_list = list() 
stk_rewards_item_91 = pro.stk_rewards(ts_code='002036.SZ,002037.SZ,002038.SZ,002039.SZ,002040.SZ,002041.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_91 返回数据 row 行数 = "+str(stk_rewards_item_91.shape[0]))
for ts_code_sh in stk_rewards_item_91['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_91_tscode_list.append(stock_name)
stk_rewards_item_91_addname_dataframe=pd.DataFrame()
stk_rewards_item_91_addname_dataframe['cname'] = stk_rewards_item_91_tscode_list
for table_name in stk_rewards_item_91.columns.values.tolist():
    stk_rewards_item_91_addname_dataframe[table_name] = stk_rewards_item_91[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_91_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_92_tscode_list = list() 
stk_rewards_item_92 = pro.stk_rewards(ts_code='002042.SZ,002043.SZ,002044.SZ,002045.SZ,002046.SZ,002047.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_92 返回数据 row 行数 = "+str(stk_rewards_item_92.shape[0]))
for ts_code_sh in stk_rewards_item_92['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_92_tscode_list.append(stock_name)
stk_rewards_item_92_addname_dataframe=pd.DataFrame()
stk_rewards_item_92_addname_dataframe['cname'] = stk_rewards_item_92_tscode_list
for table_name in stk_rewards_item_92.columns.values.tolist():
    stk_rewards_item_92_addname_dataframe[table_name] = stk_rewards_item_92[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_92_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_93_tscode_list = list() 
stk_rewards_item_93 = pro.stk_rewards(ts_code='002048.SZ,002049.SZ,002050.SZ,002051.SZ,002052.SZ,002053.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_93 返回数据 row 行数 = "+str(stk_rewards_item_93.shape[0]))
for ts_code_sh in stk_rewards_item_93['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_93_tscode_list.append(stock_name)
stk_rewards_item_93_addname_dataframe=pd.DataFrame()
stk_rewards_item_93_addname_dataframe['cname'] = stk_rewards_item_93_tscode_list
for table_name in stk_rewards_item_93.columns.values.tolist():
    stk_rewards_item_93_addname_dataframe[table_name] = stk_rewards_item_93[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_93_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_94_tscode_list = list() 
stk_rewards_item_94 = pro.stk_rewards(ts_code='002054.SZ,002055.SZ,002056.SZ,002057.SZ,002058.SZ,002059.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_94 返回数据 row 行数 = "+str(stk_rewards_item_94.shape[0]))
for ts_code_sh in stk_rewards_item_94['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_94_tscode_list.append(stock_name)
stk_rewards_item_94_addname_dataframe=pd.DataFrame()
stk_rewards_item_94_addname_dataframe['cname'] = stk_rewards_item_94_tscode_list
for table_name in stk_rewards_item_94.columns.values.tolist():
    stk_rewards_item_94_addname_dataframe[table_name] = stk_rewards_item_94[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_94_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_95_tscode_list = list() 
stk_rewards_item_95 = pro.stk_rewards(ts_code='002060.SZ,002061.SZ,002062.SZ,002063.SZ,002064.SZ,002065.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_95 返回数据 row 行数 = "+str(stk_rewards_item_95.shape[0]))
for ts_code_sh in stk_rewards_item_95['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_95_tscode_list.append(stock_name)
stk_rewards_item_95_addname_dataframe=pd.DataFrame()
stk_rewards_item_95_addname_dataframe['cname'] = stk_rewards_item_95_tscode_list
for table_name in stk_rewards_item_95.columns.values.tolist():
    stk_rewards_item_95_addname_dataframe[table_name] = stk_rewards_item_95[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_95_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_96_tscode_list = list() 
stk_rewards_item_96 = pro.stk_rewards(ts_code='002066.SZ,002067.SZ,002068.SZ,002069.SZ,002070.SZ,002071.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_96 返回数据 row 行数 = "+str(stk_rewards_item_96.shape[0]))
for ts_code_sh in stk_rewards_item_96['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_96_tscode_list.append(stock_name)
stk_rewards_item_96_addname_dataframe=pd.DataFrame()
stk_rewards_item_96_addname_dataframe['cname'] = stk_rewards_item_96_tscode_list
for table_name in stk_rewards_item_96.columns.values.tolist():
    stk_rewards_item_96_addname_dataframe[table_name] = stk_rewards_item_96[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_96_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_97_tscode_list = list() 
stk_rewards_item_97 = pro.stk_rewards(ts_code='002072.SZ,002073.SZ,002074.SZ,002075.SZ,002076.SZ,002077.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_97 返回数据 row 行数 = "+str(stk_rewards_item_97.shape[0]))
for ts_code_sh in stk_rewards_item_97['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_97_tscode_list.append(stock_name)
stk_rewards_item_97_addname_dataframe=pd.DataFrame()
stk_rewards_item_97_addname_dataframe['cname'] = stk_rewards_item_97_tscode_list
for table_name in stk_rewards_item_97.columns.values.tolist():
    stk_rewards_item_97_addname_dataframe[table_name] = stk_rewards_item_97[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_97_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_98_tscode_list = list() 
stk_rewards_item_98 = pro.stk_rewards(ts_code='002078.SZ,002079.SZ,002080.SZ,002081.SZ,002082.SZ,002083.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_98 返回数据 row 行数 = "+str(stk_rewards_item_98.shape[0]))
for ts_code_sh in stk_rewards_item_98['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_98_tscode_list.append(stock_name)
stk_rewards_item_98_addname_dataframe=pd.DataFrame()
stk_rewards_item_98_addname_dataframe['cname'] = stk_rewards_item_98_tscode_list
for table_name in stk_rewards_item_98.columns.values.tolist():
    stk_rewards_item_98_addname_dataframe[table_name] = stk_rewards_item_98[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_98_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_99_tscode_list = list() 
stk_rewards_item_99 = pro.stk_rewards(ts_code='002084.SZ,002085.SZ,002086.SZ,002087.SZ,002088.SZ,002089.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_99 返回数据 row 行数 = "+str(stk_rewards_item_99.shape[0]))
for ts_code_sh in stk_rewards_item_99['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_99_tscode_list.append(stock_name)
stk_rewards_item_99_addname_dataframe=pd.DataFrame()
stk_rewards_item_99_addname_dataframe['cname'] = stk_rewards_item_99_tscode_list
for table_name in stk_rewards_item_99.columns.values.tolist():
    stk_rewards_item_99_addname_dataframe[table_name] = stk_rewards_item_99[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_99_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_100_tscode_list = list() 
stk_rewards_item_100 = pro.stk_rewards(ts_code='002090.SZ,002091.SZ,002092.SZ,002093.SZ,002094.SZ,002095.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_100 返回数据 row 行数 = "+str(stk_rewards_item_100.shape[0]))
for ts_code_sh in stk_rewards_item_100['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_100_tscode_list.append(stock_name)
stk_rewards_item_100_addname_dataframe=pd.DataFrame()
stk_rewards_item_100_addname_dataframe['cname'] = stk_rewards_item_100_tscode_list
for table_name in stk_rewards_item_100.columns.values.tolist():
    stk_rewards_item_100_addname_dataframe[table_name] = stk_rewards_item_100[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_100_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_101_tscode_list = list() 
stk_rewards_item_101 = pro.stk_rewards(ts_code='002096.SZ,002097.SZ,002098.SZ,002099.SZ,002100.SZ,002101.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_101 返回数据 row 行数 = "+str(stk_rewards_item_101.shape[0]))
for ts_code_sh in stk_rewards_item_101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_101_tscode_list.append(stock_name)
stk_rewards_item_101_addname_dataframe=pd.DataFrame()
stk_rewards_item_101_addname_dataframe['cname'] = stk_rewards_item_101_tscode_list
for table_name in stk_rewards_item_101.columns.values.tolist():
    stk_rewards_item_101_addname_dataframe[table_name] = stk_rewards_item_101[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_101_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_102_tscode_list = list() 
stk_rewards_item_102 = pro.stk_rewards(ts_code='002102.SZ,002103.SZ,002104.SZ,002105.SZ,002106.SZ,002107.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_102 返回数据 row 行数 = "+str(stk_rewards_item_102.shape[0]))
for ts_code_sh in stk_rewards_item_102['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_102_tscode_list.append(stock_name)
stk_rewards_item_102_addname_dataframe=pd.DataFrame()
stk_rewards_item_102_addname_dataframe['cname'] = stk_rewards_item_102_tscode_list
for table_name in stk_rewards_item_102.columns.values.tolist():
    stk_rewards_item_102_addname_dataframe[table_name] = stk_rewards_item_102[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_102_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_103_tscode_list = list() 
stk_rewards_item_103 = pro.stk_rewards(ts_code='002108.SZ,002109.SZ,002110.SZ,002111.SZ,002112.SZ,002113.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_103 返回数据 row 行数 = "+str(stk_rewards_item_103.shape[0]))
for ts_code_sh in stk_rewards_item_103['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_103_tscode_list.append(stock_name)
stk_rewards_item_103_addname_dataframe=pd.DataFrame()
stk_rewards_item_103_addname_dataframe['cname'] = stk_rewards_item_103_tscode_list
for table_name in stk_rewards_item_103.columns.values.tolist():
    stk_rewards_item_103_addname_dataframe[table_name] = stk_rewards_item_103[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_103_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_104_tscode_list = list() 
stk_rewards_item_104 = pro.stk_rewards(ts_code='002114.SZ,002115.SZ,002116.SZ,002117.SZ,002118.SZ,002119.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_104 返回数据 row 行数 = "+str(stk_rewards_item_104.shape[0]))
for ts_code_sh in stk_rewards_item_104['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_104_tscode_list.append(stock_name)
stk_rewards_item_104_addname_dataframe=pd.DataFrame()
stk_rewards_item_104_addname_dataframe['cname'] = stk_rewards_item_104_tscode_list
for table_name in stk_rewards_item_104.columns.values.tolist():
    stk_rewards_item_104_addname_dataframe[table_name] = stk_rewards_item_104[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_104_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_105_tscode_list = list() 
stk_rewards_item_105 = pro.stk_rewards(ts_code='002120.SZ,002121.SZ,002122.SZ,002123.SZ,002124.SZ,002125.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_105 返回数据 row 行数 = "+str(stk_rewards_item_105.shape[0]))
for ts_code_sh in stk_rewards_item_105['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_105_tscode_list.append(stock_name)
stk_rewards_item_105_addname_dataframe=pd.DataFrame()
stk_rewards_item_105_addname_dataframe['cname'] = stk_rewards_item_105_tscode_list
for table_name in stk_rewards_item_105.columns.values.tolist():
    stk_rewards_item_105_addname_dataframe[table_name] = stk_rewards_item_105[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_105_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_106_tscode_list = list() 
stk_rewards_item_106 = pro.stk_rewards(ts_code='002126.SZ,002127.SZ,002128.SZ,002129.SZ,002130.SZ,002131.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_106 返回数据 row 行数 = "+str(stk_rewards_item_106.shape[0]))
for ts_code_sh in stk_rewards_item_106['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_106_tscode_list.append(stock_name)
stk_rewards_item_106_addname_dataframe=pd.DataFrame()
stk_rewards_item_106_addname_dataframe['cname'] = stk_rewards_item_106_tscode_list
for table_name in stk_rewards_item_106.columns.values.tolist():
    stk_rewards_item_106_addname_dataframe[table_name] = stk_rewards_item_106[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_106_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_107_tscode_list = list() 
stk_rewards_item_107 = pro.stk_rewards(ts_code='002132.SZ,002133.SZ,002134.SZ,002135.SZ,002136.SZ,002137.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_107 返回数据 row 行数 = "+str(stk_rewards_item_107.shape[0]))
for ts_code_sh in stk_rewards_item_107['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_107_tscode_list.append(stock_name)
stk_rewards_item_107_addname_dataframe=pd.DataFrame()
stk_rewards_item_107_addname_dataframe['cname'] = stk_rewards_item_107_tscode_list
for table_name in stk_rewards_item_107.columns.values.tolist():
    stk_rewards_item_107_addname_dataframe[table_name] = stk_rewards_item_107[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_107_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_108_tscode_list = list() 
stk_rewards_item_108 = pro.stk_rewards(ts_code='002138.SZ,002139.SZ,002140.SZ,002141.SZ,002142.SZ,002143.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_108 返回数据 row 行数 = "+str(stk_rewards_item_108.shape[0]))
for ts_code_sh in stk_rewards_item_108['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_108_tscode_list.append(stock_name)
stk_rewards_item_108_addname_dataframe=pd.DataFrame()
stk_rewards_item_108_addname_dataframe['cname'] = stk_rewards_item_108_tscode_list
for table_name in stk_rewards_item_108.columns.values.tolist():
    stk_rewards_item_108_addname_dataframe[table_name] = stk_rewards_item_108[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_108_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_109_tscode_list = list() 
stk_rewards_item_109 = pro.stk_rewards(ts_code='002144.SZ,002145.SZ,002146.SZ,002147.SZ,002148.SZ,002149.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_109 返回数据 row 行数 = "+str(stk_rewards_item_109.shape[0]))
for ts_code_sh in stk_rewards_item_109['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_109_tscode_list.append(stock_name)
stk_rewards_item_109_addname_dataframe=pd.DataFrame()
stk_rewards_item_109_addname_dataframe['cname'] = stk_rewards_item_109_tscode_list
for table_name in stk_rewards_item_109.columns.values.tolist():
    stk_rewards_item_109_addname_dataframe[table_name] = stk_rewards_item_109[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_109_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_110_tscode_list = list() 
stk_rewards_item_110 = pro.stk_rewards(ts_code='002150.SZ,002151.SZ,002152.SZ,002153.SZ,002154.SZ,002155.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_110 返回数据 row 行数 = "+str(stk_rewards_item_110.shape[0]))
for ts_code_sh in stk_rewards_item_110['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_110_tscode_list.append(stock_name)
stk_rewards_item_110_addname_dataframe=pd.DataFrame()
stk_rewards_item_110_addname_dataframe['cname'] = stk_rewards_item_110_tscode_list
for table_name in stk_rewards_item_110.columns.values.tolist():
    stk_rewards_item_110_addname_dataframe[table_name] = stk_rewards_item_110[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_110_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_111_tscode_list = list() 
stk_rewards_item_111 = pro.stk_rewards(ts_code='002156.SZ,002157.SZ,002158.SZ,002159.SZ,002160.SZ,002161.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_111 返回数据 row 行数 = "+str(stk_rewards_item_111.shape[0]))
for ts_code_sh in stk_rewards_item_111['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_111_tscode_list.append(stock_name)
stk_rewards_item_111_addname_dataframe=pd.DataFrame()
stk_rewards_item_111_addname_dataframe['cname'] = stk_rewards_item_111_tscode_list
for table_name in stk_rewards_item_111.columns.values.tolist():
    stk_rewards_item_111_addname_dataframe[table_name] = stk_rewards_item_111[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_111_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_112_tscode_list = list() 
stk_rewards_item_112 = pro.stk_rewards(ts_code='002162.SZ,002163.SZ,002164.SZ,002165.SZ,002166.SZ,002167.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_112 返回数据 row 行数 = "+str(stk_rewards_item_112.shape[0]))
for ts_code_sh in stk_rewards_item_112['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_112_tscode_list.append(stock_name)
stk_rewards_item_112_addname_dataframe=pd.DataFrame()
stk_rewards_item_112_addname_dataframe['cname'] = stk_rewards_item_112_tscode_list
for table_name in stk_rewards_item_112.columns.values.tolist():
    stk_rewards_item_112_addname_dataframe[table_name] = stk_rewards_item_112[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_112_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_113_tscode_list = list() 
stk_rewards_item_113 = pro.stk_rewards(ts_code='002168.SZ,002169.SZ,002170.SZ,002171.SZ,002172.SZ,002173.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_113 返回数据 row 行数 = "+str(stk_rewards_item_113.shape[0]))
for ts_code_sh in stk_rewards_item_113['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_113_tscode_list.append(stock_name)
stk_rewards_item_113_addname_dataframe=pd.DataFrame()
stk_rewards_item_113_addname_dataframe['cname'] = stk_rewards_item_113_tscode_list
for table_name in stk_rewards_item_113.columns.values.tolist():
    stk_rewards_item_113_addname_dataframe[table_name] = stk_rewards_item_113[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_113_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_114_tscode_list = list() 
stk_rewards_item_114 = pro.stk_rewards(ts_code='002174.SZ,002175.SZ,002176.SZ,002177.SZ,002178.SZ,002179.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_114 返回数据 row 行数 = "+str(stk_rewards_item_114.shape[0]))
for ts_code_sh in stk_rewards_item_114['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_114_tscode_list.append(stock_name)
stk_rewards_item_114_addname_dataframe=pd.DataFrame()
stk_rewards_item_114_addname_dataframe['cname'] = stk_rewards_item_114_tscode_list
for table_name in stk_rewards_item_114.columns.values.tolist():
    stk_rewards_item_114_addname_dataframe[table_name] = stk_rewards_item_114[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_114_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_115_tscode_list = list() 
stk_rewards_item_115 = pro.stk_rewards(ts_code='002180.SZ,002181.SZ,002182.SZ,002183.SZ,002184.SZ,002185.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_115 返回数据 row 行数 = "+str(stk_rewards_item_115.shape[0]))
for ts_code_sh in stk_rewards_item_115['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_115_tscode_list.append(stock_name)
stk_rewards_item_115_addname_dataframe=pd.DataFrame()
stk_rewards_item_115_addname_dataframe['cname'] = stk_rewards_item_115_tscode_list
for table_name in stk_rewards_item_115.columns.values.tolist():
    stk_rewards_item_115_addname_dataframe[table_name] = stk_rewards_item_115[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_115_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_116_tscode_list = list() 
stk_rewards_item_116 = pro.stk_rewards(ts_code='002186.SZ,002187.SZ,002188.SZ,002189.SZ,002190.SZ,002191.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_116 返回数据 row 行数 = "+str(stk_rewards_item_116.shape[0]))
for ts_code_sh in stk_rewards_item_116['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_116_tscode_list.append(stock_name)
stk_rewards_item_116_addname_dataframe=pd.DataFrame()
stk_rewards_item_116_addname_dataframe['cname'] = stk_rewards_item_116_tscode_list
for table_name in stk_rewards_item_116.columns.values.tolist():
    stk_rewards_item_116_addname_dataframe[table_name] = stk_rewards_item_116[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_116_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_117_tscode_list = list() 
stk_rewards_item_117 = pro.stk_rewards(ts_code='002192.SZ,002193.SZ,002194.SZ,002195.SZ,002196.SZ,002197.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_117 返回数据 row 行数 = "+str(stk_rewards_item_117.shape[0]))
for ts_code_sh in stk_rewards_item_117['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_117_tscode_list.append(stock_name)
stk_rewards_item_117_addname_dataframe=pd.DataFrame()
stk_rewards_item_117_addname_dataframe['cname'] = stk_rewards_item_117_tscode_list
for table_name in stk_rewards_item_117.columns.values.tolist():
    stk_rewards_item_117_addname_dataframe[table_name] = stk_rewards_item_117[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_117_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_118_tscode_list = list() 
stk_rewards_item_118 = pro.stk_rewards(ts_code='002198.SZ,002199.SZ,002200.SZ,002201.SZ,002202.SZ,002203.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_118 返回数据 row 行数 = "+str(stk_rewards_item_118.shape[0]))
for ts_code_sh in stk_rewards_item_118['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_118_tscode_list.append(stock_name)
stk_rewards_item_118_addname_dataframe=pd.DataFrame()
stk_rewards_item_118_addname_dataframe['cname'] = stk_rewards_item_118_tscode_list
for table_name in stk_rewards_item_118.columns.values.tolist():
    stk_rewards_item_118_addname_dataframe[table_name] = stk_rewards_item_118[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_118_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_119_tscode_list = list() 
stk_rewards_item_119 = pro.stk_rewards(ts_code='002204.SZ,002205.SZ,002206.SZ,002207.SZ,002208.SZ,002209.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_119 返回数据 row 行数 = "+str(stk_rewards_item_119.shape[0]))
for ts_code_sh in stk_rewards_item_119['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_119_tscode_list.append(stock_name)
stk_rewards_item_119_addname_dataframe=pd.DataFrame()
stk_rewards_item_119_addname_dataframe['cname'] = stk_rewards_item_119_tscode_list
for table_name in stk_rewards_item_119.columns.values.tolist():
    stk_rewards_item_119_addname_dataframe[table_name] = stk_rewards_item_119[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_119_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_120_tscode_list = list() 
stk_rewards_item_120 = pro.stk_rewards(ts_code='002210.SZ,002211.SZ,002212.SZ,002213.SZ,002214.SZ,002215.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_120 返回数据 row 行数 = "+str(stk_rewards_item_120.shape[0]))
for ts_code_sh in stk_rewards_item_120['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_120_tscode_list.append(stock_name)
stk_rewards_item_120_addname_dataframe=pd.DataFrame()
stk_rewards_item_120_addname_dataframe['cname'] = stk_rewards_item_120_tscode_list
for table_name in stk_rewards_item_120.columns.values.tolist():
    stk_rewards_item_120_addname_dataframe[table_name] = stk_rewards_item_120[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_120_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_121_tscode_list = list() 
stk_rewards_item_121 = pro.stk_rewards(ts_code='002216.SZ,002217.SZ,002218.SZ,002219.SZ,002220.SZ,002221.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_121 返回数据 row 行数 = "+str(stk_rewards_item_121.shape[0]))
for ts_code_sh in stk_rewards_item_121['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_121_tscode_list.append(stock_name)
stk_rewards_item_121_addname_dataframe=pd.DataFrame()
stk_rewards_item_121_addname_dataframe['cname'] = stk_rewards_item_121_tscode_list
for table_name in stk_rewards_item_121.columns.values.tolist():
    stk_rewards_item_121_addname_dataframe[table_name] = stk_rewards_item_121[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_121_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_122_tscode_list = list() 
stk_rewards_item_122 = pro.stk_rewards(ts_code='002222.SZ,002223.SZ,002224.SZ,002225.SZ,002226.SZ,002227.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_122 返回数据 row 行数 = "+str(stk_rewards_item_122.shape[0]))
for ts_code_sh in stk_rewards_item_122['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_122_tscode_list.append(stock_name)
stk_rewards_item_122_addname_dataframe=pd.DataFrame()
stk_rewards_item_122_addname_dataframe['cname'] = stk_rewards_item_122_tscode_list
for table_name in stk_rewards_item_122.columns.values.tolist():
    stk_rewards_item_122_addname_dataframe[table_name] = stk_rewards_item_122[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_122_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_123_tscode_list = list() 
stk_rewards_item_123 = pro.stk_rewards(ts_code='002228.SZ,002229.SZ,002230.SZ,002231.SZ,002232.SZ,002233.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_123 返回数据 row 行数 = "+str(stk_rewards_item_123.shape[0]))
for ts_code_sh in stk_rewards_item_123['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_123_tscode_list.append(stock_name)
stk_rewards_item_123_addname_dataframe=pd.DataFrame()
stk_rewards_item_123_addname_dataframe['cname'] = stk_rewards_item_123_tscode_list
for table_name in stk_rewards_item_123.columns.values.tolist():
    stk_rewards_item_123_addname_dataframe[table_name] = stk_rewards_item_123[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_123_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_124_tscode_list = list() 
stk_rewards_item_124 = pro.stk_rewards(ts_code='002234.SZ,002235.SZ,002236.SZ,002237.SZ,002238.SZ,002239.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_124 返回数据 row 行数 = "+str(stk_rewards_item_124.shape[0]))
for ts_code_sh in stk_rewards_item_124['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_124_tscode_list.append(stock_name)
stk_rewards_item_124_addname_dataframe=pd.DataFrame()
stk_rewards_item_124_addname_dataframe['cname'] = stk_rewards_item_124_tscode_list
for table_name in stk_rewards_item_124.columns.values.tolist():
    stk_rewards_item_124_addname_dataframe[table_name] = stk_rewards_item_124[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_124_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_125_tscode_list = list() 
stk_rewards_item_125 = pro.stk_rewards(ts_code='002240.SZ,002241.SZ,002242.SZ,002243.SZ,002244.SZ,002245.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_125 返回数据 row 行数 = "+str(stk_rewards_item_125.shape[0]))
for ts_code_sh in stk_rewards_item_125['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_125_tscode_list.append(stock_name)
stk_rewards_item_125_addname_dataframe=pd.DataFrame()
stk_rewards_item_125_addname_dataframe['cname'] = stk_rewards_item_125_tscode_list
for table_name in stk_rewards_item_125.columns.values.tolist():
    stk_rewards_item_125_addname_dataframe[table_name] = stk_rewards_item_125[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_125_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_126_tscode_list = list() 
stk_rewards_item_126 = pro.stk_rewards(ts_code='002246.SZ,002247.SZ,002248.SZ,002249.SZ,002250.SZ,002251.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_126 返回数据 row 行数 = "+str(stk_rewards_item_126.shape[0]))
for ts_code_sh in stk_rewards_item_126['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_126_tscode_list.append(stock_name)
stk_rewards_item_126_addname_dataframe=pd.DataFrame()
stk_rewards_item_126_addname_dataframe['cname'] = stk_rewards_item_126_tscode_list
for table_name in stk_rewards_item_126.columns.values.tolist():
    stk_rewards_item_126_addname_dataframe[table_name] = stk_rewards_item_126[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_126_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_127_tscode_list = list() 
stk_rewards_item_127 = pro.stk_rewards(ts_code='002252.SZ,002253.SZ,002254.SZ,002255.SZ,002256.SZ,002258.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_127 返回数据 row 行数 = "+str(stk_rewards_item_127.shape[0]))
for ts_code_sh in stk_rewards_item_127['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_127_tscode_list.append(stock_name)
stk_rewards_item_127_addname_dataframe=pd.DataFrame()
stk_rewards_item_127_addname_dataframe['cname'] = stk_rewards_item_127_tscode_list
for table_name in stk_rewards_item_127.columns.values.tolist():
    stk_rewards_item_127_addname_dataframe[table_name] = stk_rewards_item_127[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_127_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_128_tscode_list = list() 
stk_rewards_item_128 = pro.stk_rewards(ts_code='002259.SZ,002260.SZ,002261.SZ,002262.SZ,002263.SZ,002264.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_128 返回数据 row 行数 = "+str(stk_rewards_item_128.shape[0]))
for ts_code_sh in stk_rewards_item_128['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_128_tscode_list.append(stock_name)
stk_rewards_item_128_addname_dataframe=pd.DataFrame()
stk_rewards_item_128_addname_dataframe['cname'] = stk_rewards_item_128_tscode_list
for table_name in stk_rewards_item_128.columns.values.tolist():
    stk_rewards_item_128_addname_dataframe[table_name] = stk_rewards_item_128[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_128_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_129_tscode_list = list() 
stk_rewards_item_129 = pro.stk_rewards(ts_code='002265.SZ,002266.SZ,002267.SZ,002268.SZ,002269.SZ,002270.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_129 返回数据 row 行数 = "+str(stk_rewards_item_129.shape[0]))
for ts_code_sh in stk_rewards_item_129['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_129_tscode_list.append(stock_name)
stk_rewards_item_129_addname_dataframe=pd.DataFrame()
stk_rewards_item_129_addname_dataframe['cname'] = stk_rewards_item_129_tscode_list
for table_name in stk_rewards_item_129.columns.values.tolist():
    stk_rewards_item_129_addname_dataframe[table_name] = stk_rewards_item_129[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_129_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_130_tscode_list = list() 
stk_rewards_item_130 = pro.stk_rewards(ts_code='002271.SZ,002272.SZ,002273.SZ,002274.SZ,002275.SZ,002276.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_130 返回数据 row 行数 = "+str(stk_rewards_item_130.shape[0]))
for ts_code_sh in stk_rewards_item_130['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_130_tscode_list.append(stock_name)
stk_rewards_item_130_addname_dataframe=pd.DataFrame()
stk_rewards_item_130_addname_dataframe['cname'] = stk_rewards_item_130_tscode_list
for table_name in stk_rewards_item_130.columns.values.tolist():
    stk_rewards_item_130_addname_dataframe[table_name] = stk_rewards_item_130[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_130_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_131_tscode_list = list() 
stk_rewards_item_131 = pro.stk_rewards(ts_code='002277.SZ,002278.SZ,002279.SZ,002280.SZ,002281.SZ,002282.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_131 返回数据 row 行数 = "+str(stk_rewards_item_131.shape[0]))
for ts_code_sh in stk_rewards_item_131['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_131_tscode_list.append(stock_name)
stk_rewards_item_131_addname_dataframe=pd.DataFrame()
stk_rewards_item_131_addname_dataframe['cname'] = stk_rewards_item_131_tscode_list
for table_name in stk_rewards_item_131.columns.values.tolist():
    stk_rewards_item_131_addname_dataframe[table_name] = stk_rewards_item_131[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_131_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_132_tscode_list = list() 
stk_rewards_item_132 = pro.stk_rewards(ts_code='002283.SZ,002284.SZ,002285.SZ,002286.SZ,002287.SZ,002288.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_132 返回数据 row 行数 = "+str(stk_rewards_item_132.shape[0]))
for ts_code_sh in stk_rewards_item_132['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_132_tscode_list.append(stock_name)
stk_rewards_item_132_addname_dataframe=pd.DataFrame()
stk_rewards_item_132_addname_dataframe['cname'] = stk_rewards_item_132_tscode_list
for table_name in stk_rewards_item_132.columns.values.tolist():
    stk_rewards_item_132_addname_dataframe[table_name] = stk_rewards_item_132[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_132_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_133_tscode_list = list() 
stk_rewards_item_133 = pro.stk_rewards(ts_code='002289.SZ,002290.SZ,002291.SZ,002292.SZ,002293.SZ,002294.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_133 返回数据 row 行数 = "+str(stk_rewards_item_133.shape[0]))
for ts_code_sh in stk_rewards_item_133['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_133_tscode_list.append(stock_name)
stk_rewards_item_133_addname_dataframe=pd.DataFrame()
stk_rewards_item_133_addname_dataframe['cname'] = stk_rewards_item_133_tscode_list
for table_name in stk_rewards_item_133.columns.values.tolist():
    stk_rewards_item_133_addname_dataframe[table_name] = stk_rewards_item_133[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_133_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_134_tscode_list = list() 
stk_rewards_item_134 = pro.stk_rewards(ts_code='002295.SZ,002296.SZ,002297.SZ,002298.SZ,002299.SZ,002300.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_134 返回数据 row 行数 = "+str(stk_rewards_item_134.shape[0]))
for ts_code_sh in stk_rewards_item_134['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_134_tscode_list.append(stock_name)
stk_rewards_item_134_addname_dataframe=pd.DataFrame()
stk_rewards_item_134_addname_dataframe['cname'] = stk_rewards_item_134_tscode_list
for table_name in stk_rewards_item_134.columns.values.tolist():
    stk_rewards_item_134_addname_dataframe[table_name] = stk_rewards_item_134[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_134_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_135_tscode_list = list() 
stk_rewards_item_135 = pro.stk_rewards(ts_code='002301.SZ,002302.SZ,002303.SZ,002304.SZ,002305.SZ,002306.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_135 返回数据 row 行数 = "+str(stk_rewards_item_135.shape[0]))
for ts_code_sh in stk_rewards_item_135['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_135_tscode_list.append(stock_name)
stk_rewards_item_135_addname_dataframe=pd.DataFrame()
stk_rewards_item_135_addname_dataframe['cname'] = stk_rewards_item_135_tscode_list
for table_name in stk_rewards_item_135.columns.values.tolist():
    stk_rewards_item_135_addname_dataframe[table_name] = stk_rewards_item_135[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_135_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_136_tscode_list = list() 
stk_rewards_item_136 = pro.stk_rewards(ts_code='002307.SZ,002308.SZ,002309.SZ,002310.SZ,002311.SZ,002312.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_136 返回数据 row 行数 = "+str(stk_rewards_item_136.shape[0]))
for ts_code_sh in stk_rewards_item_136['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_136_tscode_list.append(stock_name)
stk_rewards_item_136_addname_dataframe=pd.DataFrame()
stk_rewards_item_136_addname_dataframe['cname'] = stk_rewards_item_136_tscode_list
for table_name in stk_rewards_item_136.columns.values.tolist():
    stk_rewards_item_136_addname_dataframe[table_name] = stk_rewards_item_136[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_136_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_137_tscode_list = list() 
stk_rewards_item_137 = pro.stk_rewards(ts_code='002313.SZ,002314.SZ,002315.SZ,002316.SZ,002317.SZ,002318.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_137 返回数据 row 行数 = "+str(stk_rewards_item_137.shape[0]))
for ts_code_sh in stk_rewards_item_137['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_137_tscode_list.append(stock_name)
stk_rewards_item_137_addname_dataframe=pd.DataFrame()
stk_rewards_item_137_addname_dataframe['cname'] = stk_rewards_item_137_tscode_list
for table_name in stk_rewards_item_137.columns.values.tolist():
    stk_rewards_item_137_addname_dataframe[table_name] = stk_rewards_item_137[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_137_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_138_tscode_list = list() 
stk_rewards_item_138 = pro.stk_rewards(ts_code='002319.SZ,002320.SZ,002321.SZ,002322.SZ,002323.SZ,002324.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_138 返回数据 row 行数 = "+str(stk_rewards_item_138.shape[0]))
for ts_code_sh in stk_rewards_item_138['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_138_tscode_list.append(stock_name)
stk_rewards_item_138_addname_dataframe=pd.DataFrame()
stk_rewards_item_138_addname_dataframe['cname'] = stk_rewards_item_138_tscode_list
for table_name in stk_rewards_item_138.columns.values.tolist():
    stk_rewards_item_138_addname_dataframe[table_name] = stk_rewards_item_138[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_138_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_139_tscode_list = list() 
stk_rewards_item_139 = pro.stk_rewards(ts_code='002325.SZ,002326.SZ,002327.SZ,002328.SZ,002329.SZ,002330.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_139 返回数据 row 行数 = "+str(stk_rewards_item_139.shape[0]))
for ts_code_sh in stk_rewards_item_139['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_139_tscode_list.append(stock_name)
stk_rewards_item_139_addname_dataframe=pd.DataFrame()
stk_rewards_item_139_addname_dataframe['cname'] = stk_rewards_item_139_tscode_list
for table_name in stk_rewards_item_139.columns.values.tolist():
    stk_rewards_item_139_addname_dataframe[table_name] = stk_rewards_item_139[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_139_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_140_tscode_list = list() 
stk_rewards_item_140 = pro.stk_rewards(ts_code='002331.SZ,002332.SZ,002333.SZ,002334.SZ,002335.SZ,002336.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_140 返回数据 row 行数 = "+str(stk_rewards_item_140.shape[0]))
for ts_code_sh in stk_rewards_item_140['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_140_tscode_list.append(stock_name)
stk_rewards_item_140_addname_dataframe=pd.DataFrame()
stk_rewards_item_140_addname_dataframe['cname'] = stk_rewards_item_140_tscode_list
for table_name in stk_rewards_item_140.columns.values.tolist():
    stk_rewards_item_140_addname_dataframe[table_name] = stk_rewards_item_140[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_140_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_141_tscode_list = list() 
stk_rewards_item_141 = pro.stk_rewards(ts_code='002337.SZ,002338.SZ,002339.SZ,002340.SZ,002341.SZ,002342.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_141 返回数据 row 行数 = "+str(stk_rewards_item_141.shape[0]))
for ts_code_sh in stk_rewards_item_141['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_141_tscode_list.append(stock_name)
stk_rewards_item_141_addname_dataframe=pd.DataFrame()
stk_rewards_item_141_addname_dataframe['cname'] = stk_rewards_item_141_tscode_list
for table_name in stk_rewards_item_141.columns.values.tolist():
    stk_rewards_item_141_addname_dataframe[table_name] = stk_rewards_item_141[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_141_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_142_tscode_list = list() 
stk_rewards_item_142 = pro.stk_rewards(ts_code='002343.SZ,002344.SZ,002345.SZ,002346.SZ,002347.SZ,002348.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_142 返回数据 row 行数 = "+str(stk_rewards_item_142.shape[0]))
for ts_code_sh in stk_rewards_item_142['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_142_tscode_list.append(stock_name)
stk_rewards_item_142_addname_dataframe=pd.DataFrame()
stk_rewards_item_142_addname_dataframe['cname'] = stk_rewards_item_142_tscode_list
for table_name in stk_rewards_item_142.columns.values.tolist():
    stk_rewards_item_142_addname_dataframe[table_name] = stk_rewards_item_142[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_142_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_143_tscode_list = list() 
stk_rewards_item_143 = pro.stk_rewards(ts_code='002349.SZ,002350.SZ,002351.SZ,002352.SZ,002353.SZ,002354.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_143 返回数据 row 行数 = "+str(stk_rewards_item_143.shape[0]))
for ts_code_sh in stk_rewards_item_143['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_143_tscode_list.append(stock_name)
stk_rewards_item_143_addname_dataframe=pd.DataFrame()
stk_rewards_item_143_addname_dataframe['cname'] = stk_rewards_item_143_tscode_list
for table_name in stk_rewards_item_143.columns.values.tolist():
    stk_rewards_item_143_addname_dataframe[table_name] = stk_rewards_item_143[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_143_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_144_tscode_list = list() 
stk_rewards_item_144 = pro.stk_rewards(ts_code='002355.SZ,002356.SZ,002357.SZ,002358.SZ,002359.SZ,002360.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_144 返回数据 row 行数 = "+str(stk_rewards_item_144.shape[0]))
for ts_code_sh in stk_rewards_item_144['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_144_tscode_list.append(stock_name)
stk_rewards_item_144_addname_dataframe=pd.DataFrame()
stk_rewards_item_144_addname_dataframe['cname'] = stk_rewards_item_144_tscode_list
for table_name in stk_rewards_item_144.columns.values.tolist():
    stk_rewards_item_144_addname_dataframe[table_name] = stk_rewards_item_144[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_144_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_145_tscode_list = list() 
stk_rewards_item_145 = pro.stk_rewards(ts_code='002361.SZ,002362.SZ,002363.SZ,002364.SZ,002365.SZ,002366.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_145 返回数据 row 行数 = "+str(stk_rewards_item_145.shape[0]))
for ts_code_sh in stk_rewards_item_145['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_145_tscode_list.append(stock_name)
stk_rewards_item_145_addname_dataframe=pd.DataFrame()
stk_rewards_item_145_addname_dataframe['cname'] = stk_rewards_item_145_tscode_list
for table_name in stk_rewards_item_145.columns.values.tolist():
    stk_rewards_item_145_addname_dataframe[table_name] = stk_rewards_item_145[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_145_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_146_tscode_list = list() 
stk_rewards_item_146 = pro.stk_rewards(ts_code='002367.SZ,002368.SZ,002369.SZ,002370.SZ,002371.SZ,002372.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_146 返回数据 row 行数 = "+str(stk_rewards_item_146.shape[0]))
for ts_code_sh in stk_rewards_item_146['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_146_tscode_list.append(stock_name)
stk_rewards_item_146_addname_dataframe=pd.DataFrame()
stk_rewards_item_146_addname_dataframe['cname'] = stk_rewards_item_146_tscode_list
for table_name in stk_rewards_item_146.columns.values.tolist():
    stk_rewards_item_146_addname_dataframe[table_name] = stk_rewards_item_146[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_146_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_147_tscode_list = list() 
stk_rewards_item_147 = pro.stk_rewards(ts_code='002373.SZ,002374.SZ,002375.SZ,002376.SZ,002377.SZ,002378.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_147 返回数据 row 行数 = "+str(stk_rewards_item_147.shape[0]))
for ts_code_sh in stk_rewards_item_147['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_147_tscode_list.append(stock_name)
stk_rewards_item_147_addname_dataframe=pd.DataFrame()
stk_rewards_item_147_addname_dataframe['cname'] = stk_rewards_item_147_tscode_list
for table_name in stk_rewards_item_147.columns.values.tolist():
    stk_rewards_item_147_addname_dataframe[table_name] = stk_rewards_item_147[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_147_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_148_tscode_list = list() 
stk_rewards_item_148 = pro.stk_rewards(ts_code='002379.SZ,002380.SZ,002381.SZ,002382.SZ,002383.SZ,002384.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_148 返回数据 row 行数 = "+str(stk_rewards_item_148.shape[0]))
for ts_code_sh in stk_rewards_item_148['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_148_tscode_list.append(stock_name)
stk_rewards_item_148_addname_dataframe=pd.DataFrame()
stk_rewards_item_148_addname_dataframe['cname'] = stk_rewards_item_148_tscode_list
for table_name in stk_rewards_item_148.columns.values.tolist():
    stk_rewards_item_148_addname_dataframe[table_name] = stk_rewards_item_148[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_148_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_149_tscode_list = list() 
stk_rewards_item_149 = pro.stk_rewards(ts_code='002385.SZ,002386.SZ,002387.SZ,002388.SZ,002389.SZ,002390.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_149 返回数据 row 行数 = "+str(stk_rewards_item_149.shape[0]))
for ts_code_sh in stk_rewards_item_149['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_149_tscode_list.append(stock_name)
stk_rewards_item_149_addname_dataframe=pd.DataFrame()
stk_rewards_item_149_addname_dataframe['cname'] = stk_rewards_item_149_tscode_list
for table_name in stk_rewards_item_149.columns.values.tolist():
    stk_rewards_item_149_addname_dataframe[table_name] = stk_rewards_item_149[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_149_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_150_tscode_list = list() 
stk_rewards_item_150 = pro.stk_rewards(ts_code='002391.SZ,002392.SZ,002393.SZ,002394.SZ,002395.SZ,002396.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_150 返回数据 row 行数 = "+str(stk_rewards_item_150.shape[0]))
for ts_code_sh in stk_rewards_item_150['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_150_tscode_list.append(stock_name)
stk_rewards_item_150_addname_dataframe=pd.DataFrame()
stk_rewards_item_150_addname_dataframe['cname'] = stk_rewards_item_150_tscode_list
for table_name in stk_rewards_item_150.columns.values.tolist():
    stk_rewards_item_150_addname_dataframe[table_name] = stk_rewards_item_150[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_150_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_151_tscode_list = list() 
stk_rewards_item_151 = pro.stk_rewards(ts_code='002397.SZ,002398.SZ,002399.SZ,002400.SZ,002401.SZ,002402.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_151 返回数据 row 行数 = "+str(stk_rewards_item_151.shape[0]))
for ts_code_sh in stk_rewards_item_151['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_151_tscode_list.append(stock_name)
stk_rewards_item_151_addname_dataframe=pd.DataFrame()
stk_rewards_item_151_addname_dataframe['cname'] = stk_rewards_item_151_tscode_list
for table_name in stk_rewards_item_151.columns.values.tolist():
    stk_rewards_item_151_addname_dataframe[table_name] = stk_rewards_item_151[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_151_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_152_tscode_list = list() 
stk_rewards_item_152 = pro.stk_rewards(ts_code='002403.SZ,002404.SZ,002405.SZ,002406.SZ,002407.SZ,002408.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_152 返回数据 row 行数 = "+str(stk_rewards_item_152.shape[0]))
for ts_code_sh in stk_rewards_item_152['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_152_tscode_list.append(stock_name)
stk_rewards_item_152_addname_dataframe=pd.DataFrame()
stk_rewards_item_152_addname_dataframe['cname'] = stk_rewards_item_152_tscode_list
for table_name in stk_rewards_item_152.columns.values.tolist():
    stk_rewards_item_152_addname_dataframe[table_name] = stk_rewards_item_152[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_152_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_153_tscode_list = list() 
stk_rewards_item_153 = pro.stk_rewards(ts_code='002409.SZ,002410.SZ,002411.SZ,002412.SZ,002413.SZ,002414.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_153 返回数据 row 行数 = "+str(stk_rewards_item_153.shape[0]))
for ts_code_sh in stk_rewards_item_153['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_153_tscode_list.append(stock_name)
stk_rewards_item_153_addname_dataframe=pd.DataFrame()
stk_rewards_item_153_addname_dataframe['cname'] = stk_rewards_item_153_tscode_list
for table_name in stk_rewards_item_153.columns.values.tolist():
    stk_rewards_item_153_addname_dataframe[table_name] = stk_rewards_item_153[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_153_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_154_tscode_list = list() 
stk_rewards_item_154 = pro.stk_rewards(ts_code='002415.SZ,002416.SZ,002417.SZ,002418.SZ,002419.SZ,002420.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_154 返回数据 row 行数 = "+str(stk_rewards_item_154.shape[0]))
for ts_code_sh in stk_rewards_item_154['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_154_tscode_list.append(stock_name)
stk_rewards_item_154_addname_dataframe=pd.DataFrame()
stk_rewards_item_154_addname_dataframe['cname'] = stk_rewards_item_154_tscode_list
for table_name in stk_rewards_item_154.columns.values.tolist():
    stk_rewards_item_154_addname_dataframe[table_name] = stk_rewards_item_154[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_154_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_155_tscode_list = list() 
stk_rewards_item_155 = pro.stk_rewards(ts_code='002421.SZ,002422.SZ,002423.SZ,002424.SZ,002425.SZ,002426.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_155 返回数据 row 行数 = "+str(stk_rewards_item_155.shape[0]))
for ts_code_sh in stk_rewards_item_155['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_155_tscode_list.append(stock_name)
stk_rewards_item_155_addname_dataframe=pd.DataFrame()
stk_rewards_item_155_addname_dataframe['cname'] = stk_rewards_item_155_tscode_list
for table_name in stk_rewards_item_155.columns.values.tolist():
    stk_rewards_item_155_addname_dataframe[table_name] = stk_rewards_item_155[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_155_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_156_tscode_list = list() 
stk_rewards_item_156 = pro.stk_rewards(ts_code='002427.SZ,002428.SZ,002429.SZ,002430.SZ,002431.SZ,002432.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_156 返回数据 row 行数 = "+str(stk_rewards_item_156.shape[0]))
for ts_code_sh in stk_rewards_item_156['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_156_tscode_list.append(stock_name)
stk_rewards_item_156_addname_dataframe=pd.DataFrame()
stk_rewards_item_156_addname_dataframe['cname'] = stk_rewards_item_156_tscode_list
for table_name in stk_rewards_item_156.columns.values.tolist():
    stk_rewards_item_156_addname_dataframe[table_name] = stk_rewards_item_156[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_156_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_157_tscode_list = list() 
stk_rewards_item_157 = pro.stk_rewards(ts_code='002433.SZ,002434.SZ,002435.SZ,002436.SZ,002437.SZ,002438.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_157 返回数据 row 行数 = "+str(stk_rewards_item_157.shape[0]))
for ts_code_sh in stk_rewards_item_157['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_157_tscode_list.append(stock_name)
stk_rewards_item_157_addname_dataframe=pd.DataFrame()
stk_rewards_item_157_addname_dataframe['cname'] = stk_rewards_item_157_tscode_list
for table_name in stk_rewards_item_157.columns.values.tolist():
    stk_rewards_item_157_addname_dataframe[table_name] = stk_rewards_item_157[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_157_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_158_tscode_list = list() 
stk_rewards_item_158 = pro.stk_rewards(ts_code='002439.SZ,002440.SZ,002441.SZ,002442.SZ,002443.SZ,002444.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_158 返回数据 row 行数 = "+str(stk_rewards_item_158.shape[0]))
for ts_code_sh in stk_rewards_item_158['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_158_tscode_list.append(stock_name)
stk_rewards_item_158_addname_dataframe=pd.DataFrame()
stk_rewards_item_158_addname_dataframe['cname'] = stk_rewards_item_158_tscode_list
for table_name in stk_rewards_item_158.columns.values.tolist():
    stk_rewards_item_158_addname_dataframe[table_name] = stk_rewards_item_158[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_158_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_159_tscode_list = list() 
stk_rewards_item_159 = pro.stk_rewards(ts_code='002445.SZ,002446.SZ,002447.SZ,002448.SZ,002449.SZ,002450.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_159 返回数据 row 行数 = "+str(stk_rewards_item_159.shape[0]))
for ts_code_sh in stk_rewards_item_159['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_159_tscode_list.append(stock_name)
stk_rewards_item_159_addname_dataframe=pd.DataFrame()
stk_rewards_item_159_addname_dataframe['cname'] = stk_rewards_item_159_tscode_list
for table_name in stk_rewards_item_159.columns.values.tolist():
    stk_rewards_item_159_addname_dataframe[table_name] = stk_rewards_item_159[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_159_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_160_tscode_list = list() 
stk_rewards_item_160 = pro.stk_rewards(ts_code='002451.SZ,002452.SZ,002453.SZ,002454.SZ,002455.SZ,002456.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_160 返回数据 row 行数 = "+str(stk_rewards_item_160.shape[0]))
for ts_code_sh in stk_rewards_item_160['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_160_tscode_list.append(stock_name)
stk_rewards_item_160_addname_dataframe=pd.DataFrame()
stk_rewards_item_160_addname_dataframe['cname'] = stk_rewards_item_160_tscode_list
for table_name in stk_rewards_item_160.columns.values.tolist():
    stk_rewards_item_160_addname_dataframe[table_name] = stk_rewards_item_160[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_160_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_161_tscode_list = list() 
stk_rewards_item_161 = pro.stk_rewards(ts_code='002457.SZ,002458.SZ,002459.SZ,002460.SZ,002461.SZ,002462.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_161 返回数据 row 行数 = "+str(stk_rewards_item_161.shape[0]))
for ts_code_sh in stk_rewards_item_161['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_161_tscode_list.append(stock_name)
stk_rewards_item_161_addname_dataframe=pd.DataFrame()
stk_rewards_item_161_addname_dataframe['cname'] = stk_rewards_item_161_tscode_list
for table_name in stk_rewards_item_161.columns.values.tolist():
    stk_rewards_item_161_addname_dataframe[table_name] = stk_rewards_item_161[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_161_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_162_tscode_list = list() 
stk_rewards_item_162 = pro.stk_rewards(ts_code='002463.SZ,002464.SZ,002465.SZ,002466.SZ,002467.SZ,002468.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_162 返回数据 row 行数 = "+str(stk_rewards_item_162.shape[0]))
for ts_code_sh in stk_rewards_item_162['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_162_tscode_list.append(stock_name)
stk_rewards_item_162_addname_dataframe=pd.DataFrame()
stk_rewards_item_162_addname_dataframe['cname'] = stk_rewards_item_162_tscode_list
for table_name in stk_rewards_item_162.columns.values.tolist():
    stk_rewards_item_162_addname_dataframe[table_name] = stk_rewards_item_162[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_162_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_163_tscode_list = list() 
stk_rewards_item_163 = pro.stk_rewards(ts_code='002469.SZ,002470.SZ,002471.SZ,002472.SZ,002473.SZ,002474.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_163 返回数据 row 行数 = "+str(stk_rewards_item_163.shape[0]))
for ts_code_sh in stk_rewards_item_163['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_163_tscode_list.append(stock_name)
stk_rewards_item_163_addname_dataframe=pd.DataFrame()
stk_rewards_item_163_addname_dataframe['cname'] = stk_rewards_item_163_tscode_list
for table_name in stk_rewards_item_163.columns.values.tolist():
    stk_rewards_item_163_addname_dataframe[table_name] = stk_rewards_item_163[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_163_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_164_tscode_list = list() 
stk_rewards_item_164 = pro.stk_rewards(ts_code='002475.SZ,002476.SZ,002477.SZ,002478.SZ,002479.SZ,002480.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_164 返回数据 row 行数 = "+str(stk_rewards_item_164.shape[0]))
for ts_code_sh in stk_rewards_item_164['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_164_tscode_list.append(stock_name)
stk_rewards_item_164_addname_dataframe=pd.DataFrame()
stk_rewards_item_164_addname_dataframe['cname'] = stk_rewards_item_164_tscode_list
for table_name in stk_rewards_item_164.columns.values.tolist():
    stk_rewards_item_164_addname_dataframe[table_name] = stk_rewards_item_164[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_164_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_165_tscode_list = list() 
stk_rewards_item_165 = pro.stk_rewards(ts_code='002481.SZ,002482.SZ,002483.SZ,002484.SZ,002485.SZ,002486.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_165 返回数据 row 行数 = "+str(stk_rewards_item_165.shape[0]))
for ts_code_sh in stk_rewards_item_165['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_165_tscode_list.append(stock_name)
stk_rewards_item_165_addname_dataframe=pd.DataFrame()
stk_rewards_item_165_addname_dataframe['cname'] = stk_rewards_item_165_tscode_list
for table_name in stk_rewards_item_165.columns.values.tolist():
    stk_rewards_item_165_addname_dataframe[table_name] = stk_rewards_item_165[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_165_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_166_tscode_list = list() 
stk_rewards_item_166 = pro.stk_rewards(ts_code='002487.SZ,002488.SZ,002489.SZ,002490.SZ,002491.SZ,002492.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_166 返回数据 row 行数 = "+str(stk_rewards_item_166.shape[0]))
for ts_code_sh in stk_rewards_item_166['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_166_tscode_list.append(stock_name)
stk_rewards_item_166_addname_dataframe=pd.DataFrame()
stk_rewards_item_166_addname_dataframe['cname'] = stk_rewards_item_166_tscode_list
for table_name in stk_rewards_item_166.columns.values.tolist():
    stk_rewards_item_166_addname_dataframe[table_name] = stk_rewards_item_166[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_166_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_167_tscode_list = list() 
stk_rewards_item_167 = pro.stk_rewards(ts_code='002493.SZ,002494.SZ,002495.SZ,002496.SZ,002497.SZ,002498.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_167 返回数据 row 行数 = "+str(stk_rewards_item_167.shape[0]))
for ts_code_sh in stk_rewards_item_167['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_167_tscode_list.append(stock_name)
stk_rewards_item_167_addname_dataframe=pd.DataFrame()
stk_rewards_item_167_addname_dataframe['cname'] = stk_rewards_item_167_tscode_list
for table_name in stk_rewards_item_167.columns.values.tolist():
    stk_rewards_item_167_addname_dataframe[table_name] = stk_rewards_item_167[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_167_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_168_tscode_list = list() 
stk_rewards_item_168 = pro.stk_rewards(ts_code='002499.SZ,002500.SZ,002501.SZ,002502.SZ,002503.SZ,002504.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_168 返回数据 row 行数 = "+str(stk_rewards_item_168.shape[0]))
for ts_code_sh in stk_rewards_item_168['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_168_tscode_list.append(stock_name)
stk_rewards_item_168_addname_dataframe=pd.DataFrame()
stk_rewards_item_168_addname_dataframe['cname'] = stk_rewards_item_168_tscode_list
for table_name in stk_rewards_item_168.columns.values.tolist():
    stk_rewards_item_168_addname_dataframe[table_name] = stk_rewards_item_168[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_168_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_169_tscode_list = list() 
stk_rewards_item_169 = pro.stk_rewards(ts_code='002505.SZ,002506.SZ,002507.SZ,002508.SZ,002509.SZ,002510.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_169 返回数据 row 行数 = "+str(stk_rewards_item_169.shape[0]))
for ts_code_sh in stk_rewards_item_169['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_169_tscode_list.append(stock_name)
stk_rewards_item_169_addname_dataframe=pd.DataFrame()
stk_rewards_item_169_addname_dataframe['cname'] = stk_rewards_item_169_tscode_list
for table_name in stk_rewards_item_169.columns.values.tolist():
    stk_rewards_item_169_addname_dataframe[table_name] = stk_rewards_item_169[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_169_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_170_tscode_list = list() 
stk_rewards_item_170 = pro.stk_rewards(ts_code='002511.SZ,002512.SZ,002513.SZ,002514.SZ,002515.SZ,002516.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_170 返回数据 row 行数 = "+str(stk_rewards_item_170.shape[0]))
for ts_code_sh in stk_rewards_item_170['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_170_tscode_list.append(stock_name)
stk_rewards_item_170_addname_dataframe=pd.DataFrame()
stk_rewards_item_170_addname_dataframe['cname'] = stk_rewards_item_170_tscode_list
for table_name in stk_rewards_item_170.columns.values.tolist():
    stk_rewards_item_170_addname_dataframe[table_name] = stk_rewards_item_170[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_170_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_171_tscode_list = list() 
stk_rewards_item_171 = pro.stk_rewards(ts_code='002517.SZ,002518.SZ,002519.SZ,002520.SZ,002521.SZ,002522.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_171 返回数据 row 行数 = "+str(stk_rewards_item_171.shape[0]))
for ts_code_sh in stk_rewards_item_171['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_171_tscode_list.append(stock_name)
stk_rewards_item_171_addname_dataframe=pd.DataFrame()
stk_rewards_item_171_addname_dataframe['cname'] = stk_rewards_item_171_tscode_list
for table_name in stk_rewards_item_171.columns.values.tolist():
    stk_rewards_item_171_addname_dataframe[table_name] = stk_rewards_item_171[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_171_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_172_tscode_list = list() 
stk_rewards_item_172 = pro.stk_rewards(ts_code='002523.SZ,002524.SZ,002526.SZ,002527.SZ,002528.SZ,002529.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_172 返回数据 row 行数 = "+str(stk_rewards_item_172.shape[0]))
for ts_code_sh in stk_rewards_item_172['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_172_tscode_list.append(stock_name)
stk_rewards_item_172_addname_dataframe=pd.DataFrame()
stk_rewards_item_172_addname_dataframe['cname'] = stk_rewards_item_172_tscode_list
for table_name in stk_rewards_item_172.columns.values.tolist():
    stk_rewards_item_172_addname_dataframe[table_name] = stk_rewards_item_172[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_172_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_173_tscode_list = list() 
stk_rewards_item_173 = pro.stk_rewards(ts_code='002530.SZ,002531.SZ,002532.SZ,002533.SZ,002534.SZ,002535.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_173 返回数据 row 行数 = "+str(stk_rewards_item_173.shape[0]))
for ts_code_sh in stk_rewards_item_173['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_173_tscode_list.append(stock_name)
stk_rewards_item_173_addname_dataframe=pd.DataFrame()
stk_rewards_item_173_addname_dataframe['cname'] = stk_rewards_item_173_tscode_list
for table_name in stk_rewards_item_173.columns.values.tolist():
    stk_rewards_item_173_addname_dataframe[table_name] = stk_rewards_item_173[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_173_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_174_tscode_list = list() 
stk_rewards_item_174 = pro.stk_rewards(ts_code='002536.SZ,002537.SZ,002538.SZ,002539.SZ,002540.SZ,002541.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_174 返回数据 row 行数 = "+str(stk_rewards_item_174.shape[0]))
for ts_code_sh in stk_rewards_item_174['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_174_tscode_list.append(stock_name)
stk_rewards_item_174_addname_dataframe=pd.DataFrame()
stk_rewards_item_174_addname_dataframe['cname'] = stk_rewards_item_174_tscode_list
for table_name in stk_rewards_item_174.columns.values.tolist():
    stk_rewards_item_174_addname_dataframe[table_name] = stk_rewards_item_174[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_174_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_175_tscode_list = list() 
stk_rewards_item_175 = pro.stk_rewards(ts_code='002542.SZ,002543.SZ,002544.SZ,002545.SZ,002546.SZ,002547.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_175 返回数据 row 行数 = "+str(stk_rewards_item_175.shape[0]))
for ts_code_sh in stk_rewards_item_175['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_175_tscode_list.append(stock_name)
stk_rewards_item_175_addname_dataframe=pd.DataFrame()
stk_rewards_item_175_addname_dataframe['cname'] = stk_rewards_item_175_tscode_list
for table_name in stk_rewards_item_175.columns.values.tolist():
    stk_rewards_item_175_addname_dataframe[table_name] = stk_rewards_item_175[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_175_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_176_tscode_list = list() 
stk_rewards_item_176 = pro.stk_rewards(ts_code='002548.SZ,002549.SZ,002550.SZ,002551.SZ,002552.SZ,002553.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_176 返回数据 row 行数 = "+str(stk_rewards_item_176.shape[0]))
for ts_code_sh in stk_rewards_item_176['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_176_tscode_list.append(stock_name)
stk_rewards_item_176_addname_dataframe=pd.DataFrame()
stk_rewards_item_176_addname_dataframe['cname'] = stk_rewards_item_176_tscode_list
for table_name in stk_rewards_item_176.columns.values.tolist():
    stk_rewards_item_176_addname_dataframe[table_name] = stk_rewards_item_176[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_176_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_177_tscode_list = list() 
stk_rewards_item_177 = pro.stk_rewards(ts_code='002554.SZ,002555.SZ,002556.SZ,002557.SZ,002558.SZ,002559.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_177 返回数据 row 行数 = "+str(stk_rewards_item_177.shape[0]))
for ts_code_sh in stk_rewards_item_177['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_177_tscode_list.append(stock_name)
stk_rewards_item_177_addname_dataframe=pd.DataFrame()
stk_rewards_item_177_addname_dataframe['cname'] = stk_rewards_item_177_tscode_list
for table_name in stk_rewards_item_177.columns.values.tolist():
    stk_rewards_item_177_addname_dataframe[table_name] = stk_rewards_item_177[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_177_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_178_tscode_list = list() 
stk_rewards_item_178 = pro.stk_rewards(ts_code='002560.SZ,002561.SZ,002562.SZ,002563.SZ,002564.SZ,002565.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_178 返回数据 row 行数 = "+str(stk_rewards_item_178.shape[0]))
for ts_code_sh in stk_rewards_item_178['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_178_tscode_list.append(stock_name)
stk_rewards_item_178_addname_dataframe=pd.DataFrame()
stk_rewards_item_178_addname_dataframe['cname'] = stk_rewards_item_178_tscode_list
for table_name in stk_rewards_item_178.columns.values.tolist():
    stk_rewards_item_178_addname_dataframe[table_name] = stk_rewards_item_178[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_178_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_179_tscode_list = list() 
stk_rewards_item_179 = pro.stk_rewards(ts_code='002566.SZ,002567.SZ,002568.SZ,002569.SZ,002570.SZ,002571.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_179 返回数据 row 行数 = "+str(stk_rewards_item_179.shape[0]))
for ts_code_sh in stk_rewards_item_179['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_179_tscode_list.append(stock_name)
stk_rewards_item_179_addname_dataframe=pd.DataFrame()
stk_rewards_item_179_addname_dataframe['cname'] = stk_rewards_item_179_tscode_list
for table_name in stk_rewards_item_179.columns.values.tolist():
    stk_rewards_item_179_addname_dataframe[table_name] = stk_rewards_item_179[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_179_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_180_tscode_list = list() 
stk_rewards_item_180 = pro.stk_rewards(ts_code='002572.SZ,002573.SZ,002574.SZ,002575.SZ,002576.SZ,002577.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_180 返回数据 row 行数 = "+str(stk_rewards_item_180.shape[0]))
for ts_code_sh in stk_rewards_item_180['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_180_tscode_list.append(stock_name)
stk_rewards_item_180_addname_dataframe=pd.DataFrame()
stk_rewards_item_180_addname_dataframe['cname'] = stk_rewards_item_180_tscode_list
for table_name in stk_rewards_item_180.columns.values.tolist():
    stk_rewards_item_180_addname_dataframe[table_name] = stk_rewards_item_180[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_180_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_181_tscode_list = list() 
stk_rewards_item_181 = pro.stk_rewards(ts_code='002578.SZ,002579.SZ,002580.SZ,002581.SZ,002582.SZ,002583.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_181 返回数据 row 行数 = "+str(stk_rewards_item_181.shape[0]))
for ts_code_sh in stk_rewards_item_181['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_181_tscode_list.append(stock_name)
stk_rewards_item_181_addname_dataframe=pd.DataFrame()
stk_rewards_item_181_addname_dataframe['cname'] = stk_rewards_item_181_tscode_list
for table_name in stk_rewards_item_181.columns.values.tolist():
    stk_rewards_item_181_addname_dataframe[table_name] = stk_rewards_item_181[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_181_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_182_tscode_list = list() 
stk_rewards_item_182 = pro.stk_rewards(ts_code='002584.SZ,002585.SZ,002586.SZ,002587.SZ,002588.SZ,002589.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_182 返回数据 row 行数 = "+str(stk_rewards_item_182.shape[0]))
for ts_code_sh in stk_rewards_item_182['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_182_tscode_list.append(stock_name)
stk_rewards_item_182_addname_dataframe=pd.DataFrame()
stk_rewards_item_182_addname_dataframe['cname'] = stk_rewards_item_182_tscode_list
for table_name in stk_rewards_item_182.columns.values.tolist():
    stk_rewards_item_182_addname_dataframe[table_name] = stk_rewards_item_182[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_182_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_183_tscode_list = list() 
stk_rewards_item_183 = pro.stk_rewards(ts_code='002590.SZ,002591.SZ,002592.SZ,002593.SZ,002594.SZ,002595.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_183 返回数据 row 行数 = "+str(stk_rewards_item_183.shape[0]))
for ts_code_sh in stk_rewards_item_183['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_183_tscode_list.append(stock_name)
stk_rewards_item_183_addname_dataframe=pd.DataFrame()
stk_rewards_item_183_addname_dataframe['cname'] = stk_rewards_item_183_tscode_list
for table_name in stk_rewards_item_183.columns.values.tolist():
    stk_rewards_item_183_addname_dataframe[table_name] = stk_rewards_item_183[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_183_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_184_tscode_list = list() 
stk_rewards_item_184 = pro.stk_rewards(ts_code='002596.SZ,002597.SZ,002598.SZ,002599.SZ,002600.SZ,002601.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_184 返回数据 row 行数 = "+str(stk_rewards_item_184.shape[0]))
for ts_code_sh in stk_rewards_item_184['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_184_tscode_list.append(stock_name)
stk_rewards_item_184_addname_dataframe=pd.DataFrame()
stk_rewards_item_184_addname_dataframe['cname'] = stk_rewards_item_184_tscode_list
for table_name in stk_rewards_item_184.columns.values.tolist():
    stk_rewards_item_184_addname_dataframe[table_name] = stk_rewards_item_184[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_184_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_185_tscode_list = list() 
stk_rewards_item_185 = pro.stk_rewards(ts_code='002602.SZ,002603.SZ,002604.SZ,002605.SZ,002606.SZ,002607.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_185 返回数据 row 行数 = "+str(stk_rewards_item_185.shape[0]))
for ts_code_sh in stk_rewards_item_185['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_185_tscode_list.append(stock_name)
stk_rewards_item_185_addname_dataframe=pd.DataFrame()
stk_rewards_item_185_addname_dataframe['cname'] = stk_rewards_item_185_tscode_list
for table_name in stk_rewards_item_185.columns.values.tolist():
    stk_rewards_item_185_addname_dataframe[table_name] = stk_rewards_item_185[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_185_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_186_tscode_list = list() 
stk_rewards_item_186 = pro.stk_rewards(ts_code='002608.SZ,002609.SZ,002610.SZ,002611.SZ,002612.SZ,002613.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_186 返回数据 row 行数 = "+str(stk_rewards_item_186.shape[0]))
for ts_code_sh in stk_rewards_item_186['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_186_tscode_list.append(stock_name)
stk_rewards_item_186_addname_dataframe=pd.DataFrame()
stk_rewards_item_186_addname_dataframe['cname'] = stk_rewards_item_186_tscode_list
for table_name in stk_rewards_item_186.columns.values.tolist():
    stk_rewards_item_186_addname_dataframe[table_name] = stk_rewards_item_186[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_186_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_187_tscode_list = list() 
stk_rewards_item_187 = pro.stk_rewards(ts_code='002614.SZ,002615.SZ,002616.SZ,002617.SZ,002618.SZ,002619.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_187 返回数据 row 行数 = "+str(stk_rewards_item_187.shape[0]))
for ts_code_sh in stk_rewards_item_187['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_187_tscode_list.append(stock_name)
stk_rewards_item_187_addname_dataframe=pd.DataFrame()
stk_rewards_item_187_addname_dataframe['cname'] = stk_rewards_item_187_tscode_list
for table_name in stk_rewards_item_187.columns.values.tolist():
    stk_rewards_item_187_addname_dataframe[table_name] = stk_rewards_item_187[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_187_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_188_tscode_list = list() 
stk_rewards_item_188 = pro.stk_rewards(ts_code='002620.SZ,002621.SZ,002622.SZ,002623.SZ,002624.SZ,002625.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_188 返回数据 row 行数 = "+str(stk_rewards_item_188.shape[0]))
for ts_code_sh in stk_rewards_item_188['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_188_tscode_list.append(stock_name)
stk_rewards_item_188_addname_dataframe=pd.DataFrame()
stk_rewards_item_188_addname_dataframe['cname'] = stk_rewards_item_188_tscode_list
for table_name in stk_rewards_item_188.columns.values.tolist():
    stk_rewards_item_188_addname_dataframe[table_name] = stk_rewards_item_188[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_188_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_189_tscode_list = list() 
stk_rewards_item_189 = pro.stk_rewards(ts_code='002626.SZ,002627.SZ,002628.SZ,002629.SZ,002630.SZ,002631.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_189 返回数据 row 行数 = "+str(stk_rewards_item_189.shape[0]))
for ts_code_sh in stk_rewards_item_189['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_189_tscode_list.append(stock_name)
stk_rewards_item_189_addname_dataframe=pd.DataFrame()
stk_rewards_item_189_addname_dataframe['cname'] = stk_rewards_item_189_tscode_list
for table_name in stk_rewards_item_189.columns.values.tolist():
    stk_rewards_item_189_addname_dataframe[table_name] = stk_rewards_item_189[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_189_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_190_tscode_list = list() 
stk_rewards_item_190 = pro.stk_rewards(ts_code='002632.SZ,002633.SZ,002634.SZ,002635.SZ,002636.SZ,002637.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_190 返回数据 row 行数 = "+str(stk_rewards_item_190.shape[0]))
for ts_code_sh in stk_rewards_item_190['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_190_tscode_list.append(stock_name)
stk_rewards_item_190_addname_dataframe=pd.DataFrame()
stk_rewards_item_190_addname_dataframe['cname'] = stk_rewards_item_190_tscode_list
for table_name in stk_rewards_item_190.columns.values.tolist():
    stk_rewards_item_190_addname_dataframe[table_name] = stk_rewards_item_190[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_190_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_191_tscode_list = list() 
stk_rewards_item_191 = pro.stk_rewards(ts_code='002638.SZ,002639.SZ,002640.SZ,002641.SZ,002642.SZ,002643.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_191 返回数据 row 行数 = "+str(stk_rewards_item_191.shape[0]))
for ts_code_sh in stk_rewards_item_191['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_191_tscode_list.append(stock_name)
stk_rewards_item_191_addname_dataframe=pd.DataFrame()
stk_rewards_item_191_addname_dataframe['cname'] = stk_rewards_item_191_tscode_list
for table_name in stk_rewards_item_191.columns.values.tolist():
    stk_rewards_item_191_addname_dataframe[table_name] = stk_rewards_item_191[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_191_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_192_tscode_list = list() 
stk_rewards_item_192 = pro.stk_rewards(ts_code='002644.SZ,002645.SZ,002646.SZ,002647.SZ,002648.SZ,002649.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_192 返回数据 row 行数 = "+str(stk_rewards_item_192.shape[0]))
for ts_code_sh in stk_rewards_item_192['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_192_tscode_list.append(stock_name)
stk_rewards_item_192_addname_dataframe=pd.DataFrame()
stk_rewards_item_192_addname_dataframe['cname'] = stk_rewards_item_192_tscode_list
for table_name in stk_rewards_item_192.columns.values.tolist():
    stk_rewards_item_192_addname_dataframe[table_name] = stk_rewards_item_192[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_192_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_193_tscode_list = list() 
stk_rewards_item_193 = pro.stk_rewards(ts_code='002650.SZ,002651.SZ,002652.SZ,002653.SZ,002654.SZ,002655.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_193 返回数据 row 行数 = "+str(stk_rewards_item_193.shape[0]))
for ts_code_sh in stk_rewards_item_193['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_193_tscode_list.append(stock_name)
stk_rewards_item_193_addname_dataframe=pd.DataFrame()
stk_rewards_item_193_addname_dataframe['cname'] = stk_rewards_item_193_tscode_list
for table_name in stk_rewards_item_193.columns.values.tolist():
    stk_rewards_item_193_addname_dataframe[table_name] = stk_rewards_item_193[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_193_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_194_tscode_list = list() 
stk_rewards_item_194 = pro.stk_rewards(ts_code='002656.SZ,002657.SZ,002658.SZ,002659.SZ,002660.SZ,002661.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_194 返回数据 row 行数 = "+str(stk_rewards_item_194.shape[0]))
for ts_code_sh in stk_rewards_item_194['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_194_tscode_list.append(stock_name)
stk_rewards_item_194_addname_dataframe=pd.DataFrame()
stk_rewards_item_194_addname_dataframe['cname'] = stk_rewards_item_194_tscode_list
for table_name in stk_rewards_item_194.columns.values.tolist():
    stk_rewards_item_194_addname_dataframe[table_name] = stk_rewards_item_194[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_194_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_195_tscode_list = list() 
stk_rewards_item_195 = pro.stk_rewards(ts_code='002662.SZ,002663.SZ,002664.SZ,002665.SZ,002666.SZ,002667.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_195 返回数据 row 行数 = "+str(stk_rewards_item_195.shape[0]))
for ts_code_sh in stk_rewards_item_195['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_195_tscode_list.append(stock_name)
stk_rewards_item_195_addname_dataframe=pd.DataFrame()
stk_rewards_item_195_addname_dataframe['cname'] = stk_rewards_item_195_tscode_list
for table_name in stk_rewards_item_195.columns.values.tolist():
    stk_rewards_item_195_addname_dataframe[table_name] = stk_rewards_item_195[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_195_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_196_tscode_list = list() 
stk_rewards_item_196 = pro.stk_rewards(ts_code='002668.SZ,002669.SZ,002670.SZ,002671.SZ,002672.SZ,002673.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_196 返回数据 row 行数 = "+str(stk_rewards_item_196.shape[0]))
for ts_code_sh in stk_rewards_item_196['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_196_tscode_list.append(stock_name)
stk_rewards_item_196_addname_dataframe=pd.DataFrame()
stk_rewards_item_196_addname_dataframe['cname'] = stk_rewards_item_196_tscode_list
for table_name in stk_rewards_item_196.columns.values.tolist():
    stk_rewards_item_196_addname_dataframe[table_name] = stk_rewards_item_196[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_196_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_197_tscode_list = list() 
stk_rewards_item_197 = pro.stk_rewards(ts_code='002674.SZ,002675.SZ,002676.SZ,002677.SZ,002678.SZ,002679.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_197 返回数据 row 行数 = "+str(stk_rewards_item_197.shape[0]))
for ts_code_sh in stk_rewards_item_197['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_197_tscode_list.append(stock_name)
stk_rewards_item_197_addname_dataframe=pd.DataFrame()
stk_rewards_item_197_addname_dataframe['cname'] = stk_rewards_item_197_tscode_list
for table_name in stk_rewards_item_197.columns.values.tolist():
    stk_rewards_item_197_addname_dataframe[table_name] = stk_rewards_item_197[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_197_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_198_tscode_list = list() 
stk_rewards_item_198 = pro.stk_rewards(ts_code='002680.SZ,002681.SZ,002682.SZ,002683.SZ,002684.SZ,002685.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_198 返回数据 row 行数 = "+str(stk_rewards_item_198.shape[0]))
for ts_code_sh in stk_rewards_item_198['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_198_tscode_list.append(stock_name)
stk_rewards_item_198_addname_dataframe=pd.DataFrame()
stk_rewards_item_198_addname_dataframe['cname'] = stk_rewards_item_198_tscode_list
for table_name in stk_rewards_item_198.columns.values.tolist():
    stk_rewards_item_198_addname_dataframe[table_name] = stk_rewards_item_198[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_198_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_199_tscode_list = list() 
stk_rewards_item_199 = pro.stk_rewards(ts_code='002686.SZ,002687.SZ,002688.SZ,002689.SZ,002690.SZ,002691.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_199 返回数据 row 行数 = "+str(stk_rewards_item_199.shape[0]))
for ts_code_sh in stk_rewards_item_199['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_199_tscode_list.append(stock_name)
stk_rewards_item_199_addname_dataframe=pd.DataFrame()
stk_rewards_item_199_addname_dataframe['cname'] = stk_rewards_item_199_tscode_list
for table_name in stk_rewards_item_199.columns.values.tolist():
    stk_rewards_item_199_addname_dataframe[table_name] = stk_rewards_item_199[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_199_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_200_tscode_list = list() 
stk_rewards_item_200 = pro.stk_rewards(ts_code='002692.SZ,002693.SZ,002694.SZ,002695.SZ,002696.SZ,002697.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_200 返回数据 row 行数 = "+str(stk_rewards_item_200.shape[0]))
for ts_code_sh in stk_rewards_item_200['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_200_tscode_list.append(stock_name)
stk_rewards_item_200_addname_dataframe=pd.DataFrame()
stk_rewards_item_200_addname_dataframe['cname'] = stk_rewards_item_200_tscode_list
for table_name in stk_rewards_item_200.columns.values.tolist():
    stk_rewards_item_200_addname_dataframe[table_name] = stk_rewards_item_200[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_200_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_201_tscode_list = list() 
stk_rewards_item_201 = pro.stk_rewards(ts_code='002698.SZ,002699.SZ,002700.SZ,002701.SZ,002702.SZ,002703.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_201 返回数据 row 行数 = "+str(stk_rewards_item_201.shape[0]))
for ts_code_sh in stk_rewards_item_201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_201_tscode_list.append(stock_name)
stk_rewards_item_201_addname_dataframe=pd.DataFrame()
stk_rewards_item_201_addname_dataframe['cname'] = stk_rewards_item_201_tscode_list
for table_name in stk_rewards_item_201.columns.values.tolist():
    stk_rewards_item_201_addname_dataframe[table_name] = stk_rewards_item_201[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_201_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_202_tscode_list = list() 
stk_rewards_item_202 = pro.stk_rewards(ts_code='002705.SZ,002706.SZ,002707.SZ,002708.SZ,002709.SZ,002711.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_202 返回数据 row 行数 = "+str(stk_rewards_item_202.shape[0]))
for ts_code_sh in stk_rewards_item_202['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_202_tscode_list.append(stock_name)
stk_rewards_item_202_addname_dataframe=pd.DataFrame()
stk_rewards_item_202_addname_dataframe['cname'] = stk_rewards_item_202_tscode_list
for table_name in stk_rewards_item_202.columns.values.tolist():
    stk_rewards_item_202_addname_dataframe[table_name] = stk_rewards_item_202[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_202_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_203_tscode_list = list() 
stk_rewards_item_203 = pro.stk_rewards(ts_code='002712.SZ,002713.SZ,002714.SZ,002715.SZ,002716.SZ,002717.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_203 返回数据 row 行数 = "+str(stk_rewards_item_203.shape[0]))
for ts_code_sh in stk_rewards_item_203['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_203_tscode_list.append(stock_name)
stk_rewards_item_203_addname_dataframe=pd.DataFrame()
stk_rewards_item_203_addname_dataframe['cname'] = stk_rewards_item_203_tscode_list
for table_name in stk_rewards_item_203.columns.values.tolist():
    stk_rewards_item_203_addname_dataframe[table_name] = stk_rewards_item_203[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_203_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_204_tscode_list = list() 
stk_rewards_item_204 = pro.stk_rewards(ts_code='002718.SZ,002719.SZ,002721.SZ,002722.SZ,002723.SZ,002724.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_204 返回数据 row 行数 = "+str(stk_rewards_item_204.shape[0]))
for ts_code_sh in stk_rewards_item_204['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_204_tscode_list.append(stock_name)
stk_rewards_item_204_addname_dataframe=pd.DataFrame()
stk_rewards_item_204_addname_dataframe['cname'] = stk_rewards_item_204_tscode_list
for table_name in stk_rewards_item_204.columns.values.tolist():
    stk_rewards_item_204_addname_dataframe[table_name] = stk_rewards_item_204[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_204_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_205_tscode_list = list() 
stk_rewards_item_205 = pro.stk_rewards(ts_code='002725.SZ,002726.SZ,002727.SZ,002728.SZ,002729.SZ,002730.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_205 返回数据 row 行数 = "+str(stk_rewards_item_205.shape[0]))
for ts_code_sh in stk_rewards_item_205['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_205_tscode_list.append(stock_name)
stk_rewards_item_205_addname_dataframe=pd.DataFrame()
stk_rewards_item_205_addname_dataframe['cname'] = stk_rewards_item_205_tscode_list
for table_name in stk_rewards_item_205.columns.values.tolist():
    stk_rewards_item_205_addname_dataframe[table_name] = stk_rewards_item_205[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_205_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_206_tscode_list = list() 
stk_rewards_item_206 = pro.stk_rewards(ts_code='002731.SZ,002732.SZ,002733.SZ,002734.SZ,002735.SZ,002736.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_206 返回数据 row 行数 = "+str(stk_rewards_item_206.shape[0]))
for ts_code_sh in stk_rewards_item_206['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_206_tscode_list.append(stock_name)
stk_rewards_item_206_addname_dataframe=pd.DataFrame()
stk_rewards_item_206_addname_dataframe['cname'] = stk_rewards_item_206_tscode_list
for table_name in stk_rewards_item_206.columns.values.tolist():
    stk_rewards_item_206_addname_dataframe[table_name] = stk_rewards_item_206[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_206_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_207_tscode_list = list() 
stk_rewards_item_207 = pro.stk_rewards(ts_code='002737.SZ,002738.SZ,002739.SZ,002740.SZ,002741.SZ,002742.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_207 返回数据 row 行数 = "+str(stk_rewards_item_207.shape[0]))
for ts_code_sh in stk_rewards_item_207['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_207_tscode_list.append(stock_name)
stk_rewards_item_207_addname_dataframe=pd.DataFrame()
stk_rewards_item_207_addname_dataframe['cname'] = stk_rewards_item_207_tscode_list
for table_name in stk_rewards_item_207.columns.values.tolist():
    stk_rewards_item_207_addname_dataframe[table_name] = stk_rewards_item_207[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_207_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_208_tscode_list = list() 
stk_rewards_item_208 = pro.stk_rewards(ts_code='002743.SZ,002745.SZ,002746.SZ,002747.SZ,002748.SZ,002749.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_208 返回数据 row 行数 = "+str(stk_rewards_item_208.shape[0]))
for ts_code_sh in stk_rewards_item_208['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_208_tscode_list.append(stock_name)
stk_rewards_item_208_addname_dataframe=pd.DataFrame()
stk_rewards_item_208_addname_dataframe['cname'] = stk_rewards_item_208_tscode_list
for table_name in stk_rewards_item_208.columns.values.tolist():
    stk_rewards_item_208_addname_dataframe[table_name] = stk_rewards_item_208[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_208_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_209_tscode_list = list() 
stk_rewards_item_209 = pro.stk_rewards(ts_code='002750.SZ,002751.SZ,002752.SZ,002753.SZ,002755.SZ,002756.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_209 返回数据 row 行数 = "+str(stk_rewards_item_209.shape[0]))
for ts_code_sh in stk_rewards_item_209['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_209_tscode_list.append(stock_name)
stk_rewards_item_209_addname_dataframe=pd.DataFrame()
stk_rewards_item_209_addname_dataframe['cname'] = stk_rewards_item_209_tscode_list
for table_name in stk_rewards_item_209.columns.values.tolist():
    stk_rewards_item_209_addname_dataframe[table_name] = stk_rewards_item_209[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_209_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_210_tscode_list = list() 
stk_rewards_item_210 = pro.stk_rewards(ts_code='002757.SZ,002758.SZ,002759.SZ,002760.SZ,002761.SZ,002762.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_210 返回数据 row 行数 = "+str(stk_rewards_item_210.shape[0]))
for ts_code_sh in stk_rewards_item_210['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_210_tscode_list.append(stock_name)
stk_rewards_item_210_addname_dataframe=pd.DataFrame()
stk_rewards_item_210_addname_dataframe['cname'] = stk_rewards_item_210_tscode_list
for table_name in stk_rewards_item_210.columns.values.tolist():
    stk_rewards_item_210_addname_dataframe[table_name] = stk_rewards_item_210[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_210_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_211_tscode_list = list() 
stk_rewards_item_211 = pro.stk_rewards(ts_code='002763.SZ,002765.SZ,002766.SZ,002767.SZ,002768.SZ,002769.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_211 返回数据 row 行数 = "+str(stk_rewards_item_211.shape[0]))
for ts_code_sh in stk_rewards_item_211['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_211_tscode_list.append(stock_name)
stk_rewards_item_211_addname_dataframe=pd.DataFrame()
stk_rewards_item_211_addname_dataframe['cname'] = stk_rewards_item_211_tscode_list
for table_name in stk_rewards_item_211.columns.values.tolist():
    stk_rewards_item_211_addname_dataframe[table_name] = stk_rewards_item_211[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_211_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_212_tscode_list = list() 
stk_rewards_item_212 = pro.stk_rewards(ts_code='002770.SZ,002771.SZ,002772.SZ,002773.SZ,002774.SZ,002775.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_212 返回数据 row 行数 = "+str(stk_rewards_item_212.shape[0]))
for ts_code_sh in stk_rewards_item_212['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_212_tscode_list.append(stock_name)
stk_rewards_item_212_addname_dataframe=pd.DataFrame()
stk_rewards_item_212_addname_dataframe['cname'] = stk_rewards_item_212_tscode_list
for table_name in stk_rewards_item_212.columns.values.tolist():
    stk_rewards_item_212_addname_dataframe[table_name] = stk_rewards_item_212[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_212_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_213_tscode_list = list() 
stk_rewards_item_213 = pro.stk_rewards(ts_code='002776.SZ,002777.SZ,002778.SZ,002779.SZ,002780.SZ,002781.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_213 返回数据 row 行数 = "+str(stk_rewards_item_213.shape[0]))
for ts_code_sh in stk_rewards_item_213['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_213_tscode_list.append(stock_name)
stk_rewards_item_213_addname_dataframe=pd.DataFrame()
stk_rewards_item_213_addname_dataframe['cname'] = stk_rewards_item_213_tscode_list
for table_name in stk_rewards_item_213.columns.values.tolist():
    stk_rewards_item_213_addname_dataframe[table_name] = stk_rewards_item_213[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_213_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_214_tscode_list = list() 
stk_rewards_item_214 = pro.stk_rewards(ts_code='002782.SZ,002783.SZ,002785.SZ,002786.SZ,002787.SZ,002788.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_214 返回数据 row 行数 = "+str(stk_rewards_item_214.shape[0]))
for ts_code_sh in stk_rewards_item_214['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_214_tscode_list.append(stock_name)
stk_rewards_item_214_addname_dataframe=pd.DataFrame()
stk_rewards_item_214_addname_dataframe['cname'] = stk_rewards_item_214_tscode_list
for table_name in stk_rewards_item_214.columns.values.tolist():
    stk_rewards_item_214_addname_dataframe[table_name] = stk_rewards_item_214[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_214_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_215_tscode_list = list() 
stk_rewards_item_215 = pro.stk_rewards(ts_code='002789.SZ,002790.SZ,002791.SZ,002792.SZ,002793.SZ,002795.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_215 返回数据 row 行数 = "+str(stk_rewards_item_215.shape[0]))
for ts_code_sh in stk_rewards_item_215['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_215_tscode_list.append(stock_name)
stk_rewards_item_215_addname_dataframe=pd.DataFrame()
stk_rewards_item_215_addname_dataframe['cname'] = stk_rewards_item_215_tscode_list
for table_name in stk_rewards_item_215.columns.values.tolist():
    stk_rewards_item_215_addname_dataframe[table_name] = stk_rewards_item_215[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_215_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_216_tscode_list = list() 
stk_rewards_item_216 = pro.stk_rewards(ts_code='002796.SZ,002797.SZ,002798.SZ,002799.SZ,002800.SZ,002801.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_216 返回数据 row 行数 = "+str(stk_rewards_item_216.shape[0]))
for ts_code_sh in stk_rewards_item_216['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_216_tscode_list.append(stock_name)
stk_rewards_item_216_addname_dataframe=pd.DataFrame()
stk_rewards_item_216_addname_dataframe['cname'] = stk_rewards_item_216_tscode_list
for table_name in stk_rewards_item_216.columns.values.tolist():
    stk_rewards_item_216_addname_dataframe[table_name] = stk_rewards_item_216[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_216_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_217_tscode_list = list() 
stk_rewards_item_217 = pro.stk_rewards(ts_code='002802.SZ,002803.SZ,002805.SZ,002806.SZ,002807.SZ,002808.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_217 返回数据 row 行数 = "+str(stk_rewards_item_217.shape[0]))
for ts_code_sh in stk_rewards_item_217['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_217_tscode_list.append(stock_name)
stk_rewards_item_217_addname_dataframe=pd.DataFrame()
stk_rewards_item_217_addname_dataframe['cname'] = stk_rewards_item_217_tscode_list
for table_name in stk_rewards_item_217.columns.values.tolist():
    stk_rewards_item_217_addname_dataframe[table_name] = stk_rewards_item_217[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_217_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_218_tscode_list = list() 
stk_rewards_item_218 = pro.stk_rewards(ts_code='002809.SZ,002810.SZ,002811.SZ,002812.SZ,002813.SZ,002815.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_218 返回数据 row 行数 = "+str(stk_rewards_item_218.shape[0]))
for ts_code_sh in stk_rewards_item_218['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_218_tscode_list.append(stock_name)
stk_rewards_item_218_addname_dataframe=pd.DataFrame()
stk_rewards_item_218_addname_dataframe['cname'] = stk_rewards_item_218_tscode_list
for table_name in stk_rewards_item_218.columns.values.tolist():
    stk_rewards_item_218_addname_dataframe[table_name] = stk_rewards_item_218[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_218_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_219_tscode_list = list() 
stk_rewards_item_219 = pro.stk_rewards(ts_code='002816.SZ,002817.SZ,002818.SZ,002819.SZ,002820.SZ,002821.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_219 返回数据 row 行数 = "+str(stk_rewards_item_219.shape[0]))
for ts_code_sh in stk_rewards_item_219['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_219_tscode_list.append(stock_name)
stk_rewards_item_219_addname_dataframe=pd.DataFrame()
stk_rewards_item_219_addname_dataframe['cname'] = stk_rewards_item_219_tscode_list
for table_name in stk_rewards_item_219.columns.values.tolist():
    stk_rewards_item_219_addname_dataframe[table_name] = stk_rewards_item_219[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_219_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_220_tscode_list = list() 
stk_rewards_item_220 = pro.stk_rewards(ts_code='002822.SZ,002823.SZ,002824.SZ,002825.SZ,002826.SZ,002827.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_220 返回数据 row 行数 = "+str(stk_rewards_item_220.shape[0]))
for ts_code_sh in stk_rewards_item_220['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_220_tscode_list.append(stock_name)
stk_rewards_item_220_addname_dataframe=pd.DataFrame()
stk_rewards_item_220_addname_dataframe['cname'] = stk_rewards_item_220_tscode_list
for table_name in stk_rewards_item_220.columns.values.tolist():
    stk_rewards_item_220_addname_dataframe[table_name] = stk_rewards_item_220[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_220_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_221_tscode_list = list() 
stk_rewards_item_221 = pro.stk_rewards(ts_code='002828.SZ,002829.SZ,002830.SZ,002831.SZ,002832.SZ,002833.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_221 返回数据 row 行数 = "+str(stk_rewards_item_221.shape[0]))
for ts_code_sh in stk_rewards_item_221['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_221_tscode_list.append(stock_name)
stk_rewards_item_221_addname_dataframe=pd.DataFrame()
stk_rewards_item_221_addname_dataframe['cname'] = stk_rewards_item_221_tscode_list
for table_name in stk_rewards_item_221.columns.values.tolist():
    stk_rewards_item_221_addname_dataframe[table_name] = stk_rewards_item_221[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_221_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_222_tscode_list = list() 
stk_rewards_item_222 = pro.stk_rewards(ts_code='002835.SZ,002836.SZ,002837.SZ,002838.SZ,002839.SZ,002840.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_222 返回数据 row 行数 = "+str(stk_rewards_item_222.shape[0]))
for ts_code_sh in stk_rewards_item_222['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_222_tscode_list.append(stock_name)
stk_rewards_item_222_addname_dataframe=pd.DataFrame()
stk_rewards_item_222_addname_dataframe['cname'] = stk_rewards_item_222_tscode_list
for table_name in stk_rewards_item_222.columns.values.tolist():
    stk_rewards_item_222_addname_dataframe[table_name] = stk_rewards_item_222[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_222_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_223_tscode_list = list() 
stk_rewards_item_223 = pro.stk_rewards(ts_code='002841.SZ,002842.SZ,002843.SZ,002845.SZ,002846.SZ,002847.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_223 返回数据 row 行数 = "+str(stk_rewards_item_223.shape[0]))
for ts_code_sh in stk_rewards_item_223['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_223_tscode_list.append(stock_name)
stk_rewards_item_223_addname_dataframe=pd.DataFrame()
stk_rewards_item_223_addname_dataframe['cname'] = stk_rewards_item_223_tscode_list
for table_name in stk_rewards_item_223.columns.values.tolist():
    stk_rewards_item_223_addname_dataframe[table_name] = stk_rewards_item_223[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_223_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_224_tscode_list = list() 
stk_rewards_item_224 = pro.stk_rewards(ts_code='002848.SZ,002849.SZ,002850.SZ,002851.SZ,002852.SZ,002853.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_224 返回数据 row 行数 = "+str(stk_rewards_item_224.shape[0]))
for ts_code_sh in stk_rewards_item_224['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_224_tscode_list.append(stock_name)
stk_rewards_item_224_addname_dataframe=pd.DataFrame()
stk_rewards_item_224_addname_dataframe['cname'] = stk_rewards_item_224_tscode_list
for table_name in stk_rewards_item_224.columns.values.tolist():
    stk_rewards_item_224_addname_dataframe[table_name] = stk_rewards_item_224[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_224_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_225_tscode_list = list() 
stk_rewards_item_225 = pro.stk_rewards(ts_code='002855.SZ,002856.SZ,002857.SZ,002858.SZ,002859.SZ,002860.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_225 返回数据 row 行数 = "+str(stk_rewards_item_225.shape[0]))
for ts_code_sh in stk_rewards_item_225['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_225_tscode_list.append(stock_name)
stk_rewards_item_225_addname_dataframe=pd.DataFrame()
stk_rewards_item_225_addname_dataframe['cname'] = stk_rewards_item_225_tscode_list
for table_name in stk_rewards_item_225.columns.values.tolist():
    stk_rewards_item_225_addname_dataframe[table_name] = stk_rewards_item_225[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_225_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_226_tscode_list = list() 
stk_rewards_item_226 = pro.stk_rewards(ts_code='002861.SZ,002862.SZ,002863.SZ,002864.SZ,002865.SZ,002866.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_226 返回数据 row 行数 = "+str(stk_rewards_item_226.shape[0]))
for ts_code_sh in stk_rewards_item_226['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_226_tscode_list.append(stock_name)
stk_rewards_item_226_addname_dataframe=pd.DataFrame()
stk_rewards_item_226_addname_dataframe['cname'] = stk_rewards_item_226_tscode_list
for table_name in stk_rewards_item_226.columns.values.tolist():
    stk_rewards_item_226_addname_dataframe[table_name] = stk_rewards_item_226[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_226_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_227_tscode_list = list() 
stk_rewards_item_227 = pro.stk_rewards(ts_code='002867.SZ,002868.SZ,002869.SZ,002870.SZ,002871.SZ,002872.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_227 返回数据 row 行数 = "+str(stk_rewards_item_227.shape[0]))
for ts_code_sh in stk_rewards_item_227['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_227_tscode_list.append(stock_name)
stk_rewards_item_227_addname_dataframe=pd.DataFrame()
stk_rewards_item_227_addname_dataframe['cname'] = stk_rewards_item_227_tscode_list
for table_name in stk_rewards_item_227.columns.values.tolist():
    stk_rewards_item_227_addname_dataframe[table_name] = stk_rewards_item_227[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_227_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_228_tscode_list = list() 
stk_rewards_item_228 = pro.stk_rewards(ts_code='002873.SZ,002875.SZ,002876.SZ,002877.SZ,002878.SZ,002879.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_228 返回数据 row 行数 = "+str(stk_rewards_item_228.shape[0]))
for ts_code_sh in stk_rewards_item_228['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_228_tscode_list.append(stock_name)
stk_rewards_item_228_addname_dataframe=pd.DataFrame()
stk_rewards_item_228_addname_dataframe['cname'] = stk_rewards_item_228_tscode_list
for table_name in stk_rewards_item_228.columns.values.tolist():
    stk_rewards_item_228_addname_dataframe[table_name] = stk_rewards_item_228[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_228_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_229_tscode_list = list() 
stk_rewards_item_229 = pro.stk_rewards(ts_code='002880.SZ,002881.SZ,002882.SZ,002883.SZ,002884.SZ,002885.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_229 返回数据 row 行数 = "+str(stk_rewards_item_229.shape[0]))
for ts_code_sh in stk_rewards_item_229['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_229_tscode_list.append(stock_name)
stk_rewards_item_229_addname_dataframe=pd.DataFrame()
stk_rewards_item_229_addname_dataframe['cname'] = stk_rewards_item_229_tscode_list
for table_name in stk_rewards_item_229.columns.values.tolist():
    stk_rewards_item_229_addname_dataframe[table_name] = stk_rewards_item_229[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_229_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_230_tscode_list = list() 
stk_rewards_item_230 = pro.stk_rewards(ts_code='002886.SZ,002887.SZ,002888.SZ,002889.SZ,002890.SZ,002891.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_230 返回数据 row 行数 = "+str(stk_rewards_item_230.shape[0]))
for ts_code_sh in stk_rewards_item_230['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_230_tscode_list.append(stock_name)
stk_rewards_item_230_addname_dataframe=pd.DataFrame()
stk_rewards_item_230_addname_dataframe['cname'] = stk_rewards_item_230_tscode_list
for table_name in stk_rewards_item_230.columns.values.tolist():
    stk_rewards_item_230_addname_dataframe[table_name] = stk_rewards_item_230[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_230_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_231_tscode_list = list() 
stk_rewards_item_231 = pro.stk_rewards(ts_code='002892.SZ,002893.SZ,002895.SZ,002896.SZ,002897.SZ,002898.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_231 返回数据 row 行数 = "+str(stk_rewards_item_231.shape[0]))
for ts_code_sh in stk_rewards_item_231['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_231_tscode_list.append(stock_name)
stk_rewards_item_231_addname_dataframe=pd.DataFrame()
stk_rewards_item_231_addname_dataframe['cname'] = stk_rewards_item_231_tscode_list
for table_name in stk_rewards_item_231.columns.values.tolist():
    stk_rewards_item_231_addname_dataframe[table_name] = stk_rewards_item_231[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_231_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_232_tscode_list = list() 
stk_rewards_item_232 = pro.stk_rewards(ts_code='002899.SZ,002900.SZ,002901.SZ,002902.SZ,002903.SZ,002905.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_232 返回数据 row 行数 = "+str(stk_rewards_item_232.shape[0]))
for ts_code_sh in stk_rewards_item_232['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_232_tscode_list.append(stock_name)
stk_rewards_item_232_addname_dataframe=pd.DataFrame()
stk_rewards_item_232_addname_dataframe['cname'] = stk_rewards_item_232_tscode_list
for table_name in stk_rewards_item_232.columns.values.tolist():
    stk_rewards_item_232_addname_dataframe[table_name] = stk_rewards_item_232[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_232_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_233_tscode_list = list() 
stk_rewards_item_233 = pro.stk_rewards(ts_code='002906.SZ,002907.SZ,002908.SZ,002909.SZ,002910.SZ,002911.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_233 返回数据 row 行数 = "+str(stk_rewards_item_233.shape[0]))
for ts_code_sh in stk_rewards_item_233['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_233_tscode_list.append(stock_name)
stk_rewards_item_233_addname_dataframe=pd.DataFrame()
stk_rewards_item_233_addname_dataframe['cname'] = stk_rewards_item_233_tscode_list
for table_name in stk_rewards_item_233.columns.values.tolist():
    stk_rewards_item_233_addname_dataframe[table_name] = stk_rewards_item_233[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_233_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_234_tscode_list = list() 
stk_rewards_item_234 = pro.stk_rewards(ts_code='002912.SZ,002913.SZ,002915.SZ,002916.SZ,002917.SZ,002918.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_234 返回数据 row 行数 = "+str(stk_rewards_item_234.shape[0]))
for ts_code_sh in stk_rewards_item_234['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_234_tscode_list.append(stock_name)
stk_rewards_item_234_addname_dataframe=pd.DataFrame()
stk_rewards_item_234_addname_dataframe['cname'] = stk_rewards_item_234_tscode_list
for table_name in stk_rewards_item_234.columns.values.tolist():
    stk_rewards_item_234_addname_dataframe[table_name] = stk_rewards_item_234[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_234_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_235_tscode_list = list() 
stk_rewards_item_235 = pro.stk_rewards(ts_code='002919.SZ,002920.SZ,002921.SZ,002922.SZ,002923.SZ,002925.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_235 返回数据 row 行数 = "+str(stk_rewards_item_235.shape[0]))
for ts_code_sh in stk_rewards_item_235['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_235_tscode_list.append(stock_name)
stk_rewards_item_235_addname_dataframe=pd.DataFrame()
stk_rewards_item_235_addname_dataframe['cname'] = stk_rewards_item_235_tscode_list
for table_name in stk_rewards_item_235.columns.values.tolist():
    stk_rewards_item_235_addname_dataframe[table_name] = stk_rewards_item_235[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_235_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_236_tscode_list = list() 
stk_rewards_item_236 = pro.stk_rewards(ts_code='002926.SZ,002927.SZ,002928.SZ,002929.SZ,002930.SZ,002931.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_236 返回数据 row 行数 = "+str(stk_rewards_item_236.shape[0]))
for ts_code_sh in stk_rewards_item_236['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_236_tscode_list.append(stock_name)
stk_rewards_item_236_addname_dataframe=pd.DataFrame()
stk_rewards_item_236_addname_dataframe['cname'] = stk_rewards_item_236_tscode_list
for table_name in stk_rewards_item_236.columns.values.tolist():
    stk_rewards_item_236_addname_dataframe[table_name] = stk_rewards_item_236[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_236_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_237_tscode_list = list() 
stk_rewards_item_237 = pro.stk_rewards(ts_code='002932.SZ,002933.SZ,002935.SZ,002936.SZ,002937.SZ,002938.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_237 返回数据 row 行数 = "+str(stk_rewards_item_237.shape[0]))
for ts_code_sh in stk_rewards_item_237['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_237_tscode_list.append(stock_name)
stk_rewards_item_237_addname_dataframe=pd.DataFrame()
stk_rewards_item_237_addname_dataframe['cname'] = stk_rewards_item_237_tscode_list
for table_name in stk_rewards_item_237.columns.values.tolist():
    stk_rewards_item_237_addname_dataframe[table_name] = stk_rewards_item_237[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_237_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_238_tscode_list = list() 
stk_rewards_item_238 = pro.stk_rewards(ts_code='002939.SZ,002940.SZ,002941.SZ,002942.SZ,002943.SZ,002945.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_238 返回数据 row 行数 = "+str(stk_rewards_item_238.shape[0]))
for ts_code_sh in stk_rewards_item_238['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_238_tscode_list.append(stock_name)
stk_rewards_item_238_addname_dataframe=pd.DataFrame()
stk_rewards_item_238_addname_dataframe['cname'] = stk_rewards_item_238_tscode_list
for table_name in stk_rewards_item_238.columns.values.tolist():
    stk_rewards_item_238_addname_dataframe[table_name] = stk_rewards_item_238[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_238_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_239_tscode_list = list() 
stk_rewards_item_239 = pro.stk_rewards(ts_code='002946.SZ,002947.SZ,002948.SZ,002949.SZ,002950.SZ,002951.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_239 返回数据 row 行数 = "+str(stk_rewards_item_239.shape[0]))
for ts_code_sh in stk_rewards_item_239['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_239_tscode_list.append(stock_name)
stk_rewards_item_239_addname_dataframe=pd.DataFrame()
stk_rewards_item_239_addname_dataframe['cname'] = stk_rewards_item_239_tscode_list
for table_name in stk_rewards_item_239.columns.values.tolist():
    stk_rewards_item_239_addname_dataframe[table_name] = stk_rewards_item_239[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_239_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_240_tscode_list = list() 
stk_rewards_item_240 = pro.stk_rewards(ts_code='002952.SZ,002953.SZ,002955.SZ,002956.SZ,002957.SZ,002958.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_240 返回数据 row 行数 = "+str(stk_rewards_item_240.shape[0]))
for ts_code_sh in stk_rewards_item_240['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_240_tscode_list.append(stock_name)
stk_rewards_item_240_addname_dataframe=pd.DataFrame()
stk_rewards_item_240_addname_dataframe['cname'] = stk_rewards_item_240_tscode_list
for table_name in stk_rewards_item_240.columns.values.tolist():
    stk_rewards_item_240_addname_dataframe[table_name] = stk_rewards_item_240[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_240_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_241_tscode_list = list() 
stk_rewards_item_241 = pro.stk_rewards(ts_code='002959.SZ,002960.SZ,002961.SZ,002962.SZ,002963.SZ,002965.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_241 返回数据 row 行数 = "+str(stk_rewards_item_241.shape[0]))
for ts_code_sh in stk_rewards_item_241['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_241_tscode_list.append(stock_name)
stk_rewards_item_241_addname_dataframe=pd.DataFrame()
stk_rewards_item_241_addname_dataframe['cname'] = stk_rewards_item_241_tscode_list
for table_name in stk_rewards_item_241.columns.values.tolist():
    stk_rewards_item_241_addname_dataframe[table_name] = stk_rewards_item_241[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_241_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_242_tscode_list = list() 
stk_rewards_item_242 = pro.stk_rewards(ts_code='002966.SZ,002967.SZ,002968.SZ,002969.SZ,002970.SZ,002971.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_242 返回数据 row 行数 = "+str(stk_rewards_item_242.shape[0]))
for ts_code_sh in stk_rewards_item_242['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_242_tscode_list.append(stock_name)
stk_rewards_item_242_addname_dataframe=pd.DataFrame()
stk_rewards_item_242_addname_dataframe['cname'] = stk_rewards_item_242_tscode_list
for table_name in stk_rewards_item_242.columns.values.tolist():
    stk_rewards_item_242_addname_dataframe[table_name] = stk_rewards_item_242[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_242_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_243_tscode_list = list() 
stk_rewards_item_243 = pro.stk_rewards(ts_code='002972.SZ,002973.SZ,002975.SZ,002976.SZ,002977.SZ,002978.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_243 返回数据 row 行数 = "+str(stk_rewards_item_243.shape[0]))
for ts_code_sh in stk_rewards_item_243['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_243_tscode_list.append(stock_name)
stk_rewards_item_243_addname_dataframe=pd.DataFrame()
stk_rewards_item_243_addname_dataframe['cname'] = stk_rewards_item_243_tscode_list
for table_name in stk_rewards_item_243.columns.values.tolist():
    stk_rewards_item_243_addname_dataframe[table_name] = stk_rewards_item_243[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_243_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_244_tscode_list = list() 
stk_rewards_item_244 = pro.stk_rewards(ts_code='002979.SZ,002980.SZ,002981.SZ,002982.SZ,002983.SZ,002985.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_244 返回数据 row 行数 = "+str(stk_rewards_item_244.shape[0]))
for ts_code_sh in stk_rewards_item_244['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_244_tscode_list.append(stock_name)
stk_rewards_item_244_addname_dataframe=pd.DataFrame()
stk_rewards_item_244_addname_dataframe['cname'] = stk_rewards_item_244_tscode_list
for table_name in stk_rewards_item_244.columns.values.tolist():
    stk_rewards_item_244_addname_dataframe[table_name] = stk_rewards_item_244[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_244_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_245_tscode_list = list() 
stk_rewards_item_245 = pro.stk_rewards(ts_code='002986.SZ,002987.SZ,002988.SZ,002989.SZ,002990.SZ,002991.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_245 返回数据 row 行数 = "+str(stk_rewards_item_245.shape[0]))
for ts_code_sh in stk_rewards_item_245['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_245_tscode_list.append(stock_name)
stk_rewards_item_245_addname_dataframe=pd.DataFrame()
stk_rewards_item_245_addname_dataframe['cname'] = stk_rewards_item_245_tscode_list
for table_name in stk_rewards_item_245.columns.values.tolist():
    stk_rewards_item_245_addname_dataframe[table_name] = stk_rewards_item_245[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_245_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_246_tscode_list = list() 
stk_rewards_item_246 = pro.stk_rewards(ts_code='002992.SZ,002993.SZ,002995.SZ,003816.SZ,300001.SZ,300002.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_246 返回数据 row 行数 = "+str(stk_rewards_item_246.shape[0]))
for ts_code_sh in stk_rewards_item_246['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_246_tscode_list.append(stock_name)
stk_rewards_item_246_addname_dataframe=pd.DataFrame()
stk_rewards_item_246_addname_dataframe['cname'] = stk_rewards_item_246_tscode_list
for table_name in stk_rewards_item_246.columns.values.tolist():
    stk_rewards_item_246_addname_dataframe[table_name] = stk_rewards_item_246[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_246_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_247_tscode_list = list() 
stk_rewards_item_247 = pro.stk_rewards(ts_code='300003.SZ,300004.SZ,300005.SZ,300006.SZ,300007.SZ,300008.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_247 返回数据 row 行数 = "+str(stk_rewards_item_247.shape[0]))
for ts_code_sh in stk_rewards_item_247['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_247_tscode_list.append(stock_name)
stk_rewards_item_247_addname_dataframe=pd.DataFrame()
stk_rewards_item_247_addname_dataframe['cname'] = stk_rewards_item_247_tscode_list
for table_name in stk_rewards_item_247.columns.values.tolist():
    stk_rewards_item_247_addname_dataframe[table_name] = stk_rewards_item_247[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_247_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_248_tscode_list = list() 
stk_rewards_item_248 = pro.stk_rewards(ts_code='300009.SZ,300010.SZ,300011.SZ,300012.SZ,300013.SZ,300014.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_248 返回数据 row 行数 = "+str(stk_rewards_item_248.shape[0]))
for ts_code_sh in stk_rewards_item_248['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_248_tscode_list.append(stock_name)
stk_rewards_item_248_addname_dataframe=pd.DataFrame()
stk_rewards_item_248_addname_dataframe['cname'] = stk_rewards_item_248_tscode_list
for table_name in stk_rewards_item_248.columns.values.tolist():
    stk_rewards_item_248_addname_dataframe[table_name] = stk_rewards_item_248[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_248_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_249_tscode_list = list() 
stk_rewards_item_249 = pro.stk_rewards(ts_code='300015.SZ,300016.SZ,300017.SZ,300018.SZ,300019.SZ,300020.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_249 返回数据 row 行数 = "+str(stk_rewards_item_249.shape[0]))
for ts_code_sh in stk_rewards_item_249['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_249_tscode_list.append(stock_name)
stk_rewards_item_249_addname_dataframe=pd.DataFrame()
stk_rewards_item_249_addname_dataframe['cname'] = stk_rewards_item_249_tscode_list
for table_name in stk_rewards_item_249.columns.values.tolist():
    stk_rewards_item_249_addname_dataframe[table_name] = stk_rewards_item_249[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_249_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_250_tscode_list = list() 
stk_rewards_item_250 = pro.stk_rewards(ts_code='300021.SZ,300022.SZ,300023.SZ,300024.SZ,300025.SZ,300026.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_250 返回数据 row 行数 = "+str(stk_rewards_item_250.shape[0]))
for ts_code_sh in stk_rewards_item_250['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_250_tscode_list.append(stock_name)
stk_rewards_item_250_addname_dataframe=pd.DataFrame()
stk_rewards_item_250_addname_dataframe['cname'] = stk_rewards_item_250_tscode_list
for table_name in stk_rewards_item_250.columns.values.tolist():
    stk_rewards_item_250_addname_dataframe[table_name] = stk_rewards_item_250[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_250_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_251_tscode_list = list() 
stk_rewards_item_251 = pro.stk_rewards(ts_code='300027.SZ,300028.SZ,300029.SZ,300030.SZ,300031.SZ,300032.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_251 返回数据 row 行数 = "+str(stk_rewards_item_251.shape[0]))
for ts_code_sh in stk_rewards_item_251['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_251_tscode_list.append(stock_name)
stk_rewards_item_251_addname_dataframe=pd.DataFrame()
stk_rewards_item_251_addname_dataframe['cname'] = stk_rewards_item_251_tscode_list
for table_name in stk_rewards_item_251.columns.values.tolist():
    stk_rewards_item_251_addname_dataframe[table_name] = stk_rewards_item_251[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_251_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_252_tscode_list = list() 
stk_rewards_item_252 = pro.stk_rewards(ts_code='300033.SZ,300034.SZ,300035.SZ,300036.SZ,300037.SZ,300038.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_252 返回数据 row 行数 = "+str(stk_rewards_item_252.shape[0]))
for ts_code_sh in stk_rewards_item_252['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_252_tscode_list.append(stock_name)
stk_rewards_item_252_addname_dataframe=pd.DataFrame()
stk_rewards_item_252_addname_dataframe['cname'] = stk_rewards_item_252_tscode_list
for table_name in stk_rewards_item_252.columns.values.tolist():
    stk_rewards_item_252_addname_dataframe[table_name] = stk_rewards_item_252[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_252_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_253_tscode_list = list() 
stk_rewards_item_253 = pro.stk_rewards(ts_code='300039.SZ,300040.SZ,300041.SZ,300042.SZ,300043.SZ,300044.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_253 返回数据 row 行数 = "+str(stk_rewards_item_253.shape[0]))
for ts_code_sh in stk_rewards_item_253['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_253_tscode_list.append(stock_name)
stk_rewards_item_253_addname_dataframe=pd.DataFrame()
stk_rewards_item_253_addname_dataframe['cname'] = stk_rewards_item_253_tscode_list
for table_name in stk_rewards_item_253.columns.values.tolist():
    stk_rewards_item_253_addname_dataframe[table_name] = stk_rewards_item_253[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_253_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_254_tscode_list = list() 
stk_rewards_item_254 = pro.stk_rewards(ts_code='300045.SZ,300046.SZ,300047.SZ,300048.SZ,300049.SZ,300050.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_254 返回数据 row 行数 = "+str(stk_rewards_item_254.shape[0]))
for ts_code_sh in stk_rewards_item_254['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_254_tscode_list.append(stock_name)
stk_rewards_item_254_addname_dataframe=pd.DataFrame()
stk_rewards_item_254_addname_dataframe['cname'] = stk_rewards_item_254_tscode_list
for table_name in stk_rewards_item_254.columns.values.tolist():
    stk_rewards_item_254_addname_dataframe[table_name] = stk_rewards_item_254[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_254_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_255_tscode_list = list() 
stk_rewards_item_255 = pro.stk_rewards(ts_code='300051.SZ,300052.SZ,300053.SZ,300054.SZ,300055.SZ,300056.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_255 返回数据 row 行数 = "+str(stk_rewards_item_255.shape[0]))
for ts_code_sh in stk_rewards_item_255['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_255_tscode_list.append(stock_name)
stk_rewards_item_255_addname_dataframe=pd.DataFrame()
stk_rewards_item_255_addname_dataframe['cname'] = stk_rewards_item_255_tscode_list
for table_name in stk_rewards_item_255.columns.values.tolist():
    stk_rewards_item_255_addname_dataframe[table_name] = stk_rewards_item_255[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_255_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_256_tscode_list = list() 
stk_rewards_item_256 = pro.stk_rewards(ts_code='300057.SZ,300058.SZ,300059.SZ,300061.SZ,300062.SZ,300063.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_256 返回数据 row 行数 = "+str(stk_rewards_item_256.shape[0]))
for ts_code_sh in stk_rewards_item_256['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_256_tscode_list.append(stock_name)
stk_rewards_item_256_addname_dataframe=pd.DataFrame()
stk_rewards_item_256_addname_dataframe['cname'] = stk_rewards_item_256_tscode_list
for table_name in stk_rewards_item_256.columns.values.tolist():
    stk_rewards_item_256_addname_dataframe[table_name] = stk_rewards_item_256[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_256_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_257_tscode_list = list() 
stk_rewards_item_257 = pro.stk_rewards(ts_code='300064.SZ,300065.SZ,300066.SZ,300067.SZ,300068.SZ,300069.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_257 返回数据 row 行数 = "+str(stk_rewards_item_257.shape[0]))
for ts_code_sh in stk_rewards_item_257['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_257_tscode_list.append(stock_name)
stk_rewards_item_257_addname_dataframe=pd.DataFrame()
stk_rewards_item_257_addname_dataframe['cname'] = stk_rewards_item_257_tscode_list
for table_name in stk_rewards_item_257.columns.values.tolist():
    stk_rewards_item_257_addname_dataframe[table_name] = stk_rewards_item_257[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_257_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_258_tscode_list = list() 
stk_rewards_item_258 = pro.stk_rewards(ts_code='300070.SZ,300071.SZ,300072.SZ,300073.SZ,300074.SZ,300075.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_258 返回数据 row 行数 = "+str(stk_rewards_item_258.shape[0]))
for ts_code_sh in stk_rewards_item_258['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_258_tscode_list.append(stock_name)
stk_rewards_item_258_addname_dataframe=pd.DataFrame()
stk_rewards_item_258_addname_dataframe['cname'] = stk_rewards_item_258_tscode_list
for table_name in stk_rewards_item_258.columns.values.tolist():
    stk_rewards_item_258_addname_dataframe[table_name] = stk_rewards_item_258[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_258_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_259_tscode_list = list() 
stk_rewards_item_259 = pro.stk_rewards(ts_code='300076.SZ,300077.SZ,300078.SZ,300079.SZ,300080.SZ,300081.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_259 返回数据 row 行数 = "+str(stk_rewards_item_259.shape[0]))
for ts_code_sh in stk_rewards_item_259['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_259_tscode_list.append(stock_name)
stk_rewards_item_259_addname_dataframe=pd.DataFrame()
stk_rewards_item_259_addname_dataframe['cname'] = stk_rewards_item_259_tscode_list
for table_name in stk_rewards_item_259.columns.values.tolist():
    stk_rewards_item_259_addname_dataframe[table_name] = stk_rewards_item_259[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_259_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_260_tscode_list = list() 
stk_rewards_item_260 = pro.stk_rewards(ts_code='300082.SZ,300083.SZ,300084.SZ,300085.SZ,300086.SZ,300087.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_260 返回数据 row 行数 = "+str(stk_rewards_item_260.shape[0]))
for ts_code_sh in stk_rewards_item_260['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_260_tscode_list.append(stock_name)
stk_rewards_item_260_addname_dataframe=pd.DataFrame()
stk_rewards_item_260_addname_dataframe['cname'] = stk_rewards_item_260_tscode_list
for table_name in stk_rewards_item_260.columns.values.tolist():
    stk_rewards_item_260_addname_dataframe[table_name] = stk_rewards_item_260[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_260_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_261_tscode_list = list() 
stk_rewards_item_261 = pro.stk_rewards(ts_code='300088.SZ,300089.SZ,300090.SZ,300091.SZ,300092.SZ,300093.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_261 返回数据 row 行数 = "+str(stk_rewards_item_261.shape[0]))
for ts_code_sh in stk_rewards_item_261['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_261_tscode_list.append(stock_name)
stk_rewards_item_261_addname_dataframe=pd.DataFrame()
stk_rewards_item_261_addname_dataframe['cname'] = stk_rewards_item_261_tscode_list
for table_name in stk_rewards_item_261.columns.values.tolist():
    stk_rewards_item_261_addname_dataframe[table_name] = stk_rewards_item_261[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_261_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_262_tscode_list = list() 
stk_rewards_item_262 = pro.stk_rewards(ts_code='300094.SZ,300095.SZ,300096.SZ,300097.SZ,300098.SZ,300099.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_262 返回数据 row 行数 = "+str(stk_rewards_item_262.shape[0]))
for ts_code_sh in stk_rewards_item_262['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_262_tscode_list.append(stock_name)
stk_rewards_item_262_addname_dataframe=pd.DataFrame()
stk_rewards_item_262_addname_dataframe['cname'] = stk_rewards_item_262_tscode_list
for table_name in stk_rewards_item_262.columns.values.tolist():
    stk_rewards_item_262_addname_dataframe[table_name] = stk_rewards_item_262[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_262_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_263_tscode_list = list() 
stk_rewards_item_263 = pro.stk_rewards(ts_code='300100.SZ,300101.SZ,300102.SZ,300103.SZ,300104.SZ,300105.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_263 返回数据 row 行数 = "+str(stk_rewards_item_263.shape[0]))
for ts_code_sh in stk_rewards_item_263['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_263_tscode_list.append(stock_name)
stk_rewards_item_263_addname_dataframe=pd.DataFrame()
stk_rewards_item_263_addname_dataframe['cname'] = stk_rewards_item_263_tscode_list
for table_name in stk_rewards_item_263.columns.values.tolist():
    stk_rewards_item_263_addname_dataframe[table_name] = stk_rewards_item_263[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_263_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_264_tscode_list = list() 
stk_rewards_item_264 = pro.stk_rewards(ts_code='300106.SZ,300107.SZ,300108.SZ,300109.SZ,300110.SZ,300111.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_264 返回数据 row 行数 = "+str(stk_rewards_item_264.shape[0]))
for ts_code_sh in stk_rewards_item_264['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_264_tscode_list.append(stock_name)
stk_rewards_item_264_addname_dataframe=pd.DataFrame()
stk_rewards_item_264_addname_dataframe['cname'] = stk_rewards_item_264_tscode_list
for table_name in stk_rewards_item_264.columns.values.tolist():
    stk_rewards_item_264_addname_dataframe[table_name] = stk_rewards_item_264[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_264_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_265_tscode_list = list() 
stk_rewards_item_265 = pro.stk_rewards(ts_code='300112.SZ,300113.SZ,300114.SZ,300115.SZ,300116.SZ,300117.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_265 返回数据 row 行数 = "+str(stk_rewards_item_265.shape[0]))
for ts_code_sh in stk_rewards_item_265['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_265_tscode_list.append(stock_name)
stk_rewards_item_265_addname_dataframe=pd.DataFrame()
stk_rewards_item_265_addname_dataframe['cname'] = stk_rewards_item_265_tscode_list
for table_name in stk_rewards_item_265.columns.values.tolist():
    stk_rewards_item_265_addname_dataframe[table_name] = stk_rewards_item_265[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_265_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_266_tscode_list = list() 
stk_rewards_item_266 = pro.stk_rewards(ts_code='300118.SZ,300119.SZ,300120.SZ,300121.SZ,300122.SZ,300123.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_266 返回数据 row 行数 = "+str(stk_rewards_item_266.shape[0]))
for ts_code_sh in stk_rewards_item_266['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_266_tscode_list.append(stock_name)
stk_rewards_item_266_addname_dataframe=pd.DataFrame()
stk_rewards_item_266_addname_dataframe['cname'] = stk_rewards_item_266_tscode_list
for table_name in stk_rewards_item_266.columns.values.tolist():
    stk_rewards_item_266_addname_dataframe[table_name] = stk_rewards_item_266[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_266_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_267_tscode_list = list() 
stk_rewards_item_267 = pro.stk_rewards(ts_code='300124.SZ,300125.SZ,300126.SZ,300127.SZ,300128.SZ,300129.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_267 返回数据 row 行数 = "+str(stk_rewards_item_267.shape[0]))
for ts_code_sh in stk_rewards_item_267['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_267_tscode_list.append(stock_name)
stk_rewards_item_267_addname_dataframe=pd.DataFrame()
stk_rewards_item_267_addname_dataframe['cname'] = stk_rewards_item_267_tscode_list
for table_name in stk_rewards_item_267.columns.values.tolist():
    stk_rewards_item_267_addname_dataframe[table_name] = stk_rewards_item_267[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_267_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_268_tscode_list = list() 
stk_rewards_item_268 = pro.stk_rewards(ts_code='300130.SZ,300131.SZ,300132.SZ,300133.SZ,300134.SZ,300135.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_268 返回数据 row 行数 = "+str(stk_rewards_item_268.shape[0]))
for ts_code_sh in stk_rewards_item_268['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_268_tscode_list.append(stock_name)
stk_rewards_item_268_addname_dataframe=pd.DataFrame()
stk_rewards_item_268_addname_dataframe['cname'] = stk_rewards_item_268_tscode_list
for table_name in stk_rewards_item_268.columns.values.tolist():
    stk_rewards_item_268_addname_dataframe[table_name] = stk_rewards_item_268[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_268_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_269_tscode_list = list() 
stk_rewards_item_269 = pro.stk_rewards(ts_code='300136.SZ,300137.SZ,300138.SZ,300139.SZ,300140.SZ,300141.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_269 返回数据 row 行数 = "+str(stk_rewards_item_269.shape[0]))
for ts_code_sh in stk_rewards_item_269['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_269_tscode_list.append(stock_name)
stk_rewards_item_269_addname_dataframe=pd.DataFrame()
stk_rewards_item_269_addname_dataframe['cname'] = stk_rewards_item_269_tscode_list
for table_name in stk_rewards_item_269.columns.values.tolist():
    stk_rewards_item_269_addname_dataframe[table_name] = stk_rewards_item_269[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_269_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_270_tscode_list = list() 
stk_rewards_item_270 = pro.stk_rewards(ts_code='300142.SZ,300143.SZ,300144.SZ,300145.SZ,300146.SZ,300147.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_270 返回数据 row 行数 = "+str(stk_rewards_item_270.shape[0]))
for ts_code_sh in stk_rewards_item_270['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_270_tscode_list.append(stock_name)
stk_rewards_item_270_addname_dataframe=pd.DataFrame()
stk_rewards_item_270_addname_dataframe['cname'] = stk_rewards_item_270_tscode_list
for table_name in stk_rewards_item_270.columns.values.tolist():
    stk_rewards_item_270_addname_dataframe[table_name] = stk_rewards_item_270[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_270_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_271_tscode_list = list() 
stk_rewards_item_271 = pro.stk_rewards(ts_code='300148.SZ,300149.SZ,300150.SZ,300151.SZ,300152.SZ,300153.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_271 返回数据 row 行数 = "+str(stk_rewards_item_271.shape[0]))
for ts_code_sh in stk_rewards_item_271['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_271_tscode_list.append(stock_name)
stk_rewards_item_271_addname_dataframe=pd.DataFrame()
stk_rewards_item_271_addname_dataframe['cname'] = stk_rewards_item_271_tscode_list
for table_name in stk_rewards_item_271.columns.values.tolist():
    stk_rewards_item_271_addname_dataframe[table_name] = stk_rewards_item_271[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_271_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_272_tscode_list = list() 
stk_rewards_item_272 = pro.stk_rewards(ts_code='300154.SZ,300155.SZ,300156.SZ,300157.SZ,300158.SZ,300159.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_272 返回数据 row 行数 = "+str(stk_rewards_item_272.shape[0]))
for ts_code_sh in stk_rewards_item_272['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_272_tscode_list.append(stock_name)
stk_rewards_item_272_addname_dataframe=pd.DataFrame()
stk_rewards_item_272_addname_dataframe['cname'] = stk_rewards_item_272_tscode_list
for table_name in stk_rewards_item_272.columns.values.tolist():
    stk_rewards_item_272_addname_dataframe[table_name] = stk_rewards_item_272[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_272_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_273_tscode_list = list() 
stk_rewards_item_273 = pro.stk_rewards(ts_code='300160.SZ,300161.SZ,300162.SZ,300163.SZ,300164.SZ,300165.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_273 返回数据 row 行数 = "+str(stk_rewards_item_273.shape[0]))
for ts_code_sh in stk_rewards_item_273['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_273_tscode_list.append(stock_name)
stk_rewards_item_273_addname_dataframe=pd.DataFrame()
stk_rewards_item_273_addname_dataframe['cname'] = stk_rewards_item_273_tscode_list
for table_name in stk_rewards_item_273.columns.values.tolist():
    stk_rewards_item_273_addname_dataframe[table_name] = stk_rewards_item_273[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_273_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_274_tscode_list = list() 
stk_rewards_item_274 = pro.stk_rewards(ts_code='300166.SZ,300167.SZ,300168.SZ,300169.SZ,300170.SZ,300171.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_274 返回数据 row 行数 = "+str(stk_rewards_item_274.shape[0]))
for ts_code_sh in stk_rewards_item_274['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_274_tscode_list.append(stock_name)
stk_rewards_item_274_addname_dataframe=pd.DataFrame()
stk_rewards_item_274_addname_dataframe['cname'] = stk_rewards_item_274_tscode_list
for table_name in stk_rewards_item_274.columns.values.tolist():
    stk_rewards_item_274_addname_dataframe[table_name] = stk_rewards_item_274[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_274_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_275_tscode_list = list() 
stk_rewards_item_275 = pro.stk_rewards(ts_code='300172.SZ,300173.SZ,300174.SZ,300175.SZ,300176.SZ,300177.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_275 返回数据 row 行数 = "+str(stk_rewards_item_275.shape[0]))
for ts_code_sh in stk_rewards_item_275['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_275_tscode_list.append(stock_name)
stk_rewards_item_275_addname_dataframe=pd.DataFrame()
stk_rewards_item_275_addname_dataframe['cname'] = stk_rewards_item_275_tscode_list
for table_name in stk_rewards_item_275.columns.values.tolist():
    stk_rewards_item_275_addname_dataframe[table_name] = stk_rewards_item_275[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_275_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_276_tscode_list = list() 
stk_rewards_item_276 = pro.stk_rewards(ts_code='300178.SZ,300179.SZ,300180.SZ,300181.SZ,300182.SZ,300183.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_276 返回数据 row 行数 = "+str(stk_rewards_item_276.shape[0]))
for ts_code_sh in stk_rewards_item_276['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_276_tscode_list.append(stock_name)
stk_rewards_item_276_addname_dataframe=pd.DataFrame()
stk_rewards_item_276_addname_dataframe['cname'] = stk_rewards_item_276_tscode_list
for table_name in stk_rewards_item_276.columns.values.tolist():
    stk_rewards_item_276_addname_dataframe[table_name] = stk_rewards_item_276[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_276_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_277_tscode_list = list() 
stk_rewards_item_277 = pro.stk_rewards(ts_code='300184.SZ,300185.SZ,300186.SZ,300187.SZ,300188.SZ,300189.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_277 返回数据 row 行数 = "+str(stk_rewards_item_277.shape[0]))
for ts_code_sh in stk_rewards_item_277['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_277_tscode_list.append(stock_name)
stk_rewards_item_277_addname_dataframe=pd.DataFrame()
stk_rewards_item_277_addname_dataframe['cname'] = stk_rewards_item_277_tscode_list
for table_name in stk_rewards_item_277.columns.values.tolist():
    stk_rewards_item_277_addname_dataframe[table_name] = stk_rewards_item_277[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_277_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_278_tscode_list = list() 
stk_rewards_item_278 = pro.stk_rewards(ts_code='300190.SZ,300191.SZ,300192.SZ,300193.SZ,300194.SZ,300195.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_278 返回数据 row 行数 = "+str(stk_rewards_item_278.shape[0]))
for ts_code_sh in stk_rewards_item_278['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_278_tscode_list.append(stock_name)
stk_rewards_item_278_addname_dataframe=pd.DataFrame()
stk_rewards_item_278_addname_dataframe['cname'] = stk_rewards_item_278_tscode_list
for table_name in stk_rewards_item_278.columns.values.tolist():
    stk_rewards_item_278_addname_dataframe[table_name] = stk_rewards_item_278[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_278_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_279_tscode_list = list() 
stk_rewards_item_279 = pro.stk_rewards(ts_code='300196.SZ,300197.SZ,300198.SZ,300199.SZ,300200.SZ,300201.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_279 返回数据 row 行数 = "+str(stk_rewards_item_279.shape[0]))
for ts_code_sh in stk_rewards_item_279['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_279_tscode_list.append(stock_name)
stk_rewards_item_279_addname_dataframe=pd.DataFrame()
stk_rewards_item_279_addname_dataframe['cname'] = stk_rewards_item_279_tscode_list
for table_name in stk_rewards_item_279.columns.values.tolist():
    stk_rewards_item_279_addname_dataframe[table_name] = stk_rewards_item_279[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_279_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_280_tscode_list = list() 
stk_rewards_item_280 = pro.stk_rewards(ts_code='300202.SZ,300203.SZ,300204.SZ,300205.SZ,300206.SZ,300207.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_280 返回数据 row 行数 = "+str(stk_rewards_item_280.shape[0]))
for ts_code_sh in stk_rewards_item_280['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_280_tscode_list.append(stock_name)
stk_rewards_item_280_addname_dataframe=pd.DataFrame()
stk_rewards_item_280_addname_dataframe['cname'] = stk_rewards_item_280_tscode_list
for table_name in stk_rewards_item_280.columns.values.tolist():
    stk_rewards_item_280_addname_dataframe[table_name] = stk_rewards_item_280[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_280_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_281_tscode_list = list() 
stk_rewards_item_281 = pro.stk_rewards(ts_code='300208.SZ,300209.SZ,300210.SZ,300211.SZ,300212.SZ,300213.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_281 返回数据 row 行数 = "+str(stk_rewards_item_281.shape[0]))
for ts_code_sh in stk_rewards_item_281['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_281_tscode_list.append(stock_name)
stk_rewards_item_281_addname_dataframe=pd.DataFrame()
stk_rewards_item_281_addname_dataframe['cname'] = stk_rewards_item_281_tscode_list
for table_name in stk_rewards_item_281.columns.values.tolist():
    stk_rewards_item_281_addname_dataframe[table_name] = stk_rewards_item_281[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_281_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_282_tscode_list = list() 
stk_rewards_item_282 = pro.stk_rewards(ts_code='300214.SZ,300215.SZ,300216.SZ,300217.SZ,300218.SZ,300219.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_282 返回数据 row 行数 = "+str(stk_rewards_item_282.shape[0]))
for ts_code_sh in stk_rewards_item_282['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_282_tscode_list.append(stock_name)
stk_rewards_item_282_addname_dataframe=pd.DataFrame()
stk_rewards_item_282_addname_dataframe['cname'] = stk_rewards_item_282_tscode_list
for table_name in stk_rewards_item_282.columns.values.tolist():
    stk_rewards_item_282_addname_dataframe[table_name] = stk_rewards_item_282[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_282_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_283_tscode_list = list() 
stk_rewards_item_283 = pro.stk_rewards(ts_code='300220.SZ,300221.SZ,300222.SZ,300223.SZ,300224.SZ,300225.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_283 返回数据 row 行数 = "+str(stk_rewards_item_283.shape[0]))
for ts_code_sh in stk_rewards_item_283['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_283_tscode_list.append(stock_name)
stk_rewards_item_283_addname_dataframe=pd.DataFrame()
stk_rewards_item_283_addname_dataframe['cname'] = stk_rewards_item_283_tscode_list
for table_name in stk_rewards_item_283.columns.values.tolist():
    stk_rewards_item_283_addname_dataframe[table_name] = stk_rewards_item_283[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_283_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_284_tscode_list = list() 
stk_rewards_item_284 = pro.stk_rewards(ts_code='300226.SZ,300227.SZ,300228.SZ,300229.SZ,300230.SZ,300231.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_284 返回数据 row 行数 = "+str(stk_rewards_item_284.shape[0]))
for ts_code_sh in stk_rewards_item_284['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_284_tscode_list.append(stock_name)
stk_rewards_item_284_addname_dataframe=pd.DataFrame()
stk_rewards_item_284_addname_dataframe['cname'] = stk_rewards_item_284_tscode_list
for table_name in stk_rewards_item_284.columns.values.tolist():
    stk_rewards_item_284_addname_dataframe[table_name] = stk_rewards_item_284[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_284_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_285_tscode_list = list() 
stk_rewards_item_285 = pro.stk_rewards(ts_code='300232.SZ,300233.SZ,300234.SZ,300235.SZ,300236.SZ,300237.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_285 返回数据 row 行数 = "+str(stk_rewards_item_285.shape[0]))
for ts_code_sh in stk_rewards_item_285['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_285_tscode_list.append(stock_name)
stk_rewards_item_285_addname_dataframe=pd.DataFrame()
stk_rewards_item_285_addname_dataframe['cname'] = stk_rewards_item_285_tscode_list
for table_name in stk_rewards_item_285.columns.values.tolist():
    stk_rewards_item_285_addname_dataframe[table_name] = stk_rewards_item_285[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_285_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_286_tscode_list = list() 
stk_rewards_item_286 = pro.stk_rewards(ts_code='300238.SZ,300239.SZ,300240.SZ,300241.SZ,300242.SZ,300243.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_286 返回数据 row 行数 = "+str(stk_rewards_item_286.shape[0]))
for ts_code_sh in stk_rewards_item_286['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_286_tscode_list.append(stock_name)
stk_rewards_item_286_addname_dataframe=pd.DataFrame()
stk_rewards_item_286_addname_dataframe['cname'] = stk_rewards_item_286_tscode_list
for table_name in stk_rewards_item_286.columns.values.tolist():
    stk_rewards_item_286_addname_dataframe[table_name] = stk_rewards_item_286[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_286_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_287_tscode_list = list() 
stk_rewards_item_287 = pro.stk_rewards(ts_code='300244.SZ,300245.SZ,300246.SZ,300247.SZ,300248.SZ,300249.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_287 返回数据 row 行数 = "+str(stk_rewards_item_287.shape[0]))
for ts_code_sh in stk_rewards_item_287['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_287_tscode_list.append(stock_name)
stk_rewards_item_287_addname_dataframe=pd.DataFrame()
stk_rewards_item_287_addname_dataframe['cname'] = stk_rewards_item_287_tscode_list
for table_name in stk_rewards_item_287.columns.values.tolist():
    stk_rewards_item_287_addname_dataframe[table_name] = stk_rewards_item_287[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_287_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_288_tscode_list = list() 
stk_rewards_item_288 = pro.stk_rewards(ts_code='300250.SZ,300251.SZ,300252.SZ,300253.SZ,300254.SZ,300255.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_288 返回数据 row 行数 = "+str(stk_rewards_item_288.shape[0]))
for ts_code_sh in stk_rewards_item_288['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_288_tscode_list.append(stock_name)
stk_rewards_item_288_addname_dataframe=pd.DataFrame()
stk_rewards_item_288_addname_dataframe['cname'] = stk_rewards_item_288_tscode_list
for table_name in stk_rewards_item_288.columns.values.tolist():
    stk_rewards_item_288_addname_dataframe[table_name] = stk_rewards_item_288[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_288_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_289_tscode_list = list() 
stk_rewards_item_289 = pro.stk_rewards(ts_code='300256.SZ,300257.SZ,300258.SZ,300259.SZ,300260.SZ,300261.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_289 返回数据 row 行数 = "+str(stk_rewards_item_289.shape[0]))
for ts_code_sh in stk_rewards_item_289['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_289_tscode_list.append(stock_name)
stk_rewards_item_289_addname_dataframe=pd.DataFrame()
stk_rewards_item_289_addname_dataframe['cname'] = stk_rewards_item_289_tscode_list
for table_name in stk_rewards_item_289.columns.values.tolist():
    stk_rewards_item_289_addname_dataframe[table_name] = stk_rewards_item_289[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_289_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_290_tscode_list = list() 
stk_rewards_item_290 = pro.stk_rewards(ts_code='300262.SZ,300263.SZ,300264.SZ,300265.SZ,300266.SZ,300267.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_290 返回数据 row 行数 = "+str(stk_rewards_item_290.shape[0]))
for ts_code_sh in stk_rewards_item_290['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_290_tscode_list.append(stock_name)
stk_rewards_item_290_addname_dataframe=pd.DataFrame()
stk_rewards_item_290_addname_dataframe['cname'] = stk_rewards_item_290_tscode_list
for table_name in stk_rewards_item_290.columns.values.tolist():
    stk_rewards_item_290_addname_dataframe[table_name] = stk_rewards_item_290[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_290_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_291_tscode_list = list() 
stk_rewards_item_291 = pro.stk_rewards(ts_code='300268.SZ,300269.SZ,300270.SZ,300271.SZ,300272.SZ,300273.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_291 返回数据 row 行数 = "+str(stk_rewards_item_291.shape[0]))
for ts_code_sh in stk_rewards_item_291['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_291_tscode_list.append(stock_name)
stk_rewards_item_291_addname_dataframe=pd.DataFrame()
stk_rewards_item_291_addname_dataframe['cname'] = stk_rewards_item_291_tscode_list
for table_name in stk_rewards_item_291.columns.values.tolist():
    stk_rewards_item_291_addname_dataframe[table_name] = stk_rewards_item_291[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_291_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_292_tscode_list = list() 
stk_rewards_item_292 = pro.stk_rewards(ts_code='300274.SZ,300275.SZ,300276.SZ,300277.SZ,300278.SZ,300279.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_292 返回数据 row 行数 = "+str(stk_rewards_item_292.shape[0]))
for ts_code_sh in stk_rewards_item_292['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_292_tscode_list.append(stock_name)
stk_rewards_item_292_addname_dataframe=pd.DataFrame()
stk_rewards_item_292_addname_dataframe['cname'] = stk_rewards_item_292_tscode_list
for table_name in stk_rewards_item_292.columns.values.tolist():
    stk_rewards_item_292_addname_dataframe[table_name] = stk_rewards_item_292[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_292_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_293_tscode_list = list() 
stk_rewards_item_293 = pro.stk_rewards(ts_code='300280.SZ,300281.SZ,300282.SZ,300283.SZ,300284.SZ,300285.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_293 返回数据 row 行数 = "+str(stk_rewards_item_293.shape[0]))
for ts_code_sh in stk_rewards_item_293['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_293_tscode_list.append(stock_name)
stk_rewards_item_293_addname_dataframe=pd.DataFrame()
stk_rewards_item_293_addname_dataframe['cname'] = stk_rewards_item_293_tscode_list
for table_name in stk_rewards_item_293.columns.values.tolist():
    stk_rewards_item_293_addname_dataframe[table_name] = stk_rewards_item_293[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_293_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_294_tscode_list = list() 
stk_rewards_item_294 = pro.stk_rewards(ts_code='300286.SZ,300287.SZ,300288.SZ,300289.SZ,300290.SZ,300291.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_294 返回数据 row 行数 = "+str(stk_rewards_item_294.shape[0]))
for ts_code_sh in stk_rewards_item_294['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_294_tscode_list.append(stock_name)
stk_rewards_item_294_addname_dataframe=pd.DataFrame()
stk_rewards_item_294_addname_dataframe['cname'] = stk_rewards_item_294_tscode_list
for table_name in stk_rewards_item_294.columns.values.tolist():
    stk_rewards_item_294_addname_dataframe[table_name] = stk_rewards_item_294[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_294_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_295_tscode_list = list() 
stk_rewards_item_295 = pro.stk_rewards(ts_code='300292.SZ,300293.SZ,300294.SZ,300295.SZ,300296.SZ,300297.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_295 返回数据 row 行数 = "+str(stk_rewards_item_295.shape[0]))
for ts_code_sh in stk_rewards_item_295['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_295_tscode_list.append(stock_name)
stk_rewards_item_295_addname_dataframe=pd.DataFrame()
stk_rewards_item_295_addname_dataframe['cname'] = stk_rewards_item_295_tscode_list
for table_name in stk_rewards_item_295.columns.values.tolist():
    stk_rewards_item_295_addname_dataframe[table_name] = stk_rewards_item_295[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_295_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_296_tscode_list = list() 
stk_rewards_item_296 = pro.stk_rewards(ts_code='300298.SZ,300299.SZ,300300.SZ,300301.SZ,300302.SZ,300303.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_296 返回数据 row 行数 = "+str(stk_rewards_item_296.shape[0]))
for ts_code_sh in stk_rewards_item_296['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_296_tscode_list.append(stock_name)
stk_rewards_item_296_addname_dataframe=pd.DataFrame()
stk_rewards_item_296_addname_dataframe['cname'] = stk_rewards_item_296_tscode_list
for table_name in stk_rewards_item_296.columns.values.tolist():
    stk_rewards_item_296_addname_dataframe[table_name] = stk_rewards_item_296[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_296_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_297_tscode_list = list() 
stk_rewards_item_297 = pro.stk_rewards(ts_code='300304.SZ,300305.SZ,300306.SZ,300307.SZ,300308.SZ,300309.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_297 返回数据 row 行数 = "+str(stk_rewards_item_297.shape[0]))
for ts_code_sh in stk_rewards_item_297['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_297_tscode_list.append(stock_name)
stk_rewards_item_297_addname_dataframe=pd.DataFrame()
stk_rewards_item_297_addname_dataframe['cname'] = stk_rewards_item_297_tscode_list
for table_name in stk_rewards_item_297.columns.values.tolist():
    stk_rewards_item_297_addname_dataframe[table_name] = stk_rewards_item_297[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_297_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_298_tscode_list = list() 
stk_rewards_item_298 = pro.stk_rewards(ts_code='300310.SZ,300311.SZ,300312.SZ,300313.SZ,300314.SZ,300315.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_298 返回数据 row 行数 = "+str(stk_rewards_item_298.shape[0]))
for ts_code_sh in stk_rewards_item_298['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_298_tscode_list.append(stock_name)
stk_rewards_item_298_addname_dataframe=pd.DataFrame()
stk_rewards_item_298_addname_dataframe['cname'] = stk_rewards_item_298_tscode_list
for table_name in stk_rewards_item_298.columns.values.tolist():
    stk_rewards_item_298_addname_dataframe[table_name] = stk_rewards_item_298[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_298_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_299_tscode_list = list() 
stk_rewards_item_299 = pro.stk_rewards(ts_code='300316.SZ,300317.SZ,300318.SZ,300319.SZ,300320.SZ,300321.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_299 返回数据 row 行数 = "+str(stk_rewards_item_299.shape[0]))
for ts_code_sh in stk_rewards_item_299['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_299_tscode_list.append(stock_name)
stk_rewards_item_299_addname_dataframe=pd.DataFrame()
stk_rewards_item_299_addname_dataframe['cname'] = stk_rewards_item_299_tscode_list
for table_name in stk_rewards_item_299.columns.values.tolist():
    stk_rewards_item_299_addname_dataframe[table_name] = stk_rewards_item_299[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_299_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_300_tscode_list = list() 
stk_rewards_item_300 = pro.stk_rewards(ts_code='300322.SZ,300323.SZ,300324.SZ,300325.SZ,300326.SZ,300327.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_300 返回数据 row 行数 = "+str(stk_rewards_item_300.shape[0]))
for ts_code_sh in stk_rewards_item_300['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_300_tscode_list.append(stock_name)
stk_rewards_item_300_addname_dataframe=pd.DataFrame()
stk_rewards_item_300_addname_dataframe['cname'] = stk_rewards_item_300_tscode_list
for table_name in stk_rewards_item_300.columns.values.tolist():
    stk_rewards_item_300_addname_dataframe[table_name] = stk_rewards_item_300[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_300_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_301_tscode_list = list() 
stk_rewards_item_301 = pro.stk_rewards(ts_code='300328.SZ,300329.SZ,300330.SZ,300331.SZ,300332.SZ,300333.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_301 返回数据 row 行数 = "+str(stk_rewards_item_301.shape[0]))
for ts_code_sh in stk_rewards_item_301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_301_tscode_list.append(stock_name)
stk_rewards_item_301_addname_dataframe=pd.DataFrame()
stk_rewards_item_301_addname_dataframe['cname'] = stk_rewards_item_301_tscode_list
for table_name in stk_rewards_item_301.columns.values.tolist():
    stk_rewards_item_301_addname_dataframe[table_name] = stk_rewards_item_301[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_301_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_302_tscode_list = list() 
stk_rewards_item_302 = pro.stk_rewards(ts_code='300334.SZ,300335.SZ,300336.SZ,300337.SZ,300338.SZ,300339.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_302 返回数据 row 行数 = "+str(stk_rewards_item_302.shape[0]))
for ts_code_sh in stk_rewards_item_302['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_302_tscode_list.append(stock_name)
stk_rewards_item_302_addname_dataframe=pd.DataFrame()
stk_rewards_item_302_addname_dataframe['cname'] = stk_rewards_item_302_tscode_list
for table_name in stk_rewards_item_302.columns.values.tolist():
    stk_rewards_item_302_addname_dataframe[table_name] = stk_rewards_item_302[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_302_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_303_tscode_list = list() 
stk_rewards_item_303 = pro.stk_rewards(ts_code='300340.SZ,300341.SZ,300342.SZ,300343.SZ,300344.SZ,300345.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_303 返回数据 row 行数 = "+str(stk_rewards_item_303.shape[0]))
for ts_code_sh in stk_rewards_item_303['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_303_tscode_list.append(stock_name)
stk_rewards_item_303_addname_dataframe=pd.DataFrame()
stk_rewards_item_303_addname_dataframe['cname'] = stk_rewards_item_303_tscode_list
for table_name in stk_rewards_item_303.columns.values.tolist():
    stk_rewards_item_303_addname_dataframe[table_name] = stk_rewards_item_303[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_303_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_304_tscode_list = list() 
stk_rewards_item_304 = pro.stk_rewards(ts_code='300346.SZ,300347.SZ,300348.SZ,300349.SZ,300350.SZ,300351.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_304 返回数据 row 行数 = "+str(stk_rewards_item_304.shape[0]))
for ts_code_sh in stk_rewards_item_304['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_304_tscode_list.append(stock_name)
stk_rewards_item_304_addname_dataframe=pd.DataFrame()
stk_rewards_item_304_addname_dataframe['cname'] = stk_rewards_item_304_tscode_list
for table_name in stk_rewards_item_304.columns.values.tolist():
    stk_rewards_item_304_addname_dataframe[table_name] = stk_rewards_item_304[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_304_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_305_tscode_list = list() 
stk_rewards_item_305 = pro.stk_rewards(ts_code='300352.SZ,300353.SZ,300354.SZ,300355.SZ,300356.SZ,300357.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_305 返回数据 row 行数 = "+str(stk_rewards_item_305.shape[0]))
for ts_code_sh in stk_rewards_item_305['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_305_tscode_list.append(stock_name)
stk_rewards_item_305_addname_dataframe=pd.DataFrame()
stk_rewards_item_305_addname_dataframe['cname'] = stk_rewards_item_305_tscode_list
for table_name in stk_rewards_item_305.columns.values.tolist():
    stk_rewards_item_305_addname_dataframe[table_name] = stk_rewards_item_305[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_305_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_306_tscode_list = list() 
stk_rewards_item_306 = pro.stk_rewards(ts_code='300358.SZ,300359.SZ,300360.SZ,300362.SZ,300363.SZ,300364.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_306 返回数据 row 行数 = "+str(stk_rewards_item_306.shape[0]))
for ts_code_sh in stk_rewards_item_306['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_306_tscode_list.append(stock_name)
stk_rewards_item_306_addname_dataframe=pd.DataFrame()
stk_rewards_item_306_addname_dataframe['cname'] = stk_rewards_item_306_tscode_list
for table_name in stk_rewards_item_306.columns.values.tolist():
    stk_rewards_item_306_addname_dataframe[table_name] = stk_rewards_item_306[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_306_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_307_tscode_list = list() 
stk_rewards_item_307 = pro.stk_rewards(ts_code='300365.SZ,300366.SZ,300367.SZ,300368.SZ,300369.SZ,300370.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_307 返回数据 row 行数 = "+str(stk_rewards_item_307.shape[0]))
for ts_code_sh in stk_rewards_item_307['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_307_tscode_list.append(stock_name)
stk_rewards_item_307_addname_dataframe=pd.DataFrame()
stk_rewards_item_307_addname_dataframe['cname'] = stk_rewards_item_307_tscode_list
for table_name in stk_rewards_item_307.columns.values.tolist():
    stk_rewards_item_307_addname_dataframe[table_name] = stk_rewards_item_307[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_307_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_308_tscode_list = list() 
stk_rewards_item_308 = pro.stk_rewards(ts_code='300371.SZ,300372.SZ,300373.SZ,300374.SZ,300375.SZ,300376.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_308 返回数据 row 行数 = "+str(stk_rewards_item_308.shape[0]))
for ts_code_sh in stk_rewards_item_308['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_308_tscode_list.append(stock_name)
stk_rewards_item_308_addname_dataframe=pd.DataFrame()
stk_rewards_item_308_addname_dataframe['cname'] = stk_rewards_item_308_tscode_list
for table_name in stk_rewards_item_308.columns.values.tolist():
    stk_rewards_item_308_addname_dataframe[table_name] = stk_rewards_item_308[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_308_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_309_tscode_list = list() 
stk_rewards_item_309 = pro.stk_rewards(ts_code='300377.SZ,300378.SZ,300379.SZ,300380.SZ,300381.SZ,300382.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_309 返回数据 row 行数 = "+str(stk_rewards_item_309.shape[0]))
for ts_code_sh in stk_rewards_item_309['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_309_tscode_list.append(stock_name)
stk_rewards_item_309_addname_dataframe=pd.DataFrame()
stk_rewards_item_309_addname_dataframe['cname'] = stk_rewards_item_309_tscode_list
for table_name in stk_rewards_item_309.columns.values.tolist():
    stk_rewards_item_309_addname_dataframe[table_name] = stk_rewards_item_309[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_309_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_310_tscode_list = list() 
stk_rewards_item_310 = pro.stk_rewards(ts_code='300383.SZ,300384.SZ,300385.SZ,300386.SZ,300387.SZ,300388.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_310 返回数据 row 行数 = "+str(stk_rewards_item_310.shape[0]))
for ts_code_sh in stk_rewards_item_310['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_310_tscode_list.append(stock_name)
stk_rewards_item_310_addname_dataframe=pd.DataFrame()
stk_rewards_item_310_addname_dataframe['cname'] = stk_rewards_item_310_tscode_list
for table_name in stk_rewards_item_310.columns.values.tolist():
    stk_rewards_item_310_addname_dataframe[table_name] = stk_rewards_item_310[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_310_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_311_tscode_list = list() 
stk_rewards_item_311 = pro.stk_rewards(ts_code='300389.SZ,300390.SZ,300391.SZ,300392.SZ,300393.SZ,300394.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_311 返回数据 row 行数 = "+str(stk_rewards_item_311.shape[0]))
for ts_code_sh in stk_rewards_item_311['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_311_tscode_list.append(stock_name)
stk_rewards_item_311_addname_dataframe=pd.DataFrame()
stk_rewards_item_311_addname_dataframe['cname'] = stk_rewards_item_311_tscode_list
for table_name in stk_rewards_item_311.columns.values.tolist():
    stk_rewards_item_311_addname_dataframe[table_name] = stk_rewards_item_311[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_311_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_312_tscode_list = list() 
stk_rewards_item_312 = pro.stk_rewards(ts_code='300395.SZ,300396.SZ,300397.SZ,300398.SZ,300399.SZ,300400.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_312 返回数据 row 行数 = "+str(stk_rewards_item_312.shape[0]))
for ts_code_sh in stk_rewards_item_312['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_312_tscode_list.append(stock_name)
stk_rewards_item_312_addname_dataframe=pd.DataFrame()
stk_rewards_item_312_addname_dataframe['cname'] = stk_rewards_item_312_tscode_list
for table_name in stk_rewards_item_312.columns.values.tolist():
    stk_rewards_item_312_addname_dataframe[table_name] = stk_rewards_item_312[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_312_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_313_tscode_list = list() 
stk_rewards_item_313 = pro.stk_rewards(ts_code='300401.SZ,300402.SZ,300403.SZ,300404.SZ,300405.SZ,300406.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_313 返回数据 row 行数 = "+str(stk_rewards_item_313.shape[0]))
for ts_code_sh in stk_rewards_item_313['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_313_tscode_list.append(stock_name)
stk_rewards_item_313_addname_dataframe=pd.DataFrame()
stk_rewards_item_313_addname_dataframe['cname'] = stk_rewards_item_313_tscode_list
for table_name in stk_rewards_item_313.columns.values.tolist():
    stk_rewards_item_313_addname_dataframe[table_name] = stk_rewards_item_313[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_313_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_314_tscode_list = list() 
stk_rewards_item_314 = pro.stk_rewards(ts_code='300407.SZ,300408.SZ,300409.SZ,300410.SZ,300411.SZ,300412.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_314 返回数据 row 行数 = "+str(stk_rewards_item_314.shape[0]))
for ts_code_sh in stk_rewards_item_314['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_314_tscode_list.append(stock_name)
stk_rewards_item_314_addname_dataframe=pd.DataFrame()
stk_rewards_item_314_addname_dataframe['cname'] = stk_rewards_item_314_tscode_list
for table_name in stk_rewards_item_314.columns.values.tolist():
    stk_rewards_item_314_addname_dataframe[table_name] = stk_rewards_item_314[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_314_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_315_tscode_list = list() 
stk_rewards_item_315 = pro.stk_rewards(ts_code='300413.SZ,300414.SZ,300415.SZ,300416.SZ,300417.SZ,300418.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_315 返回数据 row 行数 = "+str(stk_rewards_item_315.shape[0]))
for ts_code_sh in stk_rewards_item_315['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_315_tscode_list.append(stock_name)
stk_rewards_item_315_addname_dataframe=pd.DataFrame()
stk_rewards_item_315_addname_dataframe['cname'] = stk_rewards_item_315_tscode_list
for table_name in stk_rewards_item_315.columns.values.tolist():
    stk_rewards_item_315_addname_dataframe[table_name] = stk_rewards_item_315[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_315_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_316_tscode_list = list() 
stk_rewards_item_316 = pro.stk_rewards(ts_code='300419.SZ,300420.SZ,300421.SZ,300422.SZ,300423.SZ,300424.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_316 返回数据 row 行数 = "+str(stk_rewards_item_316.shape[0]))
for ts_code_sh in stk_rewards_item_316['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_316_tscode_list.append(stock_name)
stk_rewards_item_316_addname_dataframe=pd.DataFrame()
stk_rewards_item_316_addname_dataframe['cname'] = stk_rewards_item_316_tscode_list
for table_name in stk_rewards_item_316.columns.values.tolist():
    stk_rewards_item_316_addname_dataframe[table_name] = stk_rewards_item_316[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_316_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_317_tscode_list = list() 
stk_rewards_item_317 = pro.stk_rewards(ts_code='300425.SZ,300426.SZ,300427.SZ,300428.SZ,300429.SZ,300430.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_317 返回数据 row 行数 = "+str(stk_rewards_item_317.shape[0]))
for ts_code_sh in stk_rewards_item_317['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_317_tscode_list.append(stock_name)
stk_rewards_item_317_addname_dataframe=pd.DataFrame()
stk_rewards_item_317_addname_dataframe['cname'] = stk_rewards_item_317_tscode_list
for table_name in stk_rewards_item_317.columns.values.tolist():
    stk_rewards_item_317_addname_dataframe[table_name] = stk_rewards_item_317[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_317_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_318_tscode_list = list() 
stk_rewards_item_318 = pro.stk_rewards(ts_code='300431.SZ,300432.SZ,300433.SZ,300434.SZ,300435.SZ,300436.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_318 返回数据 row 行数 = "+str(stk_rewards_item_318.shape[0]))
for ts_code_sh in stk_rewards_item_318['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_318_tscode_list.append(stock_name)
stk_rewards_item_318_addname_dataframe=pd.DataFrame()
stk_rewards_item_318_addname_dataframe['cname'] = stk_rewards_item_318_tscode_list
for table_name in stk_rewards_item_318.columns.values.tolist():
    stk_rewards_item_318_addname_dataframe[table_name] = stk_rewards_item_318[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_318_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_319_tscode_list = list() 
stk_rewards_item_319 = pro.stk_rewards(ts_code='300437.SZ,300438.SZ,300439.SZ,300440.SZ,300441.SZ,300442.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_319 返回数据 row 行数 = "+str(stk_rewards_item_319.shape[0]))
for ts_code_sh in stk_rewards_item_319['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_319_tscode_list.append(stock_name)
stk_rewards_item_319_addname_dataframe=pd.DataFrame()
stk_rewards_item_319_addname_dataframe['cname'] = stk_rewards_item_319_tscode_list
for table_name in stk_rewards_item_319.columns.values.tolist():
    stk_rewards_item_319_addname_dataframe[table_name] = stk_rewards_item_319[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_319_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_320_tscode_list = list() 
stk_rewards_item_320 = pro.stk_rewards(ts_code='300443.SZ,300444.SZ,300445.SZ,300446.SZ,300447.SZ,300448.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_320 返回数据 row 行数 = "+str(stk_rewards_item_320.shape[0]))
for ts_code_sh in stk_rewards_item_320['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_320_tscode_list.append(stock_name)
stk_rewards_item_320_addname_dataframe=pd.DataFrame()
stk_rewards_item_320_addname_dataframe['cname'] = stk_rewards_item_320_tscode_list
for table_name in stk_rewards_item_320.columns.values.tolist():
    stk_rewards_item_320_addname_dataframe[table_name] = stk_rewards_item_320[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_320_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_321_tscode_list = list() 
stk_rewards_item_321 = pro.stk_rewards(ts_code='300449.SZ,300450.SZ,300451.SZ,300452.SZ,300453.SZ,300454.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_321 返回数据 row 行数 = "+str(stk_rewards_item_321.shape[0]))
for ts_code_sh in stk_rewards_item_321['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_321_tscode_list.append(stock_name)
stk_rewards_item_321_addname_dataframe=pd.DataFrame()
stk_rewards_item_321_addname_dataframe['cname'] = stk_rewards_item_321_tscode_list
for table_name in stk_rewards_item_321.columns.values.tolist():
    stk_rewards_item_321_addname_dataframe[table_name] = stk_rewards_item_321[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_321_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_322_tscode_list = list() 
stk_rewards_item_322 = pro.stk_rewards(ts_code='300455.SZ,300456.SZ,300457.SZ,300458.SZ,300459.SZ,300460.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_322 返回数据 row 行数 = "+str(stk_rewards_item_322.shape[0]))
for ts_code_sh in stk_rewards_item_322['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_322_tscode_list.append(stock_name)
stk_rewards_item_322_addname_dataframe=pd.DataFrame()
stk_rewards_item_322_addname_dataframe['cname'] = stk_rewards_item_322_tscode_list
for table_name in stk_rewards_item_322.columns.values.tolist():
    stk_rewards_item_322_addname_dataframe[table_name] = stk_rewards_item_322[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_322_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_323_tscode_list = list() 
stk_rewards_item_323 = pro.stk_rewards(ts_code='300461.SZ,300462.SZ,300463.SZ,300464.SZ,300465.SZ,300466.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_323 返回数据 row 行数 = "+str(stk_rewards_item_323.shape[0]))
for ts_code_sh in stk_rewards_item_323['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_323_tscode_list.append(stock_name)
stk_rewards_item_323_addname_dataframe=pd.DataFrame()
stk_rewards_item_323_addname_dataframe['cname'] = stk_rewards_item_323_tscode_list
for table_name in stk_rewards_item_323.columns.values.tolist():
    stk_rewards_item_323_addname_dataframe[table_name] = stk_rewards_item_323[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_323_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_324_tscode_list = list() 
stk_rewards_item_324 = pro.stk_rewards(ts_code='300467.SZ,300468.SZ,300469.SZ,300470.SZ,300471.SZ,300472.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_324 返回数据 row 行数 = "+str(stk_rewards_item_324.shape[0]))
for ts_code_sh in stk_rewards_item_324['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_324_tscode_list.append(stock_name)
stk_rewards_item_324_addname_dataframe=pd.DataFrame()
stk_rewards_item_324_addname_dataframe['cname'] = stk_rewards_item_324_tscode_list
for table_name in stk_rewards_item_324.columns.values.tolist():
    stk_rewards_item_324_addname_dataframe[table_name] = stk_rewards_item_324[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_324_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_325_tscode_list = list() 
stk_rewards_item_325 = pro.stk_rewards(ts_code='300473.SZ,300474.SZ,300475.SZ,300476.SZ,300477.SZ,300478.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_325 返回数据 row 行数 = "+str(stk_rewards_item_325.shape[0]))
for ts_code_sh in stk_rewards_item_325['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_325_tscode_list.append(stock_name)
stk_rewards_item_325_addname_dataframe=pd.DataFrame()
stk_rewards_item_325_addname_dataframe['cname'] = stk_rewards_item_325_tscode_list
for table_name in stk_rewards_item_325.columns.values.tolist():
    stk_rewards_item_325_addname_dataframe[table_name] = stk_rewards_item_325[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_325_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_326_tscode_list = list() 
stk_rewards_item_326 = pro.stk_rewards(ts_code='300479.SZ,300480.SZ,300481.SZ,300482.SZ,300483.SZ,300484.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_326 返回数据 row 行数 = "+str(stk_rewards_item_326.shape[0]))
for ts_code_sh in stk_rewards_item_326['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_326_tscode_list.append(stock_name)
stk_rewards_item_326_addname_dataframe=pd.DataFrame()
stk_rewards_item_326_addname_dataframe['cname'] = stk_rewards_item_326_tscode_list
for table_name in stk_rewards_item_326.columns.values.tolist():
    stk_rewards_item_326_addname_dataframe[table_name] = stk_rewards_item_326[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_326_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_327_tscode_list = list() 
stk_rewards_item_327 = pro.stk_rewards(ts_code='300485.SZ,300486.SZ,300487.SZ,300488.SZ,300489.SZ,300490.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_327 返回数据 row 行数 = "+str(stk_rewards_item_327.shape[0]))
for ts_code_sh in stk_rewards_item_327['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_327_tscode_list.append(stock_name)
stk_rewards_item_327_addname_dataframe=pd.DataFrame()
stk_rewards_item_327_addname_dataframe['cname'] = stk_rewards_item_327_tscode_list
for table_name in stk_rewards_item_327.columns.values.tolist():
    stk_rewards_item_327_addname_dataframe[table_name] = stk_rewards_item_327[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_327_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_328_tscode_list = list() 
stk_rewards_item_328 = pro.stk_rewards(ts_code='300491.SZ,300492.SZ,300493.SZ,300494.SZ,300495.SZ,300496.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_328 返回数据 row 行数 = "+str(stk_rewards_item_328.shape[0]))
for ts_code_sh in stk_rewards_item_328['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_328_tscode_list.append(stock_name)
stk_rewards_item_328_addname_dataframe=pd.DataFrame()
stk_rewards_item_328_addname_dataframe['cname'] = stk_rewards_item_328_tscode_list
for table_name in stk_rewards_item_328.columns.values.tolist():
    stk_rewards_item_328_addname_dataframe[table_name] = stk_rewards_item_328[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_328_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_329_tscode_list = list() 
stk_rewards_item_329 = pro.stk_rewards(ts_code='300497.SZ,300498.SZ,300499.SZ,300500.SZ,300501.SZ,300502.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_329 返回数据 row 行数 = "+str(stk_rewards_item_329.shape[0]))
for ts_code_sh in stk_rewards_item_329['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_329_tscode_list.append(stock_name)
stk_rewards_item_329_addname_dataframe=pd.DataFrame()
stk_rewards_item_329_addname_dataframe['cname'] = stk_rewards_item_329_tscode_list
for table_name in stk_rewards_item_329.columns.values.tolist():
    stk_rewards_item_329_addname_dataframe[table_name] = stk_rewards_item_329[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_329_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_330_tscode_list = list() 
stk_rewards_item_330 = pro.stk_rewards(ts_code='300503.SZ,300504.SZ,300505.SZ,300506.SZ,300507.SZ,300508.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_330 返回数据 row 行数 = "+str(stk_rewards_item_330.shape[0]))
for ts_code_sh in stk_rewards_item_330['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_330_tscode_list.append(stock_name)
stk_rewards_item_330_addname_dataframe=pd.DataFrame()
stk_rewards_item_330_addname_dataframe['cname'] = stk_rewards_item_330_tscode_list
for table_name in stk_rewards_item_330.columns.values.tolist():
    stk_rewards_item_330_addname_dataframe[table_name] = stk_rewards_item_330[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_330_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_331_tscode_list = list() 
stk_rewards_item_331 = pro.stk_rewards(ts_code='300509.SZ,300510.SZ,300511.SZ,300512.SZ,300513.SZ,300514.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_331 返回数据 row 行数 = "+str(stk_rewards_item_331.shape[0]))
for ts_code_sh in stk_rewards_item_331['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_331_tscode_list.append(stock_name)
stk_rewards_item_331_addname_dataframe=pd.DataFrame()
stk_rewards_item_331_addname_dataframe['cname'] = stk_rewards_item_331_tscode_list
for table_name in stk_rewards_item_331.columns.values.tolist():
    stk_rewards_item_331_addname_dataframe[table_name] = stk_rewards_item_331[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_331_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_332_tscode_list = list() 
stk_rewards_item_332 = pro.stk_rewards(ts_code='300515.SZ,300516.SZ,300517.SZ,300518.SZ,300519.SZ,300520.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_332 返回数据 row 行数 = "+str(stk_rewards_item_332.shape[0]))
for ts_code_sh in stk_rewards_item_332['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_332_tscode_list.append(stock_name)
stk_rewards_item_332_addname_dataframe=pd.DataFrame()
stk_rewards_item_332_addname_dataframe['cname'] = stk_rewards_item_332_tscode_list
for table_name in stk_rewards_item_332.columns.values.tolist():
    stk_rewards_item_332_addname_dataframe[table_name] = stk_rewards_item_332[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_332_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_333_tscode_list = list() 
stk_rewards_item_333 = pro.stk_rewards(ts_code='300521.SZ,300522.SZ,300523.SZ,300525.SZ,300526.SZ,300527.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_333 返回数据 row 行数 = "+str(stk_rewards_item_333.shape[0]))
for ts_code_sh in stk_rewards_item_333['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_333_tscode_list.append(stock_name)
stk_rewards_item_333_addname_dataframe=pd.DataFrame()
stk_rewards_item_333_addname_dataframe['cname'] = stk_rewards_item_333_tscode_list
for table_name in stk_rewards_item_333.columns.values.tolist():
    stk_rewards_item_333_addname_dataframe[table_name] = stk_rewards_item_333[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_333_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_334_tscode_list = list() 
stk_rewards_item_334 = pro.stk_rewards(ts_code='300528.SZ,300529.SZ,300530.SZ,300531.SZ,300532.SZ,300533.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_334 返回数据 row 行数 = "+str(stk_rewards_item_334.shape[0]))
for ts_code_sh in stk_rewards_item_334['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_334_tscode_list.append(stock_name)
stk_rewards_item_334_addname_dataframe=pd.DataFrame()
stk_rewards_item_334_addname_dataframe['cname'] = stk_rewards_item_334_tscode_list
for table_name in stk_rewards_item_334.columns.values.tolist():
    stk_rewards_item_334_addname_dataframe[table_name] = stk_rewards_item_334[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_334_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_335_tscode_list = list() 
stk_rewards_item_335 = pro.stk_rewards(ts_code='300534.SZ,300535.SZ,300536.SZ,300537.SZ,300538.SZ,300539.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_335 返回数据 row 行数 = "+str(stk_rewards_item_335.shape[0]))
for ts_code_sh in stk_rewards_item_335['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_335_tscode_list.append(stock_name)
stk_rewards_item_335_addname_dataframe=pd.DataFrame()
stk_rewards_item_335_addname_dataframe['cname'] = stk_rewards_item_335_tscode_list
for table_name in stk_rewards_item_335.columns.values.tolist():
    stk_rewards_item_335_addname_dataframe[table_name] = stk_rewards_item_335[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_335_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_336_tscode_list = list() 
stk_rewards_item_336 = pro.stk_rewards(ts_code='300540.SZ,300541.SZ,300542.SZ,300543.SZ,300545.SZ,300546.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_336 返回数据 row 行数 = "+str(stk_rewards_item_336.shape[0]))
for ts_code_sh in stk_rewards_item_336['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_336_tscode_list.append(stock_name)
stk_rewards_item_336_addname_dataframe=pd.DataFrame()
stk_rewards_item_336_addname_dataframe['cname'] = stk_rewards_item_336_tscode_list
for table_name in stk_rewards_item_336.columns.values.tolist():
    stk_rewards_item_336_addname_dataframe[table_name] = stk_rewards_item_336[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_336_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_337_tscode_list = list() 
stk_rewards_item_337 = pro.stk_rewards(ts_code='300547.SZ,300548.SZ,300549.SZ,300550.SZ,300551.SZ,300552.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_337 返回数据 row 行数 = "+str(stk_rewards_item_337.shape[0]))
for ts_code_sh in stk_rewards_item_337['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_337_tscode_list.append(stock_name)
stk_rewards_item_337_addname_dataframe=pd.DataFrame()
stk_rewards_item_337_addname_dataframe['cname'] = stk_rewards_item_337_tscode_list
for table_name in stk_rewards_item_337.columns.values.tolist():
    stk_rewards_item_337_addname_dataframe[table_name] = stk_rewards_item_337[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_337_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_338_tscode_list = list() 
stk_rewards_item_338 = pro.stk_rewards(ts_code='300553.SZ,300554.SZ,300555.SZ,300556.SZ,300557.SZ,300558.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_338 返回数据 row 行数 = "+str(stk_rewards_item_338.shape[0]))
for ts_code_sh in stk_rewards_item_338['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_338_tscode_list.append(stock_name)
stk_rewards_item_338_addname_dataframe=pd.DataFrame()
stk_rewards_item_338_addname_dataframe['cname'] = stk_rewards_item_338_tscode_list
for table_name in stk_rewards_item_338.columns.values.tolist():
    stk_rewards_item_338_addname_dataframe[table_name] = stk_rewards_item_338[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_338_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_339_tscode_list = list() 
stk_rewards_item_339 = pro.stk_rewards(ts_code='300559.SZ,300560.SZ,300561.SZ,300562.SZ,300563.SZ,300564.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_339 返回数据 row 行数 = "+str(stk_rewards_item_339.shape[0]))
for ts_code_sh in stk_rewards_item_339['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_339_tscode_list.append(stock_name)
stk_rewards_item_339_addname_dataframe=pd.DataFrame()
stk_rewards_item_339_addname_dataframe['cname'] = stk_rewards_item_339_tscode_list
for table_name in stk_rewards_item_339.columns.values.tolist():
    stk_rewards_item_339_addname_dataframe[table_name] = stk_rewards_item_339[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_339_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_340_tscode_list = list() 
stk_rewards_item_340 = pro.stk_rewards(ts_code='300565.SZ,300566.SZ,300567.SZ,300568.SZ,300569.SZ,300570.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_340 返回数据 row 行数 = "+str(stk_rewards_item_340.shape[0]))
for ts_code_sh in stk_rewards_item_340['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_340_tscode_list.append(stock_name)
stk_rewards_item_340_addname_dataframe=pd.DataFrame()
stk_rewards_item_340_addname_dataframe['cname'] = stk_rewards_item_340_tscode_list
for table_name in stk_rewards_item_340.columns.values.tolist():
    stk_rewards_item_340_addname_dataframe[table_name] = stk_rewards_item_340[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_340_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_341_tscode_list = list() 
stk_rewards_item_341 = pro.stk_rewards(ts_code='300571.SZ,300572.SZ,300573.SZ,300575.SZ,300576.SZ,300577.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_341 返回数据 row 行数 = "+str(stk_rewards_item_341.shape[0]))
for ts_code_sh in stk_rewards_item_341['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_341_tscode_list.append(stock_name)
stk_rewards_item_341_addname_dataframe=pd.DataFrame()
stk_rewards_item_341_addname_dataframe['cname'] = stk_rewards_item_341_tscode_list
for table_name in stk_rewards_item_341.columns.values.tolist():
    stk_rewards_item_341_addname_dataframe[table_name] = stk_rewards_item_341[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_341_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_342_tscode_list = list() 
stk_rewards_item_342 = pro.stk_rewards(ts_code='300578.SZ,300579.SZ,300580.SZ,300581.SZ,300582.SZ,300583.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_342 返回数据 row 行数 = "+str(stk_rewards_item_342.shape[0]))
for ts_code_sh in stk_rewards_item_342['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_342_tscode_list.append(stock_name)
stk_rewards_item_342_addname_dataframe=pd.DataFrame()
stk_rewards_item_342_addname_dataframe['cname'] = stk_rewards_item_342_tscode_list
for table_name in stk_rewards_item_342.columns.values.tolist():
    stk_rewards_item_342_addname_dataframe[table_name] = stk_rewards_item_342[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_342_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_343_tscode_list = list() 
stk_rewards_item_343 = pro.stk_rewards(ts_code='300584.SZ,300585.SZ,300586.SZ,300587.SZ,300588.SZ,300589.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_343 返回数据 row 行数 = "+str(stk_rewards_item_343.shape[0]))
for ts_code_sh in stk_rewards_item_343['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_343_tscode_list.append(stock_name)
stk_rewards_item_343_addname_dataframe=pd.DataFrame()
stk_rewards_item_343_addname_dataframe['cname'] = stk_rewards_item_343_tscode_list
for table_name in stk_rewards_item_343.columns.values.tolist():
    stk_rewards_item_343_addname_dataframe[table_name] = stk_rewards_item_343[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_343_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_344_tscode_list = list() 
stk_rewards_item_344 = pro.stk_rewards(ts_code='300590.SZ,300591.SZ,300592.SZ,300593.SZ,300594.SZ,300595.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_344 返回数据 row 行数 = "+str(stk_rewards_item_344.shape[0]))
for ts_code_sh in stk_rewards_item_344['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_344_tscode_list.append(stock_name)
stk_rewards_item_344_addname_dataframe=pd.DataFrame()
stk_rewards_item_344_addname_dataframe['cname'] = stk_rewards_item_344_tscode_list
for table_name in stk_rewards_item_344.columns.values.tolist():
    stk_rewards_item_344_addname_dataframe[table_name] = stk_rewards_item_344[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_344_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_345_tscode_list = list() 
stk_rewards_item_345 = pro.stk_rewards(ts_code='300596.SZ,300597.SZ,300598.SZ,300599.SZ,300600.SZ,300601.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_345 返回数据 row 行数 = "+str(stk_rewards_item_345.shape[0]))
for ts_code_sh in stk_rewards_item_345['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_345_tscode_list.append(stock_name)
stk_rewards_item_345_addname_dataframe=pd.DataFrame()
stk_rewards_item_345_addname_dataframe['cname'] = stk_rewards_item_345_tscode_list
for table_name in stk_rewards_item_345.columns.values.tolist():
    stk_rewards_item_345_addname_dataframe[table_name] = stk_rewards_item_345[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_345_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_346_tscode_list = list() 
stk_rewards_item_346 = pro.stk_rewards(ts_code='300602.SZ,300603.SZ,300604.SZ,300605.SZ,300606.SZ,300607.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_346 返回数据 row 行数 = "+str(stk_rewards_item_346.shape[0]))
for ts_code_sh in stk_rewards_item_346['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_346_tscode_list.append(stock_name)
stk_rewards_item_346_addname_dataframe=pd.DataFrame()
stk_rewards_item_346_addname_dataframe['cname'] = stk_rewards_item_346_tscode_list
for table_name in stk_rewards_item_346.columns.values.tolist():
    stk_rewards_item_346_addname_dataframe[table_name] = stk_rewards_item_346[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_346_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_347_tscode_list = list() 
stk_rewards_item_347 = pro.stk_rewards(ts_code='300608.SZ,300609.SZ,300610.SZ,300611.SZ,300612.SZ,300613.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_347 返回数据 row 行数 = "+str(stk_rewards_item_347.shape[0]))
for ts_code_sh in stk_rewards_item_347['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_347_tscode_list.append(stock_name)
stk_rewards_item_347_addname_dataframe=pd.DataFrame()
stk_rewards_item_347_addname_dataframe['cname'] = stk_rewards_item_347_tscode_list
for table_name in stk_rewards_item_347.columns.values.tolist():
    stk_rewards_item_347_addname_dataframe[table_name] = stk_rewards_item_347[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_347_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_348_tscode_list = list() 
stk_rewards_item_348 = pro.stk_rewards(ts_code='300615.SZ,300616.SZ,300617.SZ,300618.SZ,300619.SZ,300620.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_348 返回数据 row 行数 = "+str(stk_rewards_item_348.shape[0]))
for ts_code_sh in stk_rewards_item_348['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_348_tscode_list.append(stock_name)
stk_rewards_item_348_addname_dataframe=pd.DataFrame()
stk_rewards_item_348_addname_dataframe['cname'] = stk_rewards_item_348_tscode_list
for table_name in stk_rewards_item_348.columns.values.tolist():
    stk_rewards_item_348_addname_dataframe[table_name] = stk_rewards_item_348[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_348_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_349_tscode_list = list() 
stk_rewards_item_349 = pro.stk_rewards(ts_code='300621.SZ,300622.SZ,300623.SZ,300624.SZ,300625.SZ,300626.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_349 返回数据 row 行数 = "+str(stk_rewards_item_349.shape[0]))
for ts_code_sh in stk_rewards_item_349['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_349_tscode_list.append(stock_name)
stk_rewards_item_349_addname_dataframe=pd.DataFrame()
stk_rewards_item_349_addname_dataframe['cname'] = stk_rewards_item_349_tscode_list
for table_name in stk_rewards_item_349.columns.values.tolist():
    stk_rewards_item_349_addname_dataframe[table_name] = stk_rewards_item_349[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_349_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_350_tscode_list = list() 
stk_rewards_item_350 = pro.stk_rewards(ts_code='300627.SZ,300628.SZ,300629.SZ,300630.SZ,300631.SZ,300632.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_350 返回数据 row 行数 = "+str(stk_rewards_item_350.shape[0]))
for ts_code_sh in stk_rewards_item_350['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_350_tscode_list.append(stock_name)
stk_rewards_item_350_addname_dataframe=pd.DataFrame()
stk_rewards_item_350_addname_dataframe['cname'] = stk_rewards_item_350_tscode_list
for table_name in stk_rewards_item_350.columns.values.tolist():
    stk_rewards_item_350_addname_dataframe[table_name] = stk_rewards_item_350[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_350_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_351_tscode_list = list() 
stk_rewards_item_351 = pro.stk_rewards(ts_code='300633.SZ,300634.SZ,300635.SZ,300636.SZ,300637.SZ,300638.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_351 返回数据 row 行数 = "+str(stk_rewards_item_351.shape[0]))
for ts_code_sh in stk_rewards_item_351['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_351_tscode_list.append(stock_name)
stk_rewards_item_351_addname_dataframe=pd.DataFrame()
stk_rewards_item_351_addname_dataframe['cname'] = stk_rewards_item_351_tscode_list
for table_name in stk_rewards_item_351.columns.values.tolist():
    stk_rewards_item_351_addname_dataframe[table_name] = stk_rewards_item_351[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_351_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_352_tscode_list = list() 
stk_rewards_item_352 = pro.stk_rewards(ts_code='300639.SZ,300640.SZ,300641.SZ,300642.SZ,300643.SZ,300644.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_352 返回数据 row 行数 = "+str(stk_rewards_item_352.shape[0]))
for ts_code_sh in stk_rewards_item_352['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_352_tscode_list.append(stock_name)
stk_rewards_item_352_addname_dataframe=pd.DataFrame()
stk_rewards_item_352_addname_dataframe['cname'] = stk_rewards_item_352_tscode_list
for table_name in stk_rewards_item_352.columns.values.tolist():
    stk_rewards_item_352_addname_dataframe[table_name] = stk_rewards_item_352[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_352_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_353_tscode_list = list() 
stk_rewards_item_353 = pro.stk_rewards(ts_code='300645.SZ,300647.SZ,300648.SZ,300649.SZ,300650.SZ,300651.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_353 返回数据 row 行数 = "+str(stk_rewards_item_353.shape[0]))
for ts_code_sh in stk_rewards_item_353['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_353_tscode_list.append(stock_name)
stk_rewards_item_353_addname_dataframe=pd.DataFrame()
stk_rewards_item_353_addname_dataframe['cname'] = stk_rewards_item_353_tscode_list
for table_name in stk_rewards_item_353.columns.values.tolist():
    stk_rewards_item_353_addname_dataframe[table_name] = stk_rewards_item_353[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_353_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_354_tscode_list = list() 
stk_rewards_item_354 = pro.stk_rewards(ts_code='300652.SZ,300653.SZ,300654.SZ,300655.SZ,300656.SZ,300657.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_354 返回数据 row 行数 = "+str(stk_rewards_item_354.shape[0]))
for ts_code_sh in stk_rewards_item_354['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_354_tscode_list.append(stock_name)
stk_rewards_item_354_addname_dataframe=pd.DataFrame()
stk_rewards_item_354_addname_dataframe['cname'] = stk_rewards_item_354_tscode_list
for table_name in stk_rewards_item_354.columns.values.tolist():
    stk_rewards_item_354_addname_dataframe[table_name] = stk_rewards_item_354[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_354_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_355_tscode_list = list() 
stk_rewards_item_355 = pro.stk_rewards(ts_code='300658.SZ,300659.SZ,300660.SZ,300661.SZ,300662.SZ,300663.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_355 返回数据 row 行数 = "+str(stk_rewards_item_355.shape[0]))
for ts_code_sh in stk_rewards_item_355['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_355_tscode_list.append(stock_name)
stk_rewards_item_355_addname_dataframe=pd.DataFrame()
stk_rewards_item_355_addname_dataframe['cname'] = stk_rewards_item_355_tscode_list
for table_name in stk_rewards_item_355.columns.values.tolist():
    stk_rewards_item_355_addname_dataframe[table_name] = stk_rewards_item_355[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_355_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_356_tscode_list = list() 
stk_rewards_item_356 = pro.stk_rewards(ts_code='300664.SZ,300665.SZ,300666.SZ,300667.SZ,300668.SZ,300669.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_356 返回数据 row 行数 = "+str(stk_rewards_item_356.shape[0]))
for ts_code_sh in stk_rewards_item_356['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_356_tscode_list.append(stock_name)
stk_rewards_item_356_addname_dataframe=pd.DataFrame()
stk_rewards_item_356_addname_dataframe['cname'] = stk_rewards_item_356_tscode_list
for table_name in stk_rewards_item_356.columns.values.tolist():
    stk_rewards_item_356_addname_dataframe[table_name] = stk_rewards_item_356[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_356_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_357_tscode_list = list() 
stk_rewards_item_357 = pro.stk_rewards(ts_code='300670.SZ,300671.SZ,300672.SZ,300673.SZ,300674.SZ,300675.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_357 返回数据 row 行数 = "+str(stk_rewards_item_357.shape[0]))
for ts_code_sh in stk_rewards_item_357['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_357_tscode_list.append(stock_name)
stk_rewards_item_357_addname_dataframe=pd.DataFrame()
stk_rewards_item_357_addname_dataframe['cname'] = stk_rewards_item_357_tscode_list
for table_name in stk_rewards_item_357.columns.values.tolist():
    stk_rewards_item_357_addname_dataframe[table_name] = stk_rewards_item_357[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_357_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_358_tscode_list = list() 
stk_rewards_item_358 = pro.stk_rewards(ts_code='300676.SZ,300677.SZ,300678.SZ,300679.SZ,300680.SZ,300681.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_358 返回数据 row 行数 = "+str(stk_rewards_item_358.shape[0]))
for ts_code_sh in stk_rewards_item_358['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_358_tscode_list.append(stock_name)
stk_rewards_item_358_addname_dataframe=pd.DataFrame()
stk_rewards_item_358_addname_dataframe['cname'] = stk_rewards_item_358_tscode_list
for table_name in stk_rewards_item_358.columns.values.tolist():
    stk_rewards_item_358_addname_dataframe[table_name] = stk_rewards_item_358[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_358_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_359_tscode_list = list() 
stk_rewards_item_359 = pro.stk_rewards(ts_code='300682.SZ,300683.SZ,300684.SZ,300685.SZ,300686.SZ,300687.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_359 返回数据 row 行数 = "+str(stk_rewards_item_359.shape[0]))
for ts_code_sh in stk_rewards_item_359['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_359_tscode_list.append(stock_name)
stk_rewards_item_359_addname_dataframe=pd.DataFrame()
stk_rewards_item_359_addname_dataframe['cname'] = stk_rewards_item_359_tscode_list
for table_name in stk_rewards_item_359.columns.values.tolist():
    stk_rewards_item_359_addname_dataframe[table_name] = stk_rewards_item_359[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_359_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_360_tscode_list = list() 
stk_rewards_item_360 = pro.stk_rewards(ts_code='300688.SZ,300689.SZ,300690.SZ,300691.SZ,300692.SZ,300693.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_360 返回数据 row 行数 = "+str(stk_rewards_item_360.shape[0]))
for ts_code_sh in stk_rewards_item_360['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_360_tscode_list.append(stock_name)
stk_rewards_item_360_addname_dataframe=pd.DataFrame()
stk_rewards_item_360_addname_dataframe['cname'] = stk_rewards_item_360_tscode_list
for table_name in stk_rewards_item_360.columns.values.tolist():
    stk_rewards_item_360_addname_dataframe[table_name] = stk_rewards_item_360[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_360_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_361_tscode_list = list() 
stk_rewards_item_361 = pro.stk_rewards(ts_code='300694.SZ,300695.SZ,300696.SZ,300697.SZ,300698.SZ,300699.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_361 返回数据 row 行数 = "+str(stk_rewards_item_361.shape[0]))
for ts_code_sh in stk_rewards_item_361['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_361_tscode_list.append(stock_name)
stk_rewards_item_361_addname_dataframe=pd.DataFrame()
stk_rewards_item_361_addname_dataframe['cname'] = stk_rewards_item_361_tscode_list
for table_name in stk_rewards_item_361.columns.values.tolist():
    stk_rewards_item_361_addname_dataframe[table_name] = stk_rewards_item_361[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_361_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_362_tscode_list = list() 
stk_rewards_item_362 = pro.stk_rewards(ts_code='300700.SZ,300701.SZ,300702.SZ,300703.SZ,300705.SZ,300706.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_362 返回数据 row 行数 = "+str(stk_rewards_item_362.shape[0]))
for ts_code_sh in stk_rewards_item_362['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_362_tscode_list.append(stock_name)
stk_rewards_item_362_addname_dataframe=pd.DataFrame()
stk_rewards_item_362_addname_dataframe['cname'] = stk_rewards_item_362_tscode_list
for table_name in stk_rewards_item_362.columns.values.tolist():
    stk_rewards_item_362_addname_dataframe[table_name] = stk_rewards_item_362[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_362_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_363_tscode_list = list() 
stk_rewards_item_363 = pro.stk_rewards(ts_code='300707.SZ,300708.SZ,300709.SZ,300710.SZ,300711.SZ,300712.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_363 返回数据 row 行数 = "+str(stk_rewards_item_363.shape[0]))
for ts_code_sh in stk_rewards_item_363['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_363_tscode_list.append(stock_name)
stk_rewards_item_363_addname_dataframe=pd.DataFrame()
stk_rewards_item_363_addname_dataframe['cname'] = stk_rewards_item_363_tscode_list
for table_name in stk_rewards_item_363.columns.values.tolist():
    stk_rewards_item_363_addname_dataframe[table_name] = stk_rewards_item_363[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_363_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_364_tscode_list = list() 
stk_rewards_item_364 = pro.stk_rewards(ts_code='300713.SZ,300715.SZ,300716.SZ,300717.SZ,300718.SZ,300719.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_364 返回数据 row 行数 = "+str(stk_rewards_item_364.shape[0]))
for ts_code_sh in stk_rewards_item_364['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_364_tscode_list.append(stock_name)
stk_rewards_item_364_addname_dataframe=pd.DataFrame()
stk_rewards_item_364_addname_dataframe['cname'] = stk_rewards_item_364_tscode_list
for table_name in stk_rewards_item_364.columns.values.tolist():
    stk_rewards_item_364_addname_dataframe[table_name] = stk_rewards_item_364[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_364_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_365_tscode_list = list() 
stk_rewards_item_365 = pro.stk_rewards(ts_code='300720.SZ,300721.SZ,300722.SZ,300723.SZ,300724.SZ,300725.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_365 返回数据 row 行数 = "+str(stk_rewards_item_365.shape[0]))
for ts_code_sh in stk_rewards_item_365['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_365_tscode_list.append(stock_name)
stk_rewards_item_365_addname_dataframe=pd.DataFrame()
stk_rewards_item_365_addname_dataframe['cname'] = stk_rewards_item_365_tscode_list
for table_name in stk_rewards_item_365.columns.values.tolist():
    stk_rewards_item_365_addname_dataframe[table_name] = stk_rewards_item_365[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_365_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_366_tscode_list = list() 
stk_rewards_item_366 = pro.stk_rewards(ts_code='300726.SZ,300727.SZ,300729.SZ,300730.SZ,300731.SZ,300732.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_366 返回数据 row 行数 = "+str(stk_rewards_item_366.shape[0]))
for ts_code_sh in stk_rewards_item_366['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_366_tscode_list.append(stock_name)
stk_rewards_item_366_addname_dataframe=pd.DataFrame()
stk_rewards_item_366_addname_dataframe['cname'] = stk_rewards_item_366_tscode_list
for table_name in stk_rewards_item_366.columns.values.tolist():
    stk_rewards_item_366_addname_dataframe[table_name] = stk_rewards_item_366[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_366_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_367_tscode_list = list() 
stk_rewards_item_367 = pro.stk_rewards(ts_code='300733.SZ,300735.SZ,300736.SZ,300737.SZ,300738.SZ,300739.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_367 返回数据 row 行数 = "+str(stk_rewards_item_367.shape[0]))
for ts_code_sh in stk_rewards_item_367['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_367_tscode_list.append(stock_name)
stk_rewards_item_367_addname_dataframe=pd.DataFrame()
stk_rewards_item_367_addname_dataframe['cname'] = stk_rewards_item_367_tscode_list
for table_name in stk_rewards_item_367.columns.values.tolist():
    stk_rewards_item_367_addname_dataframe[table_name] = stk_rewards_item_367[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_367_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_368_tscode_list = list() 
stk_rewards_item_368 = pro.stk_rewards(ts_code='300740.SZ,300741.SZ,300742.SZ,300743.SZ,300745.SZ,300746.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_368 返回数据 row 行数 = "+str(stk_rewards_item_368.shape[0]))
for ts_code_sh in stk_rewards_item_368['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_368_tscode_list.append(stock_name)
stk_rewards_item_368_addname_dataframe=pd.DataFrame()
stk_rewards_item_368_addname_dataframe['cname'] = stk_rewards_item_368_tscode_list
for table_name in stk_rewards_item_368.columns.values.tolist():
    stk_rewards_item_368_addname_dataframe[table_name] = stk_rewards_item_368[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_368_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_369_tscode_list = list() 
stk_rewards_item_369 = pro.stk_rewards(ts_code='300747.SZ,300748.SZ,300749.SZ,300750.SZ,300751.SZ,300752.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_369 返回数据 row 行数 = "+str(stk_rewards_item_369.shape[0]))
for ts_code_sh in stk_rewards_item_369['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_369_tscode_list.append(stock_name)
stk_rewards_item_369_addname_dataframe=pd.DataFrame()
stk_rewards_item_369_addname_dataframe['cname'] = stk_rewards_item_369_tscode_list
for table_name in stk_rewards_item_369.columns.values.tolist():
    stk_rewards_item_369_addname_dataframe[table_name] = stk_rewards_item_369[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_369_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_370_tscode_list = list() 
stk_rewards_item_370 = pro.stk_rewards(ts_code='300753.SZ,300755.SZ,300756.SZ,300757.SZ,300758.SZ,300759.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_370 返回数据 row 行数 = "+str(stk_rewards_item_370.shape[0]))
for ts_code_sh in stk_rewards_item_370['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_370_tscode_list.append(stock_name)
stk_rewards_item_370_addname_dataframe=pd.DataFrame()
stk_rewards_item_370_addname_dataframe['cname'] = stk_rewards_item_370_tscode_list
for table_name in stk_rewards_item_370.columns.values.tolist():
    stk_rewards_item_370_addname_dataframe[table_name] = stk_rewards_item_370[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_370_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_371_tscode_list = list() 
stk_rewards_item_371 = pro.stk_rewards(ts_code='300760.SZ,300761.SZ,300762.SZ,300763.SZ,300765.SZ,300766.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_371 返回数据 row 行数 = "+str(stk_rewards_item_371.shape[0]))
for ts_code_sh in stk_rewards_item_371['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_371_tscode_list.append(stock_name)
stk_rewards_item_371_addname_dataframe=pd.DataFrame()
stk_rewards_item_371_addname_dataframe['cname'] = stk_rewards_item_371_tscode_list
for table_name in stk_rewards_item_371.columns.values.tolist():
    stk_rewards_item_371_addname_dataframe[table_name] = stk_rewards_item_371[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_371_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_372_tscode_list = list() 
stk_rewards_item_372 = pro.stk_rewards(ts_code='300767.SZ,300768.SZ,300769.SZ,300770.SZ,300771.SZ,300772.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_372 返回数据 row 行数 = "+str(stk_rewards_item_372.shape[0]))
for ts_code_sh in stk_rewards_item_372['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_372_tscode_list.append(stock_name)
stk_rewards_item_372_addname_dataframe=pd.DataFrame()
stk_rewards_item_372_addname_dataframe['cname'] = stk_rewards_item_372_tscode_list
for table_name in stk_rewards_item_372.columns.values.tolist():
    stk_rewards_item_372_addname_dataframe[table_name] = stk_rewards_item_372[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_372_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_373_tscode_list = list() 
stk_rewards_item_373 = pro.stk_rewards(ts_code='300773.SZ,300775.SZ,300776.SZ,300777.SZ,300778.SZ,300779.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_373 返回数据 row 行数 = "+str(stk_rewards_item_373.shape[0]))
for ts_code_sh in stk_rewards_item_373['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_373_tscode_list.append(stock_name)
stk_rewards_item_373_addname_dataframe=pd.DataFrame()
stk_rewards_item_373_addname_dataframe['cname'] = stk_rewards_item_373_tscode_list
for table_name in stk_rewards_item_373.columns.values.tolist():
    stk_rewards_item_373_addname_dataframe[table_name] = stk_rewards_item_373[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_373_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_374_tscode_list = list() 
stk_rewards_item_374 = pro.stk_rewards(ts_code='300780.SZ,300781.SZ,300782.SZ,300783.SZ,300785.SZ,300786.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_374 返回数据 row 行数 = "+str(stk_rewards_item_374.shape[0]))
for ts_code_sh in stk_rewards_item_374['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_374_tscode_list.append(stock_name)
stk_rewards_item_374_addname_dataframe=pd.DataFrame()
stk_rewards_item_374_addname_dataframe['cname'] = stk_rewards_item_374_tscode_list
for table_name in stk_rewards_item_374.columns.values.tolist():
    stk_rewards_item_374_addname_dataframe[table_name] = stk_rewards_item_374[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_374_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_375_tscode_list = list() 
stk_rewards_item_375 = pro.stk_rewards(ts_code='300787.SZ,300788.SZ,300789.SZ,300790.SZ,300791.SZ,300792.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_375 返回数据 row 行数 = "+str(stk_rewards_item_375.shape[0]))
for ts_code_sh in stk_rewards_item_375['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_375_tscode_list.append(stock_name)
stk_rewards_item_375_addname_dataframe=pd.DataFrame()
stk_rewards_item_375_addname_dataframe['cname'] = stk_rewards_item_375_tscode_list
for table_name in stk_rewards_item_375.columns.values.tolist():
    stk_rewards_item_375_addname_dataframe[table_name] = stk_rewards_item_375[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_375_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_376_tscode_list = list() 
stk_rewards_item_376 = pro.stk_rewards(ts_code='300793.SZ,300795.SZ,300796.SZ,300797.SZ,300798.SZ,300799.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_376 返回数据 row 行数 = "+str(stk_rewards_item_376.shape[0]))
for ts_code_sh in stk_rewards_item_376['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_376_tscode_list.append(stock_name)
stk_rewards_item_376_addname_dataframe=pd.DataFrame()
stk_rewards_item_376_addname_dataframe['cname'] = stk_rewards_item_376_tscode_list
for table_name in stk_rewards_item_376.columns.values.tolist():
    stk_rewards_item_376_addname_dataframe[table_name] = stk_rewards_item_376[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_376_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_377_tscode_list = list() 
stk_rewards_item_377 = pro.stk_rewards(ts_code='300800.SZ,300801.SZ,300802.SZ,300803.SZ,300805.SZ,300806.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_377 返回数据 row 行数 = "+str(stk_rewards_item_377.shape[0]))
for ts_code_sh in stk_rewards_item_377['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_377_tscode_list.append(stock_name)
stk_rewards_item_377_addname_dataframe=pd.DataFrame()
stk_rewards_item_377_addname_dataframe['cname'] = stk_rewards_item_377_tscode_list
for table_name in stk_rewards_item_377.columns.values.tolist():
    stk_rewards_item_377_addname_dataframe[table_name] = stk_rewards_item_377[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_377_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_378_tscode_list = list() 
stk_rewards_item_378 = pro.stk_rewards(ts_code='300807.SZ,300808.SZ,300809.SZ,300810.SZ,300811.SZ,300812.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_378 返回数据 row 行数 = "+str(stk_rewards_item_378.shape[0]))
for ts_code_sh in stk_rewards_item_378['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_378_tscode_list.append(stock_name)
stk_rewards_item_378_addname_dataframe=pd.DataFrame()
stk_rewards_item_378_addname_dataframe['cname'] = stk_rewards_item_378_tscode_list
for table_name in stk_rewards_item_378.columns.values.tolist():
    stk_rewards_item_378_addname_dataframe[table_name] = stk_rewards_item_378[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_378_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_379_tscode_list = list() 
stk_rewards_item_379 = pro.stk_rewards(ts_code='300813.SZ,300815.SZ,300816.SZ,300817.SZ,300818.SZ,300819.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_379 返回数据 row 行数 = "+str(stk_rewards_item_379.shape[0]))
for ts_code_sh in stk_rewards_item_379['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_379_tscode_list.append(stock_name)
stk_rewards_item_379_addname_dataframe=pd.DataFrame()
stk_rewards_item_379_addname_dataframe['cname'] = stk_rewards_item_379_tscode_list
for table_name in stk_rewards_item_379.columns.values.tolist():
    stk_rewards_item_379_addname_dataframe[table_name] = stk_rewards_item_379[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_379_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_380_tscode_list = list() 
stk_rewards_item_380 = pro.stk_rewards(ts_code='300820.SZ,300821.SZ,300822.SZ,300823.SZ,300824.SZ,300825.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_380 返回数据 row 行数 = "+str(stk_rewards_item_380.shape[0]))
for ts_code_sh in stk_rewards_item_380['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_380_tscode_list.append(stock_name)
stk_rewards_item_380_addname_dataframe=pd.DataFrame()
stk_rewards_item_380_addname_dataframe['cname'] = stk_rewards_item_380_tscode_list
for table_name in stk_rewards_item_380.columns.values.tolist():
    stk_rewards_item_380_addname_dataframe[table_name] = stk_rewards_item_380[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_380_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_381_tscode_list = list() 
stk_rewards_item_381 = pro.stk_rewards(ts_code='300826.SZ,300827.SZ,300828.SZ,300829.SZ,300830.SZ,300831.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_381 返回数据 row 行数 = "+str(stk_rewards_item_381.shape[0]))
for ts_code_sh in stk_rewards_item_381['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_381_tscode_list.append(stock_name)
stk_rewards_item_381_addname_dataframe=pd.DataFrame()
stk_rewards_item_381_addname_dataframe['cname'] = stk_rewards_item_381_tscode_list
for table_name in stk_rewards_item_381.columns.values.tolist():
    stk_rewards_item_381_addname_dataframe[table_name] = stk_rewards_item_381[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_381_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_382_tscode_list = list() 
stk_rewards_item_382 = pro.stk_rewards(ts_code='300832.SZ,300833.SZ,300835.SZ,300836.SZ,300837.SZ,300838.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_382 返回数据 row 行数 = "+str(stk_rewards_item_382.shape[0]))
for ts_code_sh in stk_rewards_item_382['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_382_tscode_list.append(stock_name)
stk_rewards_item_382_addname_dataframe=pd.DataFrame()
stk_rewards_item_382_addname_dataframe['cname'] = stk_rewards_item_382_tscode_list
for table_name in stk_rewards_item_382.columns.values.tolist():
    stk_rewards_item_382_addname_dataframe[table_name] = stk_rewards_item_382[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_382_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_383_tscode_list = list() 
stk_rewards_item_383 = pro.stk_rewards(ts_code='300839.SZ,300840.SZ,300841.SZ,300842.SZ,300843.SZ,300845.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_383 返回数据 row 行数 = "+str(stk_rewards_item_383.shape[0]))
for ts_code_sh in stk_rewards_item_383['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_383_tscode_list.append(stock_name)
stk_rewards_item_383_addname_dataframe=pd.DataFrame()
stk_rewards_item_383_addname_dataframe['cname'] = stk_rewards_item_383_tscode_list
for table_name in stk_rewards_item_383.columns.values.tolist():
    stk_rewards_item_383_addname_dataframe[table_name] = stk_rewards_item_383[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_383_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_384_tscode_list = list() 
stk_rewards_item_384 = pro.stk_rewards(ts_code='300846.SZ,300847.SZ,300848.SZ,300849.SZ,300850.SZ,300851.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_384 返回数据 row 行数 = "+str(stk_rewards_item_384.shape[0]))
for ts_code_sh in stk_rewards_item_384['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_384_tscode_list.append(stock_name)
stk_rewards_item_384_addname_dataframe=pd.DataFrame()
stk_rewards_item_384_addname_dataframe['cname'] = stk_rewards_item_384_tscode_list
for table_name in stk_rewards_item_384.columns.values.tolist():
    stk_rewards_item_384_addname_dataframe[table_name] = stk_rewards_item_384[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_384_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_385_tscode_list = list() 
stk_rewards_item_385 = pro.stk_rewards(ts_code='300852.SZ,300853.SZ,300855.SZ,300856.SZ,300857.SZ,300858.SZ', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_385 返回数据 row 行数 = "+str(stk_rewards_item_385.shape[0]))
for ts_code_sh in stk_rewards_item_385['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_385_tscode_list.append(stock_name)
stk_rewards_item_385_addname_dataframe=pd.DataFrame()
stk_rewards_item_385_addname_dataframe['cname'] = stk_rewards_item_385_tscode_list
for table_name in stk_rewards_item_385.columns.values.tolist():
    stk_rewards_item_385_addname_dataframe[table_name] = stk_rewards_item_385[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_385_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_386_tscode_list = list() 
stk_rewards_item_386 = pro.stk_rewards(ts_code='300859.SZ,600000.SH,600001.SH,600002.SH,600003.SH,600004.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_386 返回数据 row 行数 = "+str(stk_rewards_item_386.shape[0]))
for ts_code_sh in stk_rewards_item_386['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_386_tscode_list.append(stock_name)
stk_rewards_item_386_addname_dataframe=pd.DataFrame()
stk_rewards_item_386_addname_dataframe['cname'] = stk_rewards_item_386_tscode_list
for table_name in stk_rewards_item_386.columns.values.tolist():
    stk_rewards_item_386_addname_dataframe[table_name] = stk_rewards_item_386[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_386_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_387_tscode_list = list() 
stk_rewards_item_387 = pro.stk_rewards(ts_code='600005.SH,600006.SH,600007.SH,600008.SH,600009.SH,600010.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_387 返回数据 row 行数 = "+str(stk_rewards_item_387.shape[0]))
for ts_code_sh in stk_rewards_item_387['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_387_tscode_list.append(stock_name)
stk_rewards_item_387_addname_dataframe=pd.DataFrame()
stk_rewards_item_387_addname_dataframe['cname'] = stk_rewards_item_387_tscode_list
for table_name in stk_rewards_item_387.columns.values.tolist():
    stk_rewards_item_387_addname_dataframe[table_name] = stk_rewards_item_387[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_387_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_388_tscode_list = list() 
stk_rewards_item_388 = pro.stk_rewards(ts_code='600011.SH,600012.SH,600015.SH,600016.SH,600017.SH,600018.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_388 返回数据 row 行数 = "+str(stk_rewards_item_388.shape[0]))
for ts_code_sh in stk_rewards_item_388['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_388_tscode_list.append(stock_name)
stk_rewards_item_388_addname_dataframe=pd.DataFrame()
stk_rewards_item_388_addname_dataframe['cname'] = stk_rewards_item_388_tscode_list
for table_name in stk_rewards_item_388.columns.values.tolist():
    stk_rewards_item_388_addname_dataframe[table_name] = stk_rewards_item_388[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_388_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_389_tscode_list = list() 
stk_rewards_item_389 = pro.stk_rewards(ts_code='600019.SH,600020.SH,600021.SH,600022.SH,600023.SH,600025.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_389 返回数据 row 行数 = "+str(stk_rewards_item_389.shape[0]))
for ts_code_sh in stk_rewards_item_389['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_389_tscode_list.append(stock_name)
stk_rewards_item_389_addname_dataframe=pd.DataFrame()
stk_rewards_item_389_addname_dataframe['cname'] = stk_rewards_item_389_tscode_list
for table_name in stk_rewards_item_389.columns.values.tolist():
    stk_rewards_item_389_addname_dataframe[table_name] = stk_rewards_item_389[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_389_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_390_tscode_list = list() 
stk_rewards_item_390 = pro.stk_rewards(ts_code='600026.SH,600027.SH,600028.SH,600029.SH,600030.SH,600031.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_390 返回数据 row 行数 = "+str(stk_rewards_item_390.shape[0]))
for ts_code_sh in stk_rewards_item_390['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_390_tscode_list.append(stock_name)
stk_rewards_item_390_addname_dataframe=pd.DataFrame()
stk_rewards_item_390_addname_dataframe['cname'] = stk_rewards_item_390_tscode_list
for table_name in stk_rewards_item_390.columns.values.tolist():
    stk_rewards_item_390_addname_dataframe[table_name] = stk_rewards_item_390[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_390_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_391_tscode_list = list() 
stk_rewards_item_391 = pro.stk_rewards(ts_code='600033.SH,600035.SH,600036.SH,600037.SH,600038.SH,600039.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_391 返回数据 row 行数 = "+str(stk_rewards_item_391.shape[0]))
for ts_code_sh in stk_rewards_item_391['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_391_tscode_list.append(stock_name)
stk_rewards_item_391_addname_dataframe=pd.DataFrame()
stk_rewards_item_391_addname_dataframe['cname'] = stk_rewards_item_391_tscode_list
for table_name in stk_rewards_item_391.columns.values.tolist():
    stk_rewards_item_391_addname_dataframe[table_name] = stk_rewards_item_391[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_391_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_392_tscode_list = list() 
stk_rewards_item_392 = pro.stk_rewards(ts_code='600048.SH,600050.SH,600051.SH,600052.SH,600053.SH,600054.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_392 返回数据 row 行数 = "+str(stk_rewards_item_392.shape[0]))
for ts_code_sh in stk_rewards_item_392['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_392_tscode_list.append(stock_name)
stk_rewards_item_392_addname_dataframe=pd.DataFrame()
stk_rewards_item_392_addname_dataframe['cname'] = stk_rewards_item_392_tscode_list
for table_name in stk_rewards_item_392.columns.values.tolist():
    stk_rewards_item_392_addname_dataframe[table_name] = stk_rewards_item_392[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_392_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_393_tscode_list = list() 
stk_rewards_item_393 = pro.stk_rewards(ts_code='600055.SH,600056.SH,600057.SH,600058.SH,600059.SH,600060.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_393 返回数据 row 行数 = "+str(stk_rewards_item_393.shape[0]))
for ts_code_sh in stk_rewards_item_393['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_393_tscode_list.append(stock_name)
stk_rewards_item_393_addname_dataframe=pd.DataFrame()
stk_rewards_item_393_addname_dataframe['cname'] = stk_rewards_item_393_tscode_list
for table_name in stk_rewards_item_393.columns.values.tolist():
    stk_rewards_item_393_addname_dataframe[table_name] = stk_rewards_item_393[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_393_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_394_tscode_list = list() 
stk_rewards_item_394 = pro.stk_rewards(ts_code='600061.SH,600062.SH,600063.SH,600064.SH,600065.SH,600066.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_394 返回数据 row 行数 = "+str(stk_rewards_item_394.shape[0]))
for ts_code_sh in stk_rewards_item_394['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_394_tscode_list.append(stock_name)
stk_rewards_item_394_addname_dataframe=pd.DataFrame()
stk_rewards_item_394_addname_dataframe['cname'] = stk_rewards_item_394_tscode_list
for table_name in stk_rewards_item_394.columns.values.tolist():
    stk_rewards_item_394_addname_dataframe[table_name] = stk_rewards_item_394[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_394_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_395_tscode_list = list() 
stk_rewards_item_395 = pro.stk_rewards(ts_code='600067.SH,600068.SH,600069.SH,600070.SH,600071.SH,600072.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_395 返回数据 row 行数 = "+str(stk_rewards_item_395.shape[0]))
for ts_code_sh in stk_rewards_item_395['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_395_tscode_list.append(stock_name)
stk_rewards_item_395_addname_dataframe=pd.DataFrame()
stk_rewards_item_395_addname_dataframe['cname'] = stk_rewards_item_395_tscode_list
for table_name in stk_rewards_item_395.columns.values.tolist():
    stk_rewards_item_395_addname_dataframe[table_name] = stk_rewards_item_395[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_395_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_396_tscode_list = list() 
stk_rewards_item_396 = pro.stk_rewards(ts_code='600073.SH,600074.SH,600075.SH,600076.SH,600077.SH,600078.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_396 返回数据 row 行数 = "+str(stk_rewards_item_396.shape[0]))
for ts_code_sh in stk_rewards_item_396['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_396_tscode_list.append(stock_name)
stk_rewards_item_396_addname_dataframe=pd.DataFrame()
stk_rewards_item_396_addname_dataframe['cname'] = stk_rewards_item_396_tscode_list
for table_name in stk_rewards_item_396.columns.values.tolist():
    stk_rewards_item_396_addname_dataframe[table_name] = stk_rewards_item_396[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_396_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_397_tscode_list = list() 
stk_rewards_item_397 = pro.stk_rewards(ts_code='600079.SH,600080.SH,600081.SH,600082.SH,600083.SH,600084.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_397 返回数据 row 行数 = "+str(stk_rewards_item_397.shape[0]))
for ts_code_sh in stk_rewards_item_397['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_397_tscode_list.append(stock_name)
stk_rewards_item_397_addname_dataframe=pd.DataFrame()
stk_rewards_item_397_addname_dataframe['cname'] = stk_rewards_item_397_tscode_list
for table_name in stk_rewards_item_397.columns.values.tolist():
    stk_rewards_item_397_addname_dataframe[table_name] = stk_rewards_item_397[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_397_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_398_tscode_list = list() 
stk_rewards_item_398 = pro.stk_rewards(ts_code='600085.SH,600086.SH,600087.SH,600088.SH,600089.SH,600090.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_398 返回数据 row 行数 = "+str(stk_rewards_item_398.shape[0]))
for ts_code_sh in stk_rewards_item_398['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_398_tscode_list.append(stock_name)
stk_rewards_item_398_addname_dataframe=pd.DataFrame()
stk_rewards_item_398_addname_dataframe['cname'] = stk_rewards_item_398_tscode_list
for table_name in stk_rewards_item_398.columns.values.tolist():
    stk_rewards_item_398_addname_dataframe[table_name] = stk_rewards_item_398[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_398_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_399_tscode_list = list() 
stk_rewards_item_399 = pro.stk_rewards(ts_code='600091.SH,600092.SH,600093.SH,600094.SH,600095.SH,600096.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_399 返回数据 row 行数 = "+str(stk_rewards_item_399.shape[0]))
for ts_code_sh in stk_rewards_item_399['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_399_tscode_list.append(stock_name)
stk_rewards_item_399_addname_dataframe=pd.DataFrame()
stk_rewards_item_399_addname_dataframe['cname'] = stk_rewards_item_399_tscode_list
for table_name in stk_rewards_item_399.columns.values.tolist():
    stk_rewards_item_399_addname_dataframe[table_name] = stk_rewards_item_399[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_399_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_400_tscode_list = list() 
stk_rewards_item_400 = pro.stk_rewards(ts_code='600097.SH,600098.SH,600099.SH,600100.SH,600101.SH,600102.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_400 返回数据 row 行数 = "+str(stk_rewards_item_400.shape[0]))
for ts_code_sh in stk_rewards_item_400['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_400_tscode_list.append(stock_name)
stk_rewards_item_400_addname_dataframe=pd.DataFrame()
stk_rewards_item_400_addname_dataframe['cname'] = stk_rewards_item_400_tscode_list
for table_name in stk_rewards_item_400.columns.values.tolist():
    stk_rewards_item_400_addname_dataframe[table_name] = stk_rewards_item_400[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_400_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_401_tscode_list = list() 
stk_rewards_item_401 = pro.stk_rewards(ts_code='600103.SH,600104.SH,600105.SH,600106.SH,600107.SH,600108.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_401 返回数据 row 行数 = "+str(stk_rewards_item_401.shape[0]))
for ts_code_sh in stk_rewards_item_401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_401_tscode_list.append(stock_name)
stk_rewards_item_401_addname_dataframe=pd.DataFrame()
stk_rewards_item_401_addname_dataframe['cname'] = stk_rewards_item_401_tscode_list
for table_name in stk_rewards_item_401.columns.values.tolist():
    stk_rewards_item_401_addname_dataframe[table_name] = stk_rewards_item_401[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_401_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_402_tscode_list = list() 
stk_rewards_item_402 = pro.stk_rewards(ts_code='600109.SH,600110.SH,600111.SH,600112.SH,600113.SH,600114.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_402 返回数据 row 行数 = "+str(stk_rewards_item_402.shape[0]))
for ts_code_sh in stk_rewards_item_402['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_402_tscode_list.append(stock_name)
stk_rewards_item_402_addname_dataframe=pd.DataFrame()
stk_rewards_item_402_addname_dataframe['cname'] = stk_rewards_item_402_tscode_list
for table_name in stk_rewards_item_402.columns.values.tolist():
    stk_rewards_item_402_addname_dataframe[table_name] = stk_rewards_item_402[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_402_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_403_tscode_list = list() 
stk_rewards_item_403 = pro.stk_rewards(ts_code='600115.SH,600116.SH,600117.SH,600118.SH,600119.SH,600120.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_403 返回数据 row 行数 = "+str(stk_rewards_item_403.shape[0]))
for ts_code_sh in stk_rewards_item_403['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_403_tscode_list.append(stock_name)
stk_rewards_item_403_addname_dataframe=pd.DataFrame()
stk_rewards_item_403_addname_dataframe['cname'] = stk_rewards_item_403_tscode_list
for table_name in stk_rewards_item_403.columns.values.tolist():
    stk_rewards_item_403_addname_dataframe[table_name] = stk_rewards_item_403[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_403_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_404_tscode_list = list() 
stk_rewards_item_404 = pro.stk_rewards(ts_code='600121.SH,600122.SH,600123.SH,600125.SH,600126.SH,600127.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_404 返回数据 row 行数 = "+str(stk_rewards_item_404.shape[0]))
for ts_code_sh in stk_rewards_item_404['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_404_tscode_list.append(stock_name)
stk_rewards_item_404_addname_dataframe=pd.DataFrame()
stk_rewards_item_404_addname_dataframe['cname'] = stk_rewards_item_404_tscode_list
for table_name in stk_rewards_item_404.columns.values.tolist():
    stk_rewards_item_404_addname_dataframe[table_name] = stk_rewards_item_404[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_404_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_405_tscode_list = list() 
stk_rewards_item_405 = pro.stk_rewards(ts_code='600128.SH,600129.SH,600130.SH,600131.SH,600132.SH,600133.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_405 返回数据 row 行数 = "+str(stk_rewards_item_405.shape[0]))
for ts_code_sh in stk_rewards_item_405['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_405_tscode_list.append(stock_name)
stk_rewards_item_405_addname_dataframe=pd.DataFrame()
stk_rewards_item_405_addname_dataframe['cname'] = stk_rewards_item_405_tscode_list
for table_name in stk_rewards_item_405.columns.values.tolist():
    stk_rewards_item_405_addname_dataframe[table_name] = stk_rewards_item_405[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_405_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_406_tscode_list = list() 
stk_rewards_item_406 = pro.stk_rewards(ts_code='600135.SH,600136.SH,600137.SH,600138.SH,600139.SH,600141.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_406 返回数据 row 行数 = "+str(stk_rewards_item_406.shape[0]))
for ts_code_sh in stk_rewards_item_406['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_406_tscode_list.append(stock_name)
stk_rewards_item_406_addname_dataframe=pd.DataFrame()
stk_rewards_item_406_addname_dataframe['cname'] = stk_rewards_item_406_tscode_list
for table_name in stk_rewards_item_406.columns.values.tolist():
    stk_rewards_item_406_addname_dataframe[table_name] = stk_rewards_item_406[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_406_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_407_tscode_list = list() 
stk_rewards_item_407 = pro.stk_rewards(ts_code='600143.SH,600145.SH,600146.SH,600148.SH,600149.SH,600150.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_407 返回数据 row 行数 = "+str(stk_rewards_item_407.shape[0]))
for ts_code_sh in stk_rewards_item_407['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_407_tscode_list.append(stock_name)
stk_rewards_item_407_addname_dataframe=pd.DataFrame()
stk_rewards_item_407_addname_dataframe['cname'] = stk_rewards_item_407_tscode_list
for table_name in stk_rewards_item_407.columns.values.tolist():
    stk_rewards_item_407_addname_dataframe[table_name] = stk_rewards_item_407[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_407_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_408_tscode_list = list() 
stk_rewards_item_408 = pro.stk_rewards(ts_code='600151.SH,600152.SH,600153.SH,600155.SH,600156.SH,600157.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_408 返回数据 row 行数 = "+str(stk_rewards_item_408.shape[0]))
for ts_code_sh in stk_rewards_item_408['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_408_tscode_list.append(stock_name)
stk_rewards_item_408_addname_dataframe=pd.DataFrame()
stk_rewards_item_408_addname_dataframe['cname'] = stk_rewards_item_408_tscode_list
for table_name in stk_rewards_item_408.columns.values.tolist():
    stk_rewards_item_408_addname_dataframe[table_name] = stk_rewards_item_408[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_408_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_409_tscode_list = list() 
stk_rewards_item_409 = pro.stk_rewards(ts_code='600158.SH,600159.SH,600160.SH,600161.SH,600162.SH,600163.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_409 返回数据 row 行数 = "+str(stk_rewards_item_409.shape[0]))
for ts_code_sh in stk_rewards_item_409['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_409_tscode_list.append(stock_name)
stk_rewards_item_409_addname_dataframe=pd.DataFrame()
stk_rewards_item_409_addname_dataframe['cname'] = stk_rewards_item_409_tscode_list
for table_name in stk_rewards_item_409.columns.values.tolist():
    stk_rewards_item_409_addname_dataframe[table_name] = stk_rewards_item_409[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_409_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_410_tscode_list = list() 
stk_rewards_item_410 = pro.stk_rewards(ts_code='600165.SH,600166.SH,600167.SH,600168.SH,600169.SH,600170.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_410 返回数据 row 行数 = "+str(stk_rewards_item_410.shape[0]))
for ts_code_sh in stk_rewards_item_410['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_410_tscode_list.append(stock_name)
stk_rewards_item_410_addname_dataframe=pd.DataFrame()
stk_rewards_item_410_addname_dataframe['cname'] = stk_rewards_item_410_tscode_list
for table_name in stk_rewards_item_410.columns.values.tolist():
    stk_rewards_item_410_addname_dataframe[table_name] = stk_rewards_item_410[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_410_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_411_tscode_list = list() 
stk_rewards_item_411 = pro.stk_rewards(ts_code='600171.SH,600172.SH,600173.SH,600175.SH,600176.SH,600177.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_411 返回数据 row 行数 = "+str(stk_rewards_item_411.shape[0]))
for ts_code_sh in stk_rewards_item_411['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_411_tscode_list.append(stock_name)
stk_rewards_item_411_addname_dataframe=pd.DataFrame()
stk_rewards_item_411_addname_dataframe['cname'] = stk_rewards_item_411_tscode_list
for table_name in stk_rewards_item_411.columns.values.tolist():
    stk_rewards_item_411_addname_dataframe[table_name] = stk_rewards_item_411[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_411_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_412_tscode_list = list() 
stk_rewards_item_412 = pro.stk_rewards(ts_code='600178.SH,600179.SH,600180.SH,600181.SH,600182.SH,600183.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_412 返回数据 row 行数 = "+str(stk_rewards_item_412.shape[0]))
for ts_code_sh in stk_rewards_item_412['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_412_tscode_list.append(stock_name)
stk_rewards_item_412_addname_dataframe=pd.DataFrame()
stk_rewards_item_412_addname_dataframe['cname'] = stk_rewards_item_412_tscode_list
for table_name in stk_rewards_item_412.columns.values.tolist():
    stk_rewards_item_412_addname_dataframe[table_name] = stk_rewards_item_412[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_412_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_413_tscode_list = list() 
stk_rewards_item_413 = pro.stk_rewards(ts_code='600184.SH,600185.SH,600186.SH,600187.SH,600188.SH,600189.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_413 返回数据 row 行数 = "+str(stk_rewards_item_413.shape[0]))
for ts_code_sh in stk_rewards_item_413['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_413_tscode_list.append(stock_name)
stk_rewards_item_413_addname_dataframe=pd.DataFrame()
stk_rewards_item_413_addname_dataframe['cname'] = stk_rewards_item_413_tscode_list
for table_name in stk_rewards_item_413.columns.values.tolist():
    stk_rewards_item_413_addname_dataframe[table_name] = stk_rewards_item_413[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_413_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_414_tscode_list = list() 
stk_rewards_item_414 = pro.stk_rewards(ts_code='600190.SH,600191.SH,600192.SH,600193.SH,600195.SH,600196.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_414 返回数据 row 行数 = "+str(stk_rewards_item_414.shape[0]))
for ts_code_sh in stk_rewards_item_414['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_414_tscode_list.append(stock_name)
stk_rewards_item_414_addname_dataframe=pd.DataFrame()
stk_rewards_item_414_addname_dataframe['cname'] = stk_rewards_item_414_tscode_list
for table_name in stk_rewards_item_414.columns.values.tolist():
    stk_rewards_item_414_addname_dataframe[table_name] = stk_rewards_item_414[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_414_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_415_tscode_list = list() 
stk_rewards_item_415 = pro.stk_rewards(ts_code='600197.SH,600198.SH,600199.SH,600200.SH,600201.SH,600202.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_415 返回数据 row 行数 = "+str(stk_rewards_item_415.shape[0]))
for ts_code_sh in stk_rewards_item_415['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_415_tscode_list.append(stock_name)
stk_rewards_item_415_addname_dataframe=pd.DataFrame()
stk_rewards_item_415_addname_dataframe['cname'] = stk_rewards_item_415_tscode_list
for table_name in stk_rewards_item_415.columns.values.tolist():
    stk_rewards_item_415_addname_dataframe[table_name] = stk_rewards_item_415[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_415_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_416_tscode_list = list() 
stk_rewards_item_416 = pro.stk_rewards(ts_code='600203.SH,600205.SH,600206.SH,600207.SH,600208.SH,600209.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_416 返回数据 row 行数 = "+str(stk_rewards_item_416.shape[0]))
for ts_code_sh in stk_rewards_item_416['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_416_tscode_list.append(stock_name)
stk_rewards_item_416_addname_dataframe=pd.DataFrame()
stk_rewards_item_416_addname_dataframe['cname'] = stk_rewards_item_416_tscode_list
for table_name in stk_rewards_item_416.columns.values.tolist():
    stk_rewards_item_416_addname_dataframe[table_name] = stk_rewards_item_416[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_416_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_417_tscode_list = list() 
stk_rewards_item_417 = pro.stk_rewards(ts_code='600210.SH,600211.SH,600212.SH,600213.SH,600215.SH,600216.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_417 返回数据 row 行数 = "+str(stk_rewards_item_417.shape[0]))
for ts_code_sh in stk_rewards_item_417['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_417_tscode_list.append(stock_name)
stk_rewards_item_417_addname_dataframe=pd.DataFrame()
stk_rewards_item_417_addname_dataframe['cname'] = stk_rewards_item_417_tscode_list
for table_name in stk_rewards_item_417.columns.values.tolist():
    stk_rewards_item_417_addname_dataframe[table_name] = stk_rewards_item_417[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_417_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_418_tscode_list = list() 
stk_rewards_item_418 = pro.stk_rewards(ts_code='600217.SH,600218.SH,600219.SH,600220.SH,600221.SH,600222.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_418 返回数据 row 行数 = "+str(stk_rewards_item_418.shape[0]))
for ts_code_sh in stk_rewards_item_418['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_418_tscode_list.append(stock_name)
stk_rewards_item_418_addname_dataframe=pd.DataFrame()
stk_rewards_item_418_addname_dataframe['cname'] = stk_rewards_item_418_tscode_list
for table_name in stk_rewards_item_418.columns.values.tolist():
    stk_rewards_item_418_addname_dataframe[table_name] = stk_rewards_item_418[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_418_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_419_tscode_list = list() 
stk_rewards_item_419 = pro.stk_rewards(ts_code='600223.SH,600225.SH,600226.SH,600227.SH,600228.SH,600229.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_419 返回数据 row 行数 = "+str(stk_rewards_item_419.shape[0]))
for ts_code_sh in stk_rewards_item_419['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_419_tscode_list.append(stock_name)
stk_rewards_item_419_addname_dataframe=pd.DataFrame()
stk_rewards_item_419_addname_dataframe['cname'] = stk_rewards_item_419_tscode_list
for table_name in stk_rewards_item_419.columns.values.tolist():
    stk_rewards_item_419_addname_dataframe[table_name] = stk_rewards_item_419[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_419_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_420_tscode_list = list() 
stk_rewards_item_420 = pro.stk_rewards(ts_code='600230.SH,600231.SH,600232.SH,600233.SH,600234.SH,600235.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_420 返回数据 row 行数 = "+str(stk_rewards_item_420.shape[0]))
for ts_code_sh in stk_rewards_item_420['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_420_tscode_list.append(stock_name)
stk_rewards_item_420_addname_dataframe=pd.DataFrame()
stk_rewards_item_420_addname_dataframe['cname'] = stk_rewards_item_420_tscode_list
for table_name in stk_rewards_item_420.columns.values.tolist():
    stk_rewards_item_420_addname_dataframe[table_name] = stk_rewards_item_420[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_420_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_421_tscode_list = list() 
stk_rewards_item_421 = pro.stk_rewards(ts_code='600236.SH,600237.SH,600238.SH,600239.SH,600240.SH,600241.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_421 返回数据 row 行数 = "+str(stk_rewards_item_421.shape[0]))
for ts_code_sh in stk_rewards_item_421['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_421_tscode_list.append(stock_name)
stk_rewards_item_421_addname_dataframe=pd.DataFrame()
stk_rewards_item_421_addname_dataframe['cname'] = stk_rewards_item_421_tscode_list
for table_name in stk_rewards_item_421.columns.values.tolist():
    stk_rewards_item_421_addname_dataframe[table_name] = stk_rewards_item_421[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_421_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_422_tscode_list = list() 
stk_rewards_item_422 = pro.stk_rewards(ts_code='600242.SH,600243.SH,600246.SH,600247.SH,600248.SH,600249.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_422 返回数据 row 行数 = "+str(stk_rewards_item_422.shape[0]))
for ts_code_sh in stk_rewards_item_422['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_422_tscode_list.append(stock_name)
stk_rewards_item_422_addname_dataframe=pd.DataFrame()
stk_rewards_item_422_addname_dataframe['cname'] = stk_rewards_item_422_tscode_list
for table_name in stk_rewards_item_422.columns.values.tolist():
    stk_rewards_item_422_addname_dataframe[table_name] = stk_rewards_item_422[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_422_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_423_tscode_list = list() 
stk_rewards_item_423 = pro.stk_rewards(ts_code='600250.SH,600251.SH,600252.SH,600253.SH,600255.SH,600256.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_423 返回数据 row 行数 = "+str(stk_rewards_item_423.shape[0]))
for ts_code_sh in stk_rewards_item_423['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_423_tscode_list.append(stock_name)
stk_rewards_item_423_addname_dataframe=pd.DataFrame()
stk_rewards_item_423_addname_dataframe['cname'] = stk_rewards_item_423_tscode_list
for table_name in stk_rewards_item_423.columns.values.tolist():
    stk_rewards_item_423_addname_dataframe[table_name] = stk_rewards_item_423[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_423_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_424_tscode_list = list() 
stk_rewards_item_424 = pro.stk_rewards(ts_code='600257.SH,600258.SH,600259.SH,600260.SH,600261.SH,600262.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_424 返回数据 row 行数 = "+str(stk_rewards_item_424.shape[0]))
for ts_code_sh in stk_rewards_item_424['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_424_tscode_list.append(stock_name)
stk_rewards_item_424_addname_dataframe=pd.DataFrame()
stk_rewards_item_424_addname_dataframe['cname'] = stk_rewards_item_424_tscode_list
for table_name in stk_rewards_item_424.columns.values.tolist():
    stk_rewards_item_424_addname_dataframe[table_name] = stk_rewards_item_424[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_424_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_425_tscode_list = list() 
stk_rewards_item_425 = pro.stk_rewards(ts_code='600263.SH,600265.SH,600266.SH,600267.SH,600268.SH,600269.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_425 返回数据 row 行数 = "+str(stk_rewards_item_425.shape[0]))
for ts_code_sh in stk_rewards_item_425['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_425_tscode_list.append(stock_name)
stk_rewards_item_425_addname_dataframe=pd.DataFrame()
stk_rewards_item_425_addname_dataframe['cname'] = stk_rewards_item_425_tscode_list
for table_name in stk_rewards_item_425.columns.values.tolist():
    stk_rewards_item_425_addname_dataframe[table_name] = stk_rewards_item_425[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_425_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_426_tscode_list = list() 
stk_rewards_item_426 = pro.stk_rewards(ts_code='600270.SH,600271.SH,600272.SH,600273.SH,600275.SH,600276.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_426 返回数据 row 行数 = "+str(stk_rewards_item_426.shape[0]))
for ts_code_sh in stk_rewards_item_426['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_426_tscode_list.append(stock_name)
stk_rewards_item_426_addname_dataframe=pd.DataFrame()
stk_rewards_item_426_addname_dataframe['cname'] = stk_rewards_item_426_tscode_list
for table_name in stk_rewards_item_426.columns.values.tolist():
    stk_rewards_item_426_addname_dataframe[table_name] = stk_rewards_item_426[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_426_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_427_tscode_list = list() 
stk_rewards_item_427 = pro.stk_rewards(ts_code='600277.SH,600278.SH,600279.SH,600280.SH,600281.SH,600282.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_427 返回数据 row 行数 = "+str(stk_rewards_item_427.shape[0]))
for ts_code_sh in stk_rewards_item_427['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_427_tscode_list.append(stock_name)
stk_rewards_item_427_addname_dataframe=pd.DataFrame()
stk_rewards_item_427_addname_dataframe['cname'] = stk_rewards_item_427_tscode_list
for table_name in stk_rewards_item_427.columns.values.tolist():
    stk_rewards_item_427_addname_dataframe[table_name] = stk_rewards_item_427[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_427_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_428_tscode_list = list() 
stk_rewards_item_428 = pro.stk_rewards(ts_code='600283.SH,600284.SH,600285.SH,600286.SH,600287.SH,600288.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_428 返回数据 row 行数 = "+str(stk_rewards_item_428.shape[0]))
for ts_code_sh in stk_rewards_item_428['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_428_tscode_list.append(stock_name)
stk_rewards_item_428_addname_dataframe=pd.DataFrame()
stk_rewards_item_428_addname_dataframe['cname'] = stk_rewards_item_428_tscode_list
for table_name in stk_rewards_item_428.columns.values.tolist():
    stk_rewards_item_428_addname_dataframe[table_name] = stk_rewards_item_428[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_428_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_429_tscode_list = list() 
stk_rewards_item_429 = pro.stk_rewards(ts_code='600289.SH,600290.SH,600291.SH,600292.SH,600293.SH,600295.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_429 返回数据 row 行数 = "+str(stk_rewards_item_429.shape[0]))
for ts_code_sh in stk_rewards_item_429['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_429_tscode_list.append(stock_name)
stk_rewards_item_429_addname_dataframe=pd.DataFrame()
stk_rewards_item_429_addname_dataframe['cname'] = stk_rewards_item_429_tscode_list
for table_name in stk_rewards_item_429.columns.values.tolist():
    stk_rewards_item_429_addname_dataframe[table_name] = stk_rewards_item_429[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_429_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_430_tscode_list = list() 
stk_rewards_item_430 = pro.stk_rewards(ts_code='600296.SH,600297.SH,600298.SH,600299.SH,600300.SH,600301.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_430 返回数据 row 行数 = "+str(stk_rewards_item_430.shape[0]))
for ts_code_sh in stk_rewards_item_430['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_430_tscode_list.append(stock_name)
stk_rewards_item_430_addname_dataframe=pd.DataFrame()
stk_rewards_item_430_addname_dataframe['cname'] = stk_rewards_item_430_tscode_list
for table_name in stk_rewards_item_430.columns.values.tolist():
    stk_rewards_item_430_addname_dataframe[table_name] = stk_rewards_item_430[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_430_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_431_tscode_list = list() 
stk_rewards_item_431 = pro.stk_rewards(ts_code='600302.SH,600303.SH,600305.SH,600306.SH,600307.SH,600308.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_431 返回数据 row 行数 = "+str(stk_rewards_item_431.shape[0]))
for ts_code_sh in stk_rewards_item_431['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_431_tscode_list.append(stock_name)
stk_rewards_item_431_addname_dataframe=pd.DataFrame()
stk_rewards_item_431_addname_dataframe['cname'] = stk_rewards_item_431_tscode_list
for table_name in stk_rewards_item_431.columns.values.tolist():
    stk_rewards_item_431_addname_dataframe[table_name] = stk_rewards_item_431[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_431_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_432_tscode_list = list() 
stk_rewards_item_432 = pro.stk_rewards(ts_code='600309.SH,600310.SH,600311.SH,600312.SH,600313.SH,600315.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_432 返回数据 row 行数 = "+str(stk_rewards_item_432.shape[0]))
for ts_code_sh in stk_rewards_item_432['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_432_tscode_list.append(stock_name)
stk_rewards_item_432_addname_dataframe=pd.DataFrame()
stk_rewards_item_432_addname_dataframe['cname'] = stk_rewards_item_432_tscode_list
for table_name in stk_rewards_item_432.columns.values.tolist():
    stk_rewards_item_432_addname_dataframe[table_name] = stk_rewards_item_432[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_432_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_433_tscode_list = list() 
stk_rewards_item_433 = pro.stk_rewards(ts_code='600316.SH,600317.SH,600318.SH,600319.SH,600320.SH,600321.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_433 返回数据 row 行数 = "+str(stk_rewards_item_433.shape[0]))
for ts_code_sh in stk_rewards_item_433['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_433_tscode_list.append(stock_name)
stk_rewards_item_433_addname_dataframe=pd.DataFrame()
stk_rewards_item_433_addname_dataframe['cname'] = stk_rewards_item_433_tscode_list
for table_name in stk_rewards_item_433.columns.values.tolist():
    stk_rewards_item_433_addname_dataframe[table_name] = stk_rewards_item_433[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_433_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_434_tscode_list = list() 
stk_rewards_item_434 = pro.stk_rewards(ts_code='600322.SH,600323.SH,600325.SH,600326.SH,600327.SH,600328.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_434 返回数据 row 行数 = "+str(stk_rewards_item_434.shape[0]))
for ts_code_sh in stk_rewards_item_434['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_434_tscode_list.append(stock_name)
stk_rewards_item_434_addname_dataframe=pd.DataFrame()
stk_rewards_item_434_addname_dataframe['cname'] = stk_rewards_item_434_tscode_list
for table_name in stk_rewards_item_434.columns.values.tolist():
    stk_rewards_item_434_addname_dataframe[table_name] = stk_rewards_item_434[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_434_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_435_tscode_list = list() 
stk_rewards_item_435 = pro.stk_rewards(ts_code='600329.SH,600330.SH,600331.SH,600332.SH,600333.SH,600335.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_435 返回数据 row 行数 = "+str(stk_rewards_item_435.shape[0]))
for ts_code_sh in stk_rewards_item_435['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_435_tscode_list.append(stock_name)
stk_rewards_item_435_addname_dataframe=pd.DataFrame()
stk_rewards_item_435_addname_dataframe['cname'] = stk_rewards_item_435_tscode_list
for table_name in stk_rewards_item_435.columns.values.tolist():
    stk_rewards_item_435_addname_dataframe[table_name] = stk_rewards_item_435[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_435_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_436_tscode_list = list() 
stk_rewards_item_436 = pro.stk_rewards(ts_code='600336.SH,600337.SH,600338.SH,600339.SH,600340.SH,600343.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_436 返回数据 row 行数 = "+str(stk_rewards_item_436.shape[0]))
for ts_code_sh in stk_rewards_item_436['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_436_tscode_list.append(stock_name)
stk_rewards_item_436_addname_dataframe=pd.DataFrame()
stk_rewards_item_436_addname_dataframe['cname'] = stk_rewards_item_436_tscode_list
for table_name in stk_rewards_item_436.columns.values.tolist():
    stk_rewards_item_436_addname_dataframe[table_name] = stk_rewards_item_436[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_436_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_437_tscode_list = list() 
stk_rewards_item_437 = pro.stk_rewards(ts_code='600345.SH,600346.SH,600348.SH,600350.SH,600351.SH,600352.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_437 返回数据 row 行数 = "+str(stk_rewards_item_437.shape[0]))
for ts_code_sh in stk_rewards_item_437['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_437_tscode_list.append(stock_name)
stk_rewards_item_437_addname_dataframe=pd.DataFrame()
stk_rewards_item_437_addname_dataframe['cname'] = stk_rewards_item_437_tscode_list
for table_name in stk_rewards_item_437.columns.values.tolist():
    stk_rewards_item_437_addname_dataframe[table_name] = stk_rewards_item_437[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_437_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_438_tscode_list = list() 
stk_rewards_item_438 = pro.stk_rewards(ts_code='600353.SH,600354.SH,600355.SH,600356.SH,600357.SH,600358.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_438 返回数据 row 行数 = "+str(stk_rewards_item_438.shape[0]))
for ts_code_sh in stk_rewards_item_438['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_438_tscode_list.append(stock_name)
stk_rewards_item_438_addname_dataframe=pd.DataFrame()
stk_rewards_item_438_addname_dataframe['cname'] = stk_rewards_item_438_tscode_list
for table_name in stk_rewards_item_438.columns.values.tolist():
    stk_rewards_item_438_addname_dataframe[table_name] = stk_rewards_item_438[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_438_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_439_tscode_list = list() 
stk_rewards_item_439 = pro.stk_rewards(ts_code='600359.SH,600360.SH,600361.SH,600362.SH,600363.SH,600365.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_439 返回数据 row 行数 = "+str(stk_rewards_item_439.shape[0]))
for ts_code_sh in stk_rewards_item_439['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_439_tscode_list.append(stock_name)
stk_rewards_item_439_addname_dataframe=pd.DataFrame()
stk_rewards_item_439_addname_dataframe['cname'] = stk_rewards_item_439_tscode_list
for table_name in stk_rewards_item_439.columns.values.tolist():
    stk_rewards_item_439_addname_dataframe[table_name] = stk_rewards_item_439[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_439_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_440_tscode_list = list() 
stk_rewards_item_440 = pro.stk_rewards(ts_code='600366.SH,600367.SH,600368.SH,600369.SH,600370.SH,600371.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_440 返回数据 row 行数 = "+str(stk_rewards_item_440.shape[0]))
for ts_code_sh in stk_rewards_item_440['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_440_tscode_list.append(stock_name)
stk_rewards_item_440_addname_dataframe=pd.DataFrame()
stk_rewards_item_440_addname_dataframe['cname'] = stk_rewards_item_440_tscode_list
for table_name in stk_rewards_item_440.columns.values.tolist():
    stk_rewards_item_440_addname_dataframe[table_name] = stk_rewards_item_440[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_440_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_441_tscode_list = list() 
stk_rewards_item_441 = pro.stk_rewards(ts_code='600372.SH,600373.SH,600375.SH,600376.SH,600377.SH,600378.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_441 返回数据 row 行数 = "+str(stk_rewards_item_441.shape[0]))
for ts_code_sh in stk_rewards_item_441['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_441_tscode_list.append(stock_name)
stk_rewards_item_441_addname_dataframe=pd.DataFrame()
stk_rewards_item_441_addname_dataframe['cname'] = stk_rewards_item_441_tscode_list
for table_name in stk_rewards_item_441.columns.values.tolist():
    stk_rewards_item_441_addname_dataframe[table_name] = stk_rewards_item_441[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_441_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_442_tscode_list = list() 
stk_rewards_item_442 = pro.stk_rewards(ts_code='600379.SH,600380.SH,600381.SH,600382.SH,600383.SH,600385.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_442 返回数据 row 行数 = "+str(stk_rewards_item_442.shape[0]))
for ts_code_sh in stk_rewards_item_442['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_442_tscode_list.append(stock_name)
stk_rewards_item_442_addname_dataframe=pd.DataFrame()
stk_rewards_item_442_addname_dataframe['cname'] = stk_rewards_item_442_tscode_list
for table_name in stk_rewards_item_442.columns.values.tolist():
    stk_rewards_item_442_addname_dataframe[table_name] = stk_rewards_item_442[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_442_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_443_tscode_list = list() 
stk_rewards_item_443 = pro.stk_rewards(ts_code='600386.SH,600387.SH,600388.SH,600389.SH,600390.SH,600391.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_443 返回数据 row 行数 = "+str(stk_rewards_item_443.shape[0]))
for ts_code_sh in stk_rewards_item_443['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_443_tscode_list.append(stock_name)
stk_rewards_item_443_addname_dataframe=pd.DataFrame()
stk_rewards_item_443_addname_dataframe['cname'] = stk_rewards_item_443_tscode_list
for table_name in stk_rewards_item_443.columns.values.tolist():
    stk_rewards_item_443_addname_dataframe[table_name] = stk_rewards_item_443[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_443_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_444_tscode_list = list() 
stk_rewards_item_444 = pro.stk_rewards(ts_code='600392.SH,600393.SH,600395.SH,600396.SH,600397.SH,600398.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_444 返回数据 row 行数 = "+str(stk_rewards_item_444.shape[0]))
for ts_code_sh in stk_rewards_item_444['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_444_tscode_list.append(stock_name)
stk_rewards_item_444_addname_dataframe=pd.DataFrame()
stk_rewards_item_444_addname_dataframe['cname'] = stk_rewards_item_444_tscode_list
for table_name in stk_rewards_item_444.columns.values.tolist():
    stk_rewards_item_444_addname_dataframe[table_name] = stk_rewards_item_444[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_444_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_445_tscode_list = list() 
stk_rewards_item_445 = pro.stk_rewards(ts_code='600399.SH,600400.SH,600401.SH,600403.SH,600405.SH,600406.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_445 返回数据 row 行数 = "+str(stk_rewards_item_445.shape[0]))
for ts_code_sh in stk_rewards_item_445['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_445_tscode_list.append(stock_name)
stk_rewards_item_445_addname_dataframe=pd.DataFrame()
stk_rewards_item_445_addname_dataframe['cname'] = stk_rewards_item_445_tscode_list
for table_name in stk_rewards_item_445.columns.values.tolist():
    stk_rewards_item_445_addname_dataframe[table_name] = stk_rewards_item_445[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_445_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_446_tscode_list = list() 
stk_rewards_item_446 = pro.stk_rewards(ts_code='600408.SH,600409.SH,600410.SH,600415.SH,600416.SH,600418.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_446 返回数据 row 行数 = "+str(stk_rewards_item_446.shape[0]))
for ts_code_sh in stk_rewards_item_446['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_446_tscode_list.append(stock_name)
stk_rewards_item_446_addname_dataframe=pd.DataFrame()
stk_rewards_item_446_addname_dataframe['cname'] = stk_rewards_item_446_tscode_list
for table_name in stk_rewards_item_446.columns.values.tolist():
    stk_rewards_item_446_addname_dataframe[table_name] = stk_rewards_item_446[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_446_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_447_tscode_list = list() 
stk_rewards_item_447 = pro.stk_rewards(ts_code='600419.SH,600420.SH,600421.SH,600422.SH,600423.SH,600425.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_447 返回数据 row 行数 = "+str(stk_rewards_item_447.shape[0]))
for ts_code_sh in stk_rewards_item_447['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_447_tscode_list.append(stock_name)
stk_rewards_item_447_addname_dataframe=pd.DataFrame()
stk_rewards_item_447_addname_dataframe['cname'] = stk_rewards_item_447_tscode_list
for table_name in stk_rewards_item_447.columns.values.tolist():
    stk_rewards_item_447_addname_dataframe[table_name] = stk_rewards_item_447[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_447_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_448_tscode_list = list() 
stk_rewards_item_448 = pro.stk_rewards(ts_code='600426.SH,600428.SH,600429.SH,600432.SH,600433.SH,600435.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_448 返回数据 row 行数 = "+str(stk_rewards_item_448.shape[0]))
for ts_code_sh in stk_rewards_item_448['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_448_tscode_list.append(stock_name)
stk_rewards_item_448_addname_dataframe=pd.DataFrame()
stk_rewards_item_448_addname_dataframe['cname'] = stk_rewards_item_448_tscode_list
for table_name in stk_rewards_item_448.columns.values.tolist():
    stk_rewards_item_448_addname_dataframe[table_name] = stk_rewards_item_448[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_448_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_449_tscode_list = list() 
stk_rewards_item_449 = pro.stk_rewards(ts_code='600436.SH,600438.SH,600439.SH,600444.SH,600446.SH,600448.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_449 返回数据 row 行数 = "+str(stk_rewards_item_449.shape[0]))
for ts_code_sh in stk_rewards_item_449['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_449_tscode_list.append(stock_name)
stk_rewards_item_449_addname_dataframe=pd.DataFrame()
stk_rewards_item_449_addname_dataframe['cname'] = stk_rewards_item_449_tscode_list
for table_name in stk_rewards_item_449.columns.values.tolist():
    stk_rewards_item_449_addname_dataframe[table_name] = stk_rewards_item_449[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_449_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_450_tscode_list = list() 
stk_rewards_item_450 = pro.stk_rewards(ts_code='600449.SH,600452.SH,600455.SH,600456.SH,600458.SH,600459.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_450 返回数据 row 行数 = "+str(stk_rewards_item_450.shape[0]))
for ts_code_sh in stk_rewards_item_450['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_450_tscode_list.append(stock_name)
stk_rewards_item_450_addname_dataframe=pd.DataFrame()
stk_rewards_item_450_addname_dataframe['cname'] = stk_rewards_item_450_tscode_list
for table_name in stk_rewards_item_450.columns.values.tolist():
    stk_rewards_item_450_addname_dataframe[table_name] = stk_rewards_item_450[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_450_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_451_tscode_list = list() 
stk_rewards_item_451 = pro.stk_rewards(ts_code='600460.SH,600461.SH,600462.SH,600463.SH,600466.SH,600467.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_451 返回数据 row 行数 = "+str(stk_rewards_item_451.shape[0]))
for ts_code_sh in stk_rewards_item_451['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_451_tscode_list.append(stock_name)
stk_rewards_item_451_addname_dataframe=pd.DataFrame()
stk_rewards_item_451_addname_dataframe['cname'] = stk_rewards_item_451_tscode_list
for table_name in stk_rewards_item_451.columns.values.tolist():
    stk_rewards_item_451_addname_dataframe[table_name] = stk_rewards_item_451[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_451_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_452_tscode_list = list() 
stk_rewards_item_452 = pro.stk_rewards(ts_code='600468.SH,600469.SH,600470.SH,600472.SH,600475.SH,600476.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_452 返回数据 row 行数 = "+str(stk_rewards_item_452.shape[0]))
for ts_code_sh in stk_rewards_item_452['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_452_tscode_list.append(stock_name)
stk_rewards_item_452_addname_dataframe=pd.DataFrame()
stk_rewards_item_452_addname_dataframe['cname'] = stk_rewards_item_452_tscode_list
for table_name in stk_rewards_item_452.columns.values.tolist():
    stk_rewards_item_452_addname_dataframe[table_name] = stk_rewards_item_452[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_452_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_453_tscode_list = list() 
stk_rewards_item_453 = pro.stk_rewards(ts_code='600477.SH,600478.SH,600479.SH,600480.SH,600481.SH,600482.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_453 返回数据 row 行数 = "+str(stk_rewards_item_453.shape[0]))
for ts_code_sh in stk_rewards_item_453['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_453_tscode_list.append(stock_name)
stk_rewards_item_453_addname_dataframe=pd.DataFrame()
stk_rewards_item_453_addname_dataframe['cname'] = stk_rewards_item_453_tscode_list
for table_name in stk_rewards_item_453.columns.values.tolist():
    stk_rewards_item_453_addname_dataframe[table_name] = stk_rewards_item_453[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_453_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_454_tscode_list = list() 
stk_rewards_item_454 = pro.stk_rewards(ts_code='600483.SH,600485.SH,600486.SH,600487.SH,600488.SH,600489.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_454 返回数据 row 行数 = "+str(stk_rewards_item_454.shape[0]))
for ts_code_sh in stk_rewards_item_454['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_454_tscode_list.append(stock_name)
stk_rewards_item_454_addname_dataframe=pd.DataFrame()
stk_rewards_item_454_addname_dataframe['cname'] = stk_rewards_item_454_tscode_list
for table_name in stk_rewards_item_454.columns.values.tolist():
    stk_rewards_item_454_addname_dataframe[table_name] = stk_rewards_item_454[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_454_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_455_tscode_list = list() 
stk_rewards_item_455 = pro.stk_rewards(ts_code='600490.SH,600491.SH,600493.SH,600495.SH,600496.SH,600497.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_455 返回数据 row 行数 = "+str(stk_rewards_item_455.shape[0]))
for ts_code_sh in stk_rewards_item_455['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_455_tscode_list.append(stock_name)
stk_rewards_item_455_addname_dataframe=pd.DataFrame()
stk_rewards_item_455_addname_dataframe['cname'] = stk_rewards_item_455_tscode_list
for table_name in stk_rewards_item_455.columns.values.tolist():
    stk_rewards_item_455_addname_dataframe[table_name] = stk_rewards_item_455[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_455_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_456_tscode_list = list() 
stk_rewards_item_456 = pro.stk_rewards(ts_code='600498.SH,600499.SH,600500.SH,600501.SH,600502.SH,600503.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_456 返回数据 row 行数 = "+str(stk_rewards_item_456.shape[0]))
for ts_code_sh in stk_rewards_item_456['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_456_tscode_list.append(stock_name)
stk_rewards_item_456_addname_dataframe=pd.DataFrame()
stk_rewards_item_456_addname_dataframe['cname'] = stk_rewards_item_456_tscode_list
for table_name in stk_rewards_item_456.columns.values.tolist():
    stk_rewards_item_456_addname_dataframe[table_name] = stk_rewards_item_456[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_456_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_457_tscode_list = list() 
stk_rewards_item_457 = pro.stk_rewards(ts_code='600505.SH,600506.SH,600507.SH,600508.SH,600509.SH,600510.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_457 返回数据 row 行数 = "+str(stk_rewards_item_457.shape[0]))
for ts_code_sh in stk_rewards_item_457['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_457_tscode_list.append(stock_name)
stk_rewards_item_457_addname_dataframe=pd.DataFrame()
stk_rewards_item_457_addname_dataframe['cname'] = stk_rewards_item_457_tscode_list
for table_name in stk_rewards_item_457.columns.values.tolist():
    stk_rewards_item_457_addname_dataframe[table_name] = stk_rewards_item_457[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_457_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_458_tscode_list = list() 
stk_rewards_item_458 = pro.stk_rewards(ts_code='600511.SH,600512.SH,600513.SH,600515.SH,600516.SH,600517.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_458 返回数据 row 行数 = "+str(stk_rewards_item_458.shape[0]))
for ts_code_sh in stk_rewards_item_458['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_458_tscode_list.append(stock_name)
stk_rewards_item_458_addname_dataframe=pd.DataFrame()
stk_rewards_item_458_addname_dataframe['cname'] = stk_rewards_item_458_tscode_list
for table_name in stk_rewards_item_458.columns.values.tolist():
    stk_rewards_item_458_addname_dataframe[table_name] = stk_rewards_item_458[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_458_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_459_tscode_list = list() 
stk_rewards_item_459 = pro.stk_rewards(ts_code='600518.SH,600519.SH,600520.SH,600521.SH,600522.SH,600523.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_459 返回数据 row 行数 = "+str(stk_rewards_item_459.shape[0]))
for ts_code_sh in stk_rewards_item_459['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_459_tscode_list.append(stock_name)
stk_rewards_item_459_addname_dataframe=pd.DataFrame()
stk_rewards_item_459_addname_dataframe['cname'] = stk_rewards_item_459_tscode_list
for table_name in stk_rewards_item_459.columns.values.tolist():
    stk_rewards_item_459_addname_dataframe[table_name] = stk_rewards_item_459[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_459_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_460_tscode_list = list() 
stk_rewards_item_460 = pro.stk_rewards(ts_code='600525.SH,600526.SH,600527.SH,600528.SH,600529.SH,600530.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_460 返回数据 row 行数 = "+str(stk_rewards_item_460.shape[0]))
for ts_code_sh in stk_rewards_item_460['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_460_tscode_list.append(stock_name)
stk_rewards_item_460_addname_dataframe=pd.DataFrame()
stk_rewards_item_460_addname_dataframe['cname'] = stk_rewards_item_460_tscode_list
for table_name in stk_rewards_item_460.columns.values.tolist():
    stk_rewards_item_460_addname_dataframe[table_name] = stk_rewards_item_460[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_460_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_461_tscode_list = list() 
stk_rewards_item_461 = pro.stk_rewards(ts_code='600531.SH,600532.SH,600533.SH,600535.SH,600536.SH,600537.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_461 返回数据 row 行数 = "+str(stk_rewards_item_461.shape[0]))
for ts_code_sh in stk_rewards_item_461['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_461_tscode_list.append(stock_name)
stk_rewards_item_461_addname_dataframe=pd.DataFrame()
stk_rewards_item_461_addname_dataframe['cname'] = stk_rewards_item_461_tscode_list
for table_name in stk_rewards_item_461.columns.values.tolist():
    stk_rewards_item_461_addname_dataframe[table_name] = stk_rewards_item_461[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_461_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_462_tscode_list = list() 
stk_rewards_item_462 = pro.stk_rewards(ts_code='600538.SH,600539.SH,600540.SH,600543.SH,600545.SH,600546.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_462 返回数据 row 行数 = "+str(stk_rewards_item_462.shape[0]))
for ts_code_sh in stk_rewards_item_462['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_462_tscode_list.append(stock_name)
stk_rewards_item_462_addname_dataframe=pd.DataFrame()
stk_rewards_item_462_addname_dataframe['cname'] = stk_rewards_item_462_tscode_list
for table_name in stk_rewards_item_462.columns.values.tolist():
    stk_rewards_item_462_addname_dataframe[table_name] = stk_rewards_item_462[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_462_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_463_tscode_list = list() 
stk_rewards_item_463 = pro.stk_rewards(ts_code='600547.SH,600548.SH,600549.SH,600550.SH,600551.SH,600552.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_463 返回数据 row 行数 = "+str(stk_rewards_item_463.shape[0]))
for ts_code_sh in stk_rewards_item_463['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_463_tscode_list.append(stock_name)
stk_rewards_item_463_addname_dataframe=pd.DataFrame()
stk_rewards_item_463_addname_dataframe['cname'] = stk_rewards_item_463_tscode_list
for table_name in stk_rewards_item_463.columns.values.tolist():
    stk_rewards_item_463_addname_dataframe[table_name] = stk_rewards_item_463[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_463_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_464_tscode_list = list() 
stk_rewards_item_464 = pro.stk_rewards(ts_code='600553.SH,600555.SH,600556.SH,600557.SH,600558.SH,600559.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_464 返回数据 row 行数 = "+str(stk_rewards_item_464.shape[0]))
for ts_code_sh in stk_rewards_item_464['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_464_tscode_list.append(stock_name)
stk_rewards_item_464_addname_dataframe=pd.DataFrame()
stk_rewards_item_464_addname_dataframe['cname'] = stk_rewards_item_464_tscode_list
for table_name in stk_rewards_item_464.columns.values.tolist():
    stk_rewards_item_464_addname_dataframe[table_name] = stk_rewards_item_464[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_464_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_465_tscode_list = list() 
stk_rewards_item_465 = pro.stk_rewards(ts_code='600560.SH,600561.SH,600562.SH,600563.SH,600565.SH,600566.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_465 返回数据 row 行数 = "+str(stk_rewards_item_465.shape[0]))
for ts_code_sh in stk_rewards_item_465['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_465_tscode_list.append(stock_name)
stk_rewards_item_465_addname_dataframe=pd.DataFrame()
stk_rewards_item_465_addname_dataframe['cname'] = stk_rewards_item_465_tscode_list
for table_name in stk_rewards_item_465.columns.values.tolist():
    stk_rewards_item_465_addname_dataframe[table_name] = stk_rewards_item_465[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_465_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_466_tscode_list = list() 
stk_rewards_item_466 = pro.stk_rewards(ts_code='600567.SH,600568.SH,600569.SH,600570.SH,600571.SH,600572.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_466 返回数据 row 行数 = "+str(stk_rewards_item_466.shape[0]))
for ts_code_sh in stk_rewards_item_466['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_466_tscode_list.append(stock_name)
stk_rewards_item_466_addname_dataframe=pd.DataFrame()
stk_rewards_item_466_addname_dataframe['cname'] = stk_rewards_item_466_tscode_list
for table_name in stk_rewards_item_466.columns.values.tolist():
    stk_rewards_item_466_addname_dataframe[table_name] = stk_rewards_item_466[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_466_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_467_tscode_list = list() 
stk_rewards_item_467 = pro.stk_rewards(ts_code='600573.SH,600575.SH,600576.SH,600577.SH,600578.SH,600579.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_467 返回数据 row 行数 = "+str(stk_rewards_item_467.shape[0]))
for ts_code_sh in stk_rewards_item_467['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_467_tscode_list.append(stock_name)
stk_rewards_item_467_addname_dataframe=pd.DataFrame()
stk_rewards_item_467_addname_dataframe['cname'] = stk_rewards_item_467_tscode_list
for table_name in stk_rewards_item_467.columns.values.tolist():
    stk_rewards_item_467_addname_dataframe[table_name] = stk_rewards_item_467[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_467_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_468_tscode_list = list() 
stk_rewards_item_468 = pro.stk_rewards(ts_code='600580.SH,600581.SH,600582.SH,600583.SH,600584.SH,600585.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_468 返回数据 row 行数 = "+str(stk_rewards_item_468.shape[0]))
for ts_code_sh in stk_rewards_item_468['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_468_tscode_list.append(stock_name)
stk_rewards_item_468_addname_dataframe=pd.DataFrame()
stk_rewards_item_468_addname_dataframe['cname'] = stk_rewards_item_468_tscode_list
for table_name in stk_rewards_item_468.columns.values.tolist():
    stk_rewards_item_468_addname_dataframe[table_name] = stk_rewards_item_468[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_468_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_469_tscode_list = list() 
stk_rewards_item_469 = pro.stk_rewards(ts_code='600586.SH,600587.SH,600588.SH,600589.SH,600590.SH,600591.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_469 返回数据 row 行数 = "+str(stk_rewards_item_469.shape[0]))
for ts_code_sh in stk_rewards_item_469['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_469_tscode_list.append(stock_name)
stk_rewards_item_469_addname_dataframe=pd.DataFrame()
stk_rewards_item_469_addname_dataframe['cname'] = stk_rewards_item_469_tscode_list
for table_name in stk_rewards_item_469.columns.values.tolist():
    stk_rewards_item_469_addname_dataframe[table_name] = stk_rewards_item_469[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_469_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_470_tscode_list = list() 
stk_rewards_item_470 = pro.stk_rewards(ts_code='600592.SH,600593.SH,600594.SH,600595.SH,600596.SH,600597.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_470 返回数据 row 行数 = "+str(stk_rewards_item_470.shape[0]))
for ts_code_sh in stk_rewards_item_470['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_470_tscode_list.append(stock_name)
stk_rewards_item_470_addname_dataframe=pd.DataFrame()
stk_rewards_item_470_addname_dataframe['cname'] = stk_rewards_item_470_tscode_list
for table_name in stk_rewards_item_470.columns.values.tolist():
    stk_rewards_item_470_addname_dataframe[table_name] = stk_rewards_item_470[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_470_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_471_tscode_list = list() 
stk_rewards_item_471 = pro.stk_rewards(ts_code='600598.SH,600599.SH,600600.SH,600601.SH,600602.SH,600603.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_471 返回数据 row 行数 = "+str(stk_rewards_item_471.shape[0]))
for ts_code_sh in stk_rewards_item_471['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_471_tscode_list.append(stock_name)
stk_rewards_item_471_addname_dataframe=pd.DataFrame()
stk_rewards_item_471_addname_dataframe['cname'] = stk_rewards_item_471_tscode_list
for table_name in stk_rewards_item_471.columns.values.tolist():
    stk_rewards_item_471_addname_dataframe[table_name] = stk_rewards_item_471[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_471_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_472_tscode_list = list() 
stk_rewards_item_472 = pro.stk_rewards(ts_code='600604.SH,600605.SH,600606.SH,600607.SH,600608.SH,600609.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_472 返回数据 row 行数 = "+str(stk_rewards_item_472.shape[0]))
for ts_code_sh in stk_rewards_item_472['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_472_tscode_list.append(stock_name)
stk_rewards_item_472_addname_dataframe=pd.DataFrame()
stk_rewards_item_472_addname_dataframe['cname'] = stk_rewards_item_472_tscode_list
for table_name in stk_rewards_item_472.columns.values.tolist():
    stk_rewards_item_472_addname_dataframe[table_name] = stk_rewards_item_472[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_472_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_473_tscode_list = list() 
stk_rewards_item_473 = pro.stk_rewards(ts_code='600610.SH,600611.SH,600612.SH,600613.SH,600614.SH,600615.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_473 返回数据 row 行数 = "+str(stk_rewards_item_473.shape[0]))
for ts_code_sh in stk_rewards_item_473['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_473_tscode_list.append(stock_name)
stk_rewards_item_473_addname_dataframe=pd.DataFrame()
stk_rewards_item_473_addname_dataframe['cname'] = stk_rewards_item_473_tscode_list
for table_name in stk_rewards_item_473.columns.values.tolist():
    stk_rewards_item_473_addname_dataframe[table_name] = stk_rewards_item_473[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_473_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_474_tscode_list = list() 
stk_rewards_item_474 = pro.stk_rewards(ts_code='600616.SH,600617.SH,600618.SH,600619.SH,600620.SH,600621.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_474 返回数据 row 行数 = "+str(stk_rewards_item_474.shape[0]))
for ts_code_sh in stk_rewards_item_474['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_474_tscode_list.append(stock_name)
stk_rewards_item_474_addname_dataframe=pd.DataFrame()
stk_rewards_item_474_addname_dataframe['cname'] = stk_rewards_item_474_tscode_list
for table_name in stk_rewards_item_474.columns.values.tolist():
    stk_rewards_item_474_addname_dataframe[table_name] = stk_rewards_item_474[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_474_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_475_tscode_list = list() 
stk_rewards_item_475 = pro.stk_rewards(ts_code='600622.SH,600623.SH,600624.SH,600625.SH,600626.SH,600627.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_475 返回数据 row 行数 = "+str(stk_rewards_item_475.shape[0]))
for ts_code_sh in stk_rewards_item_475['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_475_tscode_list.append(stock_name)
stk_rewards_item_475_addname_dataframe=pd.DataFrame()
stk_rewards_item_475_addname_dataframe['cname'] = stk_rewards_item_475_tscode_list
for table_name in stk_rewards_item_475.columns.values.tolist():
    stk_rewards_item_475_addname_dataframe[table_name] = stk_rewards_item_475[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_475_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_476_tscode_list = list() 
stk_rewards_item_476 = pro.stk_rewards(ts_code='600628.SH,600629.SH,600630.SH,600631.SH,600632.SH,600633.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_476 返回数据 row 行数 = "+str(stk_rewards_item_476.shape[0]))
for ts_code_sh in stk_rewards_item_476['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_476_tscode_list.append(stock_name)
stk_rewards_item_476_addname_dataframe=pd.DataFrame()
stk_rewards_item_476_addname_dataframe['cname'] = stk_rewards_item_476_tscode_list
for table_name in stk_rewards_item_476.columns.values.tolist():
    stk_rewards_item_476_addname_dataframe[table_name] = stk_rewards_item_476[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_476_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_477_tscode_list = list() 
stk_rewards_item_477 = pro.stk_rewards(ts_code='600634.SH,600635.SH,600636.SH,600637.SH,600638.SH,600639.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_477 返回数据 row 行数 = "+str(stk_rewards_item_477.shape[0]))
for ts_code_sh in stk_rewards_item_477['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_477_tscode_list.append(stock_name)
stk_rewards_item_477_addname_dataframe=pd.DataFrame()
stk_rewards_item_477_addname_dataframe['cname'] = stk_rewards_item_477_tscode_list
for table_name in stk_rewards_item_477.columns.values.tolist():
    stk_rewards_item_477_addname_dataframe[table_name] = stk_rewards_item_477[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_477_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_478_tscode_list = list() 
stk_rewards_item_478 = pro.stk_rewards(ts_code='600640.SH,600641.SH,600642.SH,600643.SH,600644.SH,600645.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_478 返回数据 row 行数 = "+str(stk_rewards_item_478.shape[0]))
for ts_code_sh in stk_rewards_item_478['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_478_tscode_list.append(stock_name)
stk_rewards_item_478_addname_dataframe=pd.DataFrame()
stk_rewards_item_478_addname_dataframe['cname'] = stk_rewards_item_478_tscode_list
for table_name in stk_rewards_item_478.columns.values.tolist():
    stk_rewards_item_478_addname_dataframe[table_name] = stk_rewards_item_478[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_478_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_479_tscode_list = list() 
stk_rewards_item_479 = pro.stk_rewards(ts_code='600646.SH,600647.SH,600648.SH,600649.SH,600650.SH,600651.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_479 返回数据 row 行数 = "+str(stk_rewards_item_479.shape[0]))
for ts_code_sh in stk_rewards_item_479['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_479_tscode_list.append(stock_name)
stk_rewards_item_479_addname_dataframe=pd.DataFrame()
stk_rewards_item_479_addname_dataframe['cname'] = stk_rewards_item_479_tscode_list
for table_name in stk_rewards_item_479.columns.values.tolist():
    stk_rewards_item_479_addname_dataframe[table_name] = stk_rewards_item_479[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_479_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_480_tscode_list = list() 
stk_rewards_item_480 = pro.stk_rewards(ts_code='600652.SH,600653.SH,600654.SH,600655.SH,600656.SH,600657.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_480 返回数据 row 行数 = "+str(stk_rewards_item_480.shape[0]))
for ts_code_sh in stk_rewards_item_480['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_480_tscode_list.append(stock_name)
stk_rewards_item_480_addname_dataframe=pd.DataFrame()
stk_rewards_item_480_addname_dataframe['cname'] = stk_rewards_item_480_tscode_list
for table_name in stk_rewards_item_480.columns.values.tolist():
    stk_rewards_item_480_addname_dataframe[table_name] = stk_rewards_item_480[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_480_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_481_tscode_list = list() 
stk_rewards_item_481 = pro.stk_rewards(ts_code='600658.SH,600659.SH,600660.SH,600661.SH,600662.SH,600663.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_481 返回数据 row 行数 = "+str(stk_rewards_item_481.shape[0]))
for ts_code_sh in stk_rewards_item_481['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_481_tscode_list.append(stock_name)
stk_rewards_item_481_addname_dataframe=pd.DataFrame()
stk_rewards_item_481_addname_dataframe['cname'] = stk_rewards_item_481_tscode_list
for table_name in stk_rewards_item_481.columns.values.tolist():
    stk_rewards_item_481_addname_dataframe[table_name] = stk_rewards_item_481[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_481_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_482_tscode_list = list() 
stk_rewards_item_482 = pro.stk_rewards(ts_code='600664.SH,600665.SH,600666.SH,600667.SH,600668.SH,600669.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_482 返回数据 row 行数 = "+str(stk_rewards_item_482.shape[0]))
for ts_code_sh in stk_rewards_item_482['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_482_tscode_list.append(stock_name)
stk_rewards_item_482_addname_dataframe=pd.DataFrame()
stk_rewards_item_482_addname_dataframe['cname'] = stk_rewards_item_482_tscode_list
for table_name in stk_rewards_item_482.columns.values.tolist():
    stk_rewards_item_482_addname_dataframe[table_name] = stk_rewards_item_482[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_482_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_483_tscode_list = list() 
stk_rewards_item_483 = pro.stk_rewards(ts_code='600670.SH,600671.SH,600672.SH,600673.SH,600674.SH,600675.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_483 返回数据 row 行数 = "+str(stk_rewards_item_483.shape[0]))
for ts_code_sh in stk_rewards_item_483['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_483_tscode_list.append(stock_name)
stk_rewards_item_483_addname_dataframe=pd.DataFrame()
stk_rewards_item_483_addname_dataframe['cname'] = stk_rewards_item_483_tscode_list
for table_name in stk_rewards_item_483.columns.values.tolist():
    stk_rewards_item_483_addname_dataframe[table_name] = stk_rewards_item_483[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_483_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_484_tscode_list = list() 
stk_rewards_item_484 = pro.stk_rewards(ts_code='600676.SH,600677.SH,600678.SH,600679.SH,600680.SH,600681.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_484 返回数据 row 行数 = "+str(stk_rewards_item_484.shape[0]))
for ts_code_sh in stk_rewards_item_484['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_484_tscode_list.append(stock_name)
stk_rewards_item_484_addname_dataframe=pd.DataFrame()
stk_rewards_item_484_addname_dataframe['cname'] = stk_rewards_item_484_tscode_list
for table_name in stk_rewards_item_484.columns.values.tolist():
    stk_rewards_item_484_addname_dataframe[table_name] = stk_rewards_item_484[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_484_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_485_tscode_list = list() 
stk_rewards_item_485 = pro.stk_rewards(ts_code='600682.SH,600683.SH,600684.SH,600685.SH,600686.SH,600687.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_485 返回数据 row 行数 = "+str(stk_rewards_item_485.shape[0]))
for ts_code_sh in stk_rewards_item_485['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_485_tscode_list.append(stock_name)
stk_rewards_item_485_addname_dataframe=pd.DataFrame()
stk_rewards_item_485_addname_dataframe['cname'] = stk_rewards_item_485_tscode_list
for table_name in stk_rewards_item_485.columns.values.tolist():
    stk_rewards_item_485_addname_dataframe[table_name] = stk_rewards_item_485[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_485_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_486_tscode_list = list() 
stk_rewards_item_486 = pro.stk_rewards(ts_code='600688.SH,600689.SH,600690.SH,600691.SH,600692.SH,600693.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_486 返回数据 row 行数 = "+str(stk_rewards_item_486.shape[0]))
for ts_code_sh in stk_rewards_item_486['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_486_tscode_list.append(stock_name)
stk_rewards_item_486_addname_dataframe=pd.DataFrame()
stk_rewards_item_486_addname_dataframe['cname'] = stk_rewards_item_486_tscode_list
for table_name in stk_rewards_item_486.columns.values.tolist():
    stk_rewards_item_486_addname_dataframe[table_name] = stk_rewards_item_486[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_486_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_487_tscode_list = list() 
stk_rewards_item_487 = pro.stk_rewards(ts_code='600694.SH,600695.SH,600696.SH,600697.SH,600698.SH,600699.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_487 返回数据 row 行数 = "+str(stk_rewards_item_487.shape[0]))
for ts_code_sh in stk_rewards_item_487['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_487_tscode_list.append(stock_name)
stk_rewards_item_487_addname_dataframe=pd.DataFrame()
stk_rewards_item_487_addname_dataframe['cname'] = stk_rewards_item_487_tscode_list
for table_name in stk_rewards_item_487.columns.values.tolist():
    stk_rewards_item_487_addname_dataframe[table_name] = stk_rewards_item_487[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_487_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_488_tscode_list = list() 
stk_rewards_item_488 = pro.stk_rewards(ts_code='600700.SH,600701.SH,600702.SH,600703.SH,600704.SH,600705.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_488 返回数据 row 行数 = "+str(stk_rewards_item_488.shape[0]))
for ts_code_sh in stk_rewards_item_488['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_488_tscode_list.append(stock_name)
stk_rewards_item_488_addname_dataframe=pd.DataFrame()
stk_rewards_item_488_addname_dataframe['cname'] = stk_rewards_item_488_tscode_list
for table_name in stk_rewards_item_488.columns.values.tolist():
    stk_rewards_item_488_addname_dataframe[table_name] = stk_rewards_item_488[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_488_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_489_tscode_list = list() 
stk_rewards_item_489 = pro.stk_rewards(ts_code='600706.SH,600707.SH,600708.SH,600709.SH,600710.SH,600711.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_489 返回数据 row 行数 = "+str(stk_rewards_item_489.shape[0]))
for ts_code_sh in stk_rewards_item_489['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_489_tscode_list.append(stock_name)
stk_rewards_item_489_addname_dataframe=pd.DataFrame()
stk_rewards_item_489_addname_dataframe['cname'] = stk_rewards_item_489_tscode_list
for table_name in stk_rewards_item_489.columns.values.tolist():
    stk_rewards_item_489_addname_dataframe[table_name] = stk_rewards_item_489[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_489_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_490_tscode_list = list() 
stk_rewards_item_490 = pro.stk_rewards(ts_code='600712.SH,600713.SH,600714.SH,600715.SH,600716.SH,600717.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_490 返回数据 row 行数 = "+str(stk_rewards_item_490.shape[0]))
for ts_code_sh in stk_rewards_item_490['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_490_tscode_list.append(stock_name)
stk_rewards_item_490_addname_dataframe=pd.DataFrame()
stk_rewards_item_490_addname_dataframe['cname'] = stk_rewards_item_490_tscode_list
for table_name in stk_rewards_item_490.columns.values.tolist():
    stk_rewards_item_490_addname_dataframe[table_name] = stk_rewards_item_490[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_490_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_491_tscode_list = list() 
stk_rewards_item_491 = pro.stk_rewards(ts_code='600718.SH,600719.SH,600720.SH,600721.SH,600722.SH,600723.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_491 返回数据 row 行数 = "+str(stk_rewards_item_491.shape[0]))
for ts_code_sh in stk_rewards_item_491['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_491_tscode_list.append(stock_name)
stk_rewards_item_491_addname_dataframe=pd.DataFrame()
stk_rewards_item_491_addname_dataframe['cname'] = stk_rewards_item_491_tscode_list
for table_name in stk_rewards_item_491.columns.values.tolist():
    stk_rewards_item_491_addname_dataframe[table_name] = stk_rewards_item_491[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_491_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_492_tscode_list = list() 
stk_rewards_item_492 = pro.stk_rewards(ts_code='600724.SH,600725.SH,600726.SH,600727.SH,600728.SH,600729.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_492 返回数据 row 行数 = "+str(stk_rewards_item_492.shape[0]))
for ts_code_sh in stk_rewards_item_492['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_492_tscode_list.append(stock_name)
stk_rewards_item_492_addname_dataframe=pd.DataFrame()
stk_rewards_item_492_addname_dataframe['cname'] = stk_rewards_item_492_tscode_list
for table_name in stk_rewards_item_492.columns.values.tolist():
    stk_rewards_item_492_addname_dataframe[table_name] = stk_rewards_item_492[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_492_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_493_tscode_list = list() 
stk_rewards_item_493 = pro.stk_rewards(ts_code='600730.SH,600731.SH,600732.SH,600733.SH,600734.SH,600735.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_493 返回数据 row 行数 = "+str(stk_rewards_item_493.shape[0]))
for ts_code_sh in stk_rewards_item_493['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_493_tscode_list.append(stock_name)
stk_rewards_item_493_addname_dataframe=pd.DataFrame()
stk_rewards_item_493_addname_dataframe['cname'] = stk_rewards_item_493_tscode_list
for table_name in stk_rewards_item_493.columns.values.tolist():
    stk_rewards_item_493_addname_dataframe[table_name] = stk_rewards_item_493[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_493_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_494_tscode_list = list() 
stk_rewards_item_494 = pro.stk_rewards(ts_code='600736.SH,600737.SH,600738.SH,600739.SH,600740.SH,600741.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_494 返回数据 row 行数 = "+str(stk_rewards_item_494.shape[0]))
for ts_code_sh in stk_rewards_item_494['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_494_tscode_list.append(stock_name)
stk_rewards_item_494_addname_dataframe=pd.DataFrame()
stk_rewards_item_494_addname_dataframe['cname'] = stk_rewards_item_494_tscode_list
for table_name in stk_rewards_item_494.columns.values.tolist():
    stk_rewards_item_494_addname_dataframe[table_name] = stk_rewards_item_494[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_494_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_495_tscode_list = list() 
stk_rewards_item_495 = pro.stk_rewards(ts_code='600742.SH,600743.SH,600744.SH,600745.SH,600746.SH,600747.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_495 返回数据 row 行数 = "+str(stk_rewards_item_495.shape[0]))
for ts_code_sh in stk_rewards_item_495['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_495_tscode_list.append(stock_name)
stk_rewards_item_495_addname_dataframe=pd.DataFrame()
stk_rewards_item_495_addname_dataframe['cname'] = stk_rewards_item_495_tscode_list
for table_name in stk_rewards_item_495.columns.values.tolist():
    stk_rewards_item_495_addname_dataframe[table_name] = stk_rewards_item_495[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_495_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_496_tscode_list = list() 
stk_rewards_item_496 = pro.stk_rewards(ts_code='600748.SH,600749.SH,600750.SH,600751.SH,600752.SH,600753.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_496 返回数据 row 行数 = "+str(stk_rewards_item_496.shape[0]))
for ts_code_sh in stk_rewards_item_496['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_496_tscode_list.append(stock_name)
stk_rewards_item_496_addname_dataframe=pd.DataFrame()
stk_rewards_item_496_addname_dataframe['cname'] = stk_rewards_item_496_tscode_list
for table_name in stk_rewards_item_496.columns.values.tolist():
    stk_rewards_item_496_addname_dataframe[table_name] = stk_rewards_item_496[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_496_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_497_tscode_list = list() 
stk_rewards_item_497 = pro.stk_rewards(ts_code='600754.SH,600755.SH,600756.SH,600757.SH,600758.SH,600759.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_497 返回数据 row 行数 = "+str(stk_rewards_item_497.shape[0]))
for ts_code_sh in stk_rewards_item_497['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_497_tscode_list.append(stock_name)
stk_rewards_item_497_addname_dataframe=pd.DataFrame()
stk_rewards_item_497_addname_dataframe['cname'] = stk_rewards_item_497_tscode_list
for table_name in stk_rewards_item_497.columns.values.tolist():
    stk_rewards_item_497_addname_dataframe[table_name] = stk_rewards_item_497[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_497_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_498_tscode_list = list() 
stk_rewards_item_498 = pro.stk_rewards(ts_code='600760.SH,600761.SH,600762.SH,600763.SH,600764.SH,600765.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_498 返回数据 row 行数 = "+str(stk_rewards_item_498.shape[0]))
for ts_code_sh in stk_rewards_item_498['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_498_tscode_list.append(stock_name)
stk_rewards_item_498_addname_dataframe=pd.DataFrame()
stk_rewards_item_498_addname_dataframe['cname'] = stk_rewards_item_498_tscode_list
for table_name in stk_rewards_item_498.columns.values.tolist():
    stk_rewards_item_498_addname_dataframe[table_name] = stk_rewards_item_498[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_498_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_499_tscode_list = list() 
stk_rewards_item_499 = pro.stk_rewards(ts_code='600766.SH,600767.SH,600768.SH,600769.SH,600770.SH,600771.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_499 返回数据 row 行数 = "+str(stk_rewards_item_499.shape[0]))
for ts_code_sh in stk_rewards_item_499['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_499_tscode_list.append(stock_name)
stk_rewards_item_499_addname_dataframe=pd.DataFrame()
stk_rewards_item_499_addname_dataframe['cname'] = stk_rewards_item_499_tscode_list
for table_name in stk_rewards_item_499.columns.values.tolist():
    stk_rewards_item_499_addname_dataframe[table_name] = stk_rewards_item_499[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_499_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_500_tscode_list = list() 
stk_rewards_item_500 = pro.stk_rewards(ts_code='600772.SH,600773.SH,600774.SH,600775.SH,600776.SH,600777.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_500 返回数据 row 行数 = "+str(stk_rewards_item_500.shape[0]))
for ts_code_sh in stk_rewards_item_500['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_500_tscode_list.append(stock_name)
stk_rewards_item_500_addname_dataframe=pd.DataFrame()
stk_rewards_item_500_addname_dataframe['cname'] = stk_rewards_item_500_tscode_list
for table_name in stk_rewards_item_500.columns.values.tolist():
    stk_rewards_item_500_addname_dataframe[table_name] = stk_rewards_item_500[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_500_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_501_tscode_list = list() 
stk_rewards_item_501 = pro.stk_rewards(ts_code='600778.SH,600779.SH,600780.SH,600781.SH,600782.SH,600783.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_501 返回数据 row 行数 = "+str(stk_rewards_item_501.shape[0]))
for ts_code_sh in stk_rewards_item_501['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_501_tscode_list.append(stock_name)
stk_rewards_item_501_addname_dataframe=pd.DataFrame()
stk_rewards_item_501_addname_dataframe['cname'] = stk_rewards_item_501_tscode_list
for table_name in stk_rewards_item_501.columns.values.tolist():
    stk_rewards_item_501_addname_dataframe[table_name] = stk_rewards_item_501[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_501_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_502_tscode_list = list() 
stk_rewards_item_502 = pro.stk_rewards(ts_code='600784.SH,600785.SH,600786.SH,600787.SH,600788.SH,600789.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_502 返回数据 row 行数 = "+str(stk_rewards_item_502.shape[0]))
for ts_code_sh in stk_rewards_item_502['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_502_tscode_list.append(stock_name)
stk_rewards_item_502_addname_dataframe=pd.DataFrame()
stk_rewards_item_502_addname_dataframe['cname'] = stk_rewards_item_502_tscode_list
for table_name in stk_rewards_item_502.columns.values.tolist():
    stk_rewards_item_502_addname_dataframe[table_name] = stk_rewards_item_502[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_502_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_503_tscode_list = list() 
stk_rewards_item_503 = pro.stk_rewards(ts_code='600790.SH,600791.SH,600792.SH,600793.SH,600794.SH,600795.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_503 返回数据 row 行数 = "+str(stk_rewards_item_503.shape[0]))
for ts_code_sh in stk_rewards_item_503['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_503_tscode_list.append(stock_name)
stk_rewards_item_503_addname_dataframe=pd.DataFrame()
stk_rewards_item_503_addname_dataframe['cname'] = stk_rewards_item_503_tscode_list
for table_name in stk_rewards_item_503.columns.values.tolist():
    stk_rewards_item_503_addname_dataframe[table_name] = stk_rewards_item_503[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_503_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_504_tscode_list = list() 
stk_rewards_item_504 = pro.stk_rewards(ts_code='600796.SH,600797.SH,600798.SH,600799.SH,600800.SH,600801.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_504 返回数据 row 行数 = "+str(stk_rewards_item_504.shape[0]))
for ts_code_sh in stk_rewards_item_504['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_504_tscode_list.append(stock_name)
stk_rewards_item_504_addname_dataframe=pd.DataFrame()
stk_rewards_item_504_addname_dataframe['cname'] = stk_rewards_item_504_tscode_list
for table_name in stk_rewards_item_504.columns.values.tolist():
    stk_rewards_item_504_addname_dataframe[table_name] = stk_rewards_item_504[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_504_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_505_tscode_list = list() 
stk_rewards_item_505 = pro.stk_rewards(ts_code='600802.SH,600803.SH,600804.SH,600805.SH,600806.SH,600807.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_505 返回数据 row 行数 = "+str(stk_rewards_item_505.shape[0]))
for ts_code_sh in stk_rewards_item_505['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_505_tscode_list.append(stock_name)
stk_rewards_item_505_addname_dataframe=pd.DataFrame()
stk_rewards_item_505_addname_dataframe['cname'] = stk_rewards_item_505_tscode_list
for table_name in stk_rewards_item_505.columns.values.tolist():
    stk_rewards_item_505_addname_dataframe[table_name] = stk_rewards_item_505[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_505_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_506_tscode_list = list() 
stk_rewards_item_506 = pro.stk_rewards(ts_code='600808.SH,600809.SH,600810.SH,600811.SH,600812.SH,600813.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_506 返回数据 row 行数 = "+str(stk_rewards_item_506.shape[0]))
for ts_code_sh in stk_rewards_item_506['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_506_tscode_list.append(stock_name)
stk_rewards_item_506_addname_dataframe=pd.DataFrame()
stk_rewards_item_506_addname_dataframe['cname'] = stk_rewards_item_506_tscode_list
for table_name in stk_rewards_item_506.columns.values.tolist():
    stk_rewards_item_506_addname_dataframe[table_name] = stk_rewards_item_506[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_506_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_507_tscode_list = list() 
stk_rewards_item_507 = pro.stk_rewards(ts_code='600814.SH,600815.SH,600816.SH,600817.SH,600818.SH,600819.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_507 返回数据 row 行数 = "+str(stk_rewards_item_507.shape[0]))
for ts_code_sh in stk_rewards_item_507['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_507_tscode_list.append(stock_name)
stk_rewards_item_507_addname_dataframe=pd.DataFrame()
stk_rewards_item_507_addname_dataframe['cname'] = stk_rewards_item_507_tscode_list
for table_name in stk_rewards_item_507.columns.values.tolist():
    stk_rewards_item_507_addname_dataframe[table_name] = stk_rewards_item_507[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_507_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_508_tscode_list = list() 
stk_rewards_item_508 = pro.stk_rewards(ts_code='600820.SH,600821.SH,600822.SH,600823.SH,600824.SH,600825.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_508 返回数据 row 行数 = "+str(stk_rewards_item_508.shape[0]))
for ts_code_sh in stk_rewards_item_508['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_508_tscode_list.append(stock_name)
stk_rewards_item_508_addname_dataframe=pd.DataFrame()
stk_rewards_item_508_addname_dataframe['cname'] = stk_rewards_item_508_tscode_list
for table_name in stk_rewards_item_508.columns.values.tolist():
    stk_rewards_item_508_addname_dataframe[table_name] = stk_rewards_item_508[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_508_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_509_tscode_list = list() 
stk_rewards_item_509 = pro.stk_rewards(ts_code='600826.SH,600827.SH,600828.SH,600829.SH,600830.SH,600831.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_509 返回数据 row 行数 = "+str(stk_rewards_item_509.shape[0]))
for ts_code_sh in stk_rewards_item_509['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_509_tscode_list.append(stock_name)
stk_rewards_item_509_addname_dataframe=pd.DataFrame()
stk_rewards_item_509_addname_dataframe['cname'] = stk_rewards_item_509_tscode_list
for table_name in stk_rewards_item_509.columns.values.tolist():
    stk_rewards_item_509_addname_dataframe[table_name] = stk_rewards_item_509[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_509_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_510_tscode_list = list() 
stk_rewards_item_510 = pro.stk_rewards(ts_code='600832.SH,600833.SH,600834.SH,600835.SH,600836.SH,600837.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_510 返回数据 row 行数 = "+str(stk_rewards_item_510.shape[0]))
for ts_code_sh in stk_rewards_item_510['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_510_tscode_list.append(stock_name)
stk_rewards_item_510_addname_dataframe=pd.DataFrame()
stk_rewards_item_510_addname_dataframe['cname'] = stk_rewards_item_510_tscode_list
for table_name in stk_rewards_item_510.columns.values.tolist():
    stk_rewards_item_510_addname_dataframe[table_name] = stk_rewards_item_510[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_510_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_511_tscode_list = list() 
stk_rewards_item_511 = pro.stk_rewards(ts_code='600838.SH,600839.SH,600840.SH,600841.SH,600842.SH,600843.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_511 返回数据 row 行数 = "+str(stk_rewards_item_511.shape[0]))
for ts_code_sh in stk_rewards_item_511['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_511_tscode_list.append(stock_name)
stk_rewards_item_511_addname_dataframe=pd.DataFrame()
stk_rewards_item_511_addname_dataframe['cname'] = stk_rewards_item_511_tscode_list
for table_name in stk_rewards_item_511.columns.values.tolist():
    stk_rewards_item_511_addname_dataframe[table_name] = stk_rewards_item_511[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_511_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_512_tscode_list = list() 
stk_rewards_item_512 = pro.stk_rewards(ts_code='600844.SH,600845.SH,600846.SH,600847.SH,600848.SH,600850.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_512 返回数据 row 行数 = "+str(stk_rewards_item_512.shape[0]))
for ts_code_sh in stk_rewards_item_512['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_512_tscode_list.append(stock_name)
stk_rewards_item_512_addname_dataframe=pd.DataFrame()
stk_rewards_item_512_addname_dataframe['cname'] = stk_rewards_item_512_tscode_list
for table_name in stk_rewards_item_512.columns.values.tolist():
    stk_rewards_item_512_addname_dataframe[table_name] = stk_rewards_item_512[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_512_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_513_tscode_list = list() 
stk_rewards_item_513 = pro.stk_rewards(ts_code='600851.SH,600852.SH,600853.SH,600854.SH,600855.SH,600856.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_513 返回数据 row 行数 = "+str(stk_rewards_item_513.shape[0]))
for ts_code_sh in stk_rewards_item_513['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_513_tscode_list.append(stock_name)
stk_rewards_item_513_addname_dataframe=pd.DataFrame()
stk_rewards_item_513_addname_dataframe['cname'] = stk_rewards_item_513_tscode_list
for table_name in stk_rewards_item_513.columns.values.tolist():
    stk_rewards_item_513_addname_dataframe[table_name] = stk_rewards_item_513[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_513_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_514_tscode_list = list() 
stk_rewards_item_514 = pro.stk_rewards(ts_code='600857.SH,600858.SH,600859.SH,600860.SH,600861.SH,600862.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_514 返回数据 row 行数 = "+str(stk_rewards_item_514.shape[0]))
for ts_code_sh in stk_rewards_item_514['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_514_tscode_list.append(stock_name)
stk_rewards_item_514_addname_dataframe=pd.DataFrame()
stk_rewards_item_514_addname_dataframe['cname'] = stk_rewards_item_514_tscode_list
for table_name in stk_rewards_item_514.columns.values.tolist():
    stk_rewards_item_514_addname_dataframe[table_name] = stk_rewards_item_514[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_514_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_515_tscode_list = list() 
stk_rewards_item_515 = pro.stk_rewards(ts_code='600863.SH,600864.SH,600865.SH,600866.SH,600867.SH,600868.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_515 返回数据 row 行数 = "+str(stk_rewards_item_515.shape[0]))
for ts_code_sh in stk_rewards_item_515['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_515_tscode_list.append(stock_name)
stk_rewards_item_515_addname_dataframe=pd.DataFrame()
stk_rewards_item_515_addname_dataframe['cname'] = stk_rewards_item_515_tscode_list
for table_name in stk_rewards_item_515.columns.values.tolist():
    stk_rewards_item_515_addname_dataframe[table_name] = stk_rewards_item_515[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_515_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_516_tscode_list = list() 
stk_rewards_item_516 = pro.stk_rewards(ts_code='600869.SH,600870.SH,600871.SH,600872.SH,600873.SH,600874.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_516 返回数据 row 行数 = "+str(stk_rewards_item_516.shape[0]))
for ts_code_sh in stk_rewards_item_516['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_516_tscode_list.append(stock_name)
stk_rewards_item_516_addname_dataframe=pd.DataFrame()
stk_rewards_item_516_addname_dataframe['cname'] = stk_rewards_item_516_tscode_list
for table_name in stk_rewards_item_516.columns.values.tolist():
    stk_rewards_item_516_addname_dataframe[table_name] = stk_rewards_item_516[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_516_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_517_tscode_list = list() 
stk_rewards_item_517 = pro.stk_rewards(ts_code='600875.SH,600876.SH,600877.SH,600878.SH,600879.SH,600880.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_517 返回数据 row 行数 = "+str(stk_rewards_item_517.shape[0]))
for ts_code_sh in stk_rewards_item_517['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_517_tscode_list.append(stock_name)
stk_rewards_item_517_addname_dataframe=pd.DataFrame()
stk_rewards_item_517_addname_dataframe['cname'] = stk_rewards_item_517_tscode_list
for table_name in stk_rewards_item_517.columns.values.tolist():
    stk_rewards_item_517_addname_dataframe[table_name] = stk_rewards_item_517[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_517_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_518_tscode_list = list() 
stk_rewards_item_518 = pro.stk_rewards(ts_code='600881.SH,600882.SH,600883.SH,600884.SH,600885.SH,600886.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_518 返回数据 row 行数 = "+str(stk_rewards_item_518.shape[0]))
for ts_code_sh in stk_rewards_item_518['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_518_tscode_list.append(stock_name)
stk_rewards_item_518_addname_dataframe=pd.DataFrame()
stk_rewards_item_518_addname_dataframe['cname'] = stk_rewards_item_518_tscode_list
for table_name in stk_rewards_item_518.columns.values.tolist():
    stk_rewards_item_518_addname_dataframe[table_name] = stk_rewards_item_518[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_518_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_519_tscode_list = list() 
stk_rewards_item_519 = pro.stk_rewards(ts_code='600887.SH,600888.SH,600889.SH,600890.SH,600891.SH,600892.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_519 返回数据 row 行数 = "+str(stk_rewards_item_519.shape[0]))
for ts_code_sh in stk_rewards_item_519['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_519_tscode_list.append(stock_name)
stk_rewards_item_519_addname_dataframe=pd.DataFrame()
stk_rewards_item_519_addname_dataframe['cname'] = stk_rewards_item_519_tscode_list
for table_name in stk_rewards_item_519.columns.values.tolist():
    stk_rewards_item_519_addname_dataframe[table_name] = stk_rewards_item_519[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_519_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_520_tscode_list = list() 
stk_rewards_item_520 = pro.stk_rewards(ts_code='600893.SH,600894.SH,600895.SH,600896.SH,600897.SH,600898.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_520 返回数据 row 行数 = "+str(stk_rewards_item_520.shape[0]))
for ts_code_sh in stk_rewards_item_520['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_520_tscode_list.append(stock_name)
stk_rewards_item_520_addname_dataframe=pd.DataFrame()
stk_rewards_item_520_addname_dataframe['cname'] = stk_rewards_item_520_tscode_list
for table_name in stk_rewards_item_520.columns.values.tolist():
    stk_rewards_item_520_addname_dataframe[table_name] = stk_rewards_item_520[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_520_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_521_tscode_list = list() 
stk_rewards_item_521 = pro.stk_rewards(ts_code='600899.SH,600900.SH,600901.SH,600903.SH,600908.SH,600909.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_521 返回数据 row 行数 = "+str(stk_rewards_item_521.shape[0]))
for ts_code_sh in stk_rewards_item_521['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_521_tscode_list.append(stock_name)
stk_rewards_item_521_addname_dataframe=pd.DataFrame()
stk_rewards_item_521_addname_dataframe['cname'] = stk_rewards_item_521_tscode_list
for table_name in stk_rewards_item_521.columns.values.tolist():
    stk_rewards_item_521_addname_dataframe[table_name] = stk_rewards_item_521[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_521_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_522_tscode_list = list() 
stk_rewards_item_522 = pro.stk_rewards(ts_code='600917.SH,600918.SH,600919.SH,600926.SH,600928.SH,600929.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_522 返回数据 row 行数 = "+str(stk_rewards_item_522.shape[0]))
for ts_code_sh in stk_rewards_item_522['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_522_tscode_list.append(stock_name)
stk_rewards_item_522_addname_dataframe=pd.DataFrame()
stk_rewards_item_522_addname_dataframe['cname'] = stk_rewards_item_522_tscode_list
for table_name in stk_rewards_item_522.columns.values.tolist():
    stk_rewards_item_522_addname_dataframe[table_name] = stk_rewards_item_522[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_522_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_523_tscode_list = list() 
stk_rewards_item_523 = pro.stk_rewards(ts_code='600933.SH,600936.SH,600939.SH,600956.SH,600958.SH,600959.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_523 返回数据 row 行数 = "+str(stk_rewards_item_523.shape[0]))
for ts_code_sh in stk_rewards_item_523['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_523_tscode_list.append(stock_name)
stk_rewards_item_523_addname_dataframe=pd.DataFrame()
stk_rewards_item_523_addname_dataframe['cname'] = stk_rewards_item_523_tscode_list
for table_name in stk_rewards_item_523.columns.values.tolist():
    stk_rewards_item_523_addname_dataframe[table_name] = stk_rewards_item_523[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_523_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_524_tscode_list = list() 
stk_rewards_item_524 = pro.stk_rewards(ts_code='600960.SH,600961.SH,600962.SH,600963.SH,600965.SH,600966.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_524 返回数据 row 行数 = "+str(stk_rewards_item_524.shape[0]))
for ts_code_sh in stk_rewards_item_524['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_524_tscode_list.append(stock_name)
stk_rewards_item_524_addname_dataframe=pd.DataFrame()
stk_rewards_item_524_addname_dataframe['cname'] = stk_rewards_item_524_tscode_list
for table_name in stk_rewards_item_524.columns.values.tolist():
    stk_rewards_item_524_addname_dataframe[table_name] = stk_rewards_item_524[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_524_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_525_tscode_list = list() 
stk_rewards_item_525 = pro.stk_rewards(ts_code='600967.SH,600968.SH,600969.SH,600970.SH,600971.SH,600973.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_525 返回数据 row 行数 = "+str(stk_rewards_item_525.shape[0]))
for ts_code_sh in stk_rewards_item_525['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_525_tscode_list.append(stock_name)
stk_rewards_item_525_addname_dataframe=pd.DataFrame()
stk_rewards_item_525_addname_dataframe['cname'] = stk_rewards_item_525_tscode_list
for table_name in stk_rewards_item_525.columns.values.tolist():
    stk_rewards_item_525_addname_dataframe[table_name] = stk_rewards_item_525[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_525_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_526_tscode_list = list() 
stk_rewards_item_526 = pro.stk_rewards(ts_code='600975.SH,600976.SH,600977.SH,600978.SH,600979.SH,600980.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_526 返回数据 row 行数 = "+str(stk_rewards_item_526.shape[0]))
for ts_code_sh in stk_rewards_item_526['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_526_tscode_list.append(stock_name)
stk_rewards_item_526_addname_dataframe=pd.DataFrame()
stk_rewards_item_526_addname_dataframe['cname'] = stk_rewards_item_526_tscode_list
for table_name in stk_rewards_item_526.columns.values.tolist():
    stk_rewards_item_526_addname_dataframe[table_name] = stk_rewards_item_526[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_526_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_527_tscode_list = list() 
stk_rewards_item_527 = pro.stk_rewards(ts_code='600981.SH,600982.SH,600983.SH,600984.SH,600985.SH,600986.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_527 返回数据 row 行数 = "+str(stk_rewards_item_527.shape[0]))
for ts_code_sh in stk_rewards_item_527['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_527_tscode_list.append(stock_name)
stk_rewards_item_527_addname_dataframe=pd.DataFrame()
stk_rewards_item_527_addname_dataframe['cname'] = stk_rewards_item_527_tscode_list
for table_name in stk_rewards_item_527.columns.values.tolist():
    stk_rewards_item_527_addname_dataframe[table_name] = stk_rewards_item_527[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_527_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_528_tscode_list = list() 
stk_rewards_item_528 = pro.stk_rewards(ts_code='600987.SH,600988.SH,600989.SH,600990.SH,600991.SH,600992.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_528 返回数据 row 行数 = "+str(stk_rewards_item_528.shape[0]))
for ts_code_sh in stk_rewards_item_528['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_528_tscode_list.append(stock_name)
stk_rewards_item_528_addname_dataframe=pd.DataFrame()
stk_rewards_item_528_addname_dataframe['cname'] = stk_rewards_item_528_tscode_list
for table_name in stk_rewards_item_528.columns.values.tolist():
    stk_rewards_item_528_addname_dataframe[table_name] = stk_rewards_item_528[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_528_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_529_tscode_list = list() 
stk_rewards_item_529 = pro.stk_rewards(ts_code='600993.SH,600995.SH,600996.SH,600997.SH,600998.SH,600999.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_529 返回数据 row 行数 = "+str(stk_rewards_item_529.shape[0]))
for ts_code_sh in stk_rewards_item_529['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_529_tscode_list.append(stock_name)
stk_rewards_item_529_addname_dataframe=pd.DataFrame()
stk_rewards_item_529_addname_dataframe['cname'] = stk_rewards_item_529_tscode_list
for table_name in stk_rewards_item_529.columns.values.tolist():
    stk_rewards_item_529_addname_dataframe[table_name] = stk_rewards_item_529[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_529_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_530_tscode_list = list() 
stk_rewards_item_530 = pro.stk_rewards(ts_code='601000.SH,601001.SH,601002.SH,601003.SH,601005.SH,601006.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_530 返回数据 row 行数 = "+str(stk_rewards_item_530.shape[0]))
for ts_code_sh in stk_rewards_item_530['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_530_tscode_list.append(stock_name)
stk_rewards_item_530_addname_dataframe=pd.DataFrame()
stk_rewards_item_530_addname_dataframe['cname'] = stk_rewards_item_530_tscode_list
for table_name in stk_rewards_item_530.columns.values.tolist():
    stk_rewards_item_530_addname_dataframe[table_name] = stk_rewards_item_530[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_530_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_531_tscode_list = list() 
stk_rewards_item_531 = pro.stk_rewards(ts_code='601007.SH,601008.SH,601009.SH,601010.SH,601011.SH,601012.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_531 返回数据 row 行数 = "+str(stk_rewards_item_531.shape[0]))
for ts_code_sh in stk_rewards_item_531['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_531_tscode_list.append(stock_name)
stk_rewards_item_531_addname_dataframe=pd.DataFrame()
stk_rewards_item_531_addname_dataframe['cname'] = stk_rewards_item_531_tscode_list
for table_name in stk_rewards_item_531.columns.values.tolist():
    stk_rewards_item_531_addname_dataframe[table_name] = stk_rewards_item_531[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_531_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_532_tscode_list = list() 
stk_rewards_item_532 = pro.stk_rewards(ts_code='601015.SH,601016.SH,601018.SH,601019.SH,601020.SH,601021.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_532 返回数据 row 行数 = "+str(stk_rewards_item_532.shape[0]))
for ts_code_sh in stk_rewards_item_532['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_532_tscode_list.append(stock_name)
stk_rewards_item_532_addname_dataframe=pd.DataFrame()
stk_rewards_item_532_addname_dataframe['cname'] = stk_rewards_item_532_tscode_list
for table_name in stk_rewards_item_532.columns.values.tolist():
    stk_rewards_item_532_addname_dataframe[table_name] = stk_rewards_item_532[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_532_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_533_tscode_list = list() 
stk_rewards_item_533 = pro.stk_rewards(ts_code='601028.SH,601038.SH,601058.SH,601066.SH,601068.SH,601069.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_533 返回数据 row 行数 = "+str(stk_rewards_item_533.shape[0]))
for ts_code_sh in stk_rewards_item_533['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_533_tscode_list.append(stock_name)
stk_rewards_item_533_addname_dataframe=pd.DataFrame()
stk_rewards_item_533_addname_dataframe['cname'] = stk_rewards_item_533_tscode_list
for table_name in stk_rewards_item_533.columns.values.tolist():
    stk_rewards_item_533_addname_dataframe[table_name] = stk_rewards_item_533[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_533_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_534_tscode_list = list() 
stk_rewards_item_534 = pro.stk_rewards(ts_code='601077.SH,601086.SH,601088.SH,601098.SH,601099.SH,601100.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_534 返回数据 row 行数 = "+str(stk_rewards_item_534.shape[0]))
for ts_code_sh in stk_rewards_item_534['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_534_tscode_list.append(stock_name)
stk_rewards_item_534_addname_dataframe=pd.DataFrame()
stk_rewards_item_534_addname_dataframe['cname'] = stk_rewards_item_534_tscode_list
for table_name in stk_rewards_item_534.columns.values.tolist():
    stk_rewards_item_534_addname_dataframe[table_name] = stk_rewards_item_534[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_534_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_535_tscode_list = list() 
stk_rewards_item_535 = pro.stk_rewards(ts_code='601101.SH,601106.SH,601107.SH,601108.SH,601111.SH,601113.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_535 返回数据 row 行数 = "+str(stk_rewards_item_535.shape[0]))
for ts_code_sh in stk_rewards_item_535['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_535_tscode_list.append(stock_name)
stk_rewards_item_535_addname_dataframe=pd.DataFrame()
stk_rewards_item_535_addname_dataframe['cname'] = stk_rewards_item_535_tscode_list
for table_name in stk_rewards_item_535.columns.values.tolist():
    stk_rewards_item_535_addname_dataframe[table_name] = stk_rewards_item_535[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_535_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_536_tscode_list = list() 
stk_rewards_item_536 = pro.stk_rewards(ts_code='601116.SH,601117.SH,601118.SH,601126.SH,601127.SH,601128.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_536 返回数据 row 行数 = "+str(stk_rewards_item_536.shape[0]))
for ts_code_sh in stk_rewards_item_536['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_536_tscode_list.append(stock_name)
stk_rewards_item_536_addname_dataframe=pd.DataFrame()
stk_rewards_item_536_addname_dataframe['cname'] = stk_rewards_item_536_tscode_list
for table_name in stk_rewards_item_536.columns.values.tolist():
    stk_rewards_item_536_addname_dataframe[table_name] = stk_rewards_item_536[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_536_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_537_tscode_list = list() 
stk_rewards_item_537 = pro.stk_rewards(ts_code='601137.SH,601138.SH,601139.SH,601155.SH,601158.SH,601162.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_537 返回数据 row 行数 = "+str(stk_rewards_item_537.shape[0]))
for ts_code_sh in stk_rewards_item_537['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_537_tscode_list.append(stock_name)
stk_rewards_item_537_addname_dataframe=pd.DataFrame()
stk_rewards_item_537_addname_dataframe['cname'] = stk_rewards_item_537_tscode_list
for table_name in stk_rewards_item_537.columns.values.tolist():
    stk_rewards_item_537_addname_dataframe[table_name] = stk_rewards_item_537[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_537_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_538_tscode_list = list() 
stk_rewards_item_538 = pro.stk_rewards(ts_code='601163.SH,601166.SH,601168.SH,601169.SH,601177.SH,601179.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_538 返回数据 row 行数 = "+str(stk_rewards_item_538.shape[0]))
for ts_code_sh in stk_rewards_item_538['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_538_tscode_list.append(stock_name)
stk_rewards_item_538_addname_dataframe=pd.DataFrame()
stk_rewards_item_538_addname_dataframe['cname'] = stk_rewards_item_538_tscode_list
for table_name in stk_rewards_item_538.columns.values.tolist():
    stk_rewards_item_538_addname_dataframe[table_name] = stk_rewards_item_538[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_538_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_539_tscode_list = list() 
stk_rewards_item_539 = pro.stk_rewards(ts_code='601186.SH,601188.SH,601198.SH,601199.SH,601200.SH,601208.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_539 返回数据 row 行数 = "+str(stk_rewards_item_539.shape[0]))
for ts_code_sh in stk_rewards_item_539['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_539_tscode_list.append(stock_name)
stk_rewards_item_539_addname_dataframe=pd.DataFrame()
stk_rewards_item_539_addname_dataframe['cname'] = stk_rewards_item_539_tscode_list
for table_name in stk_rewards_item_539.columns.values.tolist():
    stk_rewards_item_539_addname_dataframe[table_name] = stk_rewards_item_539[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_539_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_540_tscode_list = list() 
stk_rewards_item_540 = pro.stk_rewards(ts_code='601211.SH,601212.SH,601216.SH,601218.SH,601222.SH,601225.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_540 返回数据 row 行数 = "+str(stk_rewards_item_540.shape[0]))
for ts_code_sh in stk_rewards_item_540['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_540_tscode_list.append(stock_name)
stk_rewards_item_540_addname_dataframe=pd.DataFrame()
stk_rewards_item_540_addname_dataframe['cname'] = stk_rewards_item_540_tscode_list
for table_name in stk_rewards_item_540.columns.values.tolist():
    stk_rewards_item_540_addname_dataframe[table_name] = stk_rewards_item_540[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_540_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_541_tscode_list = list() 
stk_rewards_item_541 = pro.stk_rewards(ts_code='601226.SH,601228.SH,601229.SH,601231.SH,601233.SH,601236.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_541 返回数据 row 行数 = "+str(stk_rewards_item_541.shape[0]))
for ts_code_sh in stk_rewards_item_541['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_541_tscode_list.append(stock_name)
stk_rewards_item_541_addname_dataframe=pd.DataFrame()
stk_rewards_item_541_addname_dataframe['cname'] = stk_rewards_item_541_tscode_list
for table_name in stk_rewards_item_541.columns.values.tolist():
    stk_rewards_item_541_addname_dataframe[table_name] = stk_rewards_item_541[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_541_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_542_tscode_list = list() 
stk_rewards_item_542 = pro.stk_rewards(ts_code='601238.SH,601258.SH,601268.SH,601288.SH,601298.SH,601299.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_542 返回数据 row 行数 = "+str(stk_rewards_item_542.shape[0]))
for ts_code_sh in stk_rewards_item_542['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_542_tscode_list.append(stock_name)
stk_rewards_item_542_addname_dataframe=pd.DataFrame()
stk_rewards_item_542_addname_dataframe['cname'] = stk_rewards_item_542_tscode_list
for table_name in stk_rewards_item_542.columns.values.tolist():
    stk_rewards_item_542_addname_dataframe[table_name] = stk_rewards_item_542[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_542_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_543_tscode_list = list() 
stk_rewards_item_543 = pro.stk_rewards(ts_code='601311.SH,601318.SH,601319.SH,601326.SH,601328.SH,601330.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_543 返回数据 row 行数 = "+str(stk_rewards_item_543.shape[0]))
for ts_code_sh in stk_rewards_item_543['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_543_tscode_list.append(stock_name)
stk_rewards_item_543_addname_dataframe=pd.DataFrame()
stk_rewards_item_543_addname_dataframe['cname'] = stk_rewards_item_543_tscode_list
for table_name in stk_rewards_item_543.columns.values.tolist():
    stk_rewards_item_543_addname_dataframe[table_name] = stk_rewards_item_543[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_543_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_544_tscode_list = list() 
stk_rewards_item_544 = pro.stk_rewards(ts_code='601333.SH,601336.SH,601339.SH,601360.SH,601366.SH,601368.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_544 返回数据 row 行数 = "+str(stk_rewards_item_544.shape[0]))
for ts_code_sh in stk_rewards_item_544['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_544_tscode_list.append(stock_name)
stk_rewards_item_544_addname_dataframe=pd.DataFrame()
stk_rewards_item_544_addname_dataframe['cname'] = stk_rewards_item_544_tscode_list
for table_name in stk_rewards_item_544.columns.values.tolist():
    stk_rewards_item_544_addname_dataframe[table_name] = stk_rewards_item_544[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_544_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_545_tscode_list = list() 
stk_rewards_item_545 = pro.stk_rewards(ts_code='601369.SH,601375.SH,601377.SH,601388.SH,601390.SH,601398.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_545 返回数据 row 行数 = "+str(stk_rewards_item_545.shape[0]))
for ts_code_sh in stk_rewards_item_545['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_545_tscode_list.append(stock_name)
stk_rewards_item_545_addname_dataframe=pd.DataFrame()
stk_rewards_item_545_addname_dataframe['cname'] = stk_rewards_item_545_tscode_list
for table_name in stk_rewards_item_545.columns.values.tolist():
    stk_rewards_item_545_addname_dataframe[table_name] = stk_rewards_item_545[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_545_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_546_tscode_list = list() 
stk_rewards_item_546 = pro.stk_rewards(ts_code='601399.SH,601456.SH,601500.SH,601512.SH,601515.SH,601518.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_546 返回数据 row 行数 = "+str(stk_rewards_item_546.shape[0]))
for ts_code_sh in stk_rewards_item_546['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_546_tscode_list.append(stock_name)
stk_rewards_item_546_addname_dataframe=pd.DataFrame()
stk_rewards_item_546_addname_dataframe['cname'] = stk_rewards_item_546_tscode_list
for table_name in stk_rewards_item_546.columns.values.tolist():
    stk_rewards_item_546_addname_dataframe[table_name] = stk_rewards_item_546[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_546_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_547_tscode_list = list() 
stk_rewards_item_547 = pro.stk_rewards(ts_code='601519.SH,601555.SH,601558.SH,601566.SH,601567.SH,601577.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_547 返回数据 row 行数 = "+str(stk_rewards_item_547.shape[0]))
for ts_code_sh in stk_rewards_item_547['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_547_tscode_list.append(stock_name)
stk_rewards_item_547_addname_dataframe=pd.DataFrame()
stk_rewards_item_547_addname_dataframe['cname'] = stk_rewards_item_547_tscode_list
for table_name in stk_rewards_item_547.columns.values.tolist():
    stk_rewards_item_547_addname_dataframe[table_name] = stk_rewards_item_547[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_547_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_548_tscode_list = list() 
stk_rewards_item_548 = pro.stk_rewards(ts_code='601579.SH,601588.SH,601595.SH,601598.SH,601599.SH,601600.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_548 返回数据 row 行数 = "+str(stk_rewards_item_548.shape[0]))
for ts_code_sh in stk_rewards_item_548['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_548_tscode_list.append(stock_name)
stk_rewards_item_548_addname_dataframe=pd.DataFrame()
stk_rewards_item_548_addname_dataframe['cname'] = stk_rewards_item_548_tscode_list
for table_name in stk_rewards_item_548.columns.values.tolist():
    stk_rewards_item_548_addname_dataframe[table_name] = stk_rewards_item_548[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_548_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_549_tscode_list = list() 
stk_rewards_item_549 = pro.stk_rewards(ts_code='601601.SH,601606.SH,601607.SH,601608.SH,601609.SH,601611.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_549 返回数据 row 行数 = "+str(stk_rewards_item_549.shape[0]))
for ts_code_sh in stk_rewards_item_549['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_549_tscode_list.append(stock_name)
stk_rewards_item_549_addname_dataframe=pd.DataFrame()
stk_rewards_item_549_addname_dataframe['cname'] = stk_rewards_item_549_tscode_list
for table_name in stk_rewards_item_549.columns.values.tolist():
    stk_rewards_item_549_addname_dataframe[table_name] = stk_rewards_item_549[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_549_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_550_tscode_list = list() 
stk_rewards_item_550 = pro.stk_rewards(ts_code='601615.SH,601616.SH,601618.SH,601619.SH,601628.SH,601633.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_550 返回数据 row 行数 = "+str(stk_rewards_item_550.shape[0]))
for ts_code_sh in stk_rewards_item_550['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_550_tscode_list.append(stock_name)
stk_rewards_item_550_addname_dataframe=pd.DataFrame()
stk_rewards_item_550_addname_dataframe['cname'] = stk_rewards_item_550_tscode_list
for table_name in stk_rewards_item_550.columns.values.tolist():
    stk_rewards_item_550_addname_dataframe[table_name] = stk_rewards_item_550[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_550_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_551_tscode_list = list() 
stk_rewards_item_551 = pro.stk_rewards(ts_code='601636.SH,601658.SH,601666.SH,601668.SH,601669.SH,601677.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_551 返回数据 row 行数 = "+str(stk_rewards_item_551.shape[0]))
for ts_code_sh in stk_rewards_item_551['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_551_tscode_list.append(stock_name)
stk_rewards_item_551_addname_dataframe=pd.DataFrame()
stk_rewards_item_551_addname_dataframe['cname'] = stk_rewards_item_551_tscode_list
for table_name in stk_rewards_item_551.columns.values.tolist():
    stk_rewards_item_551_addname_dataframe[table_name] = stk_rewards_item_551[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_551_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_552_tscode_list = list() 
stk_rewards_item_552 = pro.stk_rewards(ts_code='601678.SH,601688.SH,601689.SH,601696.SH,601698.SH,601699.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_552 返回数据 row 行数 = "+str(stk_rewards_item_552.shape[0]))
for ts_code_sh in stk_rewards_item_552['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_552_tscode_list.append(stock_name)
stk_rewards_item_552_addname_dataframe=pd.DataFrame()
stk_rewards_item_552_addname_dataframe['cname'] = stk_rewards_item_552_tscode_list
for table_name in stk_rewards_item_552.columns.values.tolist():
    stk_rewards_item_552_addname_dataframe[table_name] = stk_rewards_item_552[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_552_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_553_tscode_list = list() 
stk_rewards_item_553 = pro.stk_rewards(ts_code='601700.SH,601717.SH,601718.SH,601727.SH,601766.SH,601777.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_553 返回数据 row 行数 = "+str(stk_rewards_item_553.shape[0]))
for ts_code_sh in stk_rewards_item_553['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_553_tscode_list.append(stock_name)
stk_rewards_item_553_addname_dataframe=pd.DataFrame()
stk_rewards_item_553_addname_dataframe['cname'] = stk_rewards_item_553_tscode_list
for table_name in stk_rewards_item_553.columns.values.tolist():
    stk_rewards_item_553_addname_dataframe[table_name] = stk_rewards_item_553[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_553_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_554_tscode_list = list() 
stk_rewards_item_554 = pro.stk_rewards(ts_code='601778.SH,601788.SH,601789.SH,601798.SH,601799.SH,601800.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_554 返回数据 row 行数 = "+str(stk_rewards_item_554.shape[0]))
for ts_code_sh in stk_rewards_item_554['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_554_tscode_list.append(stock_name)
stk_rewards_item_554_addname_dataframe=pd.DataFrame()
stk_rewards_item_554_addname_dataframe['cname'] = stk_rewards_item_554_tscode_list
for table_name in stk_rewards_item_554.columns.values.tolist():
    stk_rewards_item_554_addname_dataframe[table_name] = stk_rewards_item_554[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_554_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_555_tscode_list = list() 
stk_rewards_item_555 = pro.stk_rewards(ts_code='601801.SH,601808.SH,601811.SH,601816.SH,601818.SH,601827.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_555 返回数据 row 行数 = "+str(stk_rewards_item_555.shape[0]))
for ts_code_sh in stk_rewards_item_555['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_555_tscode_list.append(stock_name)
stk_rewards_item_555_addname_dataframe=pd.DataFrame()
stk_rewards_item_555_addname_dataframe['cname'] = stk_rewards_item_555_tscode_list
for table_name in stk_rewards_item_555.columns.values.tolist():
    stk_rewards_item_555_addname_dataframe[table_name] = stk_rewards_item_555[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_555_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_556_tscode_list = list() 
stk_rewards_item_556 = pro.stk_rewards(ts_code='601828.SH,601838.SH,601857.SH,601858.SH,601860.SH,601865.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_556 返回数据 row 行数 = "+str(stk_rewards_item_556.shape[0]))
for ts_code_sh in stk_rewards_item_556['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_556_tscode_list.append(stock_name)
stk_rewards_item_556_addname_dataframe=pd.DataFrame()
stk_rewards_item_556_addname_dataframe['cname'] = stk_rewards_item_556_tscode_list
for table_name in stk_rewards_item_556.columns.values.tolist():
    stk_rewards_item_556_addname_dataframe[table_name] = stk_rewards_item_556[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_556_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_557_tscode_list = list() 
stk_rewards_item_557 = pro.stk_rewards(ts_code='601866.SH,601869.SH,601872.SH,601877.SH,601878.SH,601880.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_557 返回数据 row 行数 = "+str(stk_rewards_item_557.shape[0]))
for ts_code_sh in stk_rewards_item_557['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_557_tscode_list.append(stock_name)
stk_rewards_item_557_addname_dataframe=pd.DataFrame()
stk_rewards_item_557_addname_dataframe['cname'] = stk_rewards_item_557_tscode_list
for table_name in stk_rewards_item_557.columns.values.tolist():
    stk_rewards_item_557_addname_dataframe[table_name] = stk_rewards_item_557[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_557_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_558_tscode_list = list() 
stk_rewards_item_558 = pro.stk_rewards(ts_code='601881.SH,601882.SH,601886.SH,601888.SH,601890.SH,601898.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_558 返回数据 row 行数 = "+str(stk_rewards_item_558.shape[0]))
for ts_code_sh in stk_rewards_item_558['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_558_tscode_list.append(stock_name)
stk_rewards_item_558_addname_dataframe=pd.DataFrame()
stk_rewards_item_558_addname_dataframe['cname'] = stk_rewards_item_558_tscode_list
for table_name in stk_rewards_item_558.columns.values.tolist():
    stk_rewards_item_558_addname_dataframe[table_name] = stk_rewards_item_558[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_558_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_559_tscode_list = list() 
stk_rewards_item_559 = pro.stk_rewards(ts_code='601899.SH,601900.SH,601901.SH,601908.SH,601916.SH,601918.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_559 返回数据 row 行数 = "+str(stk_rewards_item_559.shape[0]))
for ts_code_sh in stk_rewards_item_559['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_559_tscode_list.append(stock_name)
stk_rewards_item_559_addname_dataframe=pd.DataFrame()
stk_rewards_item_559_addname_dataframe['cname'] = stk_rewards_item_559_tscode_list
for table_name in stk_rewards_item_559.columns.values.tolist():
    stk_rewards_item_559_addname_dataframe[table_name] = stk_rewards_item_559[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_559_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_560_tscode_list = list() 
stk_rewards_item_560 = pro.stk_rewards(ts_code='601919.SH,601928.SH,601929.SH,601933.SH,601939.SH,601949.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_560 返回数据 row 行数 = "+str(stk_rewards_item_560.shape[0]))
for ts_code_sh in stk_rewards_item_560['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_560_tscode_list.append(stock_name)
stk_rewards_item_560_addname_dataframe=pd.DataFrame()
stk_rewards_item_560_addname_dataframe['cname'] = stk_rewards_item_560_tscode_list
for table_name in stk_rewards_item_560.columns.values.tolist():
    stk_rewards_item_560_addname_dataframe[table_name] = stk_rewards_item_560[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_560_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_561_tscode_list = list() 
stk_rewards_item_561 = pro.stk_rewards(ts_code='601952.SH,601958.SH,601965.SH,601966.SH,601968.SH,601969.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_561 返回数据 row 行数 = "+str(stk_rewards_item_561.shape[0]))
for ts_code_sh in stk_rewards_item_561['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_561_tscode_list.append(stock_name)
stk_rewards_item_561_addname_dataframe=pd.DataFrame()
stk_rewards_item_561_addname_dataframe['cname'] = stk_rewards_item_561_tscode_list
for table_name in stk_rewards_item_561.columns.values.tolist():
    stk_rewards_item_561_addname_dataframe[table_name] = stk_rewards_item_561[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_561_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_562_tscode_list = list() 
stk_rewards_item_562 = pro.stk_rewards(ts_code='601975.SH,601985.SH,601988.SH,601989.SH,601990.SH,601991.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_562 返回数据 row 行数 = "+str(stk_rewards_item_562.shape[0]))
for ts_code_sh in stk_rewards_item_562['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_562_tscode_list.append(stock_name)
stk_rewards_item_562_addname_dataframe=pd.DataFrame()
stk_rewards_item_562_addname_dataframe['cname'] = stk_rewards_item_562_tscode_list
for table_name in stk_rewards_item_562.columns.values.tolist():
    stk_rewards_item_562_addname_dataframe[table_name] = stk_rewards_item_562[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_562_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_563_tscode_list = list() 
stk_rewards_item_563 = pro.stk_rewards(ts_code='601992.SH,601996.SH,601997.SH,601998.SH,601999.SH,603000.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_563 返回数据 row 行数 = "+str(stk_rewards_item_563.shape[0]))
for ts_code_sh in stk_rewards_item_563['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_563_tscode_list.append(stock_name)
stk_rewards_item_563_addname_dataframe=pd.DataFrame()
stk_rewards_item_563_addname_dataframe['cname'] = stk_rewards_item_563_tscode_list
for table_name in stk_rewards_item_563.columns.values.tolist():
    stk_rewards_item_563_addname_dataframe[table_name] = stk_rewards_item_563[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_563_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_564_tscode_list = list() 
stk_rewards_item_564 = pro.stk_rewards(ts_code='603001.SH,603002.SH,603003.SH,603005.SH,603006.SH,603007.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_564 返回数据 row 行数 = "+str(stk_rewards_item_564.shape[0]))
for ts_code_sh in stk_rewards_item_564['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_564_tscode_list.append(stock_name)
stk_rewards_item_564_addname_dataframe=pd.DataFrame()
stk_rewards_item_564_addname_dataframe['cname'] = stk_rewards_item_564_tscode_list
for table_name in stk_rewards_item_564.columns.values.tolist():
    stk_rewards_item_564_addname_dataframe[table_name] = stk_rewards_item_564[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_564_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_565_tscode_list = list() 
stk_rewards_item_565 = pro.stk_rewards(ts_code='603008.SH,603009.SH,603010.SH,603011.SH,603012.SH,603013.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_565 返回数据 row 行数 = "+str(stk_rewards_item_565.shape[0]))
for ts_code_sh in stk_rewards_item_565['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_565_tscode_list.append(stock_name)
stk_rewards_item_565_addname_dataframe=pd.DataFrame()
stk_rewards_item_565_addname_dataframe['cname'] = stk_rewards_item_565_tscode_list
for table_name in stk_rewards_item_565.columns.values.tolist():
    stk_rewards_item_565_addname_dataframe[table_name] = stk_rewards_item_565[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_565_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_566_tscode_list = list() 
stk_rewards_item_566 = pro.stk_rewards(ts_code='603015.SH,603016.SH,603017.SH,603018.SH,603019.SH,603020.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_566 返回数据 row 行数 = "+str(stk_rewards_item_566.shape[0]))
for ts_code_sh in stk_rewards_item_566['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_566_tscode_list.append(stock_name)
stk_rewards_item_566_addname_dataframe=pd.DataFrame()
stk_rewards_item_566_addname_dataframe['cname'] = stk_rewards_item_566_tscode_list
for table_name in stk_rewards_item_566.columns.values.tolist():
    stk_rewards_item_566_addname_dataframe[table_name] = stk_rewards_item_566[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_566_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_567_tscode_list = list() 
stk_rewards_item_567 = pro.stk_rewards(ts_code='603021.SH,603022.SH,603023.SH,603025.SH,603026.SH,603027.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_567 返回数据 row 行数 = "+str(stk_rewards_item_567.shape[0]))
for ts_code_sh in stk_rewards_item_567['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_567_tscode_list.append(stock_name)
stk_rewards_item_567_addname_dataframe=pd.DataFrame()
stk_rewards_item_567_addname_dataframe['cname'] = stk_rewards_item_567_tscode_list
for table_name in stk_rewards_item_567.columns.values.tolist():
    stk_rewards_item_567_addname_dataframe[table_name] = stk_rewards_item_567[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_567_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_568_tscode_list = list() 
stk_rewards_item_568 = pro.stk_rewards(ts_code='603028.SH,603029.SH,603030.SH,603031.SH,603032.SH,603033.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_568 返回数据 row 行数 = "+str(stk_rewards_item_568.shape[0]))
for ts_code_sh in stk_rewards_item_568['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_568_tscode_list.append(stock_name)
stk_rewards_item_568_addname_dataframe=pd.DataFrame()
stk_rewards_item_568_addname_dataframe['cname'] = stk_rewards_item_568_tscode_list
for table_name in stk_rewards_item_568.columns.values.tolist():
    stk_rewards_item_568_addname_dataframe[table_name] = stk_rewards_item_568[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_568_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_569_tscode_list = list() 
stk_rewards_item_569 = pro.stk_rewards(ts_code='603035.SH,603036.SH,603037.SH,603038.SH,603039.SH,603040.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_569 返回数据 row 行数 = "+str(stk_rewards_item_569.shape[0]))
for ts_code_sh in stk_rewards_item_569['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_569_tscode_list.append(stock_name)
stk_rewards_item_569_addname_dataframe=pd.DataFrame()
stk_rewards_item_569_addname_dataframe['cname'] = stk_rewards_item_569_tscode_list
for table_name in stk_rewards_item_569.columns.values.tolist():
    stk_rewards_item_569_addname_dataframe[table_name] = stk_rewards_item_569[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_569_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_570_tscode_list = list() 
stk_rewards_item_570 = pro.stk_rewards(ts_code='603041.SH,603042.SH,603043.SH,603045.SH,603050.SH,603053.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_570 返回数据 row 行数 = "+str(stk_rewards_item_570.shape[0]))
for ts_code_sh in stk_rewards_item_570['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_570_tscode_list.append(stock_name)
stk_rewards_item_570_addname_dataframe=pd.DataFrame()
stk_rewards_item_570_addname_dataframe['cname'] = stk_rewards_item_570_tscode_list
for table_name in stk_rewards_item_570.columns.values.tolist():
    stk_rewards_item_570_addname_dataframe[table_name] = stk_rewards_item_570[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_570_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_571_tscode_list = list() 
stk_rewards_item_571 = pro.stk_rewards(ts_code='603055.SH,603056.SH,603058.SH,603059.SH,603060.SH,603063.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_571 返回数据 row 行数 = "+str(stk_rewards_item_571.shape[0]))
for ts_code_sh in stk_rewards_item_571['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_571_tscode_list.append(stock_name)
stk_rewards_item_571_addname_dataframe=pd.DataFrame()
stk_rewards_item_571_addname_dataframe['cname'] = stk_rewards_item_571_tscode_list
for table_name in stk_rewards_item_571.columns.values.tolist():
    stk_rewards_item_571_addname_dataframe[table_name] = stk_rewards_item_571[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_571_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_572_tscode_list = list() 
stk_rewards_item_572 = pro.stk_rewards(ts_code='603066.SH,603067.SH,603068.SH,603069.SH,603076.SH,603077.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_572 返回数据 row 行数 = "+str(stk_rewards_item_572.shape[0]))
for ts_code_sh in stk_rewards_item_572['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_572_tscode_list.append(stock_name)
stk_rewards_item_572_addname_dataframe=pd.DataFrame()
stk_rewards_item_572_addname_dataframe['cname'] = stk_rewards_item_572_tscode_list
for table_name in stk_rewards_item_572.columns.values.tolist():
    stk_rewards_item_572_addname_dataframe[table_name] = stk_rewards_item_572[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_572_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_573_tscode_list = list() 
stk_rewards_item_573 = pro.stk_rewards(ts_code='603078.SH,603079.SH,603080.SH,603081.SH,603083.SH,603085.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_573 返回数据 row 行数 = "+str(stk_rewards_item_573.shape[0]))
for ts_code_sh in stk_rewards_item_573['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_573_tscode_list.append(stock_name)
stk_rewards_item_573_addname_dataframe=pd.DataFrame()
stk_rewards_item_573_addname_dataframe['cname'] = stk_rewards_item_573_tscode_list
for table_name in stk_rewards_item_573.columns.values.tolist():
    stk_rewards_item_573_addname_dataframe[table_name] = stk_rewards_item_573[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_573_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_574_tscode_list = list() 
stk_rewards_item_574 = pro.stk_rewards(ts_code='603086.SH,603087.SH,603088.SH,603089.SH,603090.SH,603093.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_574 返回数据 row 行数 = "+str(stk_rewards_item_574.shape[0]))
for ts_code_sh in stk_rewards_item_574['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_574_tscode_list.append(stock_name)
stk_rewards_item_574_addname_dataframe=pd.DataFrame()
stk_rewards_item_574_addname_dataframe['cname'] = stk_rewards_item_574_tscode_list
for table_name in stk_rewards_item_574.columns.values.tolist():
    stk_rewards_item_574_addname_dataframe[table_name] = stk_rewards_item_574[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_574_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_575_tscode_list = list() 
stk_rewards_item_575 = pro.stk_rewards(ts_code='603095.SH,603096.SH,603098.SH,603099.SH,603100.SH,603101.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_575 返回数据 row 行数 = "+str(stk_rewards_item_575.shape[0]))
for ts_code_sh in stk_rewards_item_575['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_575_tscode_list.append(stock_name)
stk_rewards_item_575_addname_dataframe=pd.DataFrame()
stk_rewards_item_575_addname_dataframe['cname'] = stk_rewards_item_575_tscode_list
for table_name in stk_rewards_item_575.columns.values.tolist():
    stk_rewards_item_575_addname_dataframe[table_name] = stk_rewards_item_575[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_575_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_576_tscode_list = list() 
stk_rewards_item_576 = pro.stk_rewards(ts_code='603103.SH,603105.SH,603106.SH,603108.SH,603109.SH,603110.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_576 返回数据 row 行数 = "+str(stk_rewards_item_576.shape[0]))
for ts_code_sh in stk_rewards_item_576['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_576_tscode_list.append(stock_name)
stk_rewards_item_576_addname_dataframe=pd.DataFrame()
stk_rewards_item_576_addname_dataframe['cname'] = stk_rewards_item_576_tscode_list
for table_name in stk_rewards_item_576.columns.values.tolist():
    stk_rewards_item_576_addname_dataframe[table_name] = stk_rewards_item_576[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_576_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_577_tscode_list = list() 
stk_rewards_item_577 = pro.stk_rewards(ts_code='603111.SH,603113.SH,603115.SH,603116.SH,603117.SH,603118.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_577 返回数据 row 行数 = "+str(stk_rewards_item_577.shape[0]))
for ts_code_sh in stk_rewards_item_577['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_577_tscode_list.append(stock_name)
stk_rewards_item_577_addname_dataframe=pd.DataFrame()
stk_rewards_item_577_addname_dataframe['cname'] = stk_rewards_item_577_tscode_list
for table_name in stk_rewards_item_577.columns.values.tolist():
    stk_rewards_item_577_addname_dataframe[table_name] = stk_rewards_item_577[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_577_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_578_tscode_list = list() 
stk_rewards_item_578 = pro.stk_rewards(ts_code='603121.SH,603123.SH,603126.SH,603127.SH,603128.SH,603129.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_578 返回数据 row 行数 = "+str(stk_rewards_item_578.shape[0]))
for ts_code_sh in stk_rewards_item_578['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_578_tscode_list.append(stock_name)
stk_rewards_item_578_addname_dataframe=pd.DataFrame()
stk_rewards_item_578_addname_dataframe['cname'] = stk_rewards_item_578_tscode_list
for table_name in stk_rewards_item_578.columns.values.tolist():
    stk_rewards_item_578_addname_dataframe[table_name] = stk_rewards_item_578[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_578_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_579_tscode_list = list() 
stk_rewards_item_579 = pro.stk_rewards(ts_code='603131.SH,603133.SH,603136.SH,603138.SH,603139.SH,603156.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_579 返回数据 row 行数 = "+str(stk_rewards_item_579.shape[0]))
for ts_code_sh in stk_rewards_item_579['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_579_tscode_list.append(stock_name)
stk_rewards_item_579_addname_dataframe=pd.DataFrame()
stk_rewards_item_579_addname_dataframe['cname'] = stk_rewards_item_579_tscode_list
for table_name in stk_rewards_item_579.columns.values.tolist():
    stk_rewards_item_579_addname_dataframe[table_name] = stk_rewards_item_579[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_579_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_580_tscode_list = list() 
stk_rewards_item_580 = pro.stk_rewards(ts_code='603157.SH,603158.SH,603159.SH,603160.SH,603161.SH,603165.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_580 返回数据 row 行数 = "+str(stk_rewards_item_580.shape[0]))
for ts_code_sh in stk_rewards_item_580['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_580_tscode_list.append(stock_name)
stk_rewards_item_580_addname_dataframe=pd.DataFrame()
stk_rewards_item_580_addname_dataframe['cname'] = stk_rewards_item_580_tscode_list
for table_name in stk_rewards_item_580.columns.values.tolist():
    stk_rewards_item_580_addname_dataframe[table_name] = stk_rewards_item_580[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_580_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_581_tscode_list = list() 
stk_rewards_item_581 = pro.stk_rewards(ts_code='603166.SH,603167.SH,603168.SH,603169.SH,603177.SH,603178.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_581 返回数据 row 行数 = "+str(stk_rewards_item_581.shape[0]))
for ts_code_sh in stk_rewards_item_581['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_581_tscode_list.append(stock_name)
stk_rewards_item_581_addname_dataframe=pd.DataFrame()
stk_rewards_item_581_addname_dataframe['cname'] = stk_rewards_item_581_tscode_list
for table_name in stk_rewards_item_581.columns.values.tolist():
    stk_rewards_item_581_addname_dataframe[table_name] = stk_rewards_item_581[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_581_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_582_tscode_list = list() 
stk_rewards_item_582 = pro.stk_rewards(ts_code='603179.SH,603180.SH,603181.SH,603183.SH,603185.SH,603186.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_582 返回数据 row 行数 = "+str(stk_rewards_item_582.shape[0]))
for ts_code_sh in stk_rewards_item_582['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_582_tscode_list.append(stock_name)
stk_rewards_item_582_addname_dataframe=pd.DataFrame()
stk_rewards_item_582_addname_dataframe['cname'] = stk_rewards_item_582_tscode_list
for table_name in stk_rewards_item_582.columns.values.tolist():
    stk_rewards_item_582_addname_dataframe[table_name] = stk_rewards_item_582[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_582_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_583_tscode_list = list() 
stk_rewards_item_583 = pro.stk_rewards(ts_code='603187.SH,603188.SH,603189.SH,603192.SH,603195.SH,603196.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_583 返回数据 row 行数 = "+str(stk_rewards_item_583.shape[0]))
for ts_code_sh in stk_rewards_item_583['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_583_tscode_list.append(stock_name)
stk_rewards_item_583_addname_dataframe=pd.DataFrame()
stk_rewards_item_583_addname_dataframe['cname'] = stk_rewards_item_583_tscode_list
for table_name in stk_rewards_item_583.columns.values.tolist():
    stk_rewards_item_583_addname_dataframe[table_name] = stk_rewards_item_583[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_583_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_584_tscode_list = list() 
stk_rewards_item_584 = pro.stk_rewards(ts_code='603197.SH,603198.SH,603199.SH,603200.SH,603203.SH,603208.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_584 返回数据 row 行数 = "+str(stk_rewards_item_584.shape[0]))
for ts_code_sh in stk_rewards_item_584['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_584_tscode_list.append(stock_name)
stk_rewards_item_584_addname_dataframe=pd.DataFrame()
stk_rewards_item_584_addname_dataframe['cname'] = stk_rewards_item_584_tscode_list
for table_name in stk_rewards_item_584.columns.values.tolist():
    stk_rewards_item_584_addname_dataframe[table_name] = stk_rewards_item_584[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_584_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_585_tscode_list = list() 
stk_rewards_item_585 = pro.stk_rewards(ts_code='603212.SH,603214.SH,603217.SH,603218.SH,603220.SH,603221.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_585 返回数据 row 行数 = "+str(stk_rewards_item_585.shape[0]))
for ts_code_sh in stk_rewards_item_585['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_585_tscode_list.append(stock_name)
stk_rewards_item_585_addname_dataframe=pd.DataFrame()
stk_rewards_item_585_addname_dataframe['cname'] = stk_rewards_item_585_tscode_list
for table_name in stk_rewards_item_585.columns.values.tolist():
    stk_rewards_item_585_addname_dataframe[table_name] = stk_rewards_item_585[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_585_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_586_tscode_list = list() 
stk_rewards_item_586 = pro.stk_rewards(ts_code='603222.SH,603223.SH,603225.SH,603226.SH,603227.SH,603228.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_586 返回数据 row 行数 = "+str(stk_rewards_item_586.shape[0]))
for ts_code_sh in stk_rewards_item_586['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_586_tscode_list.append(stock_name)
stk_rewards_item_586_addname_dataframe=pd.DataFrame()
stk_rewards_item_586_addname_dataframe['cname'] = stk_rewards_item_586_tscode_list
for table_name in stk_rewards_item_586.columns.values.tolist():
    stk_rewards_item_586_addname_dataframe[table_name] = stk_rewards_item_586[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_586_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_587_tscode_list = list() 
stk_rewards_item_587 = pro.stk_rewards(ts_code='603229.SH,603232.SH,603233.SH,603236.SH,603238.SH,603239.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_587 返回数据 row 行数 = "+str(stk_rewards_item_587.shape[0]))
for ts_code_sh in stk_rewards_item_587['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_587_tscode_list.append(stock_name)
stk_rewards_item_587_addname_dataframe=pd.DataFrame()
stk_rewards_item_587_addname_dataframe['cname'] = stk_rewards_item_587_tscode_list
for table_name in stk_rewards_item_587.columns.values.tolist():
    stk_rewards_item_587_addname_dataframe[table_name] = stk_rewards_item_587[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_587_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_588_tscode_list = list() 
stk_rewards_item_588 = pro.stk_rewards(ts_code='603256.SH,603258.SH,603259.SH,603260.SH,603266.SH,603267.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_588 返回数据 row 行数 = "+str(stk_rewards_item_588.shape[0]))
for ts_code_sh in stk_rewards_item_588['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_588_tscode_list.append(stock_name)
stk_rewards_item_588_addname_dataframe=pd.DataFrame()
stk_rewards_item_588_addname_dataframe['cname'] = stk_rewards_item_588_tscode_list
for table_name in stk_rewards_item_588.columns.values.tolist():
    stk_rewards_item_588_addname_dataframe[table_name] = stk_rewards_item_588[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_588_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_589_tscode_list = list() 
stk_rewards_item_589 = pro.stk_rewards(ts_code='603268.SH,603269.SH,603277.SH,603278.SH,603279.SH,603283.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_589 返回数据 row 行数 = "+str(stk_rewards_item_589.shape[0]))
for ts_code_sh in stk_rewards_item_589['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_589_tscode_list.append(stock_name)
stk_rewards_item_589_addname_dataframe=pd.DataFrame()
stk_rewards_item_589_addname_dataframe['cname'] = stk_rewards_item_589_tscode_list
for table_name in stk_rewards_item_589.columns.values.tolist():
    stk_rewards_item_589_addname_dataframe[table_name] = stk_rewards_item_589[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_589_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_590_tscode_list = list() 
stk_rewards_item_590 = pro.stk_rewards(ts_code='603286.SH,603288.SH,603289.SH,603290.SH,603297.SH,603298.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_590 返回数据 row 行数 = "+str(stk_rewards_item_590.shape[0]))
for ts_code_sh in stk_rewards_item_590['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_590_tscode_list.append(stock_name)
stk_rewards_item_590_addname_dataframe=pd.DataFrame()
stk_rewards_item_590_addname_dataframe['cname'] = stk_rewards_item_590_tscode_list
for table_name in stk_rewards_item_590.columns.values.tolist():
    stk_rewards_item_590_addname_dataframe[table_name] = stk_rewards_item_590[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_590_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_591_tscode_list = list() 
stk_rewards_item_591 = pro.stk_rewards(ts_code='603299.SH,603300.SH,603301.SH,603303.SH,603305.SH,603306.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_591 返回数据 row 行数 = "+str(stk_rewards_item_591.shape[0]))
for ts_code_sh in stk_rewards_item_591['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_591_tscode_list.append(stock_name)
stk_rewards_item_591_addname_dataframe=pd.DataFrame()
stk_rewards_item_591_addname_dataframe['cname'] = stk_rewards_item_591_tscode_list
for table_name in stk_rewards_item_591.columns.values.tolist():
    stk_rewards_item_591_addname_dataframe[table_name] = stk_rewards_item_591[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_591_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_592_tscode_list = list() 
stk_rewards_item_592 = pro.stk_rewards(ts_code='603308.SH,603309.SH,603311.SH,603313.SH,603315.SH,603316.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_592 返回数据 row 行数 = "+str(stk_rewards_item_592.shape[0]))
for ts_code_sh in stk_rewards_item_592['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_592_tscode_list.append(stock_name)
stk_rewards_item_592_addname_dataframe=pd.DataFrame()
stk_rewards_item_592_addname_dataframe['cname'] = stk_rewards_item_592_tscode_list
for table_name in stk_rewards_item_592.columns.values.tolist():
    stk_rewards_item_592_addname_dataframe[table_name] = stk_rewards_item_592[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_592_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_593_tscode_list = list() 
stk_rewards_item_593 = pro.stk_rewards(ts_code='603317.SH,603318.SH,603319.SH,603320.SH,603321.SH,603322.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_593 返回数据 row 行数 = "+str(stk_rewards_item_593.shape[0]))
for ts_code_sh in stk_rewards_item_593['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_593_tscode_list.append(stock_name)
stk_rewards_item_593_addname_dataframe=pd.DataFrame()
stk_rewards_item_593_addname_dataframe['cname'] = stk_rewards_item_593_tscode_list
for table_name in stk_rewards_item_593.columns.values.tolist():
    stk_rewards_item_593_addname_dataframe[table_name] = stk_rewards_item_593[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_593_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_594_tscode_list = list() 
stk_rewards_item_594 = pro.stk_rewards(ts_code='603323.SH,603326.SH,603327.SH,603328.SH,603329.SH,603330.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_594 返回数据 row 行数 = "+str(stk_rewards_item_594.shape[0]))
for ts_code_sh in stk_rewards_item_594['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_594_tscode_list.append(stock_name)
stk_rewards_item_594_addname_dataframe=pd.DataFrame()
stk_rewards_item_594_addname_dataframe['cname'] = stk_rewards_item_594_tscode_list
for table_name in stk_rewards_item_594.columns.values.tolist():
    stk_rewards_item_594_addname_dataframe[table_name] = stk_rewards_item_594[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_594_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_595_tscode_list = list() 
stk_rewards_item_595 = pro.stk_rewards(ts_code='603331.SH,603332.SH,603333.SH,603335.SH,603336.SH,603337.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_595 返回数据 row 行数 = "+str(stk_rewards_item_595.shape[0]))
for ts_code_sh in stk_rewards_item_595['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_595_tscode_list.append(stock_name)
stk_rewards_item_595_addname_dataframe=pd.DataFrame()
stk_rewards_item_595_addname_dataframe['cname'] = stk_rewards_item_595_tscode_list
for table_name in stk_rewards_item_595.columns.values.tolist():
    stk_rewards_item_595_addname_dataframe[table_name] = stk_rewards_item_595[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_595_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_596_tscode_list = list() 
stk_rewards_item_596 = pro.stk_rewards(ts_code='603338.SH,603339.SH,603345.SH,603348.SH,603351.SH,603353.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_596 返回数据 row 行数 = "+str(stk_rewards_item_596.shape[0]))
for ts_code_sh in stk_rewards_item_596['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_596_tscode_list.append(stock_name)
stk_rewards_item_596_addname_dataframe=pd.DataFrame()
stk_rewards_item_596_addname_dataframe['cname'] = stk_rewards_item_596_tscode_list
for table_name in stk_rewards_item_596.columns.values.tolist():
    stk_rewards_item_596_addname_dataframe[table_name] = stk_rewards_item_596[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_596_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_597_tscode_list = list() 
stk_rewards_item_597 = pro.stk_rewards(ts_code='603355.SH,603356.SH,603357.SH,603358.SH,603359.SH,603360.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_597 返回数据 row 行数 = "+str(stk_rewards_item_597.shape[0]))
for ts_code_sh in stk_rewards_item_597['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_597_tscode_list.append(stock_name)
stk_rewards_item_597_addname_dataframe=pd.DataFrame()
stk_rewards_item_597_addname_dataframe['cname'] = stk_rewards_item_597_tscode_list
for table_name in stk_rewards_item_597.columns.values.tolist():
    stk_rewards_item_597_addname_dataframe[table_name] = stk_rewards_item_597[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_597_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_598_tscode_list = list() 
stk_rewards_item_598 = pro.stk_rewards(ts_code='603363.SH,603365.SH,603366.SH,603367.SH,603368.SH,603369.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_598 返回数据 row 行数 = "+str(stk_rewards_item_598.shape[0]))
for ts_code_sh in stk_rewards_item_598['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_598_tscode_list.append(stock_name)
stk_rewards_item_598_addname_dataframe=pd.DataFrame()
stk_rewards_item_598_addname_dataframe['cname'] = stk_rewards_item_598_tscode_list
for table_name in stk_rewards_item_598.columns.values.tolist():
    stk_rewards_item_598_addname_dataframe[table_name] = stk_rewards_item_598[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_598_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_599_tscode_list = list() 
stk_rewards_item_599 = pro.stk_rewards(ts_code='603377.SH,603378.SH,603379.SH,603380.SH,603383.SH,603385.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_599 返回数据 row 行数 = "+str(stk_rewards_item_599.shape[0]))
for ts_code_sh in stk_rewards_item_599['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_599_tscode_list.append(stock_name)
stk_rewards_item_599_addname_dataframe=pd.DataFrame()
stk_rewards_item_599_addname_dataframe['cname'] = stk_rewards_item_599_tscode_list
for table_name in stk_rewards_item_599.columns.values.tolist():
    stk_rewards_item_599_addname_dataframe[table_name] = stk_rewards_item_599[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_599_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_600_tscode_list = list() 
stk_rewards_item_600 = pro.stk_rewards(ts_code='603386.SH,603387.SH,603388.SH,603389.SH,603390.SH,603392.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_600 返回数据 row 行数 = "+str(stk_rewards_item_600.shape[0]))
for ts_code_sh in stk_rewards_item_600['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_600_tscode_list.append(stock_name)
stk_rewards_item_600_addname_dataframe=pd.DataFrame()
stk_rewards_item_600_addname_dataframe['cname'] = stk_rewards_item_600_tscode_list
for table_name in stk_rewards_item_600.columns.values.tolist():
    stk_rewards_item_600_addname_dataframe[table_name] = stk_rewards_item_600[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_600_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_601_tscode_list = list() 
stk_rewards_item_601 = pro.stk_rewards(ts_code='603393.SH,603396.SH,603398.SH,603399.SH,603408.SH,603416.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_601 返回数据 row 行数 = "+str(stk_rewards_item_601.shape[0]))
for ts_code_sh in stk_rewards_item_601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_601_tscode_list.append(stock_name)
stk_rewards_item_601_addname_dataframe=pd.DataFrame()
stk_rewards_item_601_addname_dataframe['cname'] = stk_rewards_item_601_tscode_list
for table_name in stk_rewards_item_601.columns.values.tolist():
    stk_rewards_item_601_addname_dataframe[table_name] = stk_rewards_item_601[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_601_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_602_tscode_list = list() 
stk_rewards_item_602 = pro.stk_rewards(ts_code='603421.SH,603429.SH,603439.SH,603444.SH,603456.SH,603458.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_602 返回数据 row 行数 = "+str(stk_rewards_item_602.shape[0]))
for ts_code_sh in stk_rewards_item_602['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_602_tscode_list.append(stock_name)
stk_rewards_item_602_addname_dataframe=pd.DataFrame()
stk_rewards_item_602_addname_dataframe['cname'] = stk_rewards_item_602_tscode_list
for table_name in stk_rewards_item_602.columns.values.tolist():
    stk_rewards_item_602_addname_dataframe[table_name] = stk_rewards_item_602[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_602_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_603_tscode_list = list() 
stk_rewards_item_603 = pro.stk_rewards(ts_code='603466.SH,603477.SH,603486.SH,603488.SH,603489.SH,603496.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_603 返回数据 row 行数 = "+str(stk_rewards_item_603.shape[0]))
for ts_code_sh in stk_rewards_item_603['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_603_tscode_list.append(stock_name)
stk_rewards_item_603_addname_dataframe=pd.DataFrame()
stk_rewards_item_603_addname_dataframe['cname'] = stk_rewards_item_603_tscode_list
for table_name in stk_rewards_item_603.columns.values.tolist():
    stk_rewards_item_603_addname_dataframe[table_name] = stk_rewards_item_603[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_603_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_604_tscode_list = list() 
stk_rewards_item_604 = pro.stk_rewards(ts_code='603499.SH,603500.SH,603501.SH,603505.SH,603506.SH,603507.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_604 返回数据 row 行数 = "+str(stk_rewards_item_604.shape[0]))
for ts_code_sh in stk_rewards_item_604['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_604_tscode_list.append(stock_name)
stk_rewards_item_604_addname_dataframe=pd.DataFrame()
stk_rewards_item_604_addname_dataframe['cname'] = stk_rewards_item_604_tscode_list
for table_name in stk_rewards_item_604.columns.values.tolist():
    stk_rewards_item_604_addname_dataframe[table_name] = stk_rewards_item_604[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_604_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_605_tscode_list = list() 
stk_rewards_item_605 = pro.stk_rewards(ts_code='603508.SH,603515.SH,603516.SH,603517.SH,603518.SH,603519.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_605 返回数据 row 行数 = "+str(stk_rewards_item_605.shape[0]))
for ts_code_sh in stk_rewards_item_605['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_605_tscode_list.append(stock_name)
stk_rewards_item_605_addname_dataframe=pd.DataFrame()
stk_rewards_item_605_addname_dataframe['cname'] = stk_rewards_item_605_tscode_list
for table_name in stk_rewards_item_605.columns.values.tolist():
    stk_rewards_item_605_addname_dataframe[table_name] = stk_rewards_item_605[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_605_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_606_tscode_list = list() 
stk_rewards_item_606 = pro.stk_rewards(ts_code='603520.SH,603527.SH,603528.SH,603530.SH,603533.SH,603535.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_606 返回数据 row 行数 = "+str(stk_rewards_item_606.shape[0]))
for ts_code_sh in stk_rewards_item_606['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_606_tscode_list.append(stock_name)
stk_rewards_item_606_addname_dataframe=pd.DataFrame()
stk_rewards_item_606_addname_dataframe['cname'] = stk_rewards_item_606_tscode_list
for table_name in stk_rewards_item_606.columns.values.tolist():
    stk_rewards_item_606_addname_dataframe[table_name] = stk_rewards_item_606[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_606_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_607_tscode_list = list() 
stk_rewards_item_607 = pro.stk_rewards(ts_code='603536.SH,603538.SH,603551.SH,603555.SH,603556.SH,603557.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_607 返回数据 row 行数 = "+str(stk_rewards_item_607.shape[0]))
for ts_code_sh in stk_rewards_item_607['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_607_tscode_list.append(stock_name)
stk_rewards_item_607_addname_dataframe=pd.DataFrame()
stk_rewards_item_607_addname_dataframe['cname'] = stk_rewards_item_607_tscode_list
for table_name in stk_rewards_item_607.columns.values.tolist():
    stk_rewards_item_607_addname_dataframe[table_name] = stk_rewards_item_607[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_607_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_608_tscode_list = list() 
stk_rewards_item_608 = pro.stk_rewards(ts_code='603558.SH,603559.SH,603566.SH,603567.SH,603568.SH,603569.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_608 返回数据 row 行数 = "+str(stk_rewards_item_608.shape[0]))
for ts_code_sh in stk_rewards_item_608['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_608_tscode_list.append(stock_name)
stk_rewards_item_608_addname_dataframe=pd.DataFrame()
stk_rewards_item_608_addname_dataframe['cname'] = stk_rewards_item_608_tscode_list
for table_name in stk_rewards_item_608.columns.values.tolist():
    stk_rewards_item_608_addname_dataframe[table_name] = stk_rewards_item_608[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_608_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_609_tscode_list = list() 
stk_rewards_item_609 = pro.stk_rewards(ts_code='603577.SH,603578.SH,603579.SH,603580.SH,603583.SH,603585.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_609 返回数据 row 行数 = "+str(stk_rewards_item_609.shape[0]))
for ts_code_sh in stk_rewards_item_609['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_609_tscode_list.append(stock_name)
stk_rewards_item_609_addname_dataframe=pd.DataFrame()
stk_rewards_item_609_addname_dataframe['cname'] = stk_rewards_item_609_tscode_list
for table_name in stk_rewards_item_609.columns.values.tolist():
    stk_rewards_item_609_addname_dataframe[table_name] = stk_rewards_item_609[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_609_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_610_tscode_list = list() 
stk_rewards_item_610 = pro.stk_rewards(ts_code='603586.SH,603587.SH,603588.SH,603589.SH,603590.SH,603595.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_610 返回数据 row 行数 = "+str(stk_rewards_item_610.shape[0]))
for ts_code_sh in stk_rewards_item_610['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_610_tscode_list.append(stock_name)
stk_rewards_item_610_addname_dataframe=pd.DataFrame()
stk_rewards_item_610_addname_dataframe['cname'] = stk_rewards_item_610_tscode_list
for table_name in stk_rewards_item_610.columns.values.tolist():
    stk_rewards_item_610_addname_dataframe[table_name] = stk_rewards_item_610[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_610_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_611_tscode_list = list() 
stk_rewards_item_611 = pro.stk_rewards(ts_code='603596.SH,603598.SH,603599.SH,603600.SH,603601.SH,603602.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_611 返回数据 row 行数 = "+str(stk_rewards_item_611.shape[0]))
for ts_code_sh in stk_rewards_item_611['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_611_tscode_list.append(stock_name)
stk_rewards_item_611_addname_dataframe=pd.DataFrame()
stk_rewards_item_611_addname_dataframe['cname'] = stk_rewards_item_611_tscode_list
for table_name in stk_rewards_item_611.columns.values.tolist():
    stk_rewards_item_611_addname_dataframe[table_name] = stk_rewards_item_611[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_611_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_612_tscode_list = list() 
stk_rewards_item_612 = pro.stk_rewards(ts_code='603603.SH,603605.SH,603606.SH,603607.SH,603608.SH,603609.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_612 返回数据 row 行数 = "+str(stk_rewards_item_612.shape[0]))
for ts_code_sh in stk_rewards_item_612['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_612_tscode_list.append(stock_name)
stk_rewards_item_612_addname_dataframe=pd.DataFrame()
stk_rewards_item_612_addname_dataframe['cname'] = stk_rewards_item_612_tscode_list
for table_name in stk_rewards_item_612.columns.values.tolist():
    stk_rewards_item_612_addname_dataframe[table_name] = stk_rewards_item_612[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_612_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_613_tscode_list = list() 
stk_rewards_item_613 = pro.stk_rewards(ts_code='603610.SH,603611.SH,603612.SH,603613.SH,603615.SH,603616.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_613 返回数据 row 行数 = "+str(stk_rewards_item_613.shape[0]))
for ts_code_sh in stk_rewards_item_613['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_613_tscode_list.append(stock_name)
stk_rewards_item_613_addname_dataframe=pd.DataFrame()
stk_rewards_item_613_addname_dataframe['cname'] = stk_rewards_item_613_tscode_list
for table_name in stk_rewards_item_613.columns.values.tolist():
    stk_rewards_item_613_addname_dataframe[table_name] = stk_rewards_item_613[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_613_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_614_tscode_list = list() 
stk_rewards_item_614 = pro.stk_rewards(ts_code='603617.SH,603618.SH,603619.SH,603626.SH,603628.SH,603629.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_614 返回数据 row 行数 = "+str(stk_rewards_item_614.shape[0]))
for ts_code_sh in stk_rewards_item_614['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_614_tscode_list.append(stock_name)
stk_rewards_item_614_addname_dataframe=pd.DataFrame()
stk_rewards_item_614_addname_dataframe['cname'] = stk_rewards_item_614_tscode_list
for table_name in stk_rewards_item_614.columns.values.tolist():
    stk_rewards_item_614_addname_dataframe[table_name] = stk_rewards_item_614[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_614_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_615_tscode_list = list() 
stk_rewards_item_615 = pro.stk_rewards(ts_code='603630.SH,603633.SH,603636.SH,603637.SH,603638.SH,603639.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_615 返回数据 row 行数 = "+str(stk_rewards_item_615.shape[0]))
for ts_code_sh in stk_rewards_item_615['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_615_tscode_list.append(stock_name)
stk_rewards_item_615_addname_dataframe=pd.DataFrame()
stk_rewards_item_615_addname_dataframe['cname'] = stk_rewards_item_615_tscode_list
for table_name in stk_rewards_item_615.columns.values.tolist():
    stk_rewards_item_615_addname_dataframe[table_name] = stk_rewards_item_615[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_615_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_616_tscode_list = list() 
stk_rewards_item_616 = pro.stk_rewards(ts_code='603648.SH,603650.SH,603655.SH,603656.SH,603657.SH,603658.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_616 返回数据 row 行数 = "+str(stk_rewards_item_616.shape[0]))
for ts_code_sh in stk_rewards_item_616['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_616_tscode_list.append(stock_name)
stk_rewards_item_616_addname_dataframe=pd.DataFrame()
stk_rewards_item_616_addname_dataframe['cname'] = stk_rewards_item_616_tscode_list
for table_name in stk_rewards_item_616.columns.values.tolist():
    stk_rewards_item_616_addname_dataframe[table_name] = stk_rewards_item_616[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_616_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_617_tscode_list = list() 
stk_rewards_item_617 = pro.stk_rewards(ts_code='603659.SH,603660.SH,603661.SH,603662.SH,603663.SH,603665.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_617 返回数据 row 行数 = "+str(stk_rewards_item_617.shape[0]))
for ts_code_sh in stk_rewards_item_617['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_617_tscode_list.append(stock_name)
stk_rewards_item_617_addname_dataframe=pd.DataFrame()
stk_rewards_item_617_addname_dataframe['cname'] = stk_rewards_item_617_tscode_list
for table_name in stk_rewards_item_617.columns.values.tolist():
    stk_rewards_item_617_addname_dataframe[table_name] = stk_rewards_item_617[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_617_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_618_tscode_list = list() 
stk_rewards_item_618 = pro.stk_rewards(ts_code='603666.SH,603667.SH,603668.SH,603669.SH,603676.SH,603677.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_618 返回数据 row 行数 = "+str(stk_rewards_item_618.shape[0]))
for ts_code_sh in stk_rewards_item_618['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_618_tscode_list.append(stock_name)
stk_rewards_item_618_addname_dataframe=pd.DataFrame()
stk_rewards_item_618_addname_dataframe['cname'] = stk_rewards_item_618_tscode_list
for table_name in stk_rewards_item_618.columns.values.tolist():
    stk_rewards_item_618_addname_dataframe[table_name] = stk_rewards_item_618[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_618_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_619_tscode_list = list() 
stk_rewards_item_619 = pro.stk_rewards(ts_code='603678.SH,603679.SH,603680.SH,603681.SH,603682.SH,603683.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_619 返回数据 row 行数 = "+str(stk_rewards_item_619.shape[0]))
for ts_code_sh in stk_rewards_item_619['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_619_tscode_list.append(stock_name)
stk_rewards_item_619_addname_dataframe=pd.DataFrame()
stk_rewards_item_619_addname_dataframe['cname'] = stk_rewards_item_619_tscode_list
for table_name in stk_rewards_item_619.columns.values.tolist():
    stk_rewards_item_619_addname_dataframe[table_name] = stk_rewards_item_619[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_619_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_620_tscode_list = list() 
stk_rewards_item_620 = pro.stk_rewards(ts_code='603685.SH,603686.SH,603687.SH,603688.SH,603689.SH,603690.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_620 返回数据 row 行数 = "+str(stk_rewards_item_620.shape[0]))
for ts_code_sh in stk_rewards_item_620['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_620_tscode_list.append(stock_name)
stk_rewards_item_620_addname_dataframe=pd.DataFrame()
stk_rewards_item_620_addname_dataframe['cname'] = stk_rewards_item_620_tscode_list
for table_name in stk_rewards_item_620.columns.values.tolist():
    stk_rewards_item_620_addname_dataframe[table_name] = stk_rewards_item_620[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_620_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_621_tscode_list = list() 
stk_rewards_item_621 = pro.stk_rewards(ts_code='603693.SH,603696.SH,603697.SH,603698.SH,603699.SH,603700.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_621 返回数据 row 行数 = "+str(stk_rewards_item_621.shape[0]))
for ts_code_sh in stk_rewards_item_621['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_621_tscode_list.append(stock_name)
stk_rewards_item_621_addname_dataframe=pd.DataFrame()
stk_rewards_item_621_addname_dataframe['cname'] = stk_rewards_item_621_tscode_list
for table_name in stk_rewards_item_621.columns.values.tolist():
    stk_rewards_item_621_addname_dataframe[table_name] = stk_rewards_item_621[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_621_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_622_tscode_list = list() 
stk_rewards_item_622 = pro.stk_rewards(ts_code='603701.SH,603703.SH,603706.SH,603707.SH,603708.SH,603709.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_622 返回数据 row 行数 = "+str(stk_rewards_item_622.shape[0]))
for ts_code_sh in stk_rewards_item_622['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_622_tscode_list.append(stock_name)
stk_rewards_item_622_addname_dataframe=pd.DataFrame()
stk_rewards_item_622_addname_dataframe['cname'] = stk_rewards_item_622_tscode_list
for table_name in stk_rewards_item_622.columns.values.tolist():
    stk_rewards_item_622_addname_dataframe[table_name] = stk_rewards_item_622[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_622_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_623_tscode_list = list() 
stk_rewards_item_623 = pro.stk_rewards(ts_code='603711.SH,603712.SH,603713.SH,603716.SH,603717.SH,603718.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_623 返回数据 row 行数 = "+str(stk_rewards_item_623.shape[0]))
for ts_code_sh in stk_rewards_item_623['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_623_tscode_list.append(stock_name)
stk_rewards_item_623_addname_dataframe=pd.DataFrame()
stk_rewards_item_623_addname_dataframe['cname'] = stk_rewards_item_623_tscode_list
for table_name in stk_rewards_item_623.columns.values.tolist():
    stk_rewards_item_623_addname_dataframe[table_name] = stk_rewards_item_623[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_623_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_624_tscode_list = list() 
stk_rewards_item_624 = pro.stk_rewards(ts_code='603719.SH,603721.SH,603722.SH,603725.SH,603726.SH,603727.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_624 返回数据 row 行数 = "+str(stk_rewards_item_624.shape[0]))
for ts_code_sh in stk_rewards_item_624['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_624_tscode_list.append(stock_name)
stk_rewards_item_624_addname_dataframe=pd.DataFrame()
stk_rewards_item_624_addname_dataframe['cname'] = stk_rewards_item_624_tscode_list
for table_name in stk_rewards_item_624.columns.values.tolist():
    stk_rewards_item_624_addname_dataframe[table_name] = stk_rewards_item_624[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_624_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_625_tscode_list = list() 
stk_rewards_item_625 = pro.stk_rewards(ts_code='603728.SH,603729.SH,603730.SH,603733.SH,603737.SH,603738.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_625 返回数据 row 行数 = "+str(stk_rewards_item_625.shape[0]))
for ts_code_sh in stk_rewards_item_625['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_625_tscode_list.append(stock_name)
stk_rewards_item_625_addname_dataframe=pd.DataFrame()
stk_rewards_item_625_addname_dataframe['cname'] = stk_rewards_item_625_tscode_list
for table_name in stk_rewards_item_625.columns.values.tolist():
    stk_rewards_item_625_addname_dataframe[table_name] = stk_rewards_item_625[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_625_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_626_tscode_list = list() 
stk_rewards_item_626 = pro.stk_rewards(ts_code='603739.SH,603755.SH,603757.SH,603758.SH,603766.SH,603767.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_626 返回数据 row 行数 = "+str(stk_rewards_item_626.shape[0]))
for ts_code_sh in stk_rewards_item_626['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_626_tscode_list.append(stock_name)
stk_rewards_item_626_addname_dataframe=pd.DataFrame()
stk_rewards_item_626_addname_dataframe['cname'] = stk_rewards_item_626_tscode_list
for table_name in stk_rewards_item_626.columns.values.tolist():
    stk_rewards_item_626_addname_dataframe[table_name] = stk_rewards_item_626[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_626_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_627_tscode_list = list() 
stk_rewards_item_627 = pro.stk_rewards(ts_code='603768.SH,603773.SH,603776.SH,603777.SH,603778.SH,603779.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_627 返回数据 row 行数 = "+str(stk_rewards_item_627.shape[0]))
for ts_code_sh in stk_rewards_item_627['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_627_tscode_list.append(stock_name)
stk_rewards_item_627_addname_dataframe=pd.DataFrame()
stk_rewards_item_627_addname_dataframe['cname'] = stk_rewards_item_627_tscode_list
for table_name in stk_rewards_item_627.columns.values.tolist():
    stk_rewards_item_627_addname_dataframe[table_name] = stk_rewards_item_627[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_627_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_628_tscode_list = list() 
stk_rewards_item_628 = pro.stk_rewards(ts_code='603786.SH,603787.SH,603788.SH,603789.SH,603790.SH,603797.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_628 返回数据 row 行数 = "+str(stk_rewards_item_628.shape[0]))
for ts_code_sh in stk_rewards_item_628['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_628_tscode_list.append(stock_name)
stk_rewards_item_628_addname_dataframe=pd.DataFrame()
stk_rewards_item_628_addname_dataframe['cname'] = stk_rewards_item_628_tscode_list
for table_name in stk_rewards_item_628.columns.values.tolist():
    stk_rewards_item_628_addname_dataframe[table_name] = stk_rewards_item_628[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_628_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_629_tscode_list = list() 
stk_rewards_item_629 = pro.stk_rewards(ts_code='603798.SH,603799.SH,603800.SH,603801.SH,603803.SH,603806.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_629 返回数据 row 行数 = "+str(stk_rewards_item_629.shape[0]))
for ts_code_sh in stk_rewards_item_629['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_629_tscode_list.append(stock_name)
stk_rewards_item_629_addname_dataframe=pd.DataFrame()
stk_rewards_item_629_addname_dataframe['cname'] = stk_rewards_item_629_tscode_list
for table_name in stk_rewards_item_629.columns.values.tolist():
    stk_rewards_item_629_addname_dataframe[table_name] = stk_rewards_item_629[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_629_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_630_tscode_list = list() 
stk_rewards_item_630 = pro.stk_rewards(ts_code='603808.SH,603809.SH,603810.SH,603811.SH,603813.SH,603815.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_630 返回数据 row 行数 = "+str(stk_rewards_item_630.shape[0]))
for ts_code_sh in stk_rewards_item_630['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_630_tscode_list.append(stock_name)
stk_rewards_item_630_addname_dataframe=pd.DataFrame()
stk_rewards_item_630_addname_dataframe['cname'] = stk_rewards_item_630_tscode_list
for table_name in stk_rewards_item_630.columns.values.tolist():
    stk_rewards_item_630_addname_dataframe[table_name] = stk_rewards_item_630[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_630_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_631_tscode_list = list() 
stk_rewards_item_631 = pro.stk_rewards(ts_code='603816.SH,603817.SH,603818.SH,603819.SH,603822.SH,603823.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_631 返回数据 row 行数 = "+str(stk_rewards_item_631.shape[0]))
for ts_code_sh in stk_rewards_item_631['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_631_tscode_list.append(stock_name)
stk_rewards_item_631_addname_dataframe=pd.DataFrame()
stk_rewards_item_631_addname_dataframe['cname'] = stk_rewards_item_631_tscode_list
for table_name in stk_rewards_item_631.columns.values.tolist():
    stk_rewards_item_631_addname_dataframe[table_name] = stk_rewards_item_631[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_631_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_632_tscode_list = list() 
stk_rewards_item_632 = pro.stk_rewards(ts_code='603825.SH,603826.SH,603828.SH,603829.SH,603833.SH,603838.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_632 返回数据 row 行数 = "+str(stk_rewards_item_632.shape[0]))
for ts_code_sh in stk_rewards_item_632['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_632_tscode_list.append(stock_name)
stk_rewards_item_632_addname_dataframe=pd.DataFrame()
stk_rewards_item_632_addname_dataframe['cname'] = stk_rewards_item_632_tscode_list
for table_name in stk_rewards_item_632.columns.values.tolist():
    stk_rewards_item_632_addname_dataframe[table_name] = stk_rewards_item_632[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_632_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_633_tscode_list = list() 
stk_rewards_item_633 = pro.stk_rewards(ts_code='603839.SH,603843.SH,603848.SH,603855.SH,603856.SH,603858.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_633 返回数据 row 行数 = "+str(stk_rewards_item_633.shape[0]))
for ts_code_sh in stk_rewards_item_633['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_633_tscode_list.append(stock_name)
stk_rewards_item_633_addname_dataframe=pd.DataFrame()
stk_rewards_item_633_addname_dataframe['cname'] = stk_rewards_item_633_tscode_list
for table_name in stk_rewards_item_633.columns.values.tolist():
    stk_rewards_item_633_addname_dataframe[table_name] = stk_rewards_item_633[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_633_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_634_tscode_list = list() 
stk_rewards_item_634 = pro.stk_rewards(ts_code='603859.SH,603860.SH,603861.SH,603863.SH,603866.SH,603867.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_634 返回数据 row 行数 = "+str(stk_rewards_item_634.shape[0]))
for ts_code_sh in stk_rewards_item_634['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_634_tscode_list.append(stock_name)
stk_rewards_item_634_addname_dataframe=pd.DataFrame()
stk_rewards_item_634_addname_dataframe['cname'] = stk_rewards_item_634_tscode_list
for table_name in stk_rewards_item_634.columns.values.tolist():
    stk_rewards_item_634_addname_dataframe[table_name] = stk_rewards_item_634[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_634_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_635_tscode_list = list() 
stk_rewards_item_635 = pro.stk_rewards(ts_code='603868.SH,603869.SH,603871.SH,603876.SH,603877.SH,603878.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_635 返回数据 row 行数 = "+str(stk_rewards_item_635.shape[0]))
for ts_code_sh in stk_rewards_item_635['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_635_tscode_list.append(stock_name)
stk_rewards_item_635_addname_dataframe=pd.DataFrame()
stk_rewards_item_635_addname_dataframe['cname'] = stk_rewards_item_635_tscode_list
for table_name in stk_rewards_item_635.columns.values.tolist():
    stk_rewards_item_635_addname_dataframe[table_name] = stk_rewards_item_635[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_635_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_636_tscode_list = list() 
stk_rewards_item_636 = pro.stk_rewards(ts_code='603879.SH,603880.SH,603881.SH,603882.SH,603883.SH,603885.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_636 返回数据 row 行数 = "+str(stk_rewards_item_636.shape[0]))
for ts_code_sh in stk_rewards_item_636['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_636_tscode_list.append(stock_name)
stk_rewards_item_636_addname_dataframe=pd.DataFrame()
stk_rewards_item_636_addname_dataframe['cname'] = stk_rewards_item_636_tscode_list
for table_name in stk_rewards_item_636.columns.values.tolist():
    stk_rewards_item_636_addname_dataframe[table_name] = stk_rewards_item_636[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_636_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_637_tscode_list = list() 
stk_rewards_item_637 = pro.stk_rewards(ts_code='603886.SH,603887.SH,603888.SH,603889.SH,603890.SH,603893.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_637 返回数据 row 行数 = "+str(stk_rewards_item_637.shape[0]))
for ts_code_sh in stk_rewards_item_637['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_637_tscode_list.append(stock_name)
stk_rewards_item_637_addname_dataframe=pd.DataFrame()
stk_rewards_item_637_addname_dataframe['cname'] = stk_rewards_item_637_tscode_list
for table_name in stk_rewards_item_637.columns.values.tolist():
    stk_rewards_item_637_addname_dataframe[table_name] = stk_rewards_item_637[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_637_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_638_tscode_list = list() 
stk_rewards_item_638 = pro.stk_rewards(ts_code='603895.SH,603896.SH,603897.SH,603898.SH,603899.SH,603900.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_638 返回数据 row 行数 = "+str(stk_rewards_item_638.shape[0]))
for ts_code_sh in stk_rewards_item_638['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_638_tscode_list.append(stock_name)
stk_rewards_item_638_addname_dataframe=pd.DataFrame()
stk_rewards_item_638_addname_dataframe['cname'] = stk_rewards_item_638_tscode_list
for table_name in stk_rewards_item_638.columns.values.tolist():
    stk_rewards_item_638_addname_dataframe[table_name] = stk_rewards_item_638[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_638_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_639_tscode_list = list() 
stk_rewards_item_639 = pro.stk_rewards(ts_code='603901.SH,603903.SH,603906.SH,603908.SH,603909.SH,603912.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_639 返回数据 row 行数 = "+str(stk_rewards_item_639.shape[0]))
for ts_code_sh in stk_rewards_item_639['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_639_tscode_list.append(stock_name)
stk_rewards_item_639_addname_dataframe=pd.DataFrame()
stk_rewards_item_639_addname_dataframe['cname'] = stk_rewards_item_639_tscode_list
for table_name in stk_rewards_item_639.columns.values.tolist():
    stk_rewards_item_639_addname_dataframe[table_name] = stk_rewards_item_639[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_639_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_640_tscode_list = list() 
stk_rewards_item_640 = pro.stk_rewards(ts_code='603915.SH,603916.SH,603917.SH,603918.SH,603919.SH,603920.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_640 返回数据 row 行数 = "+str(stk_rewards_item_640.shape[0]))
for ts_code_sh in stk_rewards_item_640['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_640_tscode_list.append(stock_name)
stk_rewards_item_640_addname_dataframe=pd.DataFrame()
stk_rewards_item_640_addname_dataframe['cname'] = stk_rewards_item_640_tscode_list
for table_name in stk_rewards_item_640.columns.values.tolist():
    stk_rewards_item_640_addname_dataframe[table_name] = stk_rewards_item_640[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_640_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_641_tscode_list = list() 
stk_rewards_item_641 = pro.stk_rewards(ts_code='603922.SH,603926.SH,603927.SH,603928.SH,603929.SH,603933.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_641 返回数据 row 行数 = "+str(stk_rewards_item_641.shape[0]))
for ts_code_sh in stk_rewards_item_641['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_641_tscode_list.append(stock_name)
stk_rewards_item_641_addname_dataframe=pd.DataFrame()
stk_rewards_item_641_addname_dataframe['cname'] = stk_rewards_item_641_tscode_list
for table_name in stk_rewards_item_641.columns.values.tolist():
    stk_rewards_item_641_addname_dataframe[table_name] = stk_rewards_item_641[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_641_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_642_tscode_list = list() 
stk_rewards_item_642 = pro.stk_rewards(ts_code='603936.SH,603937.SH,603938.SH,603939.SH,603948.SH,603949.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_642 返回数据 row 行数 = "+str(stk_rewards_item_642.shape[0]))
for ts_code_sh in stk_rewards_item_642['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_642_tscode_list.append(stock_name)
stk_rewards_item_642_addname_dataframe=pd.DataFrame()
stk_rewards_item_642_addname_dataframe['cname'] = stk_rewards_item_642_tscode_list
for table_name in stk_rewards_item_642.columns.values.tolist():
    stk_rewards_item_642_addname_dataframe[table_name] = stk_rewards_item_642[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_642_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_643_tscode_list = list() 
stk_rewards_item_643 = pro.stk_rewards(ts_code='603950.SH,603955.SH,603956.SH,603958.SH,603959.SH,603960.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_643 返回数据 row 行数 = "+str(stk_rewards_item_643.shape[0]))
for ts_code_sh in stk_rewards_item_643['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_643_tscode_list.append(stock_name)
stk_rewards_item_643_addname_dataframe=pd.DataFrame()
stk_rewards_item_643_addname_dataframe['cname'] = stk_rewards_item_643_tscode_list
for table_name in stk_rewards_item_643.columns.values.tolist():
    stk_rewards_item_643_addname_dataframe[table_name] = stk_rewards_item_643[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_643_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_644_tscode_list = list() 
stk_rewards_item_644 = pro.stk_rewards(ts_code='603963.SH,603966.SH,603967.SH,603968.SH,603969.SH,603970.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_644 返回数据 row 行数 = "+str(stk_rewards_item_644.shape[0]))
for ts_code_sh in stk_rewards_item_644['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_644_tscode_list.append(stock_name)
stk_rewards_item_644_addname_dataframe=pd.DataFrame()
stk_rewards_item_644_addname_dataframe['cname'] = stk_rewards_item_644_tscode_list
for table_name in stk_rewards_item_644.columns.values.tolist():
    stk_rewards_item_644_addname_dataframe[table_name] = stk_rewards_item_644[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_644_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_645_tscode_list = list() 
stk_rewards_item_645 = pro.stk_rewards(ts_code='603976.SH,603977.SH,603978.SH,603979.SH,603980.SH,603982.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_645 返回数据 row 行数 = "+str(stk_rewards_item_645.shape[0]))
for ts_code_sh in stk_rewards_item_645['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_645_tscode_list.append(stock_name)
stk_rewards_item_645_addname_dataframe=pd.DataFrame()
stk_rewards_item_645_addname_dataframe['cname'] = stk_rewards_item_645_tscode_list
for table_name in stk_rewards_item_645.columns.values.tolist():
    stk_rewards_item_645_addname_dataframe[table_name] = stk_rewards_item_645[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_645_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_646_tscode_list = list() 
stk_rewards_item_646 = pro.stk_rewards(ts_code='603983.SH,603985.SH,603986.SH,603987.SH,603988.SH,603989.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_646 返回数据 row 行数 = "+str(stk_rewards_item_646.shape[0]))
for ts_code_sh in stk_rewards_item_646['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_646_tscode_list.append(stock_name)
stk_rewards_item_646_addname_dataframe=pd.DataFrame()
stk_rewards_item_646_addname_dataframe['cname'] = stk_rewards_item_646_tscode_list
for table_name in stk_rewards_item_646.columns.values.tolist():
    stk_rewards_item_646_addname_dataframe[table_name] = stk_rewards_item_646[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_646_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_647_tscode_list = list() 
stk_rewards_item_647 = pro.stk_rewards(ts_code='603990.SH,603991.SH,603992.SH,603993.SH,603995.SH,603996.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_647 返回数据 row 行数 = "+str(stk_rewards_item_647.shape[0]))
for ts_code_sh in stk_rewards_item_647['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_647_tscode_list.append(stock_name)
stk_rewards_item_647_addname_dataframe=pd.DataFrame()
stk_rewards_item_647_addname_dataframe['cname'] = stk_rewards_item_647_tscode_list
for table_name in stk_rewards_item_647.columns.values.tolist():
    stk_rewards_item_647_addname_dataframe[table_name] = stk_rewards_item_647[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_647_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_648_tscode_list = list() 
stk_rewards_item_648 = pro.stk_rewards(ts_code='603997.SH,603998.SH,603999.SH,605001.SH,605066.SH,605088.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_648 返回数据 row 行数 = "+str(stk_rewards_item_648.shape[0]))
for ts_code_sh in stk_rewards_item_648['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_648_tscode_list.append(stock_name)
stk_rewards_item_648_addname_dataframe=pd.DataFrame()
stk_rewards_item_648_addname_dataframe['cname'] = stk_rewards_item_648_tscode_list
for table_name in stk_rewards_item_648.columns.values.tolist():
    stk_rewards_item_648_addname_dataframe[table_name] = stk_rewards_item_648[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_648_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_649_tscode_list = list() 
stk_rewards_item_649 = pro.stk_rewards(ts_code='605100.SH,605108.SH,605118.SH,605158.SH,605166.SH,605168.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_649 返回数据 row 行数 = "+str(stk_rewards_item_649.shape[0]))
for ts_code_sh in stk_rewards_item_649['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_649_tscode_list.append(stock_name)
stk_rewards_item_649_addname_dataframe=pd.DataFrame()
stk_rewards_item_649_addname_dataframe['cname'] = stk_rewards_item_649_tscode_list
for table_name in stk_rewards_item_649.columns.values.tolist():
    stk_rewards_item_649_addname_dataframe[table_name] = stk_rewards_item_649[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_649_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_650_tscode_list = list() 
stk_rewards_item_650 = pro.stk_rewards(ts_code='605188.SH,605199.SH,605222.SH,605288.SH,605318.SH,605366.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_650 返回数据 row 行数 = "+str(stk_rewards_item_650.shape[0]))
for ts_code_sh in stk_rewards_item_650['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_650_tscode_list.append(stock_name)
stk_rewards_item_650_addname_dataframe=pd.DataFrame()
stk_rewards_item_650_addname_dataframe['cname'] = stk_rewards_item_650_tscode_list
for table_name in stk_rewards_item_650.columns.values.tolist():
    stk_rewards_item_650_addname_dataframe[table_name] = stk_rewards_item_650[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_650_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_651_tscode_list = list() 
stk_rewards_item_651 = pro.stk_rewards(ts_code='605399.SH,688001.SH,688002.SH,688003.SH,688004.SH,688005.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_651 返回数据 row 行数 = "+str(stk_rewards_item_651.shape[0]))
for ts_code_sh in stk_rewards_item_651['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_651_tscode_list.append(stock_name)
stk_rewards_item_651_addname_dataframe=pd.DataFrame()
stk_rewards_item_651_addname_dataframe['cname'] = stk_rewards_item_651_tscode_list
for table_name in stk_rewards_item_651.columns.values.tolist():
    stk_rewards_item_651_addname_dataframe[table_name] = stk_rewards_item_651[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_651_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_652_tscode_list = list() 
stk_rewards_item_652 = pro.stk_rewards(ts_code='688006.SH,688007.SH,688008.SH,688009.SH,688010.SH,688011.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_652 返回数据 row 行数 = "+str(stk_rewards_item_652.shape[0]))
for ts_code_sh in stk_rewards_item_652['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_652_tscode_list.append(stock_name)
stk_rewards_item_652_addname_dataframe=pd.DataFrame()
stk_rewards_item_652_addname_dataframe['cname'] = stk_rewards_item_652_tscode_list
for table_name in stk_rewards_item_652.columns.values.tolist():
    stk_rewards_item_652_addname_dataframe[table_name] = stk_rewards_item_652[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_652_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_653_tscode_list = list() 
stk_rewards_item_653 = pro.stk_rewards(ts_code='688012.SH,688015.SH,688016.SH,688018.SH,688019.SH,688020.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_653 返回数据 row 行数 = "+str(stk_rewards_item_653.shape[0]))
for ts_code_sh in stk_rewards_item_653['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_653_tscode_list.append(stock_name)
stk_rewards_item_653_addname_dataframe=pd.DataFrame()
stk_rewards_item_653_addname_dataframe['cname'] = stk_rewards_item_653_tscode_list
for table_name in stk_rewards_item_653.columns.values.tolist():
    stk_rewards_item_653_addname_dataframe[table_name] = stk_rewards_item_653[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_653_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_654_tscode_list = list() 
stk_rewards_item_654 = pro.stk_rewards(ts_code='688021.SH,688022.SH,688023.SH,688025.SH,688026.SH,688027.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_654 返回数据 row 行数 = "+str(stk_rewards_item_654.shape[0]))
for ts_code_sh in stk_rewards_item_654['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_654_tscode_list.append(stock_name)
stk_rewards_item_654_addname_dataframe=pd.DataFrame()
stk_rewards_item_654_addname_dataframe['cname'] = stk_rewards_item_654_tscode_list
for table_name in stk_rewards_item_654.columns.values.tolist():
    stk_rewards_item_654_addname_dataframe[table_name] = stk_rewards_item_654[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_654_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_655_tscode_list = list() 
stk_rewards_item_655 = pro.stk_rewards(ts_code='688028.SH,688029.SH,688030.SH,688033.SH,688036.SH,688037.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_655 返回数据 row 行数 = "+str(stk_rewards_item_655.shape[0]))
for ts_code_sh in stk_rewards_item_655['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_655_tscode_list.append(stock_name)
stk_rewards_item_655_addname_dataframe=pd.DataFrame()
stk_rewards_item_655_addname_dataframe['cname'] = stk_rewards_item_655_tscode_list
for table_name in stk_rewards_item_655.columns.values.tolist():
    stk_rewards_item_655_addname_dataframe[table_name] = stk_rewards_item_655[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_655_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_656_tscode_list = list() 
stk_rewards_item_656 = pro.stk_rewards(ts_code='688039.SH,688050.SH,688051.SH,688055.SH,688058.SH,688060.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_656 返回数据 row 行数 = "+str(stk_rewards_item_656.shape[0]))
for ts_code_sh in stk_rewards_item_656['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_656_tscode_list.append(stock_name)
stk_rewards_item_656_addname_dataframe=pd.DataFrame()
stk_rewards_item_656_addname_dataframe['cname'] = stk_rewards_item_656_tscode_list
for table_name in stk_rewards_item_656.columns.values.tolist():
    stk_rewards_item_656_addname_dataframe[table_name] = stk_rewards_item_656[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_656_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_657_tscode_list = list() 
stk_rewards_item_657 = pro.stk_rewards(ts_code='688065.SH,688066.SH,688068.SH,688069.SH,688077.SH,688078.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_657 返回数据 row 行数 = "+str(stk_rewards_item_657.shape[0]))
for ts_code_sh in stk_rewards_item_657['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_657_tscode_list.append(stock_name)
stk_rewards_item_657_addname_dataframe=pd.DataFrame()
stk_rewards_item_657_addname_dataframe['cname'] = stk_rewards_item_657_tscode_list
for table_name in stk_rewards_item_657.columns.values.tolist():
    stk_rewards_item_657_addname_dataframe[table_name] = stk_rewards_item_657[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_657_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_658_tscode_list = list() 
stk_rewards_item_658 = pro.stk_rewards(ts_code='688080.SH,688081.SH,688085.SH,688086.SH,688088.SH,688089.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_658 返回数据 row 行数 = "+str(stk_rewards_item_658.shape[0]))
for ts_code_sh in stk_rewards_item_658['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_658_tscode_list.append(stock_name)
stk_rewards_item_658_addname_dataframe=pd.DataFrame()
stk_rewards_item_658_addname_dataframe['cname'] = stk_rewards_item_658_tscode_list
for table_name in stk_rewards_item_658.columns.values.tolist():
    stk_rewards_item_658_addname_dataframe[table_name] = stk_rewards_item_658[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_658_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_659_tscode_list = list() 
stk_rewards_item_659 = pro.stk_rewards(ts_code='688090.SH,688096.SH,688098.SH,688099.SH,688100.SH,688101.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_659 返回数据 row 行数 = "+str(stk_rewards_item_659.shape[0]))
for ts_code_sh in stk_rewards_item_659['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_659_tscode_list.append(stock_name)
stk_rewards_item_659_addname_dataframe=pd.DataFrame()
stk_rewards_item_659_addname_dataframe['cname'] = stk_rewards_item_659_tscode_list
for table_name in stk_rewards_item_659.columns.values.tolist():
    stk_rewards_item_659_addname_dataframe[table_name] = stk_rewards_item_659[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_659_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_660_tscode_list = list() 
stk_rewards_item_660 = pro.stk_rewards(ts_code='688106.SH,688108.SH,688111.SH,688116.SH,688118.SH,688122.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_660 返回数据 row 行数 = "+str(stk_rewards_item_660.shape[0]))
for ts_code_sh in stk_rewards_item_660['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_660_tscode_list.append(stock_name)
stk_rewards_item_660_addname_dataframe=pd.DataFrame()
stk_rewards_item_660_addname_dataframe['cname'] = stk_rewards_item_660_tscode_list
for table_name in stk_rewards_item_660.columns.values.tolist():
    stk_rewards_item_660_addname_dataframe[table_name] = stk_rewards_item_660[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_660_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_661_tscode_list = list() 
stk_rewards_item_661 = pro.stk_rewards(ts_code='688123.SH,688126.SH,688128.SH,688138.SH,688139.SH,688155.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_661 返回数据 row 行数 = "+str(stk_rewards_item_661.shape[0]))
for ts_code_sh in stk_rewards_item_661['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_661_tscode_list.append(stock_name)
stk_rewards_item_661_addname_dataframe=pd.DataFrame()
stk_rewards_item_661_addname_dataframe['cname'] = stk_rewards_item_661_tscode_list
for table_name in stk_rewards_item_661.columns.values.tolist():
    stk_rewards_item_661_addname_dataframe[table_name] = stk_rewards_item_661[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_661_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_662_tscode_list = list() 
stk_rewards_item_662 = pro.stk_rewards(ts_code='688157.SH,688158.SH,688159.SH,688165.SH,688166.SH,688168.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_662 返回数据 row 行数 = "+str(stk_rewards_item_662.shape[0]))
for ts_code_sh in stk_rewards_item_662['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_662_tscode_list.append(stock_name)
stk_rewards_item_662_addname_dataframe=pd.DataFrame()
stk_rewards_item_662_addname_dataframe['cname'] = stk_rewards_item_662_tscode_list
for table_name in stk_rewards_item_662.columns.values.tolist():
    stk_rewards_item_662_addname_dataframe[table_name] = stk_rewards_item_662[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_662_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_663_tscode_list = list() 
stk_rewards_item_663 = pro.stk_rewards(ts_code='688169.SH,688177.SH,688178.SH,688180.SH,688181.SH,688185.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_663 返回数据 row 行数 = "+str(stk_rewards_item_663.shape[0]))
for ts_code_sh in stk_rewards_item_663['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_663_tscode_list.append(stock_name)
stk_rewards_item_663_addname_dataframe=pd.DataFrame()
stk_rewards_item_663_addname_dataframe['cname'] = stk_rewards_item_663_tscode_list
for table_name in stk_rewards_item_663.columns.values.tolist():
    stk_rewards_item_663_addname_dataframe[table_name] = stk_rewards_item_663[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_663_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_664_tscode_list = list() 
stk_rewards_item_664 = pro.stk_rewards(ts_code='688186.SH,688188.SH,688189.SH,688196.SH,688198.SH,688199.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_664 返回数据 row 行数 = "+str(stk_rewards_item_664.shape[0]))
for ts_code_sh in stk_rewards_item_664['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_664_tscode_list.append(stock_name)
stk_rewards_item_664_addname_dataframe=pd.DataFrame()
stk_rewards_item_664_addname_dataframe['cname'] = stk_rewards_item_664_tscode_list
for table_name in stk_rewards_item_664.columns.values.tolist():
    stk_rewards_item_664_addname_dataframe[table_name] = stk_rewards_item_664[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_664_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_665_tscode_list = list() 
stk_rewards_item_665 = pro.stk_rewards(ts_code='688200.SH,688202.SH,688208.SH,688218.SH,688222.SH,688228.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_665 返回数据 row 行数 = "+str(stk_rewards_item_665.shape[0]))
for ts_code_sh in stk_rewards_item_665['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_665_tscode_list.append(stock_name)
stk_rewards_item_665_addname_dataframe=pd.DataFrame()
stk_rewards_item_665_addname_dataframe['cname'] = stk_rewards_item_665_tscode_list
for table_name in stk_rewards_item_665.columns.values.tolist():
    stk_rewards_item_665_addname_dataframe[table_name] = stk_rewards_item_665[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_665_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_666_tscode_list = list() 
stk_rewards_item_666 = pro.stk_rewards(ts_code='688229.SH,688233.SH,688256.SH,688258.SH,688266.SH,688268.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_666 返回数据 row 行数 = "+str(stk_rewards_item_666.shape[0]))
for ts_code_sh in stk_rewards_item_666['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_666_tscode_list.append(stock_name)
stk_rewards_item_666_addname_dataframe=pd.DataFrame()
stk_rewards_item_666_addname_dataframe['cname'] = stk_rewards_item_666_tscode_list
for table_name in stk_rewards_item_666.columns.values.tolist():
    stk_rewards_item_666_addname_dataframe[table_name] = stk_rewards_item_666[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_666_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_667_tscode_list = list() 
stk_rewards_item_667 = pro.stk_rewards(ts_code='688277.SH,688278.SH,688286.SH,688288.SH,688298.SH,688299.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_667 返回数据 row 行数 = "+str(stk_rewards_item_667.shape[0]))
for ts_code_sh in stk_rewards_item_667['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_667_tscode_list.append(stock_name)
stk_rewards_item_667_addname_dataframe=pd.DataFrame()
stk_rewards_item_667_addname_dataframe['cname'] = stk_rewards_item_667_tscode_list
for table_name in stk_rewards_item_667.columns.values.tolist():
    stk_rewards_item_667_addname_dataframe[table_name] = stk_rewards_item_667[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_667_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_668_tscode_list = list() 
stk_rewards_item_668 = pro.stk_rewards(ts_code='688300.SH,688309.SH,688310.SH,688311.SH,688312.SH,688313.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_668 返回数据 row 行数 = "+str(stk_rewards_item_668.shape[0]))
for ts_code_sh in stk_rewards_item_668['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_668_tscode_list.append(stock_name)
stk_rewards_item_668_addname_dataframe=pd.DataFrame()
stk_rewards_item_668_addname_dataframe['cname'] = stk_rewards_item_668_tscode_list
for table_name in stk_rewards_item_668.columns.values.tolist():
    stk_rewards_item_668_addname_dataframe[table_name] = stk_rewards_item_668[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_668_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_669_tscode_list = list() 
stk_rewards_item_669 = pro.stk_rewards(ts_code='688318.SH,688321.SH,688333.SH,688335.SH,688336.SH,688338.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_669 返回数据 row 行数 = "+str(stk_rewards_item_669.shape[0]))
for ts_code_sh in stk_rewards_item_669['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_669_tscode_list.append(stock_name)
stk_rewards_item_669_addname_dataframe=pd.DataFrame()
stk_rewards_item_669_addname_dataframe['cname'] = stk_rewards_item_669_tscode_list
for table_name in stk_rewards_item_669.columns.values.tolist():
    stk_rewards_item_669_addname_dataframe[table_name] = stk_rewards_item_669[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_669_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_670_tscode_list = list() 
stk_rewards_item_670 = pro.stk_rewards(ts_code='688339.SH,688357.SH,688358.SH,688360.SH,688363.SH,688365.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_670 返回数据 row 行数 = "+str(stk_rewards_item_670.shape[0]))
for ts_code_sh in stk_rewards_item_670['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_670_tscode_list.append(stock_name)
stk_rewards_item_670_addname_dataframe=pd.DataFrame()
stk_rewards_item_670_addname_dataframe['cname'] = stk_rewards_item_670_tscode_list
for table_name in stk_rewards_item_670.columns.values.tolist():
    stk_rewards_item_670_addname_dataframe[table_name] = stk_rewards_item_670[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_670_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_671_tscode_list = list() 
stk_rewards_item_671 = pro.stk_rewards(ts_code='688366.SH,688368.SH,688369.SH,688377.SH,688388.SH,688389.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_671 返回数据 row 行数 = "+str(stk_rewards_item_671.shape[0]))
for ts_code_sh in stk_rewards_item_671['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_671_tscode_list.append(stock_name)
stk_rewards_item_671_addname_dataframe=pd.DataFrame()
stk_rewards_item_671_addname_dataframe['cname'] = stk_rewards_item_671_tscode_list
for table_name in stk_rewards_item_671.columns.values.tolist():
    stk_rewards_item_671_addname_dataframe[table_name] = stk_rewards_item_671[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_671_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_672_tscode_list = list() 
stk_rewards_item_672 = pro.stk_rewards(ts_code='688396.SH,688398.SH,688399.SH,688418.SH,688466.SH,688488.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_672 返回数据 row 行数 = "+str(stk_rewards_item_672.shape[0]))
for ts_code_sh in stk_rewards_item_672['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_672_tscode_list.append(stock_name)
stk_rewards_item_672_addname_dataframe=pd.DataFrame()
stk_rewards_item_672_addname_dataframe['cname'] = stk_rewards_item_672_tscode_list
for table_name in stk_rewards_item_672.columns.values.tolist():
    stk_rewards_item_672_addname_dataframe[table_name] = stk_rewards_item_672[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_672_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_673_tscode_list = list() 
stk_rewards_item_673 = pro.stk_rewards(ts_code='688500.SH,688505.SH,688508.SH,688516.SH,688518.SH,688520.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_673 返回数据 row 行数 = "+str(stk_rewards_item_673.shape[0]))
for ts_code_sh in stk_rewards_item_673['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_673_tscode_list.append(stock_name)
stk_rewards_item_673_addname_dataframe=pd.DataFrame()
stk_rewards_item_673_addname_dataframe['cname'] = stk_rewards_item_673_tscode_list
for table_name in stk_rewards_item_673.columns.values.tolist():
    stk_rewards_item_673_addname_dataframe[table_name] = stk_rewards_item_673[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_673_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_674_tscode_list = list() 
stk_rewards_item_674 = pro.stk_rewards(ts_code='688528.SH,688555.SH,688556.SH,688558.SH,688561.SH,688566.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_674 返回数据 row 行数 = "+str(stk_rewards_item_674.shape[0]))
for ts_code_sh in stk_rewards_item_674['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_674_tscode_list.append(stock_name)
stk_rewards_item_674_addname_dataframe=pd.DataFrame()
stk_rewards_item_674_addname_dataframe['cname'] = stk_rewards_item_674_tscode_list
for table_name in stk_rewards_item_674.columns.values.tolist():
    stk_rewards_item_674_addname_dataframe[table_name] = stk_rewards_item_674[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_674_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_675_tscode_list = list() 
stk_rewards_item_675 = pro.stk_rewards(ts_code='688567.SH,688568.SH,688579.SH,688580.SH,688586.SH,688588.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_675 返回数据 row 行数 = "+str(stk_rewards_item_675.shape[0]))
for ts_code_sh in stk_rewards_item_675['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_675_tscode_list.append(stock_name)
stk_rewards_item_675_addname_dataframe=pd.DataFrame()
stk_rewards_item_675_addname_dataframe['cname'] = stk_rewards_item_675_tscode_list
for table_name in stk_rewards_item_675.columns.values.tolist():
    stk_rewards_item_675_addname_dataframe[table_name] = stk_rewards_item_675[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_675_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


stk_rewards_item_676_tscode_list = list() 
stk_rewards_item_676 = pro.stk_rewards(ts_code='688589.SH,688598.SH,688599.SH,688600.SH,688981.SH,T00018.SH', fields='ts_code,ann_date,end_date,name,title,reward,hold_vol')
print("stk_rewards_item_676 返回数据 row 行数 = "+str(stk_rewards_item_676.shape[0]))
for ts_code_sh in stk_rewards_item_676['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    stk_rewards_item_676_tscode_list.append(stock_name)
stk_rewards_item_676_addname_dataframe=pd.DataFrame()
stk_rewards_item_676_addname_dataframe['cname'] = stk_rewards_item_676_tscode_list
for table_name in stk_rewards_item_676.columns.values.tolist():
    stk_rewards_item_676_addname_dataframe[table_name] = stk_rewards_item_676[table_name]
stk_rewards_data_List = stk_rewards_data_List.append(stk_rewards_item_676_addname_dataframe)
if stk_rewards_data_List.__len__() > 1160000:
    print("stk_rewards_data_List.__len__() = " + str(stk_rewards_data_List.__len__()))
    stk_rewards_data_List.to_excel(stk_rewards_excel_writer, '管理层薪酬和持股')
    stk_rewards_data_List.save()
    exit();


print("stk_rewards_data_List.__len__() = "+str(stk_rewards_data_List.__len__()))
stk_rewards_data_List.to_excel(stk_rewards_excel_writer,'管理层薪酬和持股,index=False')
stk_rewards_excel_writer.save()
