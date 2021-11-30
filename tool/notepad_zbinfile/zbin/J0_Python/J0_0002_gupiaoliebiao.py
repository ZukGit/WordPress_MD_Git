#  encoding: utf-8

import tushare as ts
import os
import pandas as pd  # https://www.jianshu.com/p/5c0aa1fa19af
import openpyxl 
from openpyxl import load_workbook,Workbook
import time,datetime  # https://www.jb51.net/article/66019.htm
import re
import os
import tempfile

############################## 全局数据初始化块 数据初始化函数  End ##############################

##############  properities 函数 Begin ##############

class Properties:

    def __init__(self, file_name):
        self.file_name = file_name
        self.properties = {}
        try:
            fopen = open(self.file_name, 'r')
            for line in fopen:
                line = line.strip()
                if line.find('=') > 0 and not line.startswith('#'):
                    strs = line.split('=')
                    self.properties[strs[0].strip()] = strs[1].strip()
        except Exception :
            raise
        else:
            fopen.close()

    def has_key(self, key):
        return key in self.properties

    def get(self, key, default_value=''):
        if key in self.properties:
            return self.properties[key]
        return default_value

    def put(self, key, value):
        self.properties[key] = value
        replace_property(self.file_name, key + '=.*', key + '=' + value, True)


def replace_property(file_name, from_regex, to_str, append_on_not_exists=True):
    tmpfile = tempfile.TemporaryFile()

    if os.path.exists(file_name):
        r_open = open(file_name, 'r')
        pattern = re.compile(r'' + from_regex)
        found = None
        for line in r_open:
            if pattern.search(line) and not line.strip().startswith('#'):
                found = True
                line = re.sub(from_regex, to_str, line)
            tmpfile.write(line.encode())
        if not found and append_on_not_exists:
            tmpfile.write(('\n' + to_str).encode())
        r_open.close()
        tmpfile.seek(0)

        content = tmpfile.read()

        if os.path.exists(file_name):
            os.remove(file_name)

        w_open = open(file_name, 'wb')
        w_open.write(content)
        w_open.close()

        tmpfile.close()
    else:
        print ("file %s not found" % file_name)

##############  properities 函数 End ##############


##############  XLSX excel 函数 Begin ##############

def createexcel(filename):    ### 创建本地文件名称为 filename的文件
    wb = Workbook()
    curPath = os.path.abspath('.')
    cur_desktop_path=os.path.join(os.path.expanduser("~"), 'Desktop')
    curDir = cur_desktop_path+os.sep+"zbin"+os.sep+"J0_Data"+os.sep  ## 创建~/Desktop/zbin/J0_Data 这个 这个工作目录
    #curDir = curPath+os.sep
    if not os.path.exists(curDir):
        os.makedirs(curDir)
    #curPath = os.path.dirname(os.path.abspath('.'))
    #t = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
    #suffix = '.xlsx' # 文件类型
    #newfile = t + suffix
    xlsxPath = curDir + filename
    print("curPath curDir=" + curPath)
    print("createexcel zbin_JO_Dir=" + curDir)
    print("createexcel path=" + xlsxPath)
    if not os.path.exists(xlsxPath):
        wb.save(xlsxPath)
        print(" wb.save(xlsxPath)  xlsxPath =" + xlsxPath)
        time.sleep(1)
    return xlsxPath

def getColumnIndex(table, columnName):   ## 返回 table 中 名称为  columnName 的 那列 的索引
    columnIndex = None
    for i in range(table.ncols):
        if(table.cell_value(0, i) == columnName):
            columnIndex = i
            break
    return columnIndex
#封装函数    https://blog.csdn.net/weixin_41267342/article/details/86634007

##############  XLSX excel 函数 End ##############



############################## Prop初始化 Begin ##############################

#
#Thu Aug 13 22:33:45 CST 2020
#rixianhangqing-time_record_date=20200707
#rixianhangqing-time_start_date=20010101

desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')
zbin_path = str(desktop_path)+os.sep+"zbin"
j0_properties_path= str(zbin_path)+os.sep+"J0.properties"
print("desktop_path = "+ str(desktop_path))
print("zbin_path = "+ str(zbin_path))
print("j0_properties_path = "+ str(j0_properties_path))
J0_PROPS =  Properties(j0_properties_path)


############################## Prop初始化 End ##############################


############## tscode_股票列表的初始化  Begin ##############
#封装函数    https://blog.csdn.net/weixin_41267342/article/details/86634007
def init_tscode_data(book_name, sheet_name,ts_code_set,tscode_name_dict):
    # 读取excel
    wb = openpyxl.load_workbook(book_name)
    # 读取sheet表
    ws = wb[sheet_name]
    for i in range(ws.max_row):
         # 获取下拉框中的所有选择值
         if (i == 0 or i == 1):
             continue
         #print("i="+str(i)+" 总的列数:" + str(ws.max_row)+"  value:"+str(ws.cell(i,2).value))
         tscode_item = str(ws.cell(i,1).value)  ##  20201010--> 2:ts_code    4_name   ##  20201116--> 1:ts_code    3_name 
         tscode_name_item = str(ws.cell(i,3).value)
         #print("index="+str(i)+" tscode_item = "+ str(tscode_item) + "   tscode_name_item="+str(tscode_name_item))
         ts_code_set.add(str(ws.cell(i,1).value))
         tscode_name_dict[tscode_item]=tscode_name_item



tscode_set = set()    #### 股票代码的集合   000001.SZ   .... 999999.SH
tscode_name_dict = dict()  #### code-name 的 map的集合
init_tscode_data(zbin_path+os.sep+"J0_Python"+os.sep+"J0_股票列表.xlsx","股票列表",tscode_set,tscode_name_dict)
############## tscode_股票列表的初始化  End  ##############


############################## 运行属性 Begin ##############################
pd.set_option('display.max_rows', None)   ##  解决纵向出现...
#pd.set_option('display.width', 1000) 
pd.set_option('expand_frame_repr', False)  ##  解决横向出现...
Cur_Abs_Path=os.path.abspath('.')   # 表示当前所处的文件夹的绝对路径
print("当前绝对路径:"+Cur_Abs_Path)
Cur_Ref_Path=os.path.abspath('..')  # 表示当前所处的文件夹上一级文件夹的绝对路径
print("当前父目录绝对路径:"+Cur_Ref_Path)
############################## 运行属性 End ##############################

############################## 时间Date Begin ##############################

now_yyyymmdd=str(time.strftime('%Y%m%d', time.localtime(time.time())))
print("now_yyyymmdd = "+ str(now_yyyymmdd))

############################## 时间Date End ##############################





pro = ts.pro_api('43acb9a5ddc2cf73c6c4ea54796748f965457ed57daaa736bb778ea2')
# print(J0_PROPS.get(tree_node_name+"record_date"))           #根据key读取value
# J0_PROPS.put(tree_node_name+"record_date", now_yyyymmdd)       ###  覆盖原有的 key 和 value
#  zukgit
# gupiaoliebiao_zukgit_website  =   https://tushare.pro/document/2?doc_id=25
createexcel('股票列表.xlsx')
stock_basic_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\股票列表.xlsx')
stock_basic_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\股票列表.xlsx', engine='openpyxl')
stock_basic_excel_writer.book = stock_basic_book
stock_basic_excel_writer.sheets = dict((ws.title, ws) for ws in stock_basic_book.worksheets)
stock_basic_data_List = pd.DataFrame()
stock_basic_item_0 = pro.stock_basic(is_hs='N',list_status='L',exchange='SSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_0 返回数据 row 行数 = "+str(stock_basic_item_0.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_0)


stock_basic_item_1 = pro.stock_basic(is_hs='H',list_status='L',exchange='SSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_1 返回数据 row 行数 = "+str(stock_basic_item_1.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_1)


stock_basic_item_2 = pro.stock_basic(is_hs='S',list_status='L',exchange='SSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_2 返回数据 row 行数 = "+str(stock_basic_item_2.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_2)


stock_basic_item_3 = pro.stock_basic(is_hs='N',list_status='D',exchange='SSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_3 返回数据 row 行数 = "+str(stock_basic_item_3.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_3)


stock_basic_item_4 = pro.stock_basic(is_hs='H',list_status='D',exchange='SSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_4 返回数据 row 行数 = "+str(stock_basic_item_4.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_4)


stock_basic_item_5 = pro.stock_basic(is_hs='S',list_status='D',exchange='SSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_5 返回数据 row 行数 = "+str(stock_basic_item_5.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_5)


stock_basic_item_6 = pro.stock_basic(is_hs='N',list_status='P',exchange='SSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_6 返回数据 row 行数 = "+str(stock_basic_item_6.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_6)


stock_basic_item_7 = pro.stock_basic(is_hs='H',list_status='P',exchange='SSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_7 返回数据 row 行数 = "+str(stock_basic_item_7.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_7)


stock_basic_item_8 = pro.stock_basic(is_hs='S',list_status='P',exchange='SSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_8 返回数据 row 行数 = "+str(stock_basic_item_8.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_8)


stock_basic_item_9 = pro.stock_basic(is_hs='N',list_status='L',exchange='SZSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_9 返回数据 row 行数 = "+str(stock_basic_item_9.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_9)


stock_basic_item_10 = pro.stock_basic(is_hs='H',list_status='L',exchange='SZSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_10 返回数据 row 行数 = "+str(stock_basic_item_10.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_10)


stock_basic_item_11 = pro.stock_basic(is_hs='S',list_status='L',exchange='SZSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_11 返回数据 row 行数 = "+str(stock_basic_item_11.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_11)


stock_basic_item_12 = pro.stock_basic(is_hs='N',list_status='D',exchange='SZSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_12 返回数据 row 行数 = "+str(stock_basic_item_12.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_12)


stock_basic_item_13 = pro.stock_basic(is_hs='H',list_status='D',exchange='SZSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_13 返回数据 row 行数 = "+str(stock_basic_item_13.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_13)


stock_basic_item_14 = pro.stock_basic(is_hs='S',list_status='D',exchange='SZSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_14 返回数据 row 行数 = "+str(stock_basic_item_14.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_14)


stock_basic_item_15 = pro.stock_basic(is_hs='N',list_status='P',exchange='SZSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_15 返回数据 row 行数 = "+str(stock_basic_item_15.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_15)


stock_basic_item_16 = pro.stock_basic(is_hs='H',list_status='P',exchange='SZSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_16 返回数据 row 行数 = "+str(stock_basic_item_16.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_16)


stock_basic_item_17 = pro.stock_basic(is_hs='S',list_status='P',exchange='SZSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_17 返回数据 row 行数 = "+str(stock_basic_item_17.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_17)


stock_basic_item_18 = pro.stock_basic(is_hs='N',list_status='L',exchange='HKEX', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_18 返回数据 row 行数 = "+str(stock_basic_item_18.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_18)


stock_basic_item_19 = pro.stock_basic(is_hs='H',list_status='L',exchange='HKEX', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_19 返回数据 row 行数 = "+str(stock_basic_item_19.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_19)


stock_basic_item_20 = pro.stock_basic(is_hs='S',list_status='L',exchange='HKEX', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_20 返回数据 row 行数 = "+str(stock_basic_item_20.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_20)


stock_basic_item_21 = pro.stock_basic(is_hs='N',list_status='D',exchange='HKEX', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_21 返回数据 row 行数 = "+str(stock_basic_item_21.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_21)


stock_basic_item_22 = pro.stock_basic(is_hs='H',list_status='D',exchange='HKEX', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_22 返回数据 row 行数 = "+str(stock_basic_item_22.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_22)


stock_basic_item_23 = pro.stock_basic(is_hs='S',list_status='D',exchange='HKEX', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_23 返回数据 row 行数 = "+str(stock_basic_item_23.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_23)


stock_basic_item_24 = pro.stock_basic(is_hs='N',list_status='P',exchange='HKEX', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_24 返回数据 row 行数 = "+str(stock_basic_item_24.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_24)


stock_basic_item_25 = pro.stock_basic(is_hs='H',list_status='P',exchange='HKEX', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_25 返回数据 row 行数 = "+str(stock_basic_item_25.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_25)


stock_basic_item_26 = pro.stock_basic(is_hs='S',list_status='P',exchange='HKEX', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_26 返回数据 row 行数 = "+str(stock_basic_item_26.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_26)


stock_basic_item_27 = pro.stock_basic(is_hs='N',list_status='L',exchange='BSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_27 返回数据 row 行数 = "+str(stock_basic_item_27.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_27)


stock_basic_item_28 = pro.stock_basic(is_hs='H',list_status='L',exchange='BSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_28 返回数据 row 行数 = "+str(stock_basic_item_28.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_28)


stock_basic_item_29 = pro.stock_basic(is_hs='S',list_status='L',exchange='BSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_29 返回数据 row 行数 = "+str(stock_basic_item_29.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_29)


stock_basic_item_30 = pro.stock_basic(is_hs='N',list_status='D',exchange='BSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_30 返回数据 row 行数 = "+str(stock_basic_item_30.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_30)


stock_basic_item_31 = pro.stock_basic(is_hs='H',list_status='D',exchange='BSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_31 返回数据 row 行数 = "+str(stock_basic_item_31.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_31)


stock_basic_item_32 = pro.stock_basic(is_hs='S',list_status='D',exchange='BSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_32 返回数据 row 行数 = "+str(stock_basic_item_32.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_32)


stock_basic_item_33 = pro.stock_basic(is_hs='N',list_status='P',exchange='BSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_33 返回数据 row 行数 = "+str(stock_basic_item_33.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_33)


stock_basic_item_34 = pro.stock_basic(is_hs='H',list_status='P',exchange='BSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_34 返回数据 row 行数 = "+str(stock_basic_item_34.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_34)


stock_basic_item_35 = pro.stock_basic(is_hs='S',list_status='P',exchange='BSE', fields='ts_code,symbol,name,area,industry,fullname,enname,market,exchange,curr_type,list_status,list_date,delist_date,is_hs')
print(" 股票列表 stock_basic_item_35 返回数据 row 行数 = "+str(stock_basic_item_35.shape[0]))
stock_basic_data_List = stock_basic_data_List.append(stock_basic_item_35)



print("stock_basic_data_List.__len__() = "+str(stock_basic_data_List.__len__()))
stock_basic_data_List.to_excel(stock_basic_excel_writer,'股票列表',index=False)
stock_basic_excel_writer.save()
